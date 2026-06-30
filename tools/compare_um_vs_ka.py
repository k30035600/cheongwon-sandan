"""05_내역서 미매칭 vs _ka.xlsx 교차 검수."""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from naeyeok_gongjong import resolve_work  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

WORK = Path("05_내역서/내역서작업")
KA = Path("08_제출내역서/청원지구_단가통합(전기제외)_ka.xlsx")

STD = [
    ("01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx", "토목(조경)"),
    ("01 조경", "01_화성 청원지구 조경_표준단가산출.xlsx", "토목(조경)"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx", "진입도로"),
    ("05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx", "회전교차로"),
    ("06 개발행위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx", "지구단위"),
]


def norm(s):
    return " ".join(str(s or "").split())


def load_std_um(path: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["통합내역"]
    hdr = [ws.cell(1, c).value for c in range(1, 20)]
    si = next(i for i, h in enumerate(hdr) if h and "상태" in str(h)) + 1
    ni = next(i for i, h in enumerate(hdr) if h and "공종명" in str(h)) + 1
    ui = next(i for i, h in enumerate(hdr) if h and str(h) == "단위") + 1
    out = []
    for r in range(2, ws.max_row + 1):
        st = str(ws.cell(r, si).value or "")
        if st not in ("미매칭", "미산출"):
            continue
        name = ws.cell(r, ni).value
        if not name or str(name).startswith("★"):
            continue
        out.append((norm(name), norm(ws.cell(r, ui).value), st))
    wb.close()
    return out


def load_ka_sheet(sheet: str) -> dict[tuple[str, str], dict]:
    wb = load_workbook(KA, data_only=True)
    ws = wb[sheet]
    hdr = next(r for r in range(1, 30) if ws.cell(r, 4).value == "공종명")
    idx = {}
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 4).value
        if not name or str(name).startswith("★") or not ws.cell(r, 3).value:
            continue
        key = (norm(name), norm(ws.cell(r, 7).value))
        idx[key] = {
            "status": norm(ws.cell(r, 16).value),
            "sum": float(ws.cell(r, 15).value or 0),
        }
    wb.close()
    return idx


print("=== 표준단가산출 미매칭 vs _ka.xlsx ===\n")
if not KA.exists():
    print("파일 없음:", KA)
    raise SystemExit(1)

ka_cache: dict[str, dict] = {}
tot = dict(in_=0, miss=0, amt=0, st_diff=0, std=0)
samples_diff = []

for label, fname, sh in STD:
    um = load_std_um(resolve_work(fname))
    if sh not in ka_cache:
        ka_cache[sh] = load_ka_sheet(sh)
    ka = ka_cache[sh]
    loc = dict(in_=0, miss=0, amt=0, st_diff=0)
    for name, unit, st in um:
        tot["std"] += 1
        h = ka.get((name, unit))
        if not h:
            loc["miss"] += 1
            tot["miss"] += 1
            continue
        loc["in_"] += 1
        tot["in_"] += 1
        if h["sum"] > 0:
            loc["amt"] += 1
            tot["amt"] += 1
        if h["status"] not in ("미매칭", "미산출"):
            loc["st_diff"] += 1
            tot["st_diff"] += 1
            if len(samples_diff) < 6:
                samples_diff.append((label, name[:24], st, h["status"], h["sum"]))
    print(f"{label}: std미매칭 {len(um)} → _ka 포함 {loc['in_']} · 미포함 {loc['miss']} · 금액>0 {loc['amt']} · 상태≠미매칭 {loc['st_diff']}")

print(f"\n【합계】 std미매칭 {tot['std']}건 → _ka 포함 {tot['in_']} · 미포함 {tot['miss']} · 금액>0 {tot['amt']} · 상태변경 {tot['st_diff']}")

wb = load_workbook(KA, data_only=True)
ka_um = ka_rv = ka_ok = 0
for sh in ("회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물"):
    ws = wb[sh]
    hdr = next(r for r in range(1, 30) if ws.cell(r, 4).value == "공종명")
    for r in range(hdr + 1, ws.max_row + 1):
        if not ws.cell(r, 3).value:
            continue
        s = str(ws.cell(r, 16).value or "")
        if s in ("미매칭", "미산출"):
            ka_um += 1
        elif s == "검토":
            ka_rv += 1
        elif s:
            ka_ok += 1
wb.close()
print(f"\n【_ka 자체】 미매칭 {ka_um} · 검토 {ka_rv} · 매칭등 {ka_ok}")

if samples_diff:
    print("\n--- std=미매칭 → _ka 상태변경 (샘플) ---")
    for row in samples_diff:
        print(f"  [{row[0]}] {row[1]} | std={row[2]} → ka={row[3]} sum={row[4]:,.0f}")

print("\n※ 이전 대화: _ka 재·노·경 검산·가영현 상태만 수행. 미매칭↔_ka 목록 대조는 본 검수가 최초.")
