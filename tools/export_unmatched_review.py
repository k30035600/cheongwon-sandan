#!/usr/bin/env python3
"""각 _표준단가산출.xlsx에서 미매칭(미산출)·검토필요를 추출해 별도 통합 엑셀 2종 생성."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
# 검토_전체·미매칭_전체 출력 위치 — 빌더(build_review_*)가 읽는 05_내역서 루트로 통일
OUT_DIR = BASE

LABELS = {
    "01": "01 토목", "01j": "01 조경", "02": "02 전기(지구내·통합)", "03": "03 전기(지구외·중복)",
    "04": "04 진입도로", "05": "05 회전교차로", "06": "06 개발행위", "07": "07 건설폐기물",
}
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
DUP_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)

def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hdr = [("" if h is None else str(h).strip()) for h in rows[0]]
    recs = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        recs.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r)))})
    return hdr, recs

def g(rec, *keys):
    for k in keys:
        if k in rec and rec[k] not in (None, ""):
            return rec[k]
    return ""

UNMATCHED_SHEETS = ("미매칭", "미산출")
REVIEW_SHEETS = ("검토필요",)

def collect(kind_sheets):
    out = []  # (no, label, hdr, recs)
    for p in sorted(WORK.glob("*_표준단가산출.xlsx")):
        no = p.name[:2]
        if no == "01" and "조경" in p.name:  # 토목/조경 분리 후 동일 접두 충돌 방지
            no = "01j"
        label = LABELS.get(no, no)
        wb = load_workbook(p, read_only=True, data_only=True)
        for sname in kind_sheets:
            if sname in wb.sheetnames:
                hdr, recs = sheet_records(wb[sname])
                if recs:
                    out.append((no, label, sname, recs))
        wb.close()
    return out

def write_workbook(path, title, cols, blocks, mapper):
    wb = Workbook()
    ws = wb.active
    ws.title = "통합"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"생성: 각 내역서 _표준단가산출.xlsx 추출 / 총 {sum(len(b[3]) for b in blocks):,}건"])
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = HEADER_FILL
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")
    total = 0
    for no, label, sname, recs in blocks:
        for rec in recs:
            ws.append([label] + mapper(rec))
            if no == "03":
                for c in range(1, len(cols) + 1):
                    ws.cell(row=ws.max_row, column=c).fill = DUP_FILL
            total += 1
    ws.freeze_panes = "A4"
    widths = [18, 7, 16, 30, 22, 6, 9, 9, 12, 14, 22, 22, 12, 14]
    for i, w in enumerate(widths[:len(cols)], 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    # 내역서별 원본 시트
    for no, label, sname, recs in blocks:
        title_sheet = f"{no}_{sname}"[:31]
        s = wb.create_sheet(title_sheet)
        if recs:
            keys = list(recs[0].keys())
            s.append(keys)
            for c in range(1, len(keys) + 1):
                s.cell(row=1, column=c).fill = HEADER_FILL
                s.cell(row=1, column=c).font = BOLD
            for rec in recs:
                s.append([rec.get(k) for k in keys])
            s.freeze_panes = "A2"
    wb.save(path)
    return total

# 미매칭
um = collect(UNMATCHED_SHEETS)
um_cols = ["내역서", "행", "공종", "명칭/공종명", "규격", "단위", "수량", "점수", "검색어/비고"]
def um_map(rec):
    return [
        g(rec, "행"), g(rec, "공종"), g(rec, "명칭", "공종명"),
        g(rec, "규격"), g(rec, "단위"), g(rec, "수량"),
        g(rec, "최고점수", "매칭점수"), g(rec, "검색어", "비고"),
    ]
um_path = OUT_DIR / "미매칭_전체.xlsx"
um_n = write_workbook(um_path, "미매칭·미산출 통합 (01~07)", um_cols, um, um_map)

# 검토
rv = collect(REVIEW_SHEETS)
rv_cols = ["내역서", "행", "공종", "명칭/공종명", "규격", "단위", "수량",
           "매칭점수", "신뢰도", "단가코드", "매칭품명", "매칭규격", "합계단가", "합계금액"]
def rv_map(rec):
    return [
        g(rec, "행"), g(rec, "공종"), g(rec, "명칭", "공종명"),
        g(rec, "규격"), g(rec, "단위"), g(rec, "수량"),
        g(rec, "매칭점수"), g(rec, "신뢰도"), g(rec, "단가코드"),
        g(rec, "매칭품명"), g(rec, "매칭규격"), g(rec, "합계단가"), g(rec, "합계금액"),
    ]
rv_path = OUT_DIR / "검토_전체.xlsx"
rv_n = write_workbook(rv_path, "검토필요 통합 (01~07)", rv_cols, rv, rv_map)

print(f"미매칭_전체.xlsx — {um_n}건  → {um_path}")
for no, label, sname, recs in um:
    print(f"   {label} [{sname}] {len(recs)}건")
print(f"검토_전체.xlsx — {rv_n}건  → {rv_path}")
for no, label, sname, recs in rv:
    print(f"   {label} [{sname}] {len(recs)}건")
