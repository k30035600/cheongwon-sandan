#!/usr/bin/env python3
"""일위대가 산출표 확정단가 → 각 _표준단가산출.xlsx · 총괄표 반영 (인-place 패치).

입력:
  - 05_내역서/미매칭_일위대가산출.xlsx (확정단가 또는 H·B·E·J 경로 제시단가)
  - 05_내역서/검토_일위대가산출.xlsx (확정단가(입력)만)
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
sys.path.insert(0, str(Path(__file__).resolve().parent))
import apply_standard_prices as asp  # noqa: E402

UNMATCHED_IL = BASE / "미매칭_일위대가산출.xlsx"
REVIEW_IL = BASE / "검토_일위대가산출.xlsx"
REVIEW_TOGONG_IL = BASE / "검토_토공_일위대가산출.xlsx"
REVIEW_SECTIONS_IL = BASE / "검토_공종별_일위대가산출.xlsx"

FILE_MAP = {
    "01 토목": "01_화성 청원지구 토목_표준단가산출.xlsx",
    "01 조경": "01_화성 청원지구 조경_표준단가산출.xlsx",
    "04 진입도로": "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
    "05 회전교차로": "05_화성 청원로(회전교차로)_표준단가산출.xlsx",
    "06 개발행위": "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
}

AUTO_ROUTES = ("H.근접단가+환산", "B.품셈 시공+시장단가", "E.조경일위/시세", "J.품셈 PE천막")
MONEY = asp.MONEY_FMT
REVIEW_FILL = asp.REVIEW_FILL
UNMATCH_FILL = asp.UNMATCH_FILL


def _norm_name(s) -> str:
    """머리표(가). 1). 등)·공백·기호 제거 후 비교용 정규화."""
    import re
    s = str(s or "")
    s = re.sub(r"^[\s0-9A-Za-z가-힣]{0,4}[).]\s*", "", s)  # 머리표 제거
    s = re.sub(r"[\s.·,()/]+", "", s)
    return s


def _name_matches(a, b) -> bool:
    na, nb = _norm_name(a), _norm_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    # 한쪽이 다른 쪽의 앞부분을 포함(규격 차이 등)하면 동일 품목으로 인정
    short, long = sorted((na, nb), key=len)
    return len(short) >= 3 and (long.startswith(short) or short in long)


def _content_key(name, spec, unit) -> tuple[str, str, str]:
    """파일 내 동일 품목 식별용 내용 키 — 품명·규격·단위 정규화."""
    return (_norm_name(name), _norm_name(spec), str(unit or "").strip())


def _num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def load_ilwidae_overrides(path: Path, *, auto_suggest: bool) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["일위대가산출"]
    col = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    out: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        file_label = ws.cell(row_idx, col.get("파일", 2)).value
        if not file_label or str(file_label).strip() not in FILE_MAP:
            continue
        name = ws.cell(row_idx, col.get("품명", 5)).value
        if not name:
            continue
        route = str(ws.cell(row_idx, col.get("경로", 1)).value or "")
        conf = _num(ws.cell(row_idx, col["확정단가(입력)"]).value) if col.get("확정단가(입력)") else None
        sug = _num(ws.cell(row_idx, col["제시단가"]).value) if col.get("제시단가") else None
        unit_total = conf
        if not unit_total and auto_suggest and any(route.startswith(p) for p in AUTO_ROUTES):
            unit_total = sug
        if not unit_total:
            continue

        def cell(key, alt=None):
            k = col.get(key) or (col.get(alt) if alt else None)
            return ws.cell(row_idx, k).value if k else None

        mat_u = _num(cell("재료단가"))
        lab_u = _num(cell("노무단가", "노무비"))
        exp_u = _num(cell("경비단가", "경비"))
        mat_u, lab_u, exp_u = _normalize_units(mat_u, lab_u, exp_u, unit_total)
        db_name = cell("DB_품명", "현재_매칭품명") or name
        db_spec = cell("DB_규격", "현재_매칭규격") or ""
        basis = cell("표준품셈·산출근거") or cell("환산·근거 / 대안후보") or f"일위대가 ({route})"
        out.append({
            "file": str(file_label).strip(),
            "row": int(cell("행")),
            "item_name": str(name).strip(),     # 검토측 품명 — 통합내역 품명과 대조용
            "item_spec": str(cell("규격") or "").strip(),  # 검토측 규격
            "item_unit": str(cell("단위") or "").strip(),  # 검토측 단위
            "total_unit": unit_total,
            "mat_unit": mat_u,
            "lab_unit": lab_u,
            "exp_unit": exp_u,
            "price_name": str(db_name).strip(),
            "price_spec": str(db_spec or "일위대가 확정").strip(),
            "basis": str(basis)[:200],
            "source": "확정" if conf else "제시",
        })
    wb.close()
    return out


def _normalize_units(mat_u, lab_u, exp_u, total_u: float) -> tuple[float, float, float]:
    mat_u = mat_u or 0.0
    lab_u = lab_u or 0.0
    exp_u = exp_u or 0.0
    s = mat_u + lab_u + exp_u
    if s <= 0:
        return 0.0, 0.0, total_u
    if abs(s - total_u) > 1:
        ratio = total_u / s
        return round(mat_u * ratio), round(lab_u * ratio), round(exp_u * ratio)
    return mat_u, lab_u, exp_u


def read_integrated_from_ws(ws) -> list[dict]:
    col = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    rows: list[dict] = []
    for row_idx in range(2, ws.max_row + 1):
        row_no = ws.cell(row_idx, col["행"]).value
        if row_no is None:
            continue
        qty = ws.cell(row_idx, col["수량"]).value
        try:
            qty_f = float(qty or 0)
        except (TypeError, ValueError):
            qty_f = 0.0
        rows.append({
            "row": int(row_no),
            "section": ws.cell(row_idx, col["공종"]).value,
            "name": ws.cell(row_idx, col["공종명"]).value,
            "spec": ws.cell(row_idx, col["규격"]).value or "",
            "qty": qty_f,
            "unit": ws.cell(row_idx, col["단위"]).value or "",
            "status": ws.cell(row_idx, col["상태"]).value or "미매칭",
            "match_score": ws.cell(row_idx, col["매칭점수"]).value,
            "price_code": ws.cell(row_idx, col["단가코드"]).value or "",
            "price_name": ws.cell(row_idx, col["매칭품명"]).value or "",
            "price_spec": ws.cell(row_idx, col["매칭규격"]).value or "",
            "mat_unit": _num(ws.cell(row_idx, col["재료단가"]).value),
            "lab_unit": _num(ws.cell(row_idx, col["노무단가"]).value),
            "exp_unit": _num(ws.cell(row_idx, col["경비단가"]).value),
            "total_unit": _num(ws.cell(row_idx, col["합계단가"]).value),
            "mat_amt": _num(ws.cell(row_idx, col["재료금액"]).value),
            "lab_amt": _num(ws.cell(row_idx, col["노무금액"]).value),
            "exp_amt": _num(ws.cell(row_idx, col["경비금액"]).value),
            "sum_amt": _num(ws.cell(row_idx, col["합계금액"]).value),
            "terms": ws.cell(row_idx, col["비고"]).value or "",
            "confidence": "",
            "_row_idx": row_idx,
        })
    return rows


def apply_override(row: dict, ov: dict) -> dict:
    qty = float(row.get("qty") or 0)
    mat_u, lab_u, exp_u = ov["mat_unit"], ov["lab_unit"], ov["exp_unit"]
    tot_u = ov["total_unit"]
    mat_amt = round(qty * mat_u)
    lab_amt = round(qty * lab_u)
    exp_amt = round(qty * exp_u)
    sum_amt = round(qty * tot_u)
    if mat_u + lab_u + exp_u == 0 and sum_amt:
        exp_amt = sum_amt
    return {
        **row,
        "status": "매칭",
        "match_score": 1.0,
        "confidence": "일위확정",
        "price_code": "일위확정",
        "price_name": ov["price_name"],
        "price_spec": ov["price_spec"],
        "mat_unit": mat_u,
        "lab_unit": lab_u,
        "exp_unit": exp_u,
        "total_unit": tot_u,
        "mat_amt": mat_amt,
        "lab_amt": lab_amt,
        "exp_amt": exp_amt,
        "sum_amt": sum_amt,
        "terms": ov["basis"],
    }


def write_integrated_row(ws, row_idx: int, r: dict) -> None:
    vals = [
        r["row"], r["section"], r["name"], r["spec"], r["qty"], r["unit"],
        r["status"], r["match_score"], r["price_code"], r["price_name"], r["price_spec"],
        r.get("mat_unit"), r.get("lab_unit"), r.get("exp_unit"), r.get("total_unit"),
        r.get("mat_amt"), r.get("lab_amt"), r.get("exp_amt"), r.get("sum_amt"),
        r.get("terms", ""),
    ]
    fill = REVIEW_FILL if r["status"] == "검토" else UNMATCH_FILL if r["status"] == "미매칭" else None
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row_idx, c, v)
        if isinstance(v, (int, float)) and c >= 12:
            cell.number_format = MONEY
        if fill:
            cell.fill = fill


def recompute_totals(integrated: list[dict]) -> tuple[dict, dict]:
    totals = {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0}
    section_totals: dict[str, dict] = {}
    for r in integrated:
        sec = r.get("section") or "(미분류)"
        section_totals.setdefault(
            sec, {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0, "matched": 0, "items": 0},
        )
        section_totals[sec]["items"] += 1
        if r.get("status") in ("매칭", "검토"):
            for k, ak in [("mat", "mat_amt"), ("lab", "lab_amt"), ("exp", "exp_amt"), ("sum", "sum_amt")]:
                v = float(r.get(ak) or 0)
                totals[k] += v
                section_totals[sec][k] += v
            section_totals[sec]["matched"] += 1
    return totals, section_totals


def patch_summary_sheet(ws, section_totals: dict, totals: dict, n_unmatched: int) -> None:
    total_items = sum(v["items"] for v in section_totals.values())
    total_matched = sum(v["matched"] for v in section_totals.values())
    sec_rows = {str(ws.cell(r, 1).value): r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
    for sec, st in section_totals.items():
        if sec not in sec_rows:
            continue
        r = sec_rows[sec]
        ws.cell(r, 2, st["matched"])
        ws.cell(r, 3, st["items"])
        ws.cell(r, 4, st["matched"] / st["items"] if st["items"] else 0)
        for c, k in enumerate(["mat", "lab", "exp", "sum"], 5):
            ws.cell(r, c, st[k]).number_format = MONEY
    for r in range(2, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "")
        if label == "★ 매칭합계":
            ws.cell(r, 2, total_matched)
            ws.cell(r, 3, total_items)
            ws.cell(r, 4, total_matched / total_items if total_items else 0)
            ws.cell(r, 5, totals["mat"]).number_format = MONEY
            ws.cell(r, 6, totals["lab"]).number_format = MONEY
            ws.cell(r, 7, totals["exp"]).number_format = MONEY
            ws.cell(r, 8, totals["sum"]).number_format = MONEY
        elif label == "※ 미매칭":
            ws.cell(r, 2, n_unmatched)
            ws.cell(r, 3, total_items)
            ws.cell(r, 4, n_unmatched / total_items if total_items else 0)


def rewrite_detail_sheet(wb, title: str, rows: list[dict], headers: list[str], fill: PatternFill | None = None) -> None:
    if title not in wb.sheetnames:
        return
    ws = wb[title]
    ws.delete_rows(2, ws.max_row - 1)
    for r in rows:
        ws.append([
            r["row"], r["section"], r["name"], r["spec"], r["qty"], r["unit"],
            r.get("match_score"), r.get("confidence", r.get("status", "")),
            r.get("price_code", ""), r.get("price_name", ""), r.get("price_spec", ""),
            r.get("mat_unit"), r.get("lab_unit"), r.get("exp_unit"), r.get("total_unit"),
            r.get("mat_amt"), r.get("lab_amt"), r.get("exp_amt"), r.get("sum_amt"),
        ])
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = fill
        for c in range(12, 20):
            if isinstance(ws.cell(ws.max_row, c).value, (int, float)):
                ws.cell(ws.max_row, c).number_format = MONEY


def rewrite_unmatched_sheet(wb, unmatched: list[dict]) -> None:
    ws = wb["미매칭"]
    ws.delete_rows(2, ws.max_row - 1)
    for r in unmatched:
        score = r.get("match_score")
        if score is not None and score != "":
            try:
                score = round(float(score), 3)
            except (TypeError, ValueError):
                score = None
        ws.append([
            r["row"], r["section"], r["name"], r["spec"], r["qty"], r["unit"],
            score, r.get("terms", ""),
        ])
        for c in range(1, 9):
            ws.cell(ws.max_row, c).fill = UNMATCH_FILL


def apply_file(xlsx_name: str, overrides: list[dict]) -> int:
    path = WORK / xlsx_name
    if not path.exists():
        print(f"  skip: {xlsx_name}", flush=True)
        return 0
    # 내용(품명+규격+단위) 키 매핑 — 분리 전/후 행 불일치 무관하게 정확 매칭.
    # 동일 키 중복 시 뒤(=tg·sc, 품셈 재산출)가 우선(앞의 um·rv를 덮어씀).
    ov_by_key: dict[tuple, dict] = {}
    for o in overrides:
        ov_by_key[_content_key(o.get("item_name", ""), o.get("item_spec", ""), o.get("item_unit", ""))] = o
    ov_by_row = {o["row"]: o for o in overrides}  # 행 fallback(이름 가드 적용)
    wb = load_workbook(path)
    ws = wb["통합내역"]
    integrated = read_integrated_from_ws(ws)
    applied = 0
    used_keys: set[tuple] = set()
    used_via_row = 0
    for i, row in enumerate(integrated):
        key = _content_key(row.get("name", ""), row.get("spec", ""), row.get("unit", ""))
        ov = ov_by_key.get(key)
        if ov is not None:
            used_keys.add(key)
        else:
            # fallback: 행으로 찾되 품명이 일치할 때만(오패치 방지)
            cand = ov_by_row.get(row["row"])
            if cand and _name_matches(cand.get("item_name", ""), row.get("name", "")):
                ov = cand
                used_via_row += 1
            else:
                continue
        integrated[i] = apply_override(row, ov)
        write_integrated_row(ws, row["_row_idx"], integrated[i])
        applied += 1
    # 매칭되지 못한 확정 항목 보고(행 불일치·품명 변형 등)
    miss = [o for o in overrides
            if _content_key(o.get("item_name", ""), o.get("item_spec", ""), o.get("item_unit", "")) not in used_keys]
    if miss:
        print(f"    내용키 미매칭 확정 {len(miss)}건(행 fallback {used_via_row}건 별도) — 예시:", flush=True)
        for o in miss[:8]:
            print(f"      · 행{o['row']} 「{str(o.get('item_name',''))[:20]}/{str(o.get('item_spec',''))[:16]}」", flush=True)
    if not applied:
        wb.close()
        return 0

    totals, section_totals = recompute_totals(integrated)
    matched = [r for r in integrated if r["status"] != "미매칭"]
    unmatched = [r for r in integrated if r["status"] == "미매칭"]
    review = [r for r in integrated if r["status"] == "검토"]
    match_only = [r for r in matched if r["status"] == "매칭"]

    patch_summary_sheet(wb["합계요약"], section_totals, totals, len(unmatched))
    detail_hdr = [
        "행", "공종", "공종명", "규격", "수량", "단위", "매칭점수", "신뢰도",
        "단가코드", "매칭품명", "매칭규격",
        "재료단가", "노무단가", "경비단가", "합계단가",
        "재료금액", "노무금액", "경비금액", "합계금액",
    ]
    rewrite_detail_sheet(wb, "매칭내역", match_only, detail_hdr)
    rewrite_detail_sheet(wb, "검토필요", review, detail_hdr, REVIEW_FILL)
    rewrite_unmatched_sheet(wb, unmatched)

    src = xlsx_name.replace("_표준단가산출.xlsx", ".XLS")
    try:
        wb.save(path)
        saved = path
    except PermissionError:
        saved = path.with_name(path.stem + "_확정반영.xlsx")
        wb.save(saved)
        print(f"  원본 사용 중 → {saved.name}", flush=True)
    wb.close()

    md_path = saved.with_name(saved.stem.replace("_확정반영", "") + "_요약.md")
    if md_path.name.endswith("_표준단가산출_요약.md"):
        asp.write_md(
            matched, unmatched, review, totals, section_totals,
            "일위대가 확정 반영", src, md_path,
        )

    n_m = len(match_only)
    print(f"  {saved.name}: {applied}건 → 매칭 {n_m} / 검토 {len(review)} / 미매칭 {len(unmatched)}", flush=True)
    return applied


def main() -> None:
    um = load_ilwidae_overrides(UNMATCHED_IL, auto_suggest=True)
    rv = load_ilwidae_overrides(REVIEW_IL, auto_suggest=False)
    tg = load_ilwidae_overrides(REVIEW_TOGONG_IL, auto_suggest=False)
    sc = load_ilwidae_overrides(REVIEW_SECTIONS_IL, auto_suggest=False)
    print(f"미매칭 {len(um)}건 · 검토 확정 {len(rv)}건 · 토공 {len(tg)}건 · 공종별 {len(sc)}건", flush=True)

    by_file: dict[str, list] = {}
    for o in um + rv + tg + sc:
        by_file.setdefault(o["file"], []).append(o)

    total = 0
    for label, fname in FILE_MAP.items():
        ovs = by_file.get(label, [])
        if not ovs:
            continue
        print(f"[{label}]", flush=True)
        total += apply_file(fname, ovs)

    print(f"\n총 {total}건 반영", flush=True)
    if total:
        print("총괄표 재생성…", flush=True)
        r = subprocess.run(
            [sys.executable, "-X", "utf8", str(ROOT / "tools" / "build_consolidated_summary.py")],
            cwd=str(ROOT),
        )
        if r.returncode != 0:
            sys.exit(r.returncode)


if __name__ == "__main__":
    main()
