#!/usr/bin/env python3
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
KEY = sys.argv[1] if len(sys.argv) > 1 else "가로지지대"

IL = {
    "미매칭_일위대가": BASE / "미매칭_일위대가산출.xlsx",
    "검토_일위대가": BASE / "검토_일위대가산출.xlsx",
    "검토_토공": BASE / "검토_토공_일위대가산출.xlsx",
    "검토_공종별": BASE / "검토_공종별_일위대가산출.xlsx",
}
for tag, p in IL.items():
    if not p.exists():
        continue
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb["일위대가산출"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c or "").strip() for c in rows[0]]
    H = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        nm = str(r[H["품명"]] or "") if "품명" in H else ""
        if KEY in nm:
            conf = r[H["확정단가(입력)"]] if "확정단가(입력)" in H else None
            sug = r[H["제시단가"]] if "제시단가" in H else None
            tot = r[H["합계단가"]] if "합계단가" in H else None
            print(f"[{tag}] 파일{r[H.get('파일',1)]} 행{r[H.get('행',2)]} 「{nm[:24]}」 "
                  f"규격{str(r[H['규격']])[:16]} 단위{r[H['단위']]} 수량{r[H['수량']]} "
                  f"| 합계단가{tot} 제시{sug} 확정{conf}")
            basis_i = H.get("표준품셈·산출근거") or H.get("환산·근거 / 대안후보")
            if basis_i is not None:
                print(f"      근거: {str(r[basis_i])[:120]}")
    wb.close()

# 통합내역(01 토목)
tok = WORK / "01_화성 청원지구 토목_표준단가산출.xlsx"
wb = load_workbook(tok, read_only=True, data_only=True)
ws = wb["통합내역"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c or "").strip() for c in rows[0]]
H = {h: i for i, h in enumerate(hdr)}
print("---통합내역(01토목)---")
for r in rows[1:]:
    nm = str(r[H["공종명"]] or "")
    if KEY in nm:
        print(f"행{r[H['행']]} 「{nm[:24]}」 규격{str(r[H['규격']])[:16]} 단위{r[H['단위']]} "
              f"수량{r[H['수량']]} 상태{r[H['상태']]} 단가코드{r[H['단가코드']]} "
              f"합계단가{r[H['합계단가']]:,} 합계금액{r[H['합계금액']]:,}" if isinstance(r[H['합계단가']], (int, float)) else
              f"행{r[H['행']]} 「{nm[:24]}」 {r[H['상태']]} {r[H['단가코드']]} 단가{r[H['합계단가']]}")
wb.close()
