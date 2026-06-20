#!/usr/bin/env python3
"""화성 청원지구 포털 — 최종 전수조사.

1) 포털 TABS 정의에서 탭별 좌측(main/placeholder)·우측(refs) 추출
2) 참조 파일 존재 여부 전수 확인
3) 좌/우 패널 역할 규칙 검증(좌 html·md·안내 / 우 xlsx·pdf·img)
4) 합계 일관성: 요약md · 총괄표md · 총괄표html · 표준단가산출 xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
PORTAL = ROOT / "청원지구_포털.html"
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"

html = PORTAL.read_text(encoding="utf-8")


def section(title: str) -> None:
    print("\n" + "=" * 64)
    print(f"■ {title}")
    print("=" * 64)


# ---------------------------------------------------------------------------
# 1) 참조 파일 경로 전수 추출
# ---------------------------------------------------------------------------
section("1. 참조 파일 존재 전수조사")

paths: set[str] = set()
# 00~04 자료 탭(상대경로)
for m in re.finditer(r'"(0\d_[^"]+\.(?:html|md|pdf|jpg|jpeg|png|txt|XLS|xlsx))"', html):
    paths.add(m.group(1))
# B + "/..." (05_내역서 하위)
for m in re.finditer(r'B \+ "(/[^"]+)"', html):
    paths.add("05_내역서" + m.group(1))
# enc("...") 직접 경로
for m in re.finditer(r'enc\("([^"]+\.(?:html|md|pdf|jpg|jpeg|png|txt|XLS|xlsx))"\)', html):
    paths.add(m.group(1))

ok = miss = 0
missing: list[str] = []
for p in sorted(paths):
    full = ROOT / p
    if full.exists():
        ok += 1
    else:
        miss += 1
        missing.append(p)
print(f"참조 경로 {len(paths)}건 · 존재 {ok} · 누락 {miss}")
for p in missing:
    print(f"  ✗ 누락: {p}")

# 동적 생성(GONG·NY·UM·RV) 파일도 직접 확인
section("1-b. 드롭다운 동적 생성 파일 확인")

GONG = [
    "01_화성 청원지구 토목.XLS", "01_화성 청원지구 조경.XLS",
    "02_화성 청원지구 전기설비.xlsx", "03_화성 청원지구 전기설비(지구외).xlsx",
    "04_화성 청원지구 진입도로 실시설계.XLS", "05_화성 청원로(회전교차로).XLS",
    "06_화성 청원지구 산업유통형 개발행위.XLS", "07_화성 청원지구 건설폐기물처리.XLS",
]
NY_STEMS = [
    "01_화성 청원지구 토목_표준단가산출", "01_화성 청원지구 조경_표준단가산출",
    "02_화성 청원지구 전기설비_표준단가산출", "04_화성 청원지구 진입도로 실시설계_표준단가산출",
    "05_화성 청원로(회전교차로)_표준단가산출", "06_화성 청원지구 산업유통형 개발행위_표준단가산출",
    "07_화성 청원지구 건설폐기물처리_표준단가산출",
]
UM_RV = [
    "미매칭_전체.xlsx", "미매칭_일위대가산출.xlsx", "토목_미매칭_수동단가입력표.xlsx",
    "조경수_미매칭점검.xlsx", "조달청_미매칭점검.xlsx", "조달청보정_매칭결과.xlsx",
    "한국물가협회_미매칭점검.xlsx",
    "검토_전체.xlsx", "검토_일위대가산출.xlsx", "검토_토공_일위대가산출.xlsx",
    "검토_공종별_일위대가산출.xlsx",
]
SUM = ["총괄표.xlsx", "총괄표_공종별.xlsx", "총괄표.md", "내역서_표준단가산출_총괄표.html"]

dyn_miss = []
for f in GONG:
    if not (BASE / "공내역서" / f).exists():
        dyn_miss.append(f"공내역서/{f}")
for stem in NY_STEMS:
    if not (WORK / f"{stem}.xlsx").exists():
        dyn_miss.append(f"내역서작업/{stem}.xlsx")
    if not (WORK / f"{stem}_요약.md").exists():
        dyn_miss.append(f"내역서작업/{stem}_요약.md")
for f in UM_RV + SUM:
    if not (BASE / f).exists():
        dyn_miss.append(f)
if dyn_miss:
    for f in dyn_miss:
        print(f"  ✗ 누락: {f}")
else:
    print("  드롭다운 동적 파일 전부 존재 (공내역서 8 · 내역서 7×2 · 미매칭/검토 11 · 집계 4)")


# ---------------------------------------------------------------------------
# 2) 좌/우 패널 역할 규칙 검증
# ---------------------------------------------------------------------------
section("2. 좌/우 패널 역할 매핑 (설계 규칙)")
print("""  좌측(본문)        우측(근거자료)
  --------------------------------------------
  진명/토지/환평/지구/인허가  html iframe   pdf·jpg·txt·링크
  공내역서 목록            html 안내     XLS 목록
  공내역서 개별            안내 placeholder  해당 XLS 1건(자동 미리보기)
  내역서(ny)              _요약.md       _표준단가산출.xlsx
  집계표 HTML             총괄표.html    xlsx·md
  집계표 xlsx/md           총괄표.md      xlsx
  미매칭/검토              안내 placeholder  해당 xlsx(자동 미리보기)""")
# 규칙 위반 패턴: 좌측 xlsx 직접 지정(switchTab은 main/placeholder만 처리)
bad_left_xlsx = re.findall(r'main:\s*B \+ "/[^"]+\.xlsx"', html)
print(f"\n  좌측 main에 xlsx 직접 지정(위반 후보): {len(bad_left_xlsx)}건")
for b in bad_left_xlsx:
    print(f"    · {b}")


# ---------------------------------------------------------------------------
# 3) 합계 일관성
# ---------------------------------------------------------------------------
section("3. 합계 일관성 대조")


def grep_won(path: Path, patterns: list[str]) -> dict:
    if not path.exists():
        return {"(없음)": str(path)}
    txt = path.read_text(encoding="utf-8", errors="ignore")
    out = {}
    for pat in patterns:
        m = re.search(pat, txt)
        out[pat] = m.group(0) if m else "—"
    return out


md = BASE / "총괄표.md"
htmlsum = BASE / "내역서_표준단가산출_총괄표.html"
print("총괄표.md  직접공사비:", grep_won(md, [r"1,372,[\d,]+", r"13722\.\d+억", r"6,247,[\d,]+"]))
print("총괄표.html 직접공사비:", grep_won(htmlsum, [r"6,247,504,709", r"62\.48", r"1,372,[\d,]+"]))

import openpyxl

def xlsx_sum(path: Path) -> str:
    if not path.exists():
        return "(없음)"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            w = wb[sn]
            for row in w.iter_rows(values_only=True):
                for c in row:
                    if isinstance(c, (int, float)) and c > 1e11:
                        wb.close()
                        return f"{sn}: 최대값 {c:,.0f}"
    finally:
        wb.close()
    return "1e11 초과 값 없음"


print("총괄표.xlsx  :", xlsx_sum(BASE / "총괄표.xlsx"))

# 01 토목 부대공 경비 점검
section("4. 01 토목 부대공 경비 폭증 점검")
p = WORK / "01_화성 청원지구 토목_표준단가산출.xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
w = wb["통합내역"]
h = [c.value for c in next(w.iter_rows(min_row=1, max_row=1))]
ci = {k: h.index(k) for k in ["행", "공종", "공종명", "규격", "수량", "단위",
                              "단가코드", "매칭품명", "경비단가", "경비금액",
                              "합계단가", "합계금액"]}
big = []
for r in w.iter_rows(min_row=2, values_only=True):
    amt = r[ci["합계금액"]]
    if isinstance(amt, (int, float)) and amt > 1e9:
        big.append(r)
big.sort(key=lambda r: -(r[ci["합계금액"]] or 0))
print(f"합계금액 10억 초과 행: {len(big)}건")
for r in big[:8]:
    print(f"  행{r[ci['행']]} [{r[ci['공종']]}] {str(r[ci['공종명']])[:18]} "
          f"규격={str(r[ci['규격']])[:18]} 수량={r[ci['수량']]}{r[ci['단위']]}")
    print(f"     코드={r[ci['단가코드']]} 매칭={str(r[ci['매칭품명']])[:24]} "
          f"경비단가={r[ci['경비단가']]:,} → 경비금액={r[ci['경비금액']]:,.0f} 합계={r[ci['합계금액']]:,.0f}")
wb.close()
