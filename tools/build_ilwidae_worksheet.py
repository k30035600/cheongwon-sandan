#!/usr/bin/env python3
"""미매칭 전체 → 일위대가·품셈 참조 산출표.

표준시장단가·표준일위대가·조경일위·forestinfo·시중노임·2026 표준품셈 PDF 조항을
경로별로 붙여 일위대가 산출·확정 작업표를 만든다.
출력: 05_내역서/미매칭_일위대가산출.xlsx
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
import apply_standard_prices as asp  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
DB = BASE / "일위대가DB"
MARKET_CSV = DB / "표준일위대가_2026" / "표준시장단가_2026.csv"
FOREST_CSV = DB / "조경수_관측시세.csv"
NOIM_CSV = DB / "시중노임_2026.csv"
WORK = BASE / "내역서작업"
OUT = BASE / "미매칭_일위대가산출.xlsx"

SRC = [
    ("01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx"),
    ("01 조경", "01_화성 청원지구 조경_표준단가산출.xlsx"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx"),
    ("05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx"),
    ("06 개발행위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx"),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ROUTE_FILLS = {
    "H.근접단가+환산": PatternFill("solid", fgColor="E2EFDA"),
    "A.품셈 자재운반": PatternFill("solid", fgColor="FFF2CC"),
    "B.품셈 시공+시장단가": PatternFill("solid", fgColor="FFF2CC"),
    "C.품셈 조경방제": PatternFill("solid", fgColor="FFF2CC"),
    "E.조경일위/시세": PatternFill("solid", fgColor="DDEBF7"),
    "D.시험(품셈外)": PatternFill("solid", fgColor="F2F2F2"),
    "F.위탁·시장단가": PatternFill("solid", fgColor="FCE4D6"),
    "G.특수품목(DB없음)": PatternFill("solid", fgColor="FCE4D6"),
    "I.수동·별도견적": PatternFill("solid", fgColor="FCE4D6"),
}
MONEY = "#,##0"
STOP_NOUN = {"조립식", "블록형", "마감형", "표준형", "일반형", "현장", "기성", "콘크리트"}

POOMSEM = [
    ("시멘트운반", "2026 건설공사 표준품셈 — 자재운반(공통) L=30km, 1대=화물차 1회(상·하차+기계경비+인력)"),
    ("철근운반", "2026 표준품셈 — 철근 운반 L=10km, Ton당(트럭+인력운반공)"),
    ("주철관운반", "2026 표준품셈 — 자재운반(공통) 주철관 트럭 ton·km"),
    ("스테인리스", "2026 표준품셈 [기계설비] 1-1-3 나사식 접합·배관 — 배관공·보통인부×시중노임, 개소→m 환산"),
    ("방제", "2026 표준품셈 [조경] 병해충 방제 — 살수·살포 인부×노임+약제"),
    ("운반", "2026 표준품셈 자재운반(공통) — 거리·중량별"),
    ("주철관", "표준시장단가 주철관 부설(CGB) — 시공조건 A/B/C"),
    ("임목", "위탁 파쇄 단가 — 장비·유류·인건비 별도 견적(품셈 外)"),
    ("시험", "시험·검사 — 발주처·협회 단가 또는 표준일위대가 시험 항목"),
]


def clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").replace("_x000D_", "").strip()


def core_name(name: str) -> str:
    s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", clean(name))
    return re.sub(r"^[.·\s]+", "", s).strip()


def norm_spec(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9가-힣]", "", (s or "").upper()).replace("Ⅹ", "X").replace("×", "X")


def dnum(spec: str, name: str) -> str | None:
    m = re.search(r"D\s*=?\s*(\d{2,4})", f"{name} {spec}", re.I)
    if not m:
        m = re.search(r"(\d{2,4})\s*mm", f"{name} {spec}", re.I)
    return m.group(1) if m else None


def classify_route(name: str, spec: str, unit: str, score: float | None) -> str:
    cn = core_name(name)
    blob = cn + " " + spec
    sc = float(score) if isinstance(score, (int, float)) and score >= 0 else -1.0
    if any(k in blob for k in ("운반", "L=")):
        return "A.품셈 자재운반"
    if any(k in blob for k in ("접합", "부설", "나사")) and "관" in blob:
        return "B.품셈 시공+시장단가"
    if any(k in blob for k in ("방제", "살수", "살포")):
        return "C.품셈 조경방제"
    if any(k in blob for k in ("시험", "검사", "통수", "수압", "공기압")):
        return "D.시험(품셈外)"
    if unit == "주" or any(k in blob for k in ("수목", "교목", "관목", "식재", "전정", "전지")):
        return "E.조경일위/시세"
    if any(k in blob for k in ("임목", "파쇄")):
        return "F.위탁·시장단가"
    if any(k in blob for k in ("조립", "저류", "PE천막", "세라믹")):
        return "G.특수품목(DB없음)"
    if sc >= 0.45:
        return "H.근접단가+환산"
    return "I.수동·별도견적"


def poomsem_ref(name: str, spec: str) -> str:
    blob = core_name(name) + " " + spec
    for kw, ref in POOMSEM:
        if kw in blob.replace(" ", "") or kw in blob:
            return ref
    return "2026 건설공사 표준품셈 — 해당 공종 조항 확인, 직종 공량×시중노임+재료"


def conversion_note(unit: str, name: str, route: str) -> str:
    u = clean(unit)
    if route != "H.근접단가+환산" and not route.startswith(("B.", "A.")):
        if route == "E.조경일위/시세":
            return "1주=식재+양식 노무비(조경일위) 또는 forestinfo 재료 참고"
        return ""
    if u in ("개소", "본") and "접합" in name:
        return "개소→m(접합 길이) 환산 후 단가 적용"
    if u in ("EA", "ea"):
        return "1ea=규격길이(m) 또는 1식 환산"
    if u.lower() in ("m2", "㎡"):
        return "㎡↔m2 동일"
    if u in ("대",):
        return "1대=운반 1회분"
    if u.lower() in ("ton", "t"):
        return "Ton=톤 운반 품셈"
    return ""


def load_market_csv() -> list[dict]:
    rows = []
    if not MARKET_CSV.exists():
        return rows
    with MARKET_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def find_market_direct(name: str, spec: str, market: list[dict]) -> list[dict]:
    cn = core_name(name)
    if "운반" in cn:
        return []
    toks = [t for t in re.split(r"[\s/()]+", cn) if len(t) >= 2 and t not in STOP_NOUN]
    if not toks:
        return []
    key = toks[0]
    d = dnum(spec, name)
    cands = []
    for p in market:
        if key not in p["품명"]:
            continue
        if d and d not in (p.get("규격") or ""):
            continue
        if not d:
            blob = f"{p['품명']} {p.get('규격', '')}"
            if len(toks) > 1 and not any(t in blob for t in toks[1:]):
                continue
        cands.append(p)
    cands.sort(key=lambda p: (0 if "연속" in (p.get("규격") or "") else 1, float(p["합계"])))
    return cands


def load_forest() -> list[dict]:
    if not FOREST_CSV.exists():
        return []
    rows = []
    with FOREST_CSV.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def find_forest(name: str, spec: str, forest: list[dict]) -> dict | None:
    species = core_name(name)
    sn = norm_spec(spec)
    best = None
    best_score = -1
    for r in forest:
        sp = r.get("species", "")
        if species not in sp and sp not in species:
            continue
        rs = norm_spec(r.get("spec_raw", "") or r.get("spec_norm", ""))
        score = 1 if sn and (sn in rs or rs in sn) else 0
        if sn and rs:
            # numeric overlap R15, H4 etc
            nums = set(re.findall(r"\d+", sn)) & set(re.findall(r"\d+", rs))
            if nums:
                score += len(nums)
        if score > best_score:
            best_score = score
            best = r
    return best if best_score >= 1 else None


def load_pool() -> list[dict]:
    market = asp.load_market_csv(asp.MARKET_2026, "표준시장단가2026")
    if not market:
        market = asp.load_prices()
    sijang = asp.load_market_csv(asp.SIJANG_2026, "시장시공가격2026")
    return asp.precompute(
        market + sijang + asp.load_ildae_prices() + asp.load_mulga()
        + asp.load_jojadang() + asp.load_landscape_ildae()
    )


def topn_db(item: dict, prices: list[dict], n: int = 3) -> list[tuple[float, dict]]:
    terms = asp.extract_search_terms(item["name"], item["spec"])
    req = asp.required_keywords(item)
    item_kw = asp.kwset(f"{item['name']} {item['spec']}")
    item["_linear"] = asp.is_linear_piece(item)
    cand = [p for p in prices if asp.unit_compatible(item["unit"], p["unit"], item["_linear"])]
    best: dict[int, tuple[float, dict]] = {}
    for term in terms:
        kw = item_kw | asp.kwset(term)
        for p in cand:
            s = asp.score_match(item, p, term, kw, req)
            if s <= 0:
                continue
            pid = id(p)
            if pid not in best or s > best[pid][0]:
                best[pid] = (s, p)
    return sorted(best.values(), key=lambda x: -x[0])[:n]


def load_unmatched(prices: list[dict], market: list[dict], forest: list[dict]) -> list[dict]:
    records: list[dict] = []
    for label, fname in SRC:
        fp = WORK / fname
        if not fp.exists():
            continue
        wb = load_workbook(fp, read_only=True, data_only=True)
        if "미매칭" not in wb.sheetnames:
            wb.close()
            continue
        for r in wb["미매칭"].iter_rows(min_row=2, values_only=True):
            if not r or r[2] is None:
                continue
            name = clean(r[2])
            spec = clean(r[3])
            qty = r[4] if isinstance(r[4], (int, float)) else None
            unit = clean(r[5])
            score = r[6]
            item = {"name": name, "spec": spec, "unit": unit, "qty": qty or 0, "section": clean(r[1])}
            route = classify_route(name, spec, unit, score)

            db_name = db_spec = db_unit = code = ""
            mat = lab = exp = tot = 0.0
            applied = None
            basis_parts: list[str] = []
            alts = ""

            if route == "H.근접단가+환산":
                ranked = topn_db(item, prices, 3)
                if ranked:
                    sc, p = ranked[0]
                    db_name, db_spec, db_unit = p["name"], p["spec"], p["unit"]
                    code = p.get("code", "")
                    mat, lab, exp, tot = p["mat"], p["lab"], p["exp"], p["total"]
                    applied = tot
                    basis_parts.append(f"표준DB {code} (점수{sc:.2f})")
                    if len(ranked) > 1:
                        alts = " · ".join(
                            f"{x[1]['name'][:16]}({x[0]:.2f})" for x in ranked[1:3]
                        )
            elif route == "B.품셈 시공+시장단가":
                mc = find_market_direct(name, spec, market)
                if mc:
                    p = mc[0]
                    db_name, db_spec = clean(p["품명"]), clean(p["규격"])
                    mat, lab, exp, tot = float(p["재료비"]), float(p["노무비"]), float(p["경비"]), float(p["합계"])
                    applied = tot
                    basis_parts.append(f"표준시장단가 {p['코드']} 직접")
                else:
                    ranked = topn_db(item, prices, 2)
                    if ranked:
                        sc, p = ranked[0]
                        db_name, db_spec, db_unit = p["name"], p["spec"], p["unit"]
                        mat, lab, exp, tot = p["mat"], p["lab"], p["exp"], p["total"]
                        basis_parts.append(f"일위대가 후보 {p.get('code','')} ({sc:.2f})")
            elif route == "E.조경일위/시세":
                fr = find_forest(name, spec, forest)
                if fr:
                    applied = float(fr.get("price_mid") or 0)
                    tot = applied
                    basis_parts.append(
                        f"forestinfo {fr['species']} {fr.get('spec_raw','')} "
                        f"참고가 {applied:,.0f}원/주(고시 아님)"
                    )
                ranked = topn_db(item, prices, 2)
                if ranked and not applied:
                    sc, p = ranked[0]
                    db_name, db_spec = p["name"], p["spec"]
                    mat, lab, exp, tot = p["mat"], p["lab"], p["exp"], p["total"]
                    basis_parts.append(f"조경일위 후보 ({sc:.2f})")
                elif ranked and applied:
                    alts = f"조경일위: {ranked[0][1]['name'][:20]}({ranked[0][0]:.2f})"

            conv = conversion_note(unit, name, route)
            if conv:
                basis_parts.append(conv)
            if route.startswith(("A.", "B.", "C.")):
                basis_parts.append(poomsem_ref(name, spec))
            elif route in ("D.시험(품셈外)", "F.위탁·시장단가", "G.특수품목(DB없음)", "I.수동·별도견적"):
                basis_parts.append(poomsem_ref(name, spec))
            if alts:
                basis_parts.append("대안: " + alts)

            records.append({
                "route": route,
                "file": label,
                "row": r[0],
                "section": clean(r[1]),
                "name": name,
                "spec": spec,
                "unit": unit,
                "qty": qty,
                "score": score,
                "db_name": db_name,
                "db_spec": db_spec,
                "db_unit": db_unit,
                "code": code,
                "mat": mat,
                "lab": lab,
                "exp": exp,
                "tot": tot,
                "applied": applied,
                "amount": (applied or tot or 0) * qty if qty and (applied or tot) else None,
                "basis": " / ".join(basis_parts),
            })
        wb.close()

    order = list(ROUTE_FILLS.keys()) + ["I.수동·별도견적"]
    records.sort(key=lambda x: (order.index(x["route"]) if x["route"] in order else 99, x["file"], x["name"]))
    return records


def load_noim() -> dict[str, int]:
    d: dict[str, int] = {}
    if not NOIM_CSV.exists():
        return d
    with NOIM_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                d[r["직종명"]] = int(r["2026.1.1"])
            except (ValueError, KeyError):
                pass
    return d


def _name_overlap(name, db_name) -> int:
    """내역 품명과 DB 매칭품명의 공유 토큰 수. DB명 없음(시세·품셈 직접)은 1로 본다."""
    import re

    def toks(s):
        s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", str(s or ""))
        return [t for t in re.split(r"[\s/()·,]+", s) if len(t) >= 2]

    db = str(db_name or "").strip()
    if not db or db == "None":
        return 1  # DB 매칭품명 없음 — 시세·품셈 직접 산출(오매칭 아님)
    blob = db.replace(" ", "")
    return sum(1 for t in toks(name) if t in db or t in blob)


def write_workbook(records: list[dict], noim: dict[str, int]) -> None:
    wb = Workbook()
    route_counts: dict[str, int] = {}
    for rec in records:
        route_counts[rec["route"]] = route_counts.get(rec["route"], 0) + 1

    sm = wb.active
    sm.title = "경로별요약"
    sm.append(["미매칭 전체 — 일위대가·품셈 참조 산출"])
    sm.append([])
    sm.append(["경로", "건수", "처리"])
    desc = {
        "H.근접단가+환산": "DB 후보+단위환산 → 확정단가",
        "A.품셈 자재운반": "2026 표준품셈 자재운반+노임",
        "B.품셈 시공+시장단가": "표준시장단가 또는 품셈+노임",
        "C.품셈 조경방제": "조경 품셈+노임+약제",
        "E.조경일위/시세": "forestinfo·조경일위2024",
        "D.시험(품셈外)": "발주처·협회 단가",
        "F.위탁·시장단가": "위탁·장비 견적",
        "G.특수품목(DB없음)": "별도 견적",
        "I.수동·별도견적": "수동 입력",
    }
    for route in list(ROUTE_FILLS.keys()) + ["I.수동·별도견적"]:
        if route_counts.get(route):
            sm.append([route, route_counts[route], desc.get(route, "")])
    sm.append([])
    sm.append(["합계", len(records), ""])
    direct = sum(1 for r in records if r["applied"] or (r["tot"] and r["route"] == "H.근접단가+환산"))
    sm.append(["단가 자동제시", direct, "확정단가 열에서 최종 기입"])
    sm["A1"].font = Font(bold=True, size=13)

    ws = wb.create_sheet("일위대가산출")
    headers = [
        "경로", "파일", "행", "공종", "품명", "규격", "단위", "수량", "최고점수",
        "DB_품명", "DB_규격", "DB_단위", "재료단가", "노무비", "경비", "합계단가",
        "제시단가", "제시금액", "확정단가(입력)", "확정금액",
        "표준품셈·산출근거",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for rec in records:
        suggested = rec["applied"] if rec["applied"] else (rec["tot"] if rec["tot"] else "")
        sug_amt = rec["amount"]
        auto_routes = ("H.근접단가+환산", "B.품셈 시공+시장단가", "E.조경일위/시세")
        auto = rec["route"] in auto_routes or str(rec["route"]).startswith("J.품셈")
        # 저신뢰 근접매칭(H)은 내역품명↔DB매칭품명 공유 토큰이 0이면 자동확정 보류(오매칭 방지)
        risky = rec["route"] == "H.근접단가+환산" and _name_overlap(rec["name"], rec.get("db_name")) == 0
        auto_confirm = auto and not risky
        basis = rec["basis"]
        if risky and suggested:
            basis = "⚠ 내역품명↔DB매칭품명 불일치(공유 토큰 0) — 자동확정 보류, 검토 요. " + str(basis or "")
        row = [
            rec["route"], rec["file"], rec["row"], rec["section"], rec["name"], rec["spec"],
            rec["unit"], rec["qty"], rec["score"],
            rec["db_name"], rec["db_spec"], rec["db_unit"],
            rec["mat"] or None, rec["lab"] or None, rec["exp"] or None, rec["tot"] or None,
            suggested or None, sug_amt,
            suggested if auto_confirm and suggested else "",
            sug_amt if auto_confirm and sug_amt else "",
            basis,
        ]
        ws.append(row)
        ridx = ws.max_row
        fill = ROUTE_FILLS.get(rec["route"], ROUTE_FILLS["I.수동·별도견적"])
        for c in range(1, len(headers) + 1):
            ws.cell(ridx, c).fill = fill
            ws.cell(ridx, c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(13, 21):
            if isinstance(ws.cell(ridx, c).value, (int, float)):
                ws.cell(ridx, c).number_format = MONEY

    ws.freeze_panes = "A2"
    widths = [14, 11, 5, 12, 24, 20, 5, 7, 7, 20, 18, 5, 9, 9, 9, 10, 10, 11, 11, 11, 52]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ref = wb.create_sheet("시중노임참고")
    ref.append(["직종", "노임(2026.1.1)", "용도"])
    use = {
        "배관공": "STS·PVC 접합", "배관공(수도)": "상수도 배관",
        "보통인부": "보조", "인력운반공": "운반", "조경공": "식재·방제",
        "특별인부": "숙련 보조", "작업반장": "관리",
    }
    for j, u in use.items():
        ref.append([j, noim.get(j, ""), u])
    for c in range(1, 4):
        ref.cell(1, c).font = Font(bold=True)
        ref.cell(1, c).fill = HEADER_FILL

    info = wb.create_sheet("안내")
    for line in [
        ["미매칭 230건 — 일위대가·품셈 참조 산출표"],
        [],
        ["색상", "연두=H DB확정 / 노랑=품셈조합 / 하늘=조경시세 / 주황·회=수동"],
        ["확정", "「확정단가(입력)」 → python -X utf8 tools/apply_confirmed_prices.py 로 표준단가산출·총괄표 반영"],
        ["품셈", "2026 건설공사 표준품셈.pdf — 조항·페이지는 산출근거 열 참고"],
        ["조경", "forestinfo=참고가(고시 아님). 조경일위2024=노무비 포함 일위대가"],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=14)
    info.column_dimensions["A"].width = 12
    info.column_dimensions["B"].width = 85

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(f"저장: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")


def main() -> None:
    prices = load_pool()
    market = load_market_csv()
    forest = load_forest()
    noim = load_noim()
    print(f"단가 풀 {len(prices):,} · 표준시장단가 {len(market):,} · forestinfo {len(forest):,} · 노임 {len(noim)}")

    records = load_unmatched(prices, market, forest)
    print(f"미매칭 {len(records)}건")
    for route in sorted({r["route"] for r in records}, key=lambda x: list(ROUTE_FILLS.keys()).index(x) if x in ROUTE_FILLS else 99):
        n = sum(1 for r in records if r["route"] == route)
        auto = sum(1 for r in records if r["route"] == route and (r["applied"] or r["tot"]))
        print(f"  {route}: {n}건 (단가제시 {auto})")

    write_workbook(records, noim)


if __name__ == "__main__":
    main()
