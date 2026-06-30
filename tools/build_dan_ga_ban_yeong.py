#!/usr/bin/env python3
"""공내역서(전기 제외) — 단가반영 통합 xlsx 1개 생성.

시트 구성:
  1) 총공사비(도급·간접비)  2) 내역집계
  3) 구분별 시트 — 공종 소계 + 내역 일괄
  내역집계·총공사비 — 구분 시트 ★소계 셀 참조

입력: 05_내역서/내역서작업/*_표준단가산출.xlsx
출력: 08_제출내역서/청원지구_단가통합(전기제외).xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_consolidated_summary import _compute_cost, _style_cost_row  # noqa: E402
from calc_overhead import CIVIL_TOMOK_RATES, PF_RATES  # noqa: E402
from naeyeok_gongjong import normalize_ka_detail_status  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
PUREUN_JO_XLS = BASE / "공내역서" / "조경시설물" / "01_260620화성시마도면청원리청원지구 조경.XLS"
OUT_DIR = ROOT / "08_제출내역서"
OUT_UNIFIED = OUT_DIR / "청원지구_단가통합(전기제외).xlsx"
OUT_KA = OUT_DIR / "청원지구_단가통합(전기제외)_ka.xlsx"
SHEET_NAERYEO = "내역집계"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
REVIEW_FILL = PatternFill("solid", fgColor="FFFF00")
MONEY_FMT = "#,##0"
PCT_FMT = "0.0%"
DEFAULT_ROW_HEIGHT = 24

GROUP_ORDER = ["회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물"]

# 구분 시트 상단 요약표 열 (No=1 … 합계=10)
SUMMARY_COLS = {
    "matched": 4, "total": 5, "rate": 6,
    "mat": 7, "lab": 8, "exp": 9, "sum": 10,
}

GROUP_DETAIL_HEADERS = [
    "내역서", "공종", "No", "공종명", "규격", "수량", "단위",
    "재료단가", "재료금액", "노무단가", "노무금액", "경비단가", "경비금액",
    "합계단가", "합계금액", "상태", "비고",
]
DETAIL_COLS = {"no": 3, "mat": 9, "lab": 11, "exp": 13, "sum": 15}
DETAIL_MONEY_COLS = (8, 9, 10, 11, 12, 13, 14, 15)
DETAIL_UNIT_INPUT_COLS = (8, 10, 12)  # 재료단가·노무단가·경비단가
REVIEW_STATUSES = frozenset({"검토", "미매칭"})

FILES: list[dict] = [
    {
        "no": "05", "short": "05", "group": "회전교차로", "sheet_prefix": "회전",
        "label": "05 회전교차로",
        "std": "회전교차로/05_화성 청원로(회전교차로)_표준단가산출.xlsx",
        "engine": "tomok",
    },
    {
        "no": "04", "short": "04", "group": "진입도로", "sheet_prefix": "진입",
        "label": "04 진입도로",
        "std": "진입도로/04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
        "engine": "tomok",
    },
    {
        "no": "01", "short": "01T", "group": "토목(조경)", "sheet_prefix": "토목",
        "label": "01 토목",
        "std": "토목/01_화성 청원지구 토목_표준단가산출.xlsx",
        "engine": "tomok",
    },
    {
        "no": "01", "short": "01J", "group": "토목(조경)", "sheet_prefix": "조경",
        "label": "01 조경",
        "std": "조경/01_화성 청원지구 조경_표준단가산출.xlsx",
        "engine": "tomok",
        "jogyeong_ref": True,
        "pureun_src": "공내역서/조경시설물/01_260620화성시마도면청원리청원지구 조경.XLS",
    },
    {
        "no": "06", "short": "06", "group": "지구단위", "sheet_prefix": "지구",
        "label": "06 개발행위",
        "std": "토목/06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
        "engine": "tomok",
    },
    {
        "no": "07", "short": "07", "group": "폐기물", "sheet_prefix": "폐기물",
        "label": "07 건설폐기물",
        "std": "폐기물/07_화성 청원지구 건설폐기물처리_표준단가산출.xlsx",
        "engine": "tomok",
        "waste": True,
    },
]


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def sanitize_sheet(name: str, used: set[str]) -> str:
    s = re.sub(r"[\\\/*?:\[\]]", "", name.strip())
    s = re.sub(r"\s+", " ", s)
    if len(s) > 31:
        s = s[:31]
    base = s
    i = 2
    while s in used:
        suffix = f"({i})"
        s = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    used.add(s)
    return s


def load_summary(std_path: Path, *, waste: bool = False) -> tuple[list[dict], dict | None]:
    wb = load_workbook(std_path, read_only=True, data_only=True)
    ws = wb["합계요약"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], None
    hdr = [str(c or "").strip() for c in rows[0]]
    sections: list[dict] = []
    grand: dict | None = None
    waste_mode = waste or "수량(Ton)" in hdr
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label.startswith("★"):
            if waste_mode:
                grand = {"mat": _num(row[2]), "lab": _num(row[3]), "exp": _num(row[4]), "sum": _num(row[5])}
            else:
                grand = {
                    "matched": int(row[1] or 0), "total_items": int(row[2] or 0),
                    "mat": _num(row[4]), "lab": _num(row[5]), "exp": _num(row[6]), "sum": _num(row[7]),
                }
            continue
        if label.startswith("※"):
            continue
        if waste_mode:
            sections.append({
                "name": label, "qty_ton": _num(row[1]),
                "mat": _num(row[2]), "lab": _num(row[3]), "exp": _num(row[4]), "sum": _num(row[5]),
            })
        else:
            sections.append({
                "name": label,
                "matched": int(row[1] or 0), "total_items": int(row[2] or 0),
                "mat": _num(row[4]), "lab": _num(row[5]), "exp": _num(row[6]), "sum": _num(row[7]),
            })
    return sections, grand


def load_integrated(std_path: Path) -> list[dict]:
    wb = load_workbook(std_path, read_only=True, data_only=True)
    ws = wb["통합내역"]
    hdr = [str(h or "").strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    items: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        d = {hdr[i]: row[i] for i in range(min(len(hdr), len(row)))}
        name = d.get("공종명") or d.get("명칭") or ""
        if not str(name).strip():
            continue
        items.append({
            "section": str(d.get("공종") or "").strip(),
            "name": str(name).strip(),
            "spec": str(d.get("규격") or "").strip(),
            "qty": _num(d.get("수량")),
            "unit": str(d.get("단위") or "").strip(),
            "status": str(d.get("상태") or "").strip(),
            "note": str(d.get("비고") or d.get("단가출처") or "").strip(),
            "mat_u": _num(d.get("재료단가")), "lab_u": _num(d.get("노무단가")),
            "exp_u": _num(d.get("경비단가")), "tot_u": _num(d.get("합계단가")),
            "mat_a": _num(d.get("재료금액")), "lab_a": _num(d.get("노무금액")),
            "exp_a": _num(d.get("경비금액")), "sum_a": _num(d.get("합계금액")),
        })
    wb.close()
    return items


def norm_section(s: str) -> str:
    return re.sub(r"\s+", "", s)


def group_items(items: list[dict], sections: list[dict], *, waste: bool) -> dict[str, list[dict]]:
    if waste:
        return {"건설폐기물처리": items}
    sec_names = [s["name"] for s in sections]
    by_sec: dict[str, list[dict]] = {n: [] for n in sec_names}
    other: list[dict] = []
    sec_norm = {norm_section(n): n for n in sec_names}
    for it in items:
        target = sec_norm.get(norm_section(it["section"]))
        if target:
            by_sec[target].append(it)
        else:
            other.append(it)
    if other:
        by_sec.setdefault("기타", []).extend(other)
    return {k: v for k, v in by_sec.items() if v}


def style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _display_width(text: str) -> float:
    """Excel 열 너비 추정 — CJK는 약 2, ASCII는 1."""
    return sum(2.0 if ord(ch) > 0x7F else 1.0 for ch in text)


def _cell_display_width(value, *, money: bool = False) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return 5.0
    if isinstance(value, (int, float)):
        if money:
            return 12.0 if abs(value) < 1e11 else 14.0
        s = f"{value:g}"
        return min(len(s) + 1, 14.0)
    s = str(value).strip()
    if not s:
        return 0.0
    if s.startswith("="):
        if money:
            return 14.0
        return min(_display_width(s[:100]), 34.0)
    return min(_display_width(s), 28.0)


_MONEY_HEADERS = frozenset({
    "재료비", "노무비", "경비", "합계", "금액(원)",
    "재료금액", "노무금액", "경비금액", "합계금액",
    "재료단가", "노무단가", "경비단가", "합계단가",
})


def _detect_header_row(ws) -> dict[int, str]:
    """시트 상단 30행에서 헤더 행 탐색 → {열: 헤더명}."""
    markers = _MONEY_HEADERS | {
        "No", "구분", "내역서", "공종", "항목", "단계", "규격", "산식", "비고", "상태", "단위", "수량",
        "매칭", "전체", "매칭률", "배율", "공종명",
    }
    for r in range(1, min(ws.max_row, 30) + 1):
        labels: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() in markers:
                for cc in range(1, ws.max_column + 1):
                    vv = ws.cell(r, cc).value
                    if isinstance(vv, str) and vv.strip():
                        labels[cc] = vv.strip()
                return labels
    return {}


def optimize_column_widths(ws, *, min_w: float = 6, max_w: float = 44) -> None:
    """열 내용·헤더 유형별로 열 너비 자동 조정."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    headers = _detect_header_row(ws)
    last = ws.max_row
    if last > 600:
        rows = sorted(set(range(1, min(last, 150) + 1)) | set(range(max(1, last - 30), last + 1)))
    else:
        rows = range(1, last + 1)

    for col_idx in range(1, ws.max_column + 1):
        hdr = headers.get(col_idx, "")
        money = hdr in _MONEY_HEADERS or hdr.endswith("비") and hdr not in ("비고",)
        pct = hdr in ("매칭률", "배율")
        spec = hdr in ("규격", "산식")

        peak = 0.0
        for r in rows:
            v = ws.cell(r, col_idx).value
            if money and isinstance(v, str) and not v.startswith("=") and _display_width(v.strip()) > 14:
                money = False
            peak = max(peak, _cell_display_width(v, money=money))

        natural = peak + 2
        if money:
            width = max(11, min(15, natural))
        elif pct or hdr in ("매칭", "전체", "수량"):
            width = max(6, min(10, natural))
        elif spec or hdr == "산식":
            width = max(14, min(36, natural))
        elif hdr in ("비고", "상태", "공종명"):
            width = max(10, min(26, natural))
        elif hdr in ("내역서", "구분", "공종", "항목"):
            width = max(10, min(20, natural))
        else:
            width = max(min_w, min(max_w, natural))

        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _optimize_workbook_columns(wb: Workbook) -> None:
    for ws in wb.worksheets:
        optimize_column_widths(ws)


def autofit(ws, max_w: int = 48) -> None:
    """하위 호환 — optimize_column_widths 사용."""
    optimize_column_widths(ws, max_w=max_w)


def _sheet_ref(sheet: str, row: int, col: int) -> str:
    q = sheet.replace("'", "''")
    return f"='{q}'!{get_column_letter(col)}{row}"


def _sum_refs(refs: list[str]) -> str:
    if not refs:
        return "=0"
    parts = [r[1:] if r.startswith("=") else r for r in refs]
    return "=" + "+".join(parts)


def _xref(meta: dict, section: str, col_key: str, *, pack_label: str | None = None) -> str:
    sec = meta[section]
    row = sec["pack_rows"][pack_label] if pack_label else sec["group_total_row"]
    return _sheet_ref(meta["sheet_name"], row, sec["cols"][col_key])


def subtotal_from_rows(rows: list[dict]) -> dict:
    return {
        "mat": sum(r["mat_a"] for r in rows),
        "lab": sum(r["lab_a"] for r in rows),
        "exp": sum(r["exp_a"] for r in rows),
        "sum": sum(r["sum_a"] for r in rows),
    }


def load_pureun_jo_totals(src: Path | None = None) -> dict | None:
    """푸른조경 XLS — 내역서 합계금액(재·노·경·직접공사비)."""
    path = src or PUREUN_JO_XLS
    if not path.exists():
        return None
    try:
        import xlrd
    except ImportError:
        return None
    sh = xlrd.open_workbook(str(path)).sheet_by_name("내역서")
    for r in range(sh.nrows - 1, 2, -1):
        if str(sh.cell_value(r, 0)).strip() == "합계금액":
            mat = _num(sh.cell_value(r, 7))
            lab = _num(sh.cell_value(r, 9))
            exp = _num(sh.cell_value(r, 11))
            return {
                "mat": mat,
                "lab": lab,
                "exp": exp,
                "sum": mat + lab + exp,
            }
    return None


def _apply_pureun_jo_pack(cfg: dict, sections: list[dict], totals: dict) -> tuple[list[dict], dict]:
    """01 조경 — 푸른조경 XLS 순공사비(직접공사비)로 합계·공종 소계 정정."""
    pu = load_pureun_jo_totals(BASE / cfg["pureun_src"]) if cfg.get("pureun_src") else load_pureun_jo_totals()
    if not pu:
        return sections, totals
    new_totals = {
        **totals,
        "mat": pu["mat"],
        "lab": pu["lab"],
        "exp": pu["exp"],
        "sum": pu["sum"],
    }
    new_sections: list[dict] = []
    for s in sections:
        ns = dict(s)
        if s.get("name", "").replace(" ", "") == "10.조경공" or "조 경 공" in s.get("name", ""):
            ns.update({
                "mat": pu["mat"],
                "lab": pu["lab"],
                "exp": pu["exp"],
                "sum": pu["sum"],
                "matched": s.get("matched", s.get("total_items", 0)),
            })
        new_sections.append(ns)
    return new_sections, new_totals


def load_all() -> list[dict]:
    packs: list[dict] = []
    for cfg in FILES:
        std_path = WORK / cfg["std"]
        if not std_path.exists():
            raise FileNotFoundError(std_path)
        waste = cfg.get("waste", False)
        sections, grand = load_summary(std_path, waste=waste)
        items = load_integrated(std_path)
        grouped = group_items(items, sections, waste=waste)
        if grand:
            totals = {"mat": grand["mat"], "lab": grand["lab"], "exp": grand["exp"], "sum": grand["sum"],
                      "matched": grand.get("matched", 0), "total_items": grand.get("total_items", 0)}
        else:
            totals = {**subtotal_from_rows(items), "matched": 0, "total_items": len(items)}
        if cfg.get("jogyeong_ref"):
            sections, totals = _apply_pureun_jo_pack(cfg, sections, totals)
        packs.append({**cfg, "sections": sections, "grand": grand, "grouped": grouped, "totals": totals})
    return packs


def write_group_sheet(ws, group_name: str, gpacks: list[dict]) -> dict:
    """구분별 시트 — 공종 소계 + 내역 일괄."""
    _, summary_refs = write_group_summary(ws, group_name, gpacks)
    detail_refs = write_group_detail(ws, group_name, gpacks)
    ws.freeze_panes = f"A{detail_refs['hdr_row'] + 1}"
    return {
        "sheet_name": ws.title,
        "summary": {**summary_refs, "cols": SUMMARY_COLS},
    }


def _set_detail_line_formulas(ws, rr: int) -> None:
    """수량(F)×단가 → 금액, 합계단가·합계금액 상호 참조."""
    q, h, j, l = f"F{rr}", f"H{rr}", f"J{rr}", f"L{rr}"
    ws.cell(rr, 9).value = f"={q}*{h}"
    ws.cell(rr, 11).value = f"={q}*{j}"
    ws.cell(rr, 13).value = f"={q}*{l}"
    ws.cell(rr, 14).value = f"={h}+{j}+{l}"
    ws.cell(rr, 15).value = f"=I{rr}+K{rr}+M{rr}"


def _detail_unit_price(r: dict, key: str) -> float | None:
    """검토·미매칭 — 단가 미확정(0)이면 빈칸(수동 입력)."""
    val = r[key]
    if r["status"] in REVIEW_STATUSES and not val:
        return None
    return val


def _style_detail_review_row(ws, rr: int) -> None:
    """이상치(가영현) — 행 전체 노란색."""
    for c in range(1, len(GROUP_DETAIL_HEADERS) + 1):
        ws.cell(rr, c).fill = REVIEW_FILL


def write_group_detail(ws, group_name: str, gpacks: list[dict]) -> dict:
    """◆ 내역 일괄 — 공종별 개별 항목 + 공종 소계 + 말미 합계."""
    ws.append([])
    ws.append([f"◆ {group_name} — 내역 일괄"])
    ws.append([])
    ws.append(GROUP_DETAIL_HEADERS)
    detail_hdr_row = ws.max_row
    style_header_row(ws, detail_hdr_row, len(GROUP_DETAIL_HEADERS))

    for cfg in gpacks:
        order = [s["name"] for s in cfg["sections"] if s["name"] in cfg["grouped"]]
        for k in cfg["grouped"]:
            if k not in order:
                order.append(k)
        for sec_name in order:
            rows = cfg["grouped"][sec_name]
            section_start = ws.max_row + 1
            for i, r in enumerate(rows, 1):
                ws.append([
                    cfg["label"], sec_name, i, r["name"], r["spec"], r["qty"], r["unit"],
                    _detail_unit_price(r, "mat_u"), None,
                    _detail_unit_price(r, "lab_u"), None,
                    _detail_unit_price(r, "exp_u"), None,
                    None, None, normalize_ka_detail_status(r["status"]), r["note"],
                ])
                rr = ws.max_row
                _set_detail_line_formulas(ws, rr)
                if str(r["status"]).startswith("가영현"):
                    _style_detail_review_row(ws, rr)
            section_end = ws.max_row
            ws.append([
                cfg["label"], sec_name, "", f"★ {sec_name} 소계", "", "", "",
                "", None, "", None, "", None, "", None, "", f"{len(rows)}건",
            ])
            sr = ws.max_row
            for col in (9, 11, 13, 15):
                letter = get_column_letter(col)
                ws.cell(sr, col).value = f"=SUM({letter}{section_start}:{letter}{section_end})"
            for c in range(1, len(GROUP_DETAIL_HEADERS) + 1):
                cell = ws.cell(sr, c)
                cell.font = Font(bold=True)
                cell.fill = SUBTOTAL_FILL

    detail_total_row = _append_detail_grand_total(ws, group_name, detail_hdr_row)

    for ri in range(detail_hdr_row + 1, ws.max_row + 1):
        for ci in DETAIL_MONEY_COLS:
            v = ws.cell(ri, ci).value
            if isinstance(v, (int, float)):
                ws.cell(ri, ci).number_format = MONEY_FMT
            if isinstance(v, str) and v.startswith("="):
                ws.cell(ri, ci).number_format = MONEY_FMT

    return {"hdr_row": detail_hdr_row, "total_row": detail_total_row}


def _preserve_sheet_layout(src_path: Path, wb: Workbook) -> None:
    """(사용 안 함) 열 너비는 optimize_column_widths, 행 높이는 DEFAULT_ROW_HEIGHT."""
    del src_path, wb


def _apply_row_heights(wb: Workbook, height: float = DEFAULT_ROW_HEIGHT) -> None:
    for ws in wb.worksheets:
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = height


def _append_detail_grand_total(ws, group_name: str, detail_hdr_row: int) -> int:
    """내역 일괄 말미 — No가 있는 행만 합산(SUMIF)."""
    data_start = detail_hdr_row + 1
    data_end = ws.max_row
    ws.append([
        "", "", "", f"★ {group_name} 내역 합계", "", "", "",
        "", None, "", None, "", None, "", None, "", "No<>인 행 합산",
    ])
    total_row = ws.max_row
    no_letter = get_column_letter(DETAIL_COLS["no"])
    for key in ("mat", "lab", "exp", "sum"):
        cl = get_column_letter(DETAIL_COLS[key])
        ws.cell(total_row, DETAIL_COLS[key]).value = (
            f'=SUMIF({no_letter}{data_start}:{no_letter}{data_end},"<>",{cl}{data_start}:{cl}{data_end})'
        )
    for c in range(1, len(GROUP_DETAIL_HEADERS) + 1):
        cell = ws.cell(total_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
    return total_row


def write_group_summary(ws, group_name: str, gpacks: list[dict]) -> tuple[dict, dict]:
    """구분(회전교차로 등)별 상단 요약 — ★ 합계 행 위치 반환."""
    ws.append([f"화성 청원지구 — {group_name}"])
    ws.append([])
    hdr = ["No", "내역서", "공종", "매칭", "전체", "매칭률", "재료비", "노무비", "경비", "합계"]
    ws.append(hdr)
    hdr_row = ws.max_row
    style_header_row(ws, hdr_row, len(hdr))

    g_mat = g_lab = g_exp = g_sum = g_matched = g_total = 0
    pack_rows: dict[str, int] = {}
    for cfg in gpacks:
        for s in cfg["sections"]:
            rate = s["matched"] / s["total_items"] if s.get("total_items") else 0
            ws.append([
                cfg["no"], cfg["label"], s["name"],
                s.get("matched", ""), s.get("total_items", ""),
                rate if s.get("total_items") else "",
                s["mat"], s["lab"], s["exp"], s["sum"],
            ])
        t = cfg["totals"]
        ws.append([
            cfg["no"], cfg["label"], f"★ {cfg['label']} 소계",
            t.get("matched", ""), t.get("total_items", ""),
            t["matched"] / t["total_items"] if t.get("total_items") else 0,
            t["mat"], t["lab"], t["exp"], t["sum"],
        ])
        pack_rows[cfg["label"]] = ws.max_row
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(ws.max_row, c)
            cell.font = Font(bold=True)
            cell.fill = SUBTOTAL_FILL
        g_mat += t["mat"]
        g_lab += t["lab"]
        g_exp += t["exp"]
        g_sum += t["sum"]
        g_matched += t.get("matched", 0)
        g_total += t.get("total_items", 0)

    ws.append([
        "", group_name, f"★ {group_name} 합계",
        g_matched, g_total, g_matched / g_total if g_total else 0,
        g_mat, g_lab, g_exp, g_sum,
    ])
    group_total_row = ws.max_row
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(ws.max_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    ws["A1"].font = Font(bold=True, size=14)
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in (7, 8, 9, 10):
                cell.number_format = MONEY_FMT
            if cell.column == 6 and isinstance(cell.value, float):
                cell.number_format = PCT_FMT
    refs = {"hdr_row": hdr_row, "group_total_row": group_total_row, "pack_rows": pack_rows}
    return {"mat": g_mat, "lab": g_lab, "exp": g_exp, "sum": g_sum}, refs


def write_unified_summary(ws, packs: list[dict], group_meta: dict[str, dict]) -> None:
    ws.append(["화성 청원지구 — 내역 집계 (01·04~07, 02 전기 제외)"])
    ws.append([])
    hdr = ["구분", "No", "내역서", "공종", "매칭", "전체", "매칭률", "재료비", "노무비", "경비", "합계"]
    ws.append(hdr)
    hdr_row = ws.max_row
    style_header_row(ws, hdr_row, len(hdr))

    by_group: dict[str, list[dict]] = {}
    for p in packs:
        by_group.setdefault(p["group"], []).append(p)

    group_total_refs: list[dict] = []

    for group_name in GROUP_ORDER:
        gpacks = by_group.get(group_name, [])
        meta = group_meta.get(group_name)
        if not gpacks or not meta:
            continue
        for cfg in gpacks:
            for s in cfg["sections"]:
                rate = s["matched"] / s["total_items"] if s.get("total_items") else 0
                ws.append([
                    group_name, cfg["no"], cfg["label"], s["name"],
                    s.get("matched", ""), s.get("total_items", ""),
                    rate if s.get("total_items") else "",
                    s["mat"], s["lab"], s["exp"], s["sum"],
                ])
            ws.append([
                group_name, cfg["no"], cfg["label"], f"★ {cfg['label']} 소계",
                _xref(meta, "summary", "matched", pack_label=cfg["label"]),
                _xref(meta, "summary", "total", pack_label=cfg["label"]),
                _xref(meta, "summary", "rate", pack_label=cfg["label"]),
                _xref(meta, "summary", "mat", pack_label=cfg["label"]),
                _xref(meta, "summary", "lab", pack_label=cfg["label"]),
                _xref(meta, "summary", "exp", pack_label=cfg["label"]),
                _xref(meta, "summary", "sum", pack_label=cfg["label"]),
            ])
            rr = ws.max_row
            for c in range(1, len(hdr) + 1):
                cell = ws.cell(rr, c)
                cell.font = Font(bold=True)
                cell.fill = SUBTOTAL_FILL

        ws.append([
            group_name, "", group_name, f"★ {group_name} 소계",
            _xref(meta, "summary", "matched"),
            _xref(meta, "summary", "total"),
            _xref(meta, "summary", "rate"),
            _xref(meta, "summary", "mat"),
            _xref(meta, "summary", "lab"),
            _xref(meta, "summary", "exp"),
            _xref(meta, "summary", "sum"),
        ])
        grr = ws.max_row
        group_total_refs.append(meta)
        for c in range(1, len(hdr) + 1):
            cell = ws.cell(grr, c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")

    grand_row = ws.max_row + 1
    ws.append([
        "합계", "", "01·04~07", "★ 전체",
        _sum_refs([_xref(m, "summary", "matched") for m in group_total_refs]),
        _sum_refs([_xref(m, "summary", "total") for m in group_total_refs]),
        None,
        _sum_refs([_xref(m, "summary", "mat") for m in group_total_refs]),
        _sum_refs([_xref(m, "summary", "lab") for m in group_total_refs]),
        _sum_refs([_xref(m, "summary", "exp") for m in group_total_refs]),
        _sum_refs([_xref(m, "summary", "sum") for m in group_total_refs]),
    ])
    grand_row = ws.max_row
    ws.cell(grand_row, 7).value = f"=IF(F{grand_row}=0,0,E{grand_row}/F{grand_row})"
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(grand_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    ws["A1"].font = Font(bold=True, size=14)
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for cell in row:
            if cell.column in (8, 9, 10, 11) and (
                isinstance(cell.value, (int, float))
                or (isinstance(cell.value, str) and cell.value.startswith("="))
            ):
                cell.number_format = MONEY_FMT
            if cell.column == 7 and isinstance(cell.value, float):
                cell.number_format = PCT_FMT
            if cell.column == 7 and isinstance(cell.value, str) and cell.value.startswith("=IF"):
                cell.number_format = PCT_FMT
    autofit(ws)


def _aggregate_cost_rows(entries: list[dict]) -> list[dict]:
    """내역서별 원가계산서 행을 항목명 기준 합산."""
    totals: dict[str, float] = {}
    meta: dict[str, dict] = {}
    order: list[str] = []
    for e in entries:
        for item in e["cs"]["rows"]:
            name = item["name"].strip()
            if name not in totals:
                order.append(name)
                meta[name] = {
                    "step": item.get("step", ""),
                    "bold": item.get("bold", False),
                    "total": item.get("total", False),
                    "indent": item.get("indent", 0),
                    "formula": item.get("formula", ""),
                }
            totals[name] = totals.get(name, 0) + item["amount"]
    return [
        {**meta[name], "name": name, "amount": round(totals[name]), "formula": meta[name]["formula"]}
        for name in order
    ]


def _aggregate_note(item: dict) -> str:
    if item.get("indent", 0):
        return ""
    return "01·04~07 내역서별 합산"


def _sanan_pack_formula(mat_cell: str, lab_cell: str) -> str:
    base = f"({mat_cell}+{lab_cell})"
    return (
        f"IF({base}<500000000,{base}*0.0315,"
        f"IF({base}<5000000000,{base}*0.0253+3300000,{base}*0.026))"
    )


def _ilban_rate_formula(est_cell: str) -> str:
    return (
        f"IF({est_cell}<5000000000,0.08,"
        f"IF({est_cell}<30000000000,0.065,"
        f"IF({est_cell}<100000000000,0.05,0.045)))"
    )


def _iyun_rate_formula(est_cell: str) -> str:
    return (
        f"IF({est_cell}<5000000000,0.15,"
        f"IF({est_cell}<30000000000,0.12,"
        f"IF({est_cell}<100000000000,0.10,0.09)))"
    )


def _pack_overhead_exprs(pack_row: int, r: dict[str, float]) -> dict[str, str]:
    """내역서 1건 — 원가계산 중간값 Excel 식(괄호 포함)."""
    m, l, e, d = f"D{pack_row}", f"E{pack_row}", f"F{pack_row}", f"G{pack_row}"
    iln = f"({l})*{r['간접노무비']}"
    n = f"({l})+({iln})"
    sanjae = f"({n})*{r['산재보험료']}"
    goyong = f"({n})*{r['고용보험료']}"
    gangang = f"({l})*{r['건강보험료']}"
    yeongeum = f"({l})*{r['연금보험료']}"
    nojang = f"({gangang})*{r['노인장기요양']}"
    sanan = _sanan_pack_formula(m, l)
    gita = f"({m})+({n})"
    gita = f"({gita})*{r['기타경비']}"
    gigye = f"({d})*{r['건설기계대여보증']}"
    hwankyung = f"({d})*{r['환경보전비']}"
    sanchul = (
        f"({sanjae})+({goyong})+({gangang})+({yeongeum})+({nojang})+"
        f"({sanan})+({gita})+({gigye})+({hwankyung})"
    )
    gyeongbi = f"({e})+({sanchul})"
    sun = f"({m})+({n})+({gyeongbi})"
    est = f"({d})*1.7"
    ilban = f"({sun})*({_ilban_rate_formula(est)})"
    iyun = f"(({n})+({gyeongbi})+({ilban}))*({_iyun_rate_formula(est)})"
    gonggeup_base = f"({sun})+({ilban})+({iyun})"
    pf_ins = f"({gonggeup_base})*{PF_RATES['이행준공보증보험료']}"
    pf_risk = f"(({d})+({iln}))*{PF_RATES['PF리스크프리미엄']}"
    gonggeup = f"({gonggeup_base})+({pf_ins})+({pf_risk})"
    vat = f"({gonggeup})*{r['부가가치세']}"
    dogeup = f"({gonggeup})+({vat})"
    return {
        "iln": iln,
        "sanjae": sanjae,
        "goyong": goyong,
        "gangang": gangang,
        "yeongeum": yeongeum,
        "nojang": nojang,
        "sanan": sanan,
        "gita": gita,
        "gigye": gigye,
        "hwankyung": hwankyung,
        "gyeongbi": gyeongbi,
        "sun": sun,
        "ilban": ilban,
        "iyun": iyun,
        "gonggeup_base": gonggeup_base,
        "pf_ins": pf_ins,
        "pf_risk": pf_risk,
        "gonggeup": gonggeup,
        "vat": vat,
        "dogeup": dogeup,
    }


def _sum_pack_exprs(pack_rows: list[int], key: str, r: dict[str, float] | None = None) -> str:
    r = r or CIVIL_TOMOK_RATES
    parts = [_pack_overhead_exprs(pr, r)[key] for pr in pack_rows]
    if len(parts) == 1:
        return f"={parts[0]}"
    return "=" + "+".join(f"({p})" for p in parts)


def _aggregate_amount_formula(
    name: str,
    *,
    grand1: int,
    pack_rows: list[int],
    cost_rows: dict[str, int],
    r: dict[str, float] | None = None,
) -> str:
    """【2】 합산표 금액(원) 열 Excel 수식."""
    r = r or CIVIL_TOMOK_RATES
    c = lambda key: f"C{cost_rows[key]}"

    if name == "직접공사비":
        return f"=G{grand1}"
    if name == "· 재료비":
        return f"=D{grand1}"
    if name == "· 직접노무비":
        return f"=E{grand1}"
    if name == "· 직접경비":
        return f"=F{grand1}"
    if name == "간접노무비":
        return _sum_pack_exprs(pack_rows, "iln", r)
    if name == "노무비계(직접+간접)":
        return f"={c('· 직접노무비')}+{c('간접노무비')}"
    if name == "산재보험료":
        return f"={c('노무비계(직접+간접)')}*{r['산재보험료']}"
    if name == "고용보험료":
        return f"={c('노무비계(직접+간접)')}*{r['고용보험료']}"
    if name == "국민건강보험료":
        return f"={c('· 직접노무비')}*{r['건강보험료']}"
    if name == "국민연금보험료":
        return f"={c('· 직접노무비')}*{r['연금보험료']}"
    if name == "노인장기요양보험료":
        return f"={c('국민건강보험료')}*{r['노인장기요양']}"
    if name == "산업안전보건관리비":
        return _sum_pack_exprs(pack_rows, "sanan", r)
    if name == "기타경비":
        return f"=({c('· 재료비')}+{c('노무비계(직접+간접)')})*{r['기타경비']}"
    if name == "건설기계대여대금 지급보증":
        return f"={c('직접공사비')}*{r['건설기계대여보증']}"
    if name == "환경보전비":
        return f"={c('직접공사비')}*{r['환경보전비']}"
    if name == "경비계(직접경비+산출경비)":
        sanchul = "+".join(
            c(n)
            for n in (
                "산재보험료", "고용보험료", "국민건강보험료", "국민연금보험료",
                "노인장기요양보험료", "산업안전보건관리비", "기타경비",
                "건설기계대여대금 지급보증", "환경보전비",
            )
        )
        return f"={c('· 직접경비')}+{sanchul}"
    if name == "순공사원가":
        return f"={c('· 재료비')}+{c('노무비계(직접+간접)')}+{c('경비계(직접경비+산출경비)')}"
    if name == "일반관리비":
        return _sum_pack_exprs(pack_rows, "ilban", r)
    if name == "이윤":
        return _sum_pack_exprs(pack_rows, "iyun", r)
    if name == "이행(준공)보증보험료":
        return _sum_pack_exprs(pack_rows, "pf_ins", r)
    if name == "PF·신용 리스크 프리미엄":
        return _sum_pack_exprs(pack_rows, "pf_risk", r)
    if name == "공급가액":
        pf_part = ""
        if "이행(준공)보증보험료" in cost_rows:
            pf_part = f"+{c('이행(준공)보증보험료')}+{c('PF·신용 리스크 프리미엄')}"
        return f"={c('순공사원가')}+{c('일반관리비')}+{c('이윤')}{pf_part}"
    if name == "부가가치세":
        return f"={c('공급가액')}*{r['부가가치세']}"
    if name == "도급액(총공사비)":
        return f"={c('공급가액')}+{c('부가가치세')}"
    return ""


# 【1】 직접비 — 재·노·경 가로(열) / 간접비 — 세로(행)
_INDIRECT_ITEMS = [
    ("간접노무비", "indirect_labor"),
    ("순공사원가", "net_cost"),
    ("일반관리비", "general_admin"),
    ("이윤", "profit"),
    ("부가세", "vat"),
    ("도급액", "contract"),
]


def _direct_note(cfg: dict) -> str:
    if cfg.get("jogyeong_ref"):
        return "푸른조경 내역(순공사비)·토목요율"
    return ""


def write_total_cost_sheet(ws, packs: list[dict], group_meta: dict[str, dict]) -> None:
    """맨 앞 시트 — 내역서별 요약(재·노·경 가로, 간접비 세로)."""
    ws.append(["화성 청원지구 — 총공사비 (도급액·간접비 포함)"])
    ws.append([])

    ws.append(["【1. 내역서별 요약】"])
    ws.append(["① 직접공사비 (재·노·경)"])
    hdr1 = ["구분", "No", "내역서", "재료비", "노무비", "경비", "직접공사비", "비고"]
    ws.append(hdr1)
    hdr1_row = ws.max_row
    style_header_row(ws, hdr1_row, len(hdr1))

    by_group: dict[str, list[dict]] = {}
    ordered: list[tuple[dict, dict]] = []
    for cfg in packs:
        by_group.setdefault(cfg["group"], []).append(cfg)
    for group_name in GROUP_ORDER:
        meta = group_meta.get(group_name)
        if not meta:
            continue
        for cfg in by_group.get(group_name, []):
            ordered.append((cfg, meta))

    data_start = hdr1_row + 1
    pack_direct_rows: list[int] = []
    pack_costs: list[dict] = []

    for cfg, meta in ordered:
        cs = _compute_cost(cfg["totals"], cfg["engine"])
        pack_costs.append(cs)
        t = cfg["totals"]
        mat, lab, exp = round(t["mat"]), round(t["lab"]), round(t["exp"])
        direct = mat + lab + exp
        ws.append([
            cfg["group"], cfg["no"], cfg["label"],
            mat, lab, exp, direct,
            _direct_note(cfg),
        ])
        rr = ws.max_row
        pack_direct_rows.append(rr)

    data_end = ws.max_row
    grand1 = data_end + 1
    sum_mat = sum(round(cfg["totals"]["mat"]) for cfg, _ in ordered)
    sum_lab = sum(round(cfg["totals"]["lab"]) for cfg, _ in ordered)
    sum_exp = sum(round(cfg["totals"]["exp"]) for cfg, _ in ordered)
    sum_direct = sum_mat + sum_lab + sum_exp
    ws.append(
        ["합계", "", "★ 01·04~07", sum_mat, sum_lab, sum_exp, sum_direct, "구분 시트 ★소계 합"]
    )
    for c in range(1, len(hdr1) + 1):
        cell = ws.cell(grand1, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    for row in ws.iter_rows(min_row=data_start, max_row=grand1):
        for cell in row:
            if cell.column in range(4, 8) and (
                isinstance(cell.value, (int, float))
                or (isinstance(cell.value, str) and str(cell.value).startswith("="))
            ):
                cell.number_format = MONEY_FMT

    ws.append([])
    ws.append(["② 간접비·도급액"])
    pack_labels = [cfg["label"] for cfg, _ in ordered]
    hdr2 = ["항목"] + pack_labels + ["합계"]
    ws.append(hdr2)
    hdr2_row = ws.max_row
    style_header_row(ws, hdr2_row, len(hdr2))
    pack_cols = list(range(2, 2 + len(ordered)))
    sum_col = 2 + len(ordered)

    indirect_row_nums: dict[str, int] = {}
    for label, key in _INDIRECT_ITEMS:
        ws.append([label] + [None] * (len(hdr2) - 1))
        rr = ws.max_row
        indirect_row_nums[label] = rr
        for i, cs in enumerate(pack_costs):
            ws.cell(rr, pack_cols[i]).value = cs[key]
        pack_letters = [get_column_letter(c) for c in pack_cols]
        ws.cell(rr, sum_col).value = "=" + "+".join(f"{L}{rr}" for L in pack_letters)
        if label == "도급액":
            for c in range(1, sum_col + 1):
                cell = ws.cell(rr, c)
                cell.font = Font(bold=True)
                cell.fill = TOTAL_FILL

    ws.append(["배율"] + [None] * (len(hdr2) - 1))
    mult_row = ws.max_row
    contract_r = indirect_row_nums["도급액"]
    for i, direct_r in enumerate(pack_direct_rows):
        col = pack_cols[i]
        letter = get_column_letter(col)
        contract_cell = f"{letter}{contract_r}"
        direct_cell = f"G{direct_r}"
        ws.cell(mult_row, col).value = f"=IF({direct_cell}=0,0,{contract_cell}/{direct_cell})"
    sum_letter = get_column_letter(sum_col)
    ws.cell(mult_row, sum_col).value = (
        f"=IF(G{grand1}=0,0,{sum_letter}{contract_r}/G{grand1})"
    )

    for row in ws.iter_rows(min_row=hdr2_row + 1, max_row=mult_row):
        for cell in row:
            if cell.column in pack_cols + [sum_col] and (
                isinstance(cell.value, (int, float))
                or (isinstance(cell.value, str) and str(cell.value).startswith("="))
            ):
                if cell.row == mult_row or ws.cell(cell.row, 1).value == "배율":
                    cell.number_format = "0.000"
                else:
                    cell.number_format = MONEY_FMT

    entries = [{**cfg, "cs": cs} for (cfg, _), cs in zip(ordered, pack_costs)]
    ws.append([])
    ws.append(["【2. 간접비·원가계산 (내역서별 합산)】"])
    ws.append(["단계", "항목", "금액(원)", "산식", None, "비고", None])
    calc_hdr = ws.max_row
    ws.merge_cells(start_row=calc_hdr, start_column=4, end_row=calc_hdr, end_column=5)
    ws.merge_cells(start_row=calc_hdr, start_column=6, end_row=calc_hdr, end_column=7)
    style_header_row(ws, calc_hdr, 7)
    agg_direct = sum(cs["direct"] for cs in pack_costs)
    agg_contract = sum(cs["contract"] for cs in pack_costs)
    agg_mult = agg_contract / agg_direct if agg_direct else 0
    cost_rows: dict[str, int] = {}
    for item in _aggregate_cost_rows(entries):
        name = ("  " * item["indent"]) + item["name"]
        ws.append([item["step"], name, None, None, None, None, None, None])
        rr = ws.max_row
        cost_rows[item["name"]] = rr
        ws.cell(rr, 3).value = item["amount"]
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=5)
        formula = item["formula"]
        if item["name"] == "도급액(총공사비)":
            formula = f"공급가액 + 부가세 (직접비의 {agg_mult:.3f}배·통합)"
        ws.cell(rr, 4).value = formula
        ws.cell(rr, 4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=7)
        ws.cell(rr, 6).value = _aggregate_note(item)
        ws.cell(rr, 6).alignment = Alignment(wrap_text=True, vertical="top")
        fill = TOTAL_FILL if item.get("total") else (SUBTOTAL_FILL if item.get("bold") else None)
        _style_cost_row(ws, rr, 7, bold=item.get("bold"), fill=fill)

    ws["A1"].font = Font(bold=True, size=14)
    autofit(ws)


def _refresh_ka_total_cost(packs: list[dict], group_meta: dict[str, dict]) -> None:
    """_ka.xlsx — 총공사비 시트만 재작성(◆ 내역 일괄 등 패치 유지)."""
    if not OUT_KA.exists():
        return
    wb = load_workbook(OUT_KA)
    if "총공사비" in wb.sheetnames:
        del wb["총공사비"]
    ws = wb.create_sheet("총공사비", 0)
    write_total_cost_sheet(ws, packs, group_meta)
    _apply_row_heights(wb)
    _optimize_workbook_columns(wb)
    try:
        wb.save(OUT_KA)
    except PermissionError:
        alt = OUT_KA.with_name(OUT_KA.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"  (주의) _ka 저장 실패 — {alt.name} 로 저장", file=sys.stderr)


def build_unified() -> Path:
    packs = load_all()
    by_group: dict[str, list[dict]] = {}
    for p in packs:
        by_group.setdefault(p["group"], []).append(p)

    wb = Workbook()
    wb.remove(wb.active)

    used: set[str] = set()
    group_meta: dict[str, dict] = {}
    for group_name in GROUP_ORDER:
        gpacks = by_group.get(group_name, [])
        if not gpacks:
            continue
        ws_g = wb.create_sheet(sanitize_sheet(group_name, used))
        group_meta[group_name] = write_group_sheet(ws_g, group_name, gpacks)

    ws_total = wb.create_sheet("총공사비", 0)
    write_total_cost_sheet(ws_total, packs, group_meta)

    ws0 = wb.create_sheet(SHEET_NAERYEO, 1)
    write_unified_summary(ws0, packs, group_meta)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _apply_row_heights(wb)
    _optimize_workbook_columns(wb)
    try:
        wb.save(OUT_UNIFIED)
        out = OUT_UNIFIED
    except PermissionError:
        out = OUT_UNIFIED.with_name(OUT_UNIFIED.stem + "_업데이트.xlsx")
        _apply_row_heights(wb)
        _optimize_workbook_columns(wb)
        wb.save(out)

    # 구 통합본(공내역서 폴더) 정리
    legacy_dir = BASE / "공내역서"
    for old in legacy_dir.glob("화성_청원지구_단가반영*.xlsx"):
        try:
            old.unlink()
        except OSError:
            pass
    for old in legacy_dir.glob("*_단가반영.xlsx"):
        try:
            old.unlink()
        except OSError:
            pass
    _refresh_ka_total_cost(packs, group_meta)
    return out


def main() -> None:
    path = build_unified()
    wb = load_workbook(path, read_only=True)
    print(f"저장: {path}")
    print(f"시트 {len(wb.sheetnames)}개 — {', '.join(wb.sheetnames)}")
    wb.close()


if __name__ == "__main__":
    main()
