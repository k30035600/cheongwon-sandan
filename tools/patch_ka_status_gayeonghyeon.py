#!/usr/bin/env python3
"""_ka ◆ 내역 일괄 — 검토·미매칭 중 합계(단가·금액) 있는 행 → 상태 「가영현」."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from naeyeok_gongjong import (  # noqa: E402
    GAYEONGHYEON_STATUS,
    detail_row_has_sum,
    is_ka_status_excluded,
    normalize_ka_detail_status,
)

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
KA = ROOT / "08_제출내역서" / "청원지구_단가통합(전기제외)_ka.xlsx"
SHEETS = ("회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물")
FROM_STATUSES = frozenset({"검토", "미매칭", "미산출"})


def _detail_hdr_row(ws) -> int | None:
    for r in range(1, min(ws.max_row + 1, 800)):
        if str(ws.cell(r, 4).value or "").strip() == "공종명":
            return r
    return None


def patch(path: Path = KA) -> tuple[int, int]:
    wb = load_workbook(path, data_only=False)
    to_gyh = cleared = 0
    for sn in SHEETS:
        ws = wb[sn]
        hdr = _detail_hdr_row(ws)
        if hdr is None:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            if not ws.cell(r, 3).value:
                continue
            name = ws.cell(r, 4).value
            if not name or str(name).startswith("★"):
                continue
            st = str(ws.cell(r, 16).value or "").strip()
            if st == "매칭":
                ws.cell(r, 16).value = normalize_ka_detail_status(st)
                cleared += 1
                continue
            if st not in FROM_STATUSES:
                continue
            has_sum = detail_row_has_sum(
                mat_u=ws.cell(r, 8).value,
                lab_u=ws.cell(r, 10).value,
                exp_u=ws.cell(r, 12).value,
                mat_amt=ws.cell(r, 9).value,
                lab_amt=ws.cell(r, 11).value,
                exp_amt=ws.cell(r, 13).value,
                unit_sum=ws.cell(r, 14).value if not isinstance(ws.cell(r, 14).value, str) else None,
                total_amt=ws.cell(r, 15).value if not isinstance(ws.cell(r, 15).value, str) else None,
            )
            if not has_sum:
                continue
            ws.cell(r, 16).value = GAYEONGHYEON_STATUS
            to_gyh += 1
    wb.save(path)
    wb.close()
    return cleared, to_gyh


if __name__ == "__main__":
    n_clear, n_gyh = patch()
    print(f"{KA.name}: 매칭→빈칸 {n_clear}건 · 검토·미매칭→{GAYEONGHYEON_STATUS} {n_gyh}건")
