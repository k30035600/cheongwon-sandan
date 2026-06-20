#!/usr/bin/env python3
"""한국물가협회 2018 PDF(스캔본) OCR → 현재 미매칭 품목 해소 가능성 점검(수정 없음).

출력: 05_내역서/한국물가협회_미매칭점검.xlsx
캐시: tools/_kpa_ocr_cache.jsonl (재실행 시 OCR 생략)
"""
from __future__ import annotations
import json, re, sys, time, warnings
from pathlib import Path

import fitz
import easyocr
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
KPA_DIR = BASE / "일위대가DB" / "한국물가협회"
CACHE = TOOLS / "_kpa_ocr_cache.jsonl"
OUT = BASE / "한국물가협회_미매칭점검.xlsx"

sys.path.insert(0, str(TOOLS))
import apply_standard_prices as asp  # noqa: E402

HDR = PatternFill("solid", fgColor="D9E1F2")
OKF = PatternFill("solid", fgColor="E2EFDA")
RVF = PatternFill("solid", fgColor="FFF2CC")
NMF = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)

LABELS = {
    "01": "01 토목", "01j": "01 조경", "02": "02 전기", "04": "04 진입도로",
    "05": "05 회전교차로", "06": "06 개발행위", "07": "07 건설폐기물",
}

PRICE_RE = re.compile(r"(\d{1,3}(?:,\d{3})+|\d{4,})")
UNIT_HINTS = ("ton", "TON", "M/T", "㎏", "kg", "KG", "m2", "㎡", "m", "M", "L", "ℓ",
              "개", "본", "EA", "식", "대", "주", "톤", "km", "hr", "시간")


def newest_xlsx(stem_prefix: str) -> Path | None:
    cands = list(BASE.glob(f"{stem_prefix}*_표준단가산출.xlsx"))
    cands += list(WORK.glob(f"{stem_prefix}*_표준단가산출.xlsx"))
    return max(cands, key=lambda p: p.stat().st_mtime) if cands else None


def load_unmatched() -> list[dict]:
    items: list[dict] = []
    for no in ("01", "02", "04", "05", "06", "07"):
        p = newest_xlsx(no)
        if not p:
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        for sname in ("미매칭", "미산출"):
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hdr = [("" if h is None else str(h).strip()) for h in rows[0]]
            idx = {h: i for i, h in enumerate(hdr)}
            def col(r, *names):
                for n in names:
                    for k, j in idx.items():
                        if n in k and j < len(r):
                            return r[j]
                return None
            for r in rows[1:]:
                if not any(c is not None and str(c).strip() for c in r):
                    continue
                name = str(col(r, "공종명", "명칭") or "").strip()
                spec = str(col(r, "규격") or "").strip()
                unit = str(col(r, "단위") or "").strip()
                if not unit or unit == "식":
                    continue
                qty = col(r, "수량")
                try:
                    qty = float(qty) if qty is not None else 0
                except (TypeError, ValueError):
                    qty = 0
                items.append({
                    "ledger": LABELS.get(no, no),
                    "row": col(r, "행"),
                    "section": str(col(r, "공종") or ""),
                    "name": name, "spec": spec, "unit": unit, "qty": qty,
                    "score_before": col(r, "최고점수", "매칭점수"),
                })
        wb.close()
    return items


def load_cache() -> dict[str, list[str]]:
    data: dict[str, list[str]] = {}
    if not CACHE.exists():
        return data
    with CACHE.open(encoding="utf-8") as f:
        for line in f:
            rec = json.loads(line)
            data[rec["key"]] = rec["lines"]
    return data


def save_cache_line(key: str, lines: list[str]):
    with CACHE.open("a", encoding="utf-8") as f:
        f.write(json.dumps({"key": key, "lines": lines}, ensure_ascii=False) + "\n")


def pixmap_rgb(pix) -> np.ndarray:
    arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    return arr[:, :, :3] if pix.n == 4 else arr


def ocr_pdfs(reader) -> list[dict]:
    """페이지별 OCR 텍스트 → 단가 후보 파싱."""
    warnings.filterwarnings("ignore", category=UserWarning)
    cache = load_cache()
    pages_meta: list[dict] = []
    pdfs = sorted(KPA_DIR.glob("*.pdf"))
    print(f"PDF {len(pdfs)}개 OCR 시작 (캐시 {len(cache)}페이지)…")
    done = 0
    for pdf in pdfs:
        doc = fitz.open(str(pdf))
        n = doc.page_count
        for pi in range(n):
            key = f"{pdf.name}::{pi}"
            if key in cache:
                lines = cache[key]
            else:
                pix = doc[pi].get_pixmap(matrix=fitz.Matrix(100 / 72, 100 / 72))
                result = reader.readtext(pixmap_rgb(pix), detail=0, paragraph=True)
                lines = [ln.strip() for ln in result if ln and ln.strip()]
                save_cache_line(key, lines)
                cache[key] = lines
            pages_meta.append({"src": pdf.name, "page": pi + 1, "lines": lines})
            done += 1
            if done % 50 == 0:
                print(f"  … {done}페이지 OCR 완료", flush=True)
        doc.close()
        print(f"  완료 {pdf.name} ({n}p)", flush=True)
    return pages_meta


def parse_prices_from_pages(pages_meta: list[dict]) -> list[dict]:
    """OCR 줄에서 (품명·규격·단위·가격) 후보 추출 — 휴리스틱."""
    prices: list[dict] = []
    seen: set[str] = set()
    for pg in pages_meta:
        lines = pg["lines"]
        blob = " ".join(lines)
        for i, line in enumerate(lines):
            m = PRICE_RE.search(line.replace(" ", ""))
            if not m:
                continue
            try:
                val = float(m.group(1).replace(",", ""))
            except ValueError:
                continue
            if val < 100 or val > 500_000_000:
                continue
            # 품명: 가격 줄 + 앞 2줄 결합
            ctx = " ".join(lines[max(0, i - 2): i + 1])
            if len(ctx) < 4:
                continue
            unit = ""
            for u in UNIT_HINTS:
                if u in ctx:
                    unit = u.replace("M/T", "ton").replace("㎏", "kg").replace("㎡", "m2").replace("ℓ", "L")
                    break
            if not unit:
                unit = "식"
            key = ctx[:80] + "|" + str(int(val))
            if key in seen:
                continue
            seen.add(key)
            prices.append({
                "code": f"{pg['src']} p{pg['page']}",
                "name": ctx[:120],
                "spec": "",
                "unit": unit,
                "mat": val, "lab": 0.0, "exp": 0.0, "total": val,
                "date": "한국물가협회2018",
                "_src": pg["src"],
            })
    return prices


def write_xlsx(items: list[dict], results: list[dict], price_count: int):
    WORK.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "점검결과"
    ws.append(["한국물가협회 2018 PDF OCR 기반 — 미매칭 해소 가능성 점검(기존 산출 미수정)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"단가 후보(OCR 추출): {price_count:,}건 / 미매칭 대상: {len(items):,}건"])
    n_ok = sum(1 for r in results if r["status"] == "매칭")
    n_rv = sum(1 for r in results if r["status"] == "검토")
    n_no = sum(1 for r in results if r["status"] == "미매칭")
    ws.append([f"신규 매칭(≥{asp.REVIEW_THRESHOLD}) {n_ok} / 검토({asp.THRESHOLD}~{asp.REVIEW_THRESHOLD}) {n_rv} / 여전히 미매칭 {n_no}"])
    cols = ["내역서", "행", "공종", "명칭", "규격", "단위", "수량", "기존최고점수",
            "점수", "상태", "PDF출처", "매칭텍스트", "단가(원)", "추정금액"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=4, column=c).fill = HDR
        ws.cell(row=4, column=c).font = BOLD
    for it, r in zip(items, results):
        ws.append([
            it["ledger"], it["row"], it["section"], it["name"], it["spec"], it["unit"], it["qty"],
            it.get("score_before"), r["score"], r["status"], r.get("src", ""),
            r.get("match_name", ""), r.get("price", ""), r.get("amt", ""),
        ])
        fill = OKF if r["status"] == "매칭" else RVF if r["status"] == "검토" else NMF
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).fill = fill
    ws.freeze_panes = "A5"
    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2.append(["구분", "건수"])
    ws2.append(["미매칭 대상", len(items)])
    ws2.append(["OCR 단가 후보", price_count])
    ws2.append(["신규 매칭(높음)", n_ok])
    ws2.append(["검토(점수 중간)", n_rv])
    ws2.append(["여전히 미매칭", n_no])
    ws2.append(["해소 가능 합계", n_ok + n_rv])
    ws2.append([])
    ws2.append(["내역서", "대상", "매칭", "검토", "미매칭"])
    from collections import defaultdict
    buck = defaultdict(lambda: {"t": 0, "ok": 0, "rv": 0, "no": 0})
    for it, r in zip(items, results):
        b = buck[it["ledger"]]
        b["t"] += 1
        if r["status"] == "매칭":
            b["ok"] += 1
        elif r["status"] == "검토":
            b["rv"] += 1
        else:
            b["no"] += 1
    for lab in sorted(buck):
        b = buck[lab]
        ws2.append([lab, b["t"], b["ok"], b["rv"], b["no"]])
    ws2.append([])
    ws2.append(["참고", "2018년 스캔 PDF OCR·휴리스틱 파싱. 기존 산출·총괄표 미반영."])
    wb.save(OUT)
    print(f"\n저장: {OUT}")
    print(f"매칭 {n_ok} / 검토 {n_rv} / 미매칭 {n_no} (해소 가능 {n_ok + n_rv}건, 단 OCR 품질·2018 시세 한계)")


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--match-only", action="store_true", help="OCR 캐시만 사용(재OCR 생략)")
    args = ap.parse_args()
    items = load_unmatched()
    print(f"미매칭 대상 {len(items):,}건")
    if not items:
        print("미매칭 없음"); return
    if args.match_only:
        cache = load_cache()
        pages = []
        for key, lines in cache.items():
            src, pi = key.rsplit("::", 1)
            pages.append({"src": src, "page": int(pi) + 1, "lines": lines})
        print(f"캐시 {len(pages)}페이지 로드")
    else:
        reader = easyocr.Reader(["ko", "en"], gpu=False, verbose=False)
        pages = ocr_pdfs(reader)
    raw_prices = parse_prices_from_pages(pages)
    prices = asp.precompute(raw_prices)
    print(f"OCR 단가 후보 {len(raw_prices):,}건 → precompute {len(prices):,}건")
    results = []
    for it in items:
        price, score, _, _ = asp.find_best_match(it, prices)
        if not price or score < asp.THRESHOLD:
            results.append({"status": "미매칭", "score": round(score, 3) if score >= 0 else None})
            continue
        amts = asp.calc_amounts(it, price)
        status = "매칭" if score >= asp.REVIEW_THRESHOLD else "검토"
        results.append({
            "status": status,
            "score": round(score, 3),
            "src": price.get("_src", price.get("code", "")),
            "match_name": price["name"][:80],
            "price": int(price["total"]),
            "amt": round(amts["sum_amt"]),
        })
    write_xlsx(items, results, len(raw_prices))


if __name__ == "__main__":
    main()
