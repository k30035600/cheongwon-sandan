#!/usr/bin/env python3
"""확정단가 출처 팩트체크 — 확정==제시(자동복사 의심) vs 확정!=제시(수동조정) 분류."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"

def analyze(path: Path, sheet="일위대가산출"):
    print("=" * 78)
    print(path.name)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c or "").strip() for c in rows[0]]
    H = {h: i for i, h in enumerate(hdr)}
    ci_conf = H.get("확정단가(입력)")
    ci_sug = H.get("제시단가")
    ci_tot = H.get("합계단가")
    ci_sc = H.get("최고점수", H.get("매칭점수"))
    has_sug = ci_sug is not None
    n_conf = n_eq_sug = n_ne_sug = n_eq_tot = 0
    low_conf = 0
    ne_examples = []
    for r in rows[1:]:
        if ci_conf is None or ci_conf >= len(r):
            continue
        conf = r[ci_conf]
        if conf in (None, "", 0):
            continue
        n_conf += 1
        try:
            confv = float(conf)
        except (TypeError, ValueError):
            continue
        sug = r[ci_sug] if has_sug and ci_sug < len(r) else None
        tot = r[ci_tot] if ci_tot is not None and ci_tot < len(r) else None
        try:
            sugv = float(sug) if sug not in (None, "") else None
        except (TypeError, ValueError):
            sugv = None
        try:
            totv = float(tot) if tot not in (None, "") else None
        except (TypeError, ValueError):
            totv = None
        if sugv is not None:
            if abs(confv - sugv) < 0.5:
                n_eq_sug += 1
            else:
                n_ne_sug += 1
                if len(ne_examples) < 8:
                    nm = r[H.get("품명", 4)] if H.get("품명") is not None else ""
                    ne_examples.append((str(nm)[:22], confv, sugv))
        if totv is not None and abs(confv - totv) < 0.5:
            n_eq_tot += 1
        sc = r[ci_sc] if ci_sc is not None and ci_sc < len(r) else None
        try:
            if float(sc) < 0.56:
                low_conf += 1
        except (TypeError, ValueError):
            pass
    print(f"  제시단가 열 존재: {has_sug}")
    print(f"  확정 채워진 행: {n_conf}")
    if has_sug:
        print(f"   - 확정==제시(자동복사 의심): {n_eq_sug}")
        print(f"   - 확정!=제시(수동조정 가능): {n_ne_sug}")
        for nm, c, s in ne_examples:
            print(f"       · 「{nm}」 확정{c:,.0f} ≠ 제시{s:,.0f}")
    print(f"   - 확정==합계단가(DB자동값 그대로): {n_eq_tot}")
    print(f"   - 점수<0.56(저신뢰) 확정: {low_conf}")
    wb.close()

analyze(BASE / "미매칭_일위대가산출.xlsx")
analyze(BASE / "검토_일위대가산출.xlsx")
