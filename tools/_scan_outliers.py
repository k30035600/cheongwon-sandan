#!/usr/bin/env python3
"""일위확정 반영 항목 중 단가/금액 이상치 스캔(읽기 전용).
저신뢰 매칭(미매칭 auto-suggest)에서 굳어진 과대 단가를 식별."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
FILE_MAP = {
    "01 토목": "01_화성 청원지구 토목_표준단가산출.xlsx",
    "01 조경": "01_화성 청원지구 조경_표준단가산출.xlsx",
    "04 진입도로": "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
    "05 회전교차로": "05_화성 청원로(회전교차로)_표준단가산출.xlsx",
    "06 개발행위": "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
}
# 단위별 '상식적 상한' 단가(원) — 초과 시 의심
UNIT_CAP = {
    "개소": 2_000_000, "개": 2_000_000, "EA": 2_000_000, "ea": 2_000_000,
    "본": 500_000, "주": 1_000_000, "m": 500_000, "ｍ": 500_000,
    "m2": 300_000, "㎡": 300_000, "m3": 500_000, "㎥": 500_000,
    "매": 200_000, "kg": 50_000, "회": 5_000_000, "대": 5_000_000,
}
print("의심 일위확정 단가(단위별 상한 초과):")
print("=" * 90)
flagged = 0
for label, fname in FILE_MAP.items():
    p = WORK / fname
    if not p.exists():
        continue
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb["통합내역"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c or "").strip() for c in rows[0]]
    H = {h: i for i, h in enumerate(hdr)}
    for r in rows[1:]:
        if r[H["행"]] is None:
            continue
        code = str(r[H["단가코드"]] or "")
        if code != "일위확정":
            continue
        unit = str(r[H["단위"]] or "").strip()
        try:
            tot = float(r[H["합계단가"]] or 0)
            amt = float(r[H["합계금액"]] or 0)
        except (TypeError, ValueError):
            continue
        cap = UNIT_CAP.get(unit)
        if cap and tot > cap:
            flagged += 1
            print(f"[{label}] 행{r[H['행']]} 「{str(r[H['공종명']])[:22]}」 "
                  f"{str(r[H['규격']])[:14]} {unit} ×{r[H['수량']]}  "
                  f"단가 {tot:,.0f}  금액 {amt:,.0f}")
    wb.close()
print("=" * 90)
print(f"의심 항목 {flagged}건")
