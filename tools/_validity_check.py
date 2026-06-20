#!/usr/bin/env python3
"""미매칭_일위대가 자동확정 항목 전수 품목검증.
내역 품명 ↔ DB 매칭품명(DB_품명) 토큰 일치 여부로 오매칭 색출."""
from __future__ import annotations
import re
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
P = BASE / "미매칭_일위대가산출.xlsx"


def core(name: str) -> str:
    s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", str(name or ""))
    return re.sub(r"^[.·\s]+", "", s).strip()


def tokens(name: str) -> list[str]:
    return [t for t in re.split(r"[\s/()·,]+", core(name)) if len(t) >= 2]


def overlap(item: str, db: str) -> int:
    """내역품명 토큰 중 DB품명에 등장하는 개수."""
    if not db:
        return 0
    blob = str(db).replace(" ", "")
    return sum(1 for t in tokens(item) if t in blob or t in str(db))


wb = load_workbook(P, read_only=True, data_only=True)
ws = wb["일위대가산출"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c or "").strip() for c in rows[0]]
H = {h: i for i, h in enumerate(hdr)}
ci = {k: H.get(k) for k in ("파일", "품명", "규격", "단위", "수량", "최고점수",
                            "DB_품명", "DB_규격", "확정단가(입력)", "합계단가", "표준품셈·산출근거")}

flagged, ok = [], 0
for r in rows[1:]:
    if ci["품명"] is None or ci["품명"] >= len(r):
        continue
    name = r[ci["품명"]]
    conf = r[ci["확정단가(입력)"]] if ci["확정단가(입력)"] is not None else None
    if not name or conf in (None, "", 0):
        continue
    dbn = r[ci["DB_품명"]] if ci["DB_품명"] is not None else ""
    ov = overlap(name, dbn)
    basis = str(r[ci["표준품셈·산출근거"]] or "") if ci["표준품셈·산출근거"] is not None else ""
    corrected = "오매칭" in basis  # fix_outlier_prices 가 남긴 표식
    try:
        confv = float(conf)
        qty = float(r[ci["수량"]] or 0)
    except (TypeError, ValueError):
        confv, qty = 0, 0
    rec = (confv * qty, str(r[ci["파일"]]), str(name)[:22], str(r[ci["규격"]])[:14],
           r[ci["단위"]], qty, confv, str(dbn)[:24], ov, corrected)
    if ov == 0 and not corrected:
        flagged.append(rec)
    else:
        ok += 1

print(f"자동확정 검증 — 품목일치 의심없음 {ok}건 / 오매칭 의심(토큰0·미교정) {len(flagged)}건")
print("=" * 100)
print("오매칭 의심(내역품명 ↔ DB품명 공통 토큰 0):")
for amt, fl, nm, sp, u, qty, cv, dbn, ov, corr in sorted(flagged, reverse=True):
    print(f"  [{fl}] 「{nm}」 {sp} {u} ×{qty:,.0f}  단가{cv:,.0f}  금액{amt:,.0f}  ← DB「{dbn}」")
wb.close()
