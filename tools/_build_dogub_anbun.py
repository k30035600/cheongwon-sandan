# -*- coding: utf-8 -*-
"""도급공사비 안분내역서(공동수급체 분담이행방식) 생성.

기준: 직접공사비 안분내역서(공동수급체 분담이행방식).xlsx
변경: 공종별 금액(D5:D20)을 직접공사비 → 도급공사비(직접+간접+부가세)로 치환.
      도급공사비는 각 블록의 도급액(VAT 포함)에 맞춰 직접공사비 비율로 안분(반올림).
모든 수식(업종별 SUMIF, 구성원별 합계/비율, F=D/H=D)은 그대로 유지.
"""
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

SRC = Path(r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
           r"\공사비 안분내역서(공동수급체 분담이행방식).xlsx")
DST = Path(r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
           r"\도급공사비 안분내역서(공동수급체 분담이행방식).xlsx")

# 블록별 (직접공사비 계, 도급액 = 직접+간접+부가세)  ※ split_groups(토목만) 기준
BLOCKS = {
    "회전진입": {"rows": range(5, 11),  "direct": 1_794_475_779, "dogub": 2_505_994_618},
    "토목":     {"rows": range(11, 16), "direct": 4_037_192_417, "dogub": 5_550_771_985},
    "지구단위": {"rows": range(16, 21), "direct": 2_522_148_203, "dogub": 3_362_111_980},
}


def R(x):
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    ws = wb["직접공사비"]
    ws.title = "도급공사비"

    # 제목 · 헤더 문구 갱신
    ws["A1"] = "도급공사비 안분내역서 (공동수급체 분담이행방식)"
    ws["D3"] = "금액\n(도급공사비)"

    total_dogub = 0
    for bname, info in BLOCKS.items():
        rows = list(info["rows"])
        ratio = info["dogub"] / info["direct"]
        scaled = {}
        for r in rows:
            d = ws.cell(r, 4).value  # 직접공사비
            scaled[r] = R(d * ratio)
        # 잔차는 블록 첫 행(토공)에서 흡수하여 블록 도급액에 정확히 정합
        resid = info["dogub"] - sum(scaled.values())
        scaled[rows[0]] += resid
        for r in rows:
            ws.cell(r, 4).value = scaled[r]
        total_dogub += info["dogub"]
        print(f"{bname}: 도급 {info['dogub']:,} (배분합 {sum(scaled.values()):,})")

    # 안내 문구(표 하단)
    note = ("※ 도급공사비 = 직접공사비 + 간접공사비 + 부가가치세. "
            "공종별 금액은 각 블록 도급액(VAT 포함)을 직접공사비 비율로 안분(반올림)한 값이다. "
            "구성원별 배분(근일건설·삼아건설)은 면허·시공역량에 따른 것으로, "
            "세부 물량·차수별 계약에서 확정한다.")
    ncell = ws.cell(29, 1)
    ncell.value = note
    ws.merge_cells("A29:I29")
    from openpyxl.styles import Alignment, Font
    ncell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ncell.font = Font(size=9, color="555555")
    ws.row_dimensions[29].height = 44

    wb.save(DST)
    print(f"\n저장: {DST}")
    print(f"도급공사비 합계(토목만): {total_dogub:,}")


if __name__ == "__main__":
    main()
