#!/usr/bin/env python3
"""_ka.xlsx → 공종별 직접·간접비 HTML · 미확정(미매칭) HTML (포털 SSOT)."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_gongjong_cost_html import (  # noqa: E402
    CSS,
    render_cost_table,
    render_direct_table,
    section_level,
    section_sort_key,
    won,
)
from calc_overhead import CIVIL_JOGYEONG_RATES, CIVIL_TOMOK_RATES, compute_cost_statement_civil  # noqa: E402
from jogyeong_crosscheck_poomsem import render_jogyeong_html, sync_ilwidae_xlsx  # noqa: E402
from naeyeok_gongjong import (  # noqa: E402
    classify_ka_mihwakjeong,
    detail_row_has_sum,
    is_jogyeong_crosscheck_item,
    is_ka_pending_excluded,
)

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "05_내역서" / "내역서작업"
COMMON = WORK / "_공통"
KA = ROOT / "08_제출내역서" / "청원지구_단가통합(전기제외)_ka.xlsx"
KA_FILE = "08_제출내역서/청원지구_단가통합(전기제외)_ka.xlsx"
_t = date.today()
GEN_DATE = f"{_t.year}. {_t.month}. {_t.day}."

PHASES: list[dict] = [
    {
        "slug": "회전교차로", "title": "05 회전교차로", "sheet": "회전교차로",
        "star": "★ 회전교차로 합계", "dogeup_col": 2,
        "rates": CIVIL_TOMOK_RATES, "rate_label": "토목",
        "src": "05_화성 청원로(회전교차로).XLS", "file_filter": None,
    },
    {
        "slug": "진입도로", "title": "04 진입도로", "sheet": "진입도로",
        "star": "★ 진입도로 합계", "dogeup_col": 3,
        "rates": CIVIL_TOMOK_RATES, "rate_label": "토목",
        "src": "04_화성 청원지구 진입도로 실시설계.XLS", "file_filter": None,
    },
    {
        "slug": "토목", "title": "01 토목", "sheet": "토목(조경)",
        "star": "★ 01 토목 소계", "dogeup_col": 4,
        "rates": CIVIL_TOMOK_RATES, "rate_label": "토목",
        "src": "01_화성 청원지구 토목.XLS", "file_filter": "01 토목",
        "out_dir": "토목",
    },
    {
        "slug": "조경", "title": "01 조경", "sheet": "토목(조경)",
        "star": "★ 01 조경 소계", "dogeup_col": 5,
        "rates": CIVIL_JOGYEONG_RATES, "rate_label": "조경",
        "src": "01_화성 청원지구 조경.XLS", "file_filter": "01 조경",
        "out_dir": "조경",
    },
    {
        "slug": "지구단위", "title": "06 산업유통형 개발행위", "sheet": "지구단위",
        "star": "★ 지구단위 합계", "dogeup_col": 6,
        "rates": CIVIL_TOMOK_RATES, "rate_label": "토목",
        "src": "06_화성 청원지구 산업유통형 개발행위.XLS", "file_filter": None,
        "out_dir": "토목",
    },
    {
        "slug": "폐기물", "title": "07 건설폐기물처리", "sheet": "폐기물",
        "star": "★ 폐기물 합계", "dogeup_col": 7,
        "rates": CIVIL_TOMOK_RATES, "rate_label": "토목",
        "src": "07_화성 청원지구 건설폐기물처리.XLS", "file_filter": None,
        "out_dir": "폐기물",
    },
]

DETAIL_SHEETS = ("회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물")


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _find_star_row(ws, star_label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if v and str(v).strip() == star_label:
            return r
    return None


def _read_star(ws, row: int) -> dict:
    matched = int(_num(ws.cell(row, 4).value))
    total = int(_num(ws.cell(row, 5).value))
    mat = _num(ws.cell(row, 7).value)
    lab = _num(ws.cell(row, 8).value)
    exp = _num(ws.cell(row, 9).value)
    direct = _num(ws.cell(row, 10).value) or (mat + lab + exp)
    return {
        "matched": matched,
        "total": total,
        "rate": matched / total if total else 0.0,
        "mat": mat,
        "lab": lab,
        "exp": exp,
        "direct": direct,
    }


def _read_sections(ws, *, file_filter: str | None = None, hdr_row: int = 3) -> list[dict]:
    out: list[dict] = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).startswith("◆"):
            break
        name = ws.cell(r, 3).value
        if not name:
            if out:
                break
            continue
        name = str(name).strip()
        if name.startswith("★"):
            continue
        naeyeok = str(ws.cell(r, 2).value or "").strip()
        if file_filter and naeyeok != file_filter:
            continue
        if not ws.cell(r, 1).value:
            continue
        out.append({
            "name": name,
            "level": section_level(name),
            "matched": int(_num(ws.cell(r, 4).value)),
            "total": int(_num(ws.cell(r, 5).value)),
            "mat": _num(ws.cell(r, 7).value),
            "lab": _num(ws.cell(r, 8).value),
            "exp": _num(ws.cell(r, 9).value),
            "sum": _num(ws.cell(r, 10).value),
        })
    out.sort(key=lambda s: section_sort_key(s["name"]))
    return out


def _detail_hdr_row(ws) -> int | None:
    for r in range(1, min(ws.max_row + 1, 800)):
        if str(ws.cell(r, 4).value or "").strip() == "공종명":
            return r
    return None


def build_phase_html(cfg: dict, ws, cost_ws, *, dogeup_row: int = 21) -> str:
    star_row = _find_star_row(ws, cfg["star"])
    if star_row is None:
        raise ValueError(f"{cfg['sheet']} — '{cfg['star']}' 없음")
    grand = _read_star(ws, star_row)
    sections = _read_sections(ws, file_filter=cfg.get("file_filter"))
    cs = compute_cost_statement_civil(grand["mat"], grand["lab"], grand["exp"], cfg["rates"])
    dogeup_ka = _num(cost_ws.cell(dogeup_row, cfg["dogeup_col"]).value)
    mult = dogeup_ka / grand["direct"] if grand["direct"] else 0.0

    grand_tbl = {**grand, "sum": grand["direct"]}
    body = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 — {cfg['title']} 직접·간접비 (_ka)</title>
<style>{CSS}</style></head><body>
<h1>화성 청원지구 — {cfg['title']}</h1>
<ul>
  <li><strong>SSOT</strong>: <code>{KA_FILE}</code> · 시트 <code>{cfg['sheet']}</code></li>
  <li><strong>원본 공내역서</strong>: <code>{cfg['src']}</code></li>
  <li><strong>작성일</strong>: {GEN_DATE}</li>
  <li><strong>직접비</strong>: _ka 「합계요약」 차수별 재·노·경</li>
  <li><strong>도급액</strong>: _ka 「총공사비」 {won(dogeup_ka)}원 · 직접의 {mult:.3f}배</li>
</ul>
<h2>{cfg['title']} — _ka {cfg['sheet']}</h2>
<p class="muted">간접비 요율: {cfg['rate_label']}공사 간접공사비 적용기준(2026.4.13) ·
  매칭 {grand['matched']}/{grand['total']}건 ({grand['rate']*100:.1f}%)</p>
<h3>1. 차수별 직접공사비</h3>
{render_direct_table(sections, grand_tbl)}
<h3>2. 간접비·도급액 (원가계산서 · 참고)</h3>
{render_cost_table(cs)}
<blockquote>
  <p><strong>제출 도급액</strong>은 위 원가계산서 산출과 별도로 _ka 「총공사비」 시트 값
  <strong>{won(dogeup_ka)}원</strong>을 따릅니다.</p>
</blockquote>
<p class="muted">끝.</p>
</body></html>"""
    return body


def _pending_row_html(row: dict) -> str:
    return (
        f"<tr><td>{row['sheet']}</td><td class='l2'>{row['section'] or ''}</td>"
        f"<td class='l2'>{row['name']}</td><td>{row['status']}</td>"
        f"<td class='num'>{won(row['sum']) if row['sum'] else '—'}</td></tr>"
    )


def build_pending_html(wb) -> str:
    rows: list[dict] = []
    jogyeong_rows: list[dict] = []
    counts = {"검토": 0, "미매칭": 0}
    for sn in DETAIL_SHEETS:
        ws = wb[sn]
        hdr = _detail_hdr_row(ws)
        if hdr is None:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            if not ws.cell(r, 3).value:
                continue
            name = ws.cell(r, 4).value
            if not name or str(name).startswith("★"):
                continue
            section = ws.cell(r, 2).value
            if is_jogyeong_crosscheck_item(section, name):
                jogyeong_rows.append({
                    "sheet": sn,
                    "section": section,
                    "name": name,
                    "status": str(ws.cell(r, 16).value or "미매칭").strip() or "미매칭",
                    "sum": _num(ws.cell(r, 15).value),
                })
                continue
            if is_ka_pending_excluded(section, name):
                continue
            has_sum = detail_row_has_sum(
                mat_u=ws.cell(r, 8).value,
                lab_u=ws.cell(r, 10).value,
                exp_u=ws.cell(r, 12).value,
                mat_amt=ws.cell(r, 9).value,
                lab_amt=ws.cell(r, 11).value,
                exp_amt=ws.cell(r, 13).value,
                unit_sum=ws.cell(r, 14).value,
                total_amt=ws.cell(r, 15).value,
            )
            cat = classify_ka_mihwakjeong(ws.cell(r, 16).value, has_sum=has_sum)
            if not cat:
                continue
            counts[cat] += 1
            rows.append({
                "sheet": sn,
                "no": ws.cell(r, 1).value,
                "naeyeok": ws.cell(r, 1).value if sn == "토목(조경)" else "",
                "section": section,
                "name": name,
                "status": cat,
                "sum": _num(ws.cell(r, 15).value),
            })

    trs = [_pending_row_html(row) for row in rows]
    table = (
        "<table><thead><tr><th>시트</th><th>공종</th><th>공종명</th><th>상태</th>"
        "<th class='num'>합계금액</th></tr></thead><tbody>"
        + "\n".join(trs)
        + f"</tbody><tfoot><tr class='total'><td colspan='3'>소계</td>"
        f"<td>검토 {counts['검토']} · 미매칭 {counts['미매칭']}</td>"
        f"<td class='num'>{sum(counts.values())}건</td></tr></tfoot></table>"
    )
    jogyeong_block = render_jogyeong_html() if jogyeong_rows else ""
    return f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 — 미확정 내역 (_ka)</title>
<style>{CSS}</style></head><body>
<h1>화성 청원지구 — 미확정 내역 (_ka SSOT)</h1>
<ul>
  <li><strong>SSOT</strong>: <code>{KA_FILE}</code> 「◆ 내역 일괄」</li>
  <li><strong>작성일</strong>: {GEN_DATE}</li>
  <li><strong>집계</strong>: 매칭 확정 제외 · 가영현 제외 · 품질관리비 제외 · 조경 확인 3건은 하단 품셈 산정표</li>
</ul>
<blockquote>
  검토 <strong>{counts['검토']}</strong> · 미매칭 <strong>{counts['미매칭']}</strong> ·
  합계 <strong>{sum(counts.values())}</strong>건
</blockquote>
{table}
{jogyeong_block}
<p class="muted">끝.</p>
</body></html>"""


def main() -> None:
    wb = load_workbook(KA, data_only=True)
    cost_ws = wb["총공사비"]
    for cfg in PHASES:
        ws = wb[cfg["sheet"]]
        html = build_phase_html(cfg, ws, cost_ws)
        out_dir = WORK / cfg.get("out_dir", cfg["slug"])
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / f"{cfg['slug']}_직접간접비.html"
        out.write_text(html, encoding="utf-8")
        print(f"  {out.relative_to(ROOT)}")

    pending = build_pending_html(wb)
    COMMON.mkdir(parents=True, exist_ok=True)
    pending_path = COMMON / "청원지구_미확정_ka.html"
    pending_path.write_text(pending, encoding="utf-8")
    print(f"  {pending_path.relative_to(ROOT)}")
    sync_ilwidae_xlsx()
    wb.close()


if __name__ == "__main__":
    print("=== _ka 공종별 HTML ===")
    main()
