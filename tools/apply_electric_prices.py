#!/usr/bin/env python3
"""전기 공내역서 — 참조파일 단가 + 표준시장단가로 재료비/노무비/경비 산출."""
from __future__ import annotations

import argparse
import csv
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 공내역서·산출물 모두 05_내역서 하위
WORK_DIR = BASE / "내역서작업"
SRC_DIR = BASE / "공내역서"
PRICE_CSV = ROOT / "tools" / "_poomsem_cache" / "표준시장단가.csv"

SRC = SRC_DIR / "02_화성 청원지구 전기설비.xlsx"
REF = None
OUT = WORK_DIR / "02_화성 청원지구 전기설비_표준단가산출.xlsx"
OUT_MD = WORK_DIR / "02_화성 청원지구 전기설비_표준단가산출_요약.md"
DOC_TITLE = "전기설비"
REF_LABEL = "동일파일 지구외"

THRESHOLD = 0.58
REVIEW_THRESHOLD = 0.72

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ORIGIN_FILL = PatternFill("solid", fgColor="E2EFDA")
CROSS_FILL = PatternFill("solid", fgColor="DDEBF7")
STD_FILL = PatternFill("solid", fgColor="FFF2CC")
UNMATCH_FILL = PatternFill("solid", fgColor="FCE4D6")
MONEY_FMT = "#,##0"

UNIT_MAP = {"m": "m", "M": "m", "ｍ": "m", "개": "개", "ea": "개", "EA": "개", "개소": "개소", "본": "본", "주": "주", "대": "대", "set": "set", "식": "식", "km": "km", "㎞": "km"}


def norm_unit(u: str) -> str:
    return UNIT_MAP.get((u or "").strip().replace(" ", ""), (u or "").strip().replace(" ", ""))


def norm_text(s: str) -> str:
    s = re.sub(r"\s+", "", s or "")
    return re.sub(r"[\(\),/\-\+\.·]", "", s).lower()


def norm_hopyo(s) -> str:
    """비고/호표 정규화: '제 11 호' → '제11호'."""
    return re.sub(r"\s+", "", str(s or ""))


def load_poomsem_tables(path: Path) -> dict:
    """파일 내장 품셈 일위대가(일대목차)·자재(합산자재) 단가표 적재.

    - 일대목차: 번호(1열)·호표(3열) → 재료/노무/경비 단가 (품셈 기반 일위대가)
    - 합산자재: 코드(3열)·명칭+규격 → 재료비 단가 (자재단가)
    반환: {ildae_no, ildae_hopyo, mat_code, mat_ns}. 시트 없으면 빈 dict.
    """
    tables = {"ildae_no": {}, "ildae_hopyo": {}, "mat_code": {}, "mat_ns": {}}
    try:
        wb = load_workbook(str(path), data_only=True)
    except Exception:
        return tables

    if "일대목차" in wb.sheetnames:
        sh = wb["일대목차"]
        for r in range(4, sh.max_row + 1):
            no = sh.cell(r, 1).value
            hop = sh.cell(r, 3).value
            name = sh.cell(r, 4).value
            if not name:
                continue
            mat = sh.cell(r, 8).value or 0
            lab = sh.cell(r, 11).value or 0
            exp = sh.cell(r, 13).value or 0
            tot = sh.cell(r, 16).value or (mat + lab + exp)
            try:
                rec = {
                    "mat_u": float(mat), "lab_u": float(lab), "exp_u": float(exp),
                    "tot_u": float(tot),
                    "price_name": name, "price_spec": sh.cell(r, 5).value or "",
                }
            except (TypeError, ValueError):
                continue
            if no:
                tables["ildae_no"][str(no).strip()] = rec
            if hop and str(hop).startswith("제"):
                tables["ildae_hopyo"][norm_hopyo(hop)] = rec

    if "합산자재" in wb.sheetnames:
        shm = wb["합산자재"]
        for r in range(4, shm.max_row + 1):
            code = str(shm.cell(r, 3).value or "").strip()
            name = shm.cell(r, 4).value
            spec = shm.cell(r, 5).value
            price = shm.cell(r, 8).value
            if not name or not isinstance(price, (int, float)):
                continue
            rec = {
                "mat_u": float(price), "lab_u": 0.0, "exp_u": 0.0, "tot_u": float(price),
                "price_name": name, "price_spec": spec or "",
            }
            if code:
                tables["mat_code"][code] = rec
            tables["mat_ns"][(norm_text(name), norm_text(spec))] = rec

    wb.close()
    return tables


def load_dangajosa(path: Path) -> dict:
    """원본 '단가조사' 시트의 중기(크레인 등) 합성 단가 적재.

    중기 행은 값→라벨 순으로 재료/노무/경비가 기입됨
    (예: col8=15976, col9='재료', col10=48663, col11='노무', col12=29487, col13='경비').
    라벨 '재료/노무/경비' 바로 왼쪽 셀의 숫자를 단가로 본다.
    단순 자재 행(라벨 없음)은 합산자재로 이미 처리되므로 제외.
    반환: {jb_ns, jb_code}.
    """
    out = {"jb_ns": {}, "jb_code": {}}
    try:
        wb = load_workbook(str(path), data_only=True)
    except Exception:
        return out
    if "단가조사" not in wb.sheetnames:
        wb.close()
        return out
    sh = wb["단가조사"]
    label_key = {"재료": "mat_u", "노무": "lab_u", "경비": "exp_u"}
    for r in range(4, sh.max_row + 1):
        name = sh.cell(r, 4).value
        if not name:
            continue
        spec = sh.cell(r, 5).value or ""
        unit = sh.cell(r, 6).value or ""
        code = str(sh.cell(r, 3).value or "").strip()
        vals = {"mat_u": 0.0, "lab_u": 0.0, "exp_u": 0.0}
        has_comp = False
        for c in range(7, 15):
            lab = sh.cell(r, c).value
            if isinstance(lab, str) and lab.strip() in label_key:
                num = sh.cell(r, c - 1).value
                if isinstance(num, (int, float)):
                    vals[label_key[lab.strip()]] = float(num)
                    has_comp = True
        if not has_comp:
            continue
        rec = {
            **vals,
            "tot_u": vals["mat_u"] + vals["lab_u"] + vals["exp_u"],
            "price_name": name,
            "price_spec": spec,
            "unit": unit,
        }
        out["jb_ns"][(norm_text(name), norm_text(spec))] = rec
        if code:
            out["jb_code"][code] = rec
    wb.close()
    return out


def load_standard_prices() -> list[dict]:
    if not PRICE_CSV.exists():
        return []
    rows = list(csv.reader(PRICE_CSV.read_text("cp949").splitlines()))
    out = []
    for row in rows[1:]:
        if len(row) < 9:
            continue
        try:
            out.append(
                {
                    "code": row[1].strip(),
                    "name": row[2].strip(),
                    "spec": row[3].strip(),
                    "unit": row[4].strip().replace(" ", ""),
                    "mat": float(row[5] or 0),
                    "lab": float(row[6] or 0),
                    "exp": float(row[7] or 0),
                    "total": float(row[8] or 0),
                    "date": row[0].strip(),
                }
            )
        except ValueError:
            continue
    return out


def is_priced_row(tot, mat_u, mat_a) -> bool:
    if tot == "#REF!" or mat_a == "#REF!":
        return False
    if isinstance(tot, (int, float)) and tot > 0:
        return True
    if isinstance(mat_u, (int, float)) and mat_u > 0:
        return True
    if isinstance(mat_a, (int, float)) and mat_a > 0:
        return True
    return False


def load_estimate_rows(path: Path) -> list[dict]:
    wb = load_workbook(str(path), data_only=True)
    sh = wb["내역서"]
    rows = []
    section = ""
    for r in range(4, sh.max_row + 1):
        rt = sh.cell(r, 2).value
        if rt == "공종줄":
            section = sh.cell(r, 4).value or ""
            continue
        if rt in ("합계줄",) or not sh.cell(r, 4).value:
            continue
        qty = sh.cell(r, 7).value
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue
        tot = sh.cell(r, 16).value
        mat_u = sh.cell(r, 8).value
        mat_a = sh.cell(r, 9).value
        priced = is_priced_row(tot, mat_u, mat_a)
        rows.append(
            {
                "row": r,
                "section": section,
                "no": sh.cell(r, 1).value,
                "gongjong_code": sh.cell(r, 2).value,
                "code": str(sh.cell(r, 3).value or "").strip(),
                "name": sh.cell(r, 4).value,
                "spec": sh.cell(r, 5).value or "",
                "unit": sh.cell(r, 6).value,
                "qty": float(qty),
                "note": sh.cell(r, 17).value or "",
                "priced": priced,
                "mat_u": mat_u,
                "mat_a": mat_a if isinstance(mat_a, (int, float)) else 0,
                "lab_u": sh.cell(r, 11).value,
                "lab_a": sh.cell(r, 12).value or 0,
                "exp_u": sh.cell(r, 13).value,
                "exp_a": sh.cell(r, 14).value or 0,
                "sum_a": tot if isinstance(tot, (int, float)) else 0,
            }
        )
    return rows


def build_cross_ref(rows: list[dict]) -> tuple[dict, dict]:
    by_code: dict[str, dict] = {}
    by_ns: dict[tuple, dict] = {}
    for item in rows:
        if not item["priced"]:
            continue
        mat_u = float(item["mat_u"] or 0)
        lab_u = float(item["lab_u"] or 0)
        exp_u = float(item["exp_u"] or 0)
        if mat_u or lab_u or exp_u:
            tot_u = mat_u + lab_u + exp_u
        elif item["sum_a"] and item["qty"]:
            tot_u = item["sum_a"] / item["qty"]
        else:
            tot_u = 0.0
        rec = {
            "mat_u": mat_u,
            "lab_u": lab_u,
            "exp_u": exp_u,
            "tot_u": tot_u,
            "price_name": item["name"],
            "price_spec": item["spec"],
            "ref_section": item["section"],
        }
        if item["code"]:
            by_code[item["code"]] = rec
        by_ns[(item["name"], item["spec"])] = rec
    return by_code, by_ns


def score_std(item: dict, price: dict) -> float:
    iu, pu = norm_unit(item["unit"]), norm_unit(price["unit"])
    if iu != pu:
        return -1.0
    blob_item = norm_text(f"{item['name']} {item['spec']}")
    blob_price = norm_text(f"{price['name']} {price['spec']}")
    name_score = SequenceMatcher(None, norm_text(item["name"]), norm_text(price["name"])).ratio()
    if norm_text(item["name"]) in blob_price or norm_text(price["name"]) in blob_item:
        name_score = max(name_score, 0.85)
    spec_score = SequenceMatcher(None, norm_text(item["spec"]), norm_text(price["spec"])).ratio()
    if not item["spec"].strip():
        spec_score = 0.4
    return 0.3 + name_score * 0.45 + spec_score * 0.25


def find_std_match(item: dict, prices: list[dict]) -> tuple[dict | None, float]:
    best, best_sc = None, -1.0
    for p in prices:
        sc = score_std(item, p)
        if sc > best_sc:
            best, best_sc = p, sc
    return best, best_sc


def calc_from_unit(qty: float, mat_u: float, lab_u: float, exp_u: float) -> dict:
    mat_a = qty * mat_u
    lab_a = qty * lab_u
    exp_a = qty * exp_u
    return {
        "mat_u": mat_u,
        "lab_u": lab_u,
        "exp_u": exp_u,
        "tot_u": mat_u + lab_u + exp_u,
        "mat_a": mat_a,
        "lab_a": lab_a,
        "exp_a": exp_a,
        "sum_a": mat_a + lab_a + exp_a,
    }


def process_rows(rows: list[dict], std_prices: list[dict], cross_rows: list[dict] | None = None, poomsem: dict | None = None, jangbi: dict | None = None):
    by_code, by_ns = build_cross_ref(cross_rows if cross_rows is not None else rows)
    poomsem = poomsem or {"ildae_no": {}, "ildae_hopyo": {}, "mat_code": {}, "mat_ns": {}}
    jangbi = jangbi or {"jb_ns": {}, "jb_code": {}}
    results = []
    totals = {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0}
    section_totals: dict[str, dict] = {}

    for item in rows:
        sec = item["section"]
        section_totals.setdefault(sec, {"mat": 0, "lab": 0, "exp": 0, "sum": 0, "items": 0, "filled": 0})
        section_totals[sec]["items"] += 1
        out = {**item, "source": "", "status": "미산출", "match_score": None, "price_ref": ""}

        ildae = None
        ildae_ref = ""
        if str(item.get("no") or "").strip() in poomsem["ildae_no"]:
            ildae = poomsem["ildae_no"][str(item["no"]).strip()]
            ildae_ref = f"번호 {item['no']}"
        elif norm_hopyo(item.get("note")) in poomsem["ildae_hopyo"]:
            ildae = poomsem["ildae_hopyo"][norm_hopyo(item["note"])]
            ildae_ref = f"호표 {norm_hopyo(item['note'])}"
        mat_rec = None
        mat_ref = ""
        if item["code"] in poomsem["mat_code"]:
            mat_rec = poomsem["mat_code"][item["code"]]
            mat_ref = f"자재코드 {item['code']}"
        elif (norm_text(item["name"]), norm_text(item["spec"])) in poomsem["mat_ns"]:
            mat_rec = poomsem["mat_ns"][(norm_text(item["name"]), norm_text(item["spec"]))]
            mat_ref = "자재 품명·규격"

        jb_rec = None
        jb_ref = ""
        if item["code"] and item["code"] in jangbi["jb_code"]:
            jb_rec = jangbi["jb_code"][item["code"]]
            jb_ref = f"단가조사 코드 {item['code']}"
        elif (norm_text(item["name"]), norm_text(item["spec"])) in jangbi["jb_ns"]:
            jb_rec = jangbi["jb_ns"][(norm_text(item["name"]), norm_text(item["spec"]))]
            jb_ref = "단가조사 품명·규격"
        if jb_rec and jb_rec.get("unit") and norm_unit(item["unit"]) != norm_unit(jb_rec["unit"]):
            jb_rec = None

        if item["priced"]:
            out.update(
                {
                    "source": "원본(기입)",
                    "status": "원본",
                    "mat_u": item["mat_u"],
                    "lab_u": item["lab_u"],
                    "exp_u": item["exp_u"],
                    "mat_a": item["mat_a"],
                    "lab_a": item["lab_a"],
                    "exp_a": item["exp_a"],
                    "sum_a": item["sum_a"],
                    "price_ref": "원본 내역",
                }
            )
        elif ildae is not None:
            vals = calc_from_unit(item["qty"], ildae["mat_u"], ildae["lab_u"], ildae["exp_u"])
            out.update({
                **vals,
                "source": "일위대가(품셈)",
                "status": "매칭",
                "match_score": 1.0,
                "price_ref": ildae_ref,
                "price_name": ildae["price_name"],
                "price_spec": ildae["price_spec"],
            })
        elif mat_rec is not None:
            vals = calc_from_unit(item["qty"], mat_rec["mat_u"], mat_rec["lab_u"], mat_rec["exp_u"])
            out.update({
                **vals,
                "source": "합산자재(자재비)",
                "status": "매칭",
                "match_score": 1.0,
                "price_ref": mat_ref,
                "price_name": mat_rec["price_name"],
                "price_spec": mat_rec["price_spec"],
            })
        elif jb_rec is not None:
            vals = calc_from_unit(item["qty"], jb_rec["mat_u"], jb_rec["lab_u"], jb_rec["exp_u"])
            out.update({
                **vals,
                "source": "단가조사(중기)",
                "status": "매칭",
                "match_score": 1.0,
                "price_ref": jb_ref,
                "price_name": jb_rec["price_name"],
                "price_spec": jb_rec["price_spec"],
            })
        elif item["code"] in by_code:
            u = by_code[item["code"]]
            vals = calc_from_unit(item["qty"], u["mat_u"], u["lab_u"], u["exp_u"])
            out.update({
                **vals,
                "source": REF_LABEL,
                "status": "매칭",
                "match_score": 1.0,
                "price_ref": f"코드 {item['code']}",
                "price_name": u["price_name"],
                "price_spec": u["price_spec"],
            })
        elif (item["name"], item["spec"]) in by_ns:
            u = by_ns[(item["name"], item["spec"])]
            vals = calc_from_unit(item["qty"], u["mat_u"], u["lab_u"], u["exp_u"])
            out.update({
                **vals,
                "source": REF_LABEL,
                "status": "매칭",
                "match_score": 0.95,
                "price_ref": "품명·규격",
                "price_name": u["price_name"],
                "price_spec": u["price_spec"],
            })
        elif std_prices:
            p, sc = find_std_match(item, std_prices)
            if p and sc >= THRESHOLD:
                vals = calc_from_unit(item["qty"], p["mat"], p["lab"], p["exp"])
                out.update({
                    **vals,
                    "source": "표준시장단가",
                    "status": "검토" if sc < REVIEW_THRESHOLD else "매칭",
                    "match_score": round(sc, 3),
                    "price_ref": p["code"],
                    "price_name": p["name"],
                    "price_spec": p["spec"],
                })

        if out["status"] not in ("미산출",):
            section_totals[sec]["filled"] += 1
            for k in ("mat", "lab", "exp", "sum"):
                key = f"{k}_a" if k != "sum" else "sum_a"
                section_totals[sec][k] += out[key]
                totals[k] += out[key]
        results.append(out)

    return results, totals, section_totals


def style_header(ws, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws, max_w=52):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(max_w, max(len(str(c.value or "")) for c in col) + 2)


def write_outputs(results, totals, section_totals, std_date: str, ref_name: str | None = None):
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "안내"
    ref_note = ref_name or "동일 파일 지구외 구간"
    for line in [
        [f"화성 청원지구 {DOC_TITLE} — 단가 산출"],
        [],
        ["원본", SRC.name],
        ["단가 참조", ref_note],
        ["산출 우선순위", f"① 원본 기입 → ② 일위대가(품셈) → ③ 합산자재(자재비) → ④ 단가조사(중기) → ⑤ {REF_LABEL}(코드) → ⑥ 표준시장단가"],
        ["품셈 적용", "파일 내장 일대목차(품셈 기반 일위대가)·합산자재 단가 우선 적용"],
        ["표준시장단가", std_date or "미사용"],
        [],
        ["#REF! 복구", "원본에 #REF!가 있으면 품셈 일위대가·참조파일·표준시장단가로 대체 산출."],
    ]:
        ws0.append(line)
    ws0["A1"].font = Font(bold=True, size=14)
    autofit(ws0, 70)

    ws = wb.create_sheet("합계요약")
    ws.append(["공종", "산출", "전체", "산출률", "재료비", "노무비", "경비", "합계"])
    style_header(ws, 8)
    for sec, st in section_totals.items():
        rate = st["filled"] / st["items"] if st["items"] else 0
        ws.append([sec, st["filled"], st["items"], rate, st["mat"], st["lab"], st["exp"], st["sum"]])
    total_items = sum(st["items"] for st in section_totals.values())
    total_filled = sum(st["filled"] for st in section_totals.values())
    ws.append([])
    ws.append(["★ 합계", total_filled, total_items, total_filled / total_items if total_items else 0, *[
        totals["mat"], totals["lab"], totals["exp"], totals["sum"]
    ]])
    for r in range(2, ws.max_row + 1):
        if isinstance(ws.cell(r, 4).value, float):
            ws.cell(r, 4).number_format = "0.0%"
        for c in range(5, 9):
            if isinstance(ws.cell(r, c).value, (int, float)):
                ws.cell(r, c).number_format = MONEY_FMT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    headers = [
        "행", "공종", "코드", "명칭", "규격", "단위", "수량", "상태", "단가출처", "매칭점수",
        "참조", "매칭품명", "매칭규격",
        "재료단가", "노무단가", "경비단가", "합계단가",
        "재료금액", "노무금액", "경비금액", "합계금액", "비고",
    ]
    ws1 = wb.create_sheet("통합내역")
    ws1.append(headers)
    style_header(ws1, len(headers))
    fill_map = {"원본": ORIGIN_FILL, "매칭": CROSS_FILL, "검토": STD_FILL, "미산출": UNMATCH_FILL}
    for r in results:
        ws1.append([
            r["row"], r["section"], r["code"], r["name"], r["spec"], r["unit"], r["qty"],
            r["status"], r["source"], r.get("match_score"), r.get("price_ref", ""),
            r.get("price_name", ""), r.get("price_spec", ""),
            r.get("mat_u", ""), r.get("lab_u", ""), r.get("exp_u", ""), r.get("tot_u", ""),
            r.get("mat_a", ""), r.get("lab_a", ""), r.get("exp_a", ""), r.get("sum_a", ""),
            r.get("note", ""),
        ])
    for ri in range(2, ws1.max_row + 1):
        st = ws1.cell(ri, 8).value
        fill = fill_map.get(st)
        if fill:
            for c in range(1, len(headers) + 1):
                ws1.cell(ri, c).fill = fill
        for c in range(14, 22):
            if isinstance(ws1.cell(ri, c).value, (int, float)):
                ws1.cell(ri, c).number_format = MONEY_FMT
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    autofit(ws1)

    extra_sheets = [
        ("지구내_산출", lambda r: "(지구내)" in r["section"]),
        ("지구외_산출", lambda r: "(지구외)" in r["section"]),
        ("미산출", lambda r: r["status"] == "미산출"),
        ("검토필요", lambda r: r["status"] == "검토"),
    ]
    for title, filt in extra_sheets:
        w = wb.create_sheet(title)
        w.append(headers)
        style_header(w, len(headers))
        for r in [x for x in results if filt(x)]:
            w.append([
                r["row"], r["section"], r["code"], r["name"], r["spec"], r["unit"], r["qty"],
                r["status"], r["source"], r.get("match_score"), r.get("price_ref", ""),
                r.get("price_name", ""), r.get("price_spec", ""),
                r.get("mat_u", ""), r.get("lab_u", ""), r.get("exp_u", ""), r.get("tot_u", ""),
                r.get("mat_a", ""), r.get("lab_a", ""), r.get("exp_a", ""), r.get("sum_a", ""),
                r.get("note", ""),
            ])
        w.freeze_panes = "A2"
        w.auto_filter.ref = w.dimensions
        autofit(w)

    try:
        wb.save(OUT)
        saved = OUT
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        saved = alt
    return saved


def write_md(results, totals, section_totals, std_date: str):
    inner = [r for r in results if "(지구내)" in r["section"]]
    inner_filled = [r for r in inner if r["status"] != "미산출"]
    lines = [
        f"# {DOC_TITLE} 공내역서 — 단가 산출 요약",
        "",
        f"- **원본**: `{SRC.name}`",
        f"- **표준시장단가**: {std_date or '미사용'}",
        "",
        "## 전체 합계 (산출·원본 포함)",
        "",
        f"| 재료비 | 노무비 | 경비 | 합계 |",
        f"|---:|---:|---:|---:|",
        f"| {totals['mat']:,.0f} | {totals['lab']:,.0f} | {totals['exp']:,.0f} | {totals['sum']:,.0f} |",
        "",
        f"- **지구내**: {len(inner_filled)}/{len(inner)}건 산출",
        "",
        "## 공종별",
        "",
        "| 공종 | 산출 | 전체 | 재료비 | 노무비 | 경비 | 합계 |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for sec, st in section_totals.items():
        lines.append(
            f"| {sec} | {st['filled']} | {st['items']} | {st['mat']:,.0f} | {st['lab']:,.0f} | {st['exp']:,.0f} | {st['sum']:,.0f} |"
        )
    lines.append("\n끝.")
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def parse_args():
    p = argparse.ArgumentParser(description="전기 공내역서 재료비·노무비·경비 산출")
    p.add_argument("--src", type=Path, default=SRC, help="대상 공내역서 xlsx")
    p.add_argument("--ref", type=Path, default=None, help="단가 참조 공내역서 (미지정 시 src 자체)")
    p.add_argument("--out", type=Path, default=None, help="결과 xlsx (기본: src명_표준단가산출.xlsx)")
    p.add_argument("--title", default=None, help="문서 제목 (예: 전기설비(지구외))")
    p.add_argument("--ref-label", default=None, help="참조 단가 출처 표기")
    return p.parse_args()


def main():
    global SRC, REF, OUT, OUT_MD, DOC_TITLE, REF_LABEL
    args = parse_args()
    SRC = args.src if args.src.is_absolute() else ROOT / args.src
    if not SRC.exists() and (SRC_DIR / args.src.name).exists():
        SRC = SRC_DIR / args.src.name
    if args.ref is None:
        REF = None
    else:
        REF = args.ref if args.ref.is_absolute() else ROOT / args.ref
        if not REF.exists() and (SRC_DIR / args.ref.name).exists():
            REF = SRC_DIR / args.ref.name
    # 출력물은 05_내역서에 저장(원본 공내역서 폴더와 분리)
    out_stem = SRC.stem.replace("-" + "공내역서", "").replace("_" + "공내역서", "")
    OUT = args.out or WORK_DIR / (out_stem + "_표준단가산출.xlsx")
    OUT_MD = OUT.with_suffix(".md").with_name(OUT.stem + "_요약.md")
    if args.title:
        DOC_TITLE = args.title
    elif "지구외" in SRC.name and "지구내" not in SRC.name:
        DOC_TITLE = "전기설비(지구외)"
    else:
        DOC_TITLE = "전기설비"
    if args.ref_label:
        REF_LABEL = args.ref_label
    elif REF and REF != SRC:
        REF_LABEL = f"참조({REF.name})"
    else:
        REF_LABEL = "동일파일 지구외"

    rows = load_estimate_rows(SRC)
    cross_rows = load_estimate_rows(REF) if REF else None
    std = load_standard_prices()
    std_date = std[0]["date"] if std else ""
    # 품셈 일위대가·자재 단가표: SRC 우선, 없으면 REF(예: 03→02)
    poomsem = load_poomsem_tables(SRC)
    if not poomsem["ildae_no"] and REF:
        poomsem = load_poomsem_tables(REF)
    jangbi = load_dangajosa(SRC)
    if not jangbi["jb_ns"] and REF:
        jangbi = load_dangajosa(REF)
    results, totals, section_totals = process_rows(rows, std, cross_rows, poomsem, jangbi)
    ref_name = REF.name if REF and REF != SRC else None
    saved = write_outputs(results, totals, section_totals, std_date, ref_name)
    write_md(results, totals, section_totals, std_date)
    ref_cnt = sum(1 for r in results if r["source"] == REF_LABEL)
    ildae_cnt = sum(1 for r in results if r["source"] == "일위대가(품셈)")
    mat_cnt = sum(1 for r in results if r["source"] == "합산자재(자재비)")
    print(f"원본 {SRC.name}")
    print(f"품셈 일위대가 {ildae_cnt} / 합산자재 {mat_cnt}")
    print(f"전체 {len(rows)} / 미산출 {sum(1 for r in results if r['status']=='미산출')} / 참조단가 {ref_cnt}")
    print(f"합계 재료 {totals['mat']:,.0f} 노무 {totals['lab']:,.0f} 경비 {totals['exp']:,.0f} 계 {totals['sum']:,.0f}")
    print(f"저장: {saved}")


if __name__ == "__main__":
    main()
