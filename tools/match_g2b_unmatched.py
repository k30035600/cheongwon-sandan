#!/usr/bin/env python3
"""미매칭 품목 × 조달청 가격정보(15129415) 시뮬레이션 — 기존 산출 미수정.

단가원: 05_내역서/일위대가DB/조달청_가격정보_2026/*.csv
출력: 05_내역서/조달청_미매칭점검.xlsx
"""
from __future__ import annotations

import csv
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
G2B_DIR = BASE / "일위대가DB" / "조달청_가격정보_2026"
OUT = BASE / "조달청_미매칭점검.xlsx"

sys.path.insert(0, str(TOOLS))
import apply_standard_prices as asp  # noqa: E402

LABELS = {
    "01": "01 토목", "01j": "01 조경", "02": "02 전기", "04": "04 진입도로",
    "05": "05 회전교차로", "06": "06 개발행위", "07": "07 건설폐기물",
}

PRICE_FILES = [
    ("시설공통자재_종합.csv", "시설공통자재(종합)"),
    ("시설공통자재_토목.csv", "시설공통자재(토목)"),
    ("시설공통자재_건축.csv", "시설공통자재(건축)"),
    ("시설공통자재_기계설비.csv", "시설공통자재(기계설비)"),
    ("시설공통자재_전기정보통신.csv", "시설공통자재(전기·정보통신)"),
    ("시장시공가격_토목.csv", "시장시공가격(토목)"),
    ("시장시공가격_건축.csv", "시장시공가격(건축)"),
    ("시장시공가격_기계설비.csv", "시장시공가격(기계설비)"),
]

HDR = PatternFill("solid", fgColor="D9E1F2")
OKF = PatternFill("solid", fgColor="E2EFDA")
RVF = PatternFill("solid", fgColor="FFF2CC")
NMF = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)


def fnum(v) -> float:
    try:
        return float(str(v).replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def newest_xlsx(stem_prefix: str) -> Path | None:
    cands = list(BASE.glob(f"{stem_prefix}*_표준단가산출.xlsx"))
    cands += list(WORK.glob(f"{stem_prefix}*_표준단가산출.xlsx"))
    return max(cands, key=lambda p: p.stat().st_mtime) if cands else None


def load_unmatched() -> list[dict]:
    items: list[dict] = []
    for no in ("01", "02", "04", "05", "06", "07"):
        p = newest_xlsx(no)
        if not p:
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        for sname in ("미매칭", "미산출"):
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hdr = [("" if h is None else str(h).strip()) for h in rows[0]]
            idx = {h: i for i, h in enumerate(hdr)}

            def col(r, *names):
                for n in names:
                    for k, j in idx.items():
                        if n in k and j < len(r):
                            return r[j]
                return None

            for r in rows[1:]:
                if not any(c is not None and str(c).strip() for c in r):
                    continue
                name = str(col(r, "공종명", "명칭") or "").strip()
                spec = str(col(r, "규격") or "").strip()
                unit = str(col(r, "단위") or "").strip()
                if not unit or unit == "식":
                    continue
                qty = col(r, "수량")
                try:
                    qty = float(qty) if qty is not None else 0
                except (TypeError, ValueError):
                    qty = 0
                items.append({
                    "ledger": LABELS.get(no, no),
                    "row": col(r, "행"),
                    "section": str(col(r, "공종") or ""),
                    "name": name, "spec": spec, "unit": unit, "qty": qty,
                })
        wb.close()
    return items


def load_g2b_prices() -> tuple[list[dict], list[str]]:
    prices: list[dict] = []
    loaded: list[str] = []
    for fname, tag in PRICE_FILES:
        path = G2B_DIR / fname
        if not path.exists():
            continue
        loaded.append(fname)
        with path.open(encoding="utf-8-sig", newline="") as f:
            for rec in csv.DictReader(f):
                name = (rec.get("prdctClsfcNoNm") or "").strip()
                spec = (rec.get("krnPrdctNm") or "").strip()
                if not name and not spec:
                    continue
                unit = (rec.get("unit") or "").strip().replace(" ", "")
                prce = fnum(rec.get("prce"))
                mat = fnum(rec.get("mtrlcst"))
                lab = fnum(rec.get("lbrcst"))
                exp = fnum(rec.get("gnrlexpns"))
                if mat == 0 and lab == 0 and exp == 0:
                    mat, total = prce, prce
                else:
                    total = prce if prce else (mat + lab + exp)
                if total <= 0:
                    continue
                prices.append({
                    "code": (rec.get("prdctClsfcNo") or rec.get("prceNticeNo") or "").strip(),
                    "name": name or spec,
                    "spec": spec,
                    "unit": unit,
                    "mat": mat, "lab": lab, "exp": exp, "total": total,
                    "date": tag, "_src": tag,
                })
    return prices, loaded


def main():
    raw, loaded = load_g2b_prices()
    prices = asp.precompute(raw)
    items = load_unmatched()
    print(f"조달청 CSV {len(loaded)}종 → 단가 {len(prices):,}건 / 미매칭 {len(items):,}건")
    if not loaded:
        print("[경고] CSV 없음 — fetch_g2b_price.py 실행 필요")
    if "시설공통자재_종합.csv" not in loaded:
        print("[참고] 시설공통자재_종합.csv 미수집 — G2B_KEY 설정 후 fetch 권장")

    WORK.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "점검결과"
    cols = ["내역서", "행", "공종", "명칭", "규격", "단위", "수량", "점수", "상태",
            "조달청출처", "매칭품명", "매칭규격", "매칭단위",
            "재료비", "노무비", "경비", "합계단가", "산출금액"]
    ws.append(["조달청 나라장터 가격정보(15129415) — 미매칭 해소 가능성 점검(기존 산출 미반영)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"단가원 {len(prices):,}건 ({', '.join(loaded)}) / 임계 매칭≥{asp.REVIEW_THRESHOLD}·검토≥{asp.THRESHOLD}"])
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = HDR
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")

    n_ok = n_rv = n_no = 0
    add_sum = 0.0
    buck = defaultdict(lambda: {"t": 0, "ok": 0, "rv": 0, "no": 0})

    for it in items:
        price, score, _, _ = asp.find_best_match(it, prices)
        b = buck[it["ledger"]]
        b["t"] += 1
        if not price or score < asp.THRESHOLD:
            n_no += 1
            b["no"] += 1
            ws.append([it["ledger"], it["row"], it["section"], it["name"], it["spec"],
                       it["unit"], it["qty"], round(score, 3) if score >= 0 else None,
                       "미매칭", "", "", "", "", "", "", "", "", ""])
            fill = NMF
        else:
            amts = asp.calc_amounts(it, price)
            status = "매칭" if score >= asp.REVIEW_THRESHOLD else "검토"
            if status == "매칭":
                n_ok += 1
                b["ok"] += 1
                add_sum += amts["sum_amt"]
            else:
                n_rv += 1
                b["rv"] += 1
            ws.append([it["ledger"], it["row"], it["section"], it["name"], it["spec"],
                       it["unit"], it["qty"], round(score, 3), status, price["_src"],
                       price["name"], price["spec"], price["unit"],
                       price["mat"], price["lab"], price["exp"], price["total"],
                       round(amts["sum_amt"])])
            fill = OKF if status == "매칭" else RVF
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).fill = fill

    ws.freeze_panes = "A4"
    ws2 = wb.create_sheet("요약")
    ws2.append(["구분", "건수"])
    ws2.append(["미매칭 대상", len(items)])
    ws2.append(["조달청 단가 풀", len(prices)])
    ws2.append(["신규 매칭(높음)", n_ok])
    ws2.append(["검토", n_rv])
    ws2.append(["여전히 미매칭", n_no])
    ws2.append(["해소 가능 합계", n_ok + n_rv])
    ws2.append(["매칭+검토 산출금액(추정)", round(add_sum)])
    ws2.append([])
    ws2.append(["내역서", "대상", "매칭", "검토", "미매칭"])
    for lab in sorted(buck):
        b = buck[lab]
        ws2.append([lab, b["t"], b["ok"], b["rv"], b["no"]])
    ws2.append([])
    ws2.append(["참고", "v1.5 조경수목 API 폐기 — 수목 구입비(주)는 본 API로 해소 불가"])

    wb.save(OUT)
    print(f"매칭 {n_ok} / 검토 {n_rv} / 미매칭 {n_no} (해소 가능 {n_ok + n_rv}건)")
    print(f"매칭+검토 산출금액(추정): {add_sum:,.0f}원")
    print(f"저장: {OUT}")


if __name__ == "__main__":
    main()
