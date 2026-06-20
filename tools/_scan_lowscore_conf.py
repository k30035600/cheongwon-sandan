#!/usr/bin/env python3
"""미매칭_일위대가산출 — 확정/제시단가가 채워졌으나 매칭점수가 낮은(저신뢰) 항목 스캔."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
P = BASE / "미매칭_일위대가산출.xlsx"
THRESH = float(sys.argv[1]) if len(sys.argv) > 1 else 0.56

wb = load_workbook(P, read_only=True, data_only=True)
ws = wb["일위대가산출"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c or "").strip() for c in rows[0]]
H = {h: i for i, h in enumerate(hdr)}
sc_col = "최고점수" if "최고점수" in H else ("매칭점수" if "매칭점수" in H else None)
print(f"미매칭_일위대가 — 점수<{THRESH} 확정/제시 항목:")
print("=" * 95)
flagged = []
for r in rows[1:]:
    if H.get("품명") is None:
        break
    name = r[H["품명"]]
    if not name:
        continue
    conf = r[H["확정단가(입력)"]] if "확정단가(입력)" in H else None
    sug = r[H["제시단가"]] if "제시단가" in H else None
    val = conf or sug
    if val in (None, "", 0):
        continue
    sc = r[H[sc_col]] if sc_col else None
    try:
        scf = float(sc)
    except (TypeError, ValueError):
        scf = None
    if scf is not None and scf < THRESH:
        qty = r[H["수량"]] if "수량" in H else None
        try:
            amt = float(val) * float(qty or 0)
        except (TypeError, ValueError):
            amt = 0
        flagged.append((amt, r[H.get("파일", 1)], r[H.get("행", 2)], str(name)[:24],
                        str(r[H["규격"]])[:14], r[H["단위"]], qty, float(val), scf))
flagged.sort(reverse=True)
tot = 0
for amt, fl, rn, nm, spec, unit, qty, val, sc in flagged:
    tot += amt
    print(f"[{fl}] 행{rn} 「{nm}」 {spec} {unit} ×{qty}  단가 {val:,.0f}  금액 {amt:,.0f}  점수{sc:.2f}")
print("=" * 95)
print(f"저신뢰 확정 {len(flagged)}건 / 금액 합계 {tot:,.0f}")
wb.close()
