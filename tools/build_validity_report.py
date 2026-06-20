#!/usr/bin/env python3
"""자동확정 품목검증 보고서 — 내역품명↔DB매칭품명 일치 여부 4분류.
출력: 05_내역서/검토_품목검증.xlsx"""
from __future__ import annotations
import re
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
SRC = BASE / "미매칭_일위대가산출.xlsx"
OUT = BASE / "검토_품목검증.xlsx"

# 라벨 불일치이나 단가가 타당해 유지 가능한 품목(품명 키워드)
KEEP_OK = ("PVC이중벽관", "닥타일", "집수정", "투수블럭", "물사용량")
# 과소평가 의심 — 수동단가 확인 권장(품명 키워드)
UNDER = ("오수정화조", "공기변실", "제수변실", "소음측정기", "플랜터형의자")


def core(name: str) -> str:
    s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", str(name or ""))
    return re.sub(r"^[.·\s]+", "", s).strip()


def toks(name: str):
    return [t for t in re.split(r"[\s/()·,]+", core(name)) if len(t) >= 2]


def overlap(item, db):
    db = str(db or "").strip()
    if not db or db == "None":
        return -1  # DB명 없음
    blob = db.replace(" ", "")
    return sum(1 for t in toks(item) if t in db or t in blob)


def classify(name, dbn, basis):
    if "오매칭" in str(basis):
        return "C. 오매칭 교정완료", "fix_outlier_prices 반영(보수적 재산출)"
    ov = overlap(name, dbn)
    if ov == -1:
        return "B. 시세·품셈 직접확정", "DB 매칭품명 없음 — 시세·품셈 직접 산출"
    if ov >= 1:
        return "A. 품목 일치", f"내역품명↔DB품명 공유 토큰 {ov}개"
    c = core(name)
    for k in UNDER:
        if k in c:
            return "D. 라벨불일치-검토요", "⚠ 과소평가 의심 — 수동단가 확인 권장"
    for k in KEEP_OK:
        if k in c:
            return "D. 라벨불일치-검토요", "단가 범위 타당(라벨만 불일치) — 유지 가능"
    return "D. 라벨불일치-검토요", "공유 토큰 0 — 검토 권장"


wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb["일위대가산출"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c or "").strip() for c in rows[0]]
H = {h: i for i, h in enumerate(hdr)}


def g(r, k):
    i = H.get(k)
    return r[i] if i is not None and i < len(r) else None


recs = []
for r in rows[1:]:
    name = g(r, "품명")
    conf = g(r, "확정단가(입력)")
    if not name or conf in (None, "", 0):
        continue
    try:
        cv = float(conf)
        qty = float(g(r, "수량") or 0)
    except (TypeError, ValueError):
        cv, qty = 0, 0
    cls, why = classify(name, g(r, "DB_품명"), g(r, "표준품셈·산출근거"))
    recs.append([cls, str(g(r, "파일") or ""), core(str(name)), str(g(r, "규격") or ""),
                 str(g(r, "단위") or ""), qty, cv, round(cv * qty),
                 str(g(r, "DB_품명") or ""), g(r, "최고점수"), why])
wb.close()

order = {"A": 0, "B": 1, "C": 2, "D": 3}
recs.sort(key=lambda x: (order.get(x[0][0], 9), -x[7]))

out = Workbook()
sm = out.active
sm.title = "요약"
cnt: dict[str, int] = {}
amt: dict[str, float] = {}
for x in recs:
    cnt[x[0]] = cnt.get(x[0], 0) + 1
    amt[x[0]] = amt.get(x[0], 0) + x[7]
sm.append(["자동확정 품목검증 요약 — 내역품명 ↔ DB매칭품명 일치도"])
sm.append([])
sm.append(["분류", "건수", "확정금액 합계", "조치"])
ACT = {
    "A. 품목 일치": "자동확정 적정 — 유지",
    "B. 시세·품셈 직접확정": "DB명 없음 — 유지(시세·품셈)",
    "C. 오매칭 교정완료": "보수적 재산출 반영 완료",
    "D. 라벨불일치-검토요": "유지 가능/과소의심 혼재 — 검토 권장",
}
for k in sorted(cnt, key=lambda c: order.get(c[0], 9)):
    sm.append([k, cnt[k], amt[k], ACT.get(k, "")])
sm.append([])
sm.append(["합계", sum(cnt.values()), sum(amt.values()), ""])
sm["A1"].font = Font(bold=True, size=13)

ws2 = out.create_sheet("품목검증")
heads = ["분류", "파일", "품명", "규격", "단위", "수량", "확정단가", "확정금액",
         "DB_매칭품명", "최고점수", "판정사유"]
ws2.append(heads)
fillmap = {
    "A": PatternFill("solid", fgColor="E8F5E9"),
    "B": PatternFill("solid", fgColor="E3F2FD"),
    "C": PatternFill("solid", fgColor="FFF9C4"),
    "D": PatternFill("solid", fgColor="FFEBEE"),
}
for c in range(1, len(heads) + 1):
    cell = ws2.cell(1, c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="455A64")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
for x in recs:
    ws2.append(x)
    ridx = ws2.max_row
    f = fillmap.get(x[0][0])
    if f:
        for c in range(1, len(heads) + 1):
            ws2.cell(ridx, c).fill = f
    for c in (7, 8):
        ws2.cell(ridx, c).number_format = '#,##0'
ws2.freeze_panes = "A2"
for i, w in enumerate([16, 11, 24, 16, 6, 9, 11, 13, 22, 8, 40], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out.save(OUT)
print(f"보고서 저장 → {OUT}")
for k in sorted(cnt, key=lambda c: order.get(c[0], 9)):
    print(f"  {k}: {cnt[k]}건  확정금액 {amt[k]:,.0f}원")
print(f"  합계: {sum(cnt.values())}건  {sum(amt.values()):,.0f}원")
