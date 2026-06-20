#!/usr/bin/env python3
"""4개 일위대가 산출 파일 + 통합내역의 행 범위·확정단가 현황 조사(읽기 전용)."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"

IL_FILES = [
    ("미매칭_일위대가산출.xlsx", BASE / "미매칭_일위대가산출.xlsx"),
    ("검토_일위대가산출.xlsx", BASE / "검토_일위대가산출.xlsx"),
    ("검토_토공_일위대가산출.xlsx", BASE / "검토_토공_일위대가산출.xlsx"),
    ("검토_공종별_일위대가산출.xlsx", BASE / "검토_공종별_일위대가산출.xlsx"),
]


def main():
    for name, p in IL_FILES:
        print("=" * 70)
        print(name, "존재" if p.exists() else "없음")
        if not p.exists():
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        print("  시트:", wb.sheetnames)
        # 메인 시트(일위대가산출) 우선
        sname = "일위대가산출" if "일위대가산출" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  (빈 시트)")
            wb.close()
            continue
        hdr = [str(c or "").strip() for c in rows[0]]
        print(f"  메인시트[{sname}] 헤더:", hdr)
        def cidx(*names):
            for n in names:
                if n in hdr:
                    return hdr.index(n)
            return None
        ci_row = cidx("행")
        ci_file = cidx("파일", "내역서")
        ci_conf = cidx("확정단가(입력)")
        ci_sug = cidx("제시단가")
        per_file = {}
        conf_cnt = 0
        sug_cnt = 0
        for r in rows[1:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            fl = str(r[ci_file]).strip() if ci_file is not None and ci_file < len(r) else ""
            rv = r[ci_row] if ci_row is not None and ci_row < len(r) else None
            try:
                rv = int(rv)
            except (TypeError, ValueError):
                rv = None
            d = per_file.setdefault(fl, {"n": 0, "min": 10**9, "max": -1})
            d["n"] += 1
            if rv is not None:
                d["min"] = min(d["min"], rv)
                d["max"] = max(d["max"], rv)
            if ci_conf is not None and ci_conf < len(r):
                v = r[ci_conf]
                if v is not None and str(v).strip() not in ("", "0"):
                    conf_cnt += 1
            if ci_sug is not None and ci_sug < len(r):
                v = r[ci_sug]
                if v is not None and str(v).strip() not in ("", "0"):
                    sug_cnt += 1
        print(f"  확정단가(입력) 채워진 행: {conf_cnt}  / 제시단가 채워진 행: {sug_cnt}")
        for fl, d in sorted(per_file.items()):
            mn = d["min"] if d["max"] >= 0 else "-"
            print(f"    파일[{fl}] {d['n']}건  행 {mn}~{d['max']}")
        wb.close()

    # 통합내역 행 범위(01 토목)
    print("=" * 70)
    tok = WORK / "01_화성 청원지구 토목_표준단가산출.xlsx"
    if tok.exists():
        wb = load_workbook(tok, read_only=True, data_only=True)
        ws = wb["통합내역"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c or "").strip() for c in rows[0]]
        ci = hdr.index("행")
        vals = []
        for r in rows[1:]:
            try:
                vals.append(int(r[ci]))
            except (TypeError, ValueError):
                pass
        print(f"01 토목 통합내역 행 범위: {min(vals)}~{max(vals)} (총 {len(vals)}행)")
        wb.close()


if __name__ == "__main__":
    main()
