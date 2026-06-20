#!/usr/bin/env python3
"""총괄표 원가계산서 요율 vs 03 전기설비 원가 시트 요율 비교."""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from calc_overhead import ELECTRIC_RATES, ELECTRIC_RATE_TABLE  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
ELECTRIC_XLSX = ROOT / "05_내역서" / "공내역서" / "03_화성 청원지구 전기설비.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
DIFF_FILL = PatternFill("solid", fgColor="FCE4D6")
SAME_FILL = PatternFill("solid", fgColor="E2EFDA")
ONLY_FILL = PatternFill("solid", fgColor="FFF2CC")
PCT_FMT = "0.00%"

# (표시명, 총괄 RATES 키, 03 원가 시트 정규화명 후보)
COMPARE_ITEMS: list[tuple[str, str | None, list[str]]] = [
    ("간접노무비", "간접노무비", ["간접노무비"]),
    ("산재보험료", "산재보험료", ["산재보험료"]),
    ("고용보험료", "고용보험료", ["고용보험료"]),
    ("국민건강보험료", "건강보험료", ["건강보험료"]),
    ("국민연금보험료", "연금보험료", ["연금보험료"]),
    ("노인장기요양보험료", "노인장기요양", ["노인장기요양보험", "노인장기요양"]),
    ("안전관리비", "안전관리비", ["안전관리비"]),
    ("기타경비", "기타경비", ["기타경비"]),
    ("일반관리비", "일반관리비", ["일반관리비"]),
    ("이윤", "이윤", ["이윤"]),
    ("부가가치세", "부가가치세", ["부가가치세"]),
    ("퇴직금(퇴직공제부금)", "퇴직공제", ["퇴직금공제"]),
    ("석면분담금", "석면분담금", ["석면분담금"]),
    ("임금채권부담금", "임금채권부담금", ["임금채권부담금"]),
]


def _norm(name: str) -> str:
    return re.sub(r"[\s()·]", "", str(name or ""))


def load_electric_rates(path: Path | None = None) -> tuple[dict[str, dict], dict[str, Any]]:
    """03 전기설비.xlsx 「원가」 시트에서 요율·산식 추출."""
    p = path or ELECTRIC_XLSX
    if not p.exists():
        return {}, {"path": str(p), "error": "파일 없음"}

    wb = load_workbook(p, read_only=True, data_only=True)
    if "원가" not in wb.sheetnames:
        wb.close()
        return {}, {"path": p.name, "error": "「원가」 시트 없음"}

    ws = wb["원가"]
    meta: dict[str, Any] = {
        "path": p.name,
        "project": ws.cell(3, 3).value,
        "period": ws.cell(3, 5).value,
    }
    rates: dict[str, dict] = {}
    extras: dict[str, Any] = {}

    for r in range(4, ws.max_row + 1):
        name_raw = ws.cell(r, 3).value or ws.cell(r, 1).value
        if not name_raw or str(name_raw).strip() in ("구  분", "비 목"):
            continue
        name = str(name_raw).strip()
        key = _norm(name)
        if key.startswith("소계") or key.startswith("순공사") or key.startswith("총원가") or key.startswith("총계"):
            continue

        rate = ws.cell(r, 9).value
        basis = ws.cell(r, 5).value
        amount = ws.cell(r, 11).value or ws.cell(r, 4).value

        if "한전수탁" in key:
            extras["한전수탁비"] = {
                "label": "한전수탁비",
                "amount": amount,
                "basis": "고정액(요율 아님)",
            }
            continue

        if rate is None:
            continue

        rates[key] = {
            "label": name,
            "rate": float(rate) / 100,
            "rate_pct": float(rate),
            "basis": str(basis or "").strip(),
            "row": r,
        }

    wb.close()
    meta["extras"] = extras
    return rates, meta


def _find_electric(rates: dict, keys: list[str]) -> dict | None:
    for k in keys:
        nk = _norm(k)
        if nk in rates:
            return rates[nk]
        for rk, rv in rates.items():
            if nk in rk or rk in nk:
                return rv
    return None


def _consolidated_item(key: str) -> dict | None:
    for label, k, basis in ELECTRIC_RATE_TABLE:
        if k == key:
            return {"label": label, "rate": ELECTRIC_RATES[k], "basis": basis}
    return None


def build_comparison(electric_path: Path | None = None) -> tuple[list[dict], dict]:
    elec, meta = load_electric_rates(electric_path)
    rows: list[dict] = []

    for display, con_key, elec_keys in COMPARE_ITEMS:
        c = _consolidated_item(con_key) if con_key else None
        e = _find_electric(elec, elec_keys) if elec_keys else None

        c_rate = c["rate"] if c else None
        e_rate = e["rate"] if e else None
        diff = (e_rate - c_rate) if c_rate is not None and e_rate is not None else None

        if c and not e:
            note = "적용만"
        elif e and not c:
            note = "03 원본만"
        elif diff is not None and abs(diff) >= 0.001:
            note = f"차이 {diff * 100:+.2f}%p"
        elif diff is not None:
            note = "동일"
        else:
            note = ""

        rows.append(
            {
                "item": display,
                "con_rate": c_rate,
                "con_basis": c["basis"] if c else "",
                "elec_rate": e_rate,
                "elec_basis": e["basis"] if e else "",
                "diff": diff,
                "note": note,
            }
        )

    hanjeon = meta.get("extras", {}).get("한전수탁비")
    if hanjeon:
        rows.append(
            {
                "item": "한전수탁비",
                "con_rate": None,
                "con_basis": "",
                "elec_rate": None,
                "elec_basis": hanjeon["basis"],
                "elec_amount": hanjeon.get("amount"),
                "diff": None,
                "note": "03 전기만 · 고정액",
            }
        )

    return rows, meta


def write_rate_compare_sheet(wb: Workbook, electric_path: Path | None = None) -> None:
    rows, meta = build_comparison(electric_path)
    ws = wb.create_sheet("요율비교")

    ws.append(["원가계산서 요율 검증 — 총괄표(03 전기 요율 적용) vs 03 원본"])
    ws.append(["작성일", "2026. 6. 19."])
    ws.append(["총괄표", "05_내역서/총괄표.xlsx · 원가계산서 · ELECTRIC_RATES 적용"])
    ws.append(["03 전기", f"05_내역서/공내역서/{meta.get('path', ELECTRIC_XLSX.name)} · 원가 시트"])
    ws.append(["03 공사명", meta.get("project", ""), meta.get("period", "")])
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["【요율 대조표】"])
    headers = [
        "No", "항목", "총괄(적용) 요율", "총괄 적용기준",
        "03 원본 요율", "03 원본 적용기준", "차이(03−적용)", "비고",
    ]
    ws.append(headers)
    hdr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for i, r in enumerate(rows, 1):
        elec_val = r["elec_rate"]
        if elec_val is None and r.get("item") == "한전수탁비":
            elec_val = r.get("elec_amount")
        ws.append([
            i, r["item"], r["con_rate"], r["con_basis"],
            elec_val, r["elec_basis"], r["diff"], r["note"],
        ])
        rr = ws.max_row
        if isinstance(r["con_rate"], float):
            ws.cell(rr, 3).number_format = PCT_FMT
        if isinstance(r["elec_rate"], float):
            ws.cell(rr, 5).number_format = PCT_FMT
        elif r.get("item") == "한전수탁비" and r.get("elec_amount"):
            ws.cell(rr, 5).number_format = "#,##0"
        if isinstance(r["diff"], float):
            ws.cell(rr, 7).number_format = PCT_FMT

        fill = None
        if r["note"] == "동일":
            fill = SAME_FILL
        elif "원본만" in r["note"] or r["note"] == "적용만" or "고정액" in r["note"]:
            fill = ONLY_FILL
        elif r["diff"] is not None and abs(r["diff"]) >= 0.005:
            fill = DIFF_FILL
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(rr, c).fill = fill

    ws.append([])
    ws.append(["【주요 차이 요약】"])
    big = [r for r in rows if r.get("diff") is not None and abs(r["diff"]) >= 0.01]
    big.sort(key=lambda x: -abs(x["diff"]))
    for r in big[:8]:
        ws.append([
            "", r["item"],
            r["con_rate"], "",
            r["elec_rate"], "",
            r["diff"], r["note"],
        ])
        rr = ws.max_row
        ws.cell(rr, 3).number_format = PCT_FMT
        ws.cell(rr, 5).number_format = PCT_FMT
        ws.cell(rr, 7).number_format = "+0.00%;-0.00%"

    ws.append([])
    nr = ws.max_row + 1
    ws.cell(
        nr,
        1,
        "※ 총괄표 원가계산서는 ELECTRIC_RATES(03 전기 원가 시트와 동일 수치)를 적용. "
        "본 시트는 03 xlsx 원본과의 일치 여부 검증용. "
        "한전수탁비(5,131,000원)는 총괄 원가계산서에 미포함.",
    )
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=8)
    ws.cell(nr, 1).alignment = Alignment(wrap_text=True, vertical="top")

    for i, w in enumerate([5, 22, 12, 28, 12, 28, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    rows, meta = build_comparison()
    print(f"03: {meta.get('project')} / {meta.get('period')}")
    print(f"{'항목':<22} {'총괄':>8} {'03전기':>8} {'차이':>10}  비고")
    for r in rows:
        c = f"{r['con_rate'] * 100:g}%" if r["con_rate"] is not None else "-"
        e = f"{r['elec_rate'] * 100:g}%" if r["elec_rate"] is not None else (
            f"{r.get('elec_amount'):,.0f}원" if r.get("elec_amount") else "-"
        )
        d = f"{r['diff'] * 100:+.2f}%p" if r["diff"] is not None else "-"
        print(f"{r['item']:<22} {c:>8} {e:>8} {d:>10}  {r['note']}")


if __name__ == "__main__":
    main()
