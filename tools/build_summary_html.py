#!/usr/bin/env python3
"""총괄표.xlsx → 내역서_표준단가산출_총괄표.html 자동 생성.
총괄·구성비·원가계산서 시트를 읽어 현행값으로 HTML을 재생성한다."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
SRC = BASE / "총괄표.xlsx"
OUT = BASE / "내역서_표준단가산출_총괄표.html"
GEN_DATE = "2026. 6. 20."

# 공종별 「주요 미매칭 원인」 정성 설명(정적)
CAUSE = {
    "토목": "화강석 경계석 보정 필요(A) · 특수관/이형관 규격 유사도 미달(B)",
    "조경": "조경 수목 자재가 표준단가 폐지(A) · 조경일위·forestinfo 별도 보완",
    "전기설비": "파일 내장 일위대가·단가조사로 209/209 산출(100%)",
    "진입도로": "경계석류·부대공 특수품목 유사도 미달(B) · 제경비·VAT 식(1) 요율행(D)",
    "화성 청원로(회전교차로)": "경계석류·표지·신호 부대품목 유사도 미달(B) · 식(1) 요율행(D)",
    "개발행위": "규격 세분·특수 품목 유사도 미달(B) · 식(1) 요율행(D)",
    "건설폐기물처리": "없음(협회단가 100% 적용)",
}


def won(v) -> str:
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return str(v or "")


def eok(v) -> str:
    return f"{float(v) / 1e8:,.2f}"


def badge(rate: float) -> str:
    cls = "b-ok" if rate >= 0.95 else ("b-mid" if rate >= 0.65 else "b-bad")
    return f'<span class="badge {cls}">{rate * 100:.1f}%</span>'


def load():
    wb = load_workbook(SRC, data_only=True)
    tw = wb["총괄"]
    secs = []
    for r in range(7, 14):
        secs.append({
            "no": tw.cell(r, 1).value, "name": tw.cell(r, 2).value,
            "src_file": tw.cell(r, 3).value, "src": tw.cell(r, 4).value,
            "m": tw.cell(r, 5).value or 0, "rv": tw.cell(r, 6).value or 0,
            "um": tw.cell(r, 7).value or 0, "all": tw.cell(r, 8).value or 0,
            "rate": tw.cell(r, 9).value or 0, "mat": tw.cell(r, 10).value or 0,
            "lab": tw.cell(r, 11).value or 0, "exp": tw.cell(r, 12).value or 0,
            "tot": tw.cell(r, 13).value or 0, "note": tw.cell(r, 15).value or "",
        })
    tot = {
        "m": tw.cell(15, 5).value, "rv": tw.cell(15, 6).value, "um": tw.cell(15, 7).value,
        "all": tw.cell(15, 8).value, "rate": tw.cell(15, 9).value,
        "mat": tw.cell(15, 10).value, "lab": tw.cell(15, 11).value,
        "exp": tw.cell(15, 12).value, "tot": tw.cell(15, 13).value,
    }
    cw = wb["구성비"]
    comp = {cw.cell(r, 1).value: (cw.cell(r, 2).value, cw.cell(r, 3).value) for r in range(2, 6)}
    ow = wb["원가계산서"]
    cost = {"direct": ow.cell(4, 2).value, "dogup": ow.cell(5, 2).value}
    cost_rows = []
    for r in range(26, 49):
        step, item, amt, _eok, formula = (ow.cell(r, c).value for c in range(1, 6))
        if item is None:
            continue
        cost_rows.append((step, item, amt, formula))
    # 공급가액·부가세
    for r in range(26, 49):
        if ow.cell(r, 2).value == "공급가액":
            cost["supply"] = ow.cell(r, 3).value
        if ow.cell(r, 2).value == "부가가치세":
            cost["vat"] = ow.cell(r, 3).value
    wb.close()
    return secs, tot, comp, cost, cost_rows


def render() -> str:
    secs, tot, comp, cost, cost_rows = load()
    mat_amt, mat_pct = comp["재료비"]
    lab_amt, lab_pct = comp["노무비"]
    exp_amt, exp_pct = comp["경비"]

    rows1 = "\n".join(
        f'    <tr><td class="c">{s["no"]}</td><td class="l">{s["name"]}</td>'
        f'<td class="r">{won(s["mat"])}</td><td class="r">{won(s["lab"])}</td>'
        f'<td class="r">{won(s["exp"])}</td><td class="r">{won(s["tot"])}</td>'
        f'<td class="r">{eok(s["tot"])}</td></tr>'
        for s in secs)

    rows3 = "\n".join(
        f'    <tr><td class="c">{s["no"]}</td><td class="l">{s["name"]}</td>'
        f'<td class="r">{s["m"]}</td><td class="r">{s["rv"]}</td><td class="r">{s["um"]}</td>'
        f'<td class="r">{s["all"]}</td><td class="c">{badge(float(s["rate"]))}</td>'
        f'<td class="l">{CAUSE.get(s["name"], "")}</td></tr>'
        for s in secs)

    rows7 = "\n".join(
        f'    <tr><td class="c">{s["no"]}</td><td class="l"><code>{s["src_file"]}</code></td>'
        f'<td class="l">{s["src"]}{(" · " + str(s["note"])) if s["note"] else ""}</td></tr>'
        for s in secs)

    cost_html = []
    for step, item, amt, formula in cost_rows:
        is_total = step in ("①", "⑥", "★")
        is_sub = item.strip().startswith("·")
        is_mid = item.startswith(("노무비계", "경비계", "공급가액", "부가가치세"))
        cls = ' class="total"' if is_total else ""
        stepc = step or ""
        itemcell = (f'<td class="l" style="padding-left:22px">{item}</td>' if is_sub
                    else (f'<td class="l"><b>{item}</b></td>' if is_mid
                          else f'<td class="l">{item}</td>'))
        amtcell = (f'<td class="r"><b>{won(amt)}</b></td>' if is_mid
                   else f'<td class="r">{won(amt)}</td>')
        cost_html.append(
            f'    <tr{cls}><td class="c">{stepc}</td>{itemcell}{amtcell}'
            f'<td class="l">{formula or ""}</td></tr>')
    cost_rows_html = "\n".join(cost_html)

    indirect = float(cost["supply"]) - float(cost["direct"])
    mult = float(cost["dogup"]) / float(cost["direct"])

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 공내역서 — 표준단가 산출 총괄표</title>
<style>
  :root{{--navy:#1f3a5f;--blue:#2c5aa0;--line:#d9e1ec;--bg:#f4f6fa;
        --ok:#1a7f37;--warn:#b25e00;--bad:#b3261e;--mid:#7a5cc0;}}
  *{{box-sizing:border-box;}}
  body{{font-family:"Malgun Gothic","맑은 고딕",system-ui,sans-serif;margin:0;
       background:var(--bg);color:#1c2430;line-height:1.6;font-size:15px;}}
  .wrap{{max-width:1080px;margin:0 auto;padding:32px 22px 64px;}}
  header.doc{{border-bottom:3px solid var(--navy);padding-bottom:16px;margin-bottom:8px;}}
  header.doc h1{{font-size:23px;color:var(--navy);margin:0 0 6px;letter-spacing:-.5px;}}
  .meta{{color:#5a6675;font-size:13px;}} .meta b{{color:#36424f;}}
  h2{{font-size:18px;color:var(--navy);margin:30px 0 10px;padding-left:10px;border-left:5px solid var(--blue);}}
  .kpis{{display:flex;flex-wrap:wrap;gap:12px;margin:18px 0;}}
  .kpi{{flex:1 1 180px;background:#fff;border:1px solid var(--line);border-radius:10px;
       padding:14px 16px;box-shadow:0 1px 3px rgba(0,0,0,.04);}}
  .kpi .lab{{font-size:12.5px;color:#6b7785;}} .kpi .val{{font-size:22px;font-weight:800;color:var(--navy);margin-top:2px;}}
  .kpi .sub{{font-size:12px;color:#8a94a3;margin-top:2px;}}
  .summary{{background:#fff;border:1px solid var(--line);border-left:5px solid var(--blue);
           border-radius:8px;padding:14px 20px;margin:16px 0;}}
  .summary ol{{margin:0;padding-left:22px;}} .summary li{{margin:5px 0;}}
  table{{width:100%;border-collapse:collapse;background:#fff;margin:8px 0 4px;
        border:1px solid var(--line);border-radius:6px;overflow:hidden;font-size:12.8px;}}
  th,td{{padding:8px 9px;border:1px solid var(--line);vertical-align:middle;}}
  th{{background:#eaf0f8;color:var(--navy);font-weight:700;text-align:center;}}
  td.l{{text-align:left;}} td.r{{text-align:right;white-space:nowrap;}} td.c{{text-align:center;white-space:nowrap;}}
  tfoot td,tr.total td{{font-weight:800;background:#fff2cc;}}
  .bar{{height:9px;border-radius:5px;background:#e6ecf5;overflow:hidden;}}
  .bar>span{{display:block;height:100%;}}
  .mat{{background:#5b8def;}} .lab{{background:#27ae60;}} .exp{{background:#e67e22;}}
  .badge{{display:inline-block;padding:1px 7px;border-radius:10px;font-size:11px;font-weight:700;color:#fff;}}
  .b-ok{{background:var(--ok);}} .b-mid{{background:var(--mid);}} .b-warn{{background:var(--warn);}} .b-bad{{background:var(--bad);}}
  .note{{background:#fff7ed;border:1px solid #f0c98a;border-radius:8px;padding:12px 18px;margin:14px 0;font-size:13.5px;}}
  .note b{{color:var(--warn);}}
  .note ol{{margin:6px 0 0;padding-left:20px;}} .note li{{margin:4px 0;}}
  .ok-box{{background:#eaf7ee;border:1px solid #a9d8b8;border-left:5px solid var(--ok);border-radius:8px;padding:12px 18px;margin:14px 0;font-size:13.5px;}}
  code{{background:#eef2f8;padding:1px 5px;border-radius:4px;font-size:12px;}}
  footer{{margin-top:40px;padding-top:14px;border-top:1px solid var(--line);font-size:12px;color:#8a94a3;}}
</style>
</head>
<body>
<div class="wrap">

<header class="doc">
  <h1>화성 청원지구 공내역서 — 표준단가 산출 총괄표</h1>
  <div class="meta">
    <b>작성일</b> {GEN_DATE} &nbsp;|&nbsp;
    <b>범위</b> 01·02·04·05·06·07 공내역서 표준단가·협회단가 산출 결과 통합 &nbsp;|&nbsp;
    <b>합계 산식</b> 01+02+04+05+06+07 (<b>03 전기(지구외)=02 중복·#REF! → 제외</b>)
  </div>
  <div class="meta" style="margin-top:4px">
    <b>금액 성격</b> 직접공사비(재료·노무·경비) 추정 — <b>제경비·부가가치세·원단위 절사 미포함</b>
  </div>
</header>

<div class="meta" style="margin-top:12px;padding:10px 14px;background:#f1f8f2;border:1px solid #bfe0c5;border-radius:8px;font-size:13.5px">
  <b>확정단가 반영({GEN_DATE}).</b> 본 문서는 <code>총괄표.xlsx</code>에서 자동 생성된 <b>최종 직접공사비</b>이다(<code>tools/build_summary_html.py</code>). 교정단가·baseline 대비 열은 포함하지 않는다.
</div>

<div class="kpis">
  <div class="kpi"><div class="lab">직접공사비 합계</div><div class="val">약 {eok(tot["tot"])}억</div><div class="sub">{won(tot["tot"])}원</div></div>
  <div class="kpi"><div class="lab">재료비</div><div class="val">{eok(mat_amt)}억</div><div class="sub">{won(mat_amt)}원 · {mat_pct * 100:.1f}%</div></div>
  <div class="kpi"><div class="lab">노무비</div><div class="val">{eok(lab_amt)}억</div><div class="sub">{won(lab_amt)}원 · {lab_pct * 100:.1f}%</div></div>
  <div class="kpi"><div class="lab">경비</div><div class="val">{eok(exp_amt)}억</div><div class="sub">{won(exp_amt)}원 · {exp_pct * 100:.1f}%</div></div>
  <div class="kpi"><div class="lab">매칭률</div><div class="val">{float(tot["rate"]) * 100:.1f}%</div><div class="sub">매칭 {tot["m"]} / 전체 {tot["all"]:,}건</div></div>
</div>

<div class="summary">
  <ol>
    <li><b>직접공사비 합계</b> — <b>약 {eok(tot["tot"])}억원</b>({won(tot["tot"])}원). <span class="meta">※ 확정단가 반영 후 최종값. 03 전기(지구외)는 02와 동일 내역서(중복·#REF!)로 합계에서 제외.</span></li>
    <li><b>재료비</b> {eok(mat_amt)}억 · <b>노무비</b> {eok(lab_amt)}억 · <b>경비</b> {eok(exp_amt)}억.</li>
    <li><b>매칭</b> {tot["m"]} / <b>검토</b> {tot["rv"]} / <b>미매칭</b> {tot["um"]}(01 조경 제외) / <b>전체</b> {tot["all"]:,}건(매칭률 <b>{float(tot["rate"]) * 100:.1f}%</b>).</li>
    <li>금액은 <b>매칭+검토+확정</b> 건 반영. 미매칭·미산출은 0(누락분이므로 실제 총액은 더 커질 수 있음).</li>
  </ol>
</div>

<h2>1. 내역서별 직접공사비</h2>
<table>
  <thead><tr>
    <th style="width:34px">No</th><th style="width:170px">구분</th>
    <th style="width:110px">재료비</th><th style="width:110px">노무비</th><th style="width:100px">경비</th>
    <th style="width:120px">합계</th><th style="width:62px">합계(억)</th>
  </tr></thead>
  <tbody>
{rows1}
  </tbody>
  <tfoot><tr class="total"><td class="c">계</td><td class="l">01·02·04~07</td><td class="r">{won(tot["mat"])}</td><td class="r">{won(tot["lab"])}</td><td class="r">{won(tot["exp"])}</td><td class="r">{won(tot["tot"])}</td><td class="r">{eok(tot["tot"])}</td></tr></tfoot>
</table>
<p class="meta" style="margin-top:4px">※ <b>03 전기(지구외) 제외</b> — 02와 품목 209개가 100% 동일한 복사본이고 값 셀이 전부 <code>#REF!</code>(깨진 수식). 02 파일이 지구내+지구외 전체를 포함하므로 03을 더하면 이중계상됨.</p>

<h2>2. 재료·노무·경비 구성</h2>
<table>
  <thead><tr><th style="width:90px">구분</th><th style="width:140px">금액(원)</th><th style="width:70px">비율</th><th>구성</th></tr></thead>
  <tbody>
    <tr><td class="c">재료비</td><td class="r">{won(mat_amt)}</td><td class="c">{mat_pct * 100:.1f}%</td><td><div class="bar"><span class="mat" style="width:{mat_pct * 100:.1f}%"></span></div></td></tr>
    <tr><td class="c">노무비</td><td class="r">{won(lab_amt)}</td><td class="c">{lab_pct * 100:.1f}%</td><td><div class="bar"><span class="lab" style="width:{lab_pct * 100:.1f}%"></span></div></td></tr>
    <tr><td class="c">경비</td><td class="r">{won(exp_amt)}</td><td class="c">{exp_pct * 100:.1f}%</td><td><div class="bar"><span class="exp" style="width:{exp_pct * 100:.1f}%"></span></div></td></tr>
  </tbody>
  <tfoot><tr class="total"><td class="c">합계</td><td class="r">{won(tot["tot"])}</td><td class="c">100%</td><td></td></tr></tfoot>
</table>

<h2>3. 매칭·검토·미매칭 현황 (내역서별)</h2>
<table>
  <thead><tr>
    <th style="width:30px">No</th><th style="width:150px">구분</th>
    <th style="width:54px">매칭</th><th style="width:54px">검토</th><th style="width:60px">미매칭</th><th style="width:50px">전체</th>
    <th style="width:64px">매칭률</th><th>주요 미매칭 원인</th>
  </tr></thead>
  <tbody>
{rows3}
  </tbody>
  <tfoot><tr class="total"><td class="c">계</td><td class="l">01·02·04~07</td><td class="r">{tot["m"]}</td><td class="r">{tot["rv"]}</td><td class="r">{tot["um"]}</td><td class="r">{tot["all"]:,}</td><td class="c">{float(tot["rate"]) * 100:.1f}%</td><td class="l">미매칭은 01 조경 20건 제외 기준(전체 미매칭 122건)</td></tr></tfoot>
</table>
<p class="meta" style="margin-top:4px">※ <b>매칭</b>=단가 매칭 확정 · <b>검토</b>=확정·환산·재산출 경로 · <b>미매칭</b>=잔여 미산출. 금액은 매칭+검토+확정 건만 반영.</p>

<h2>4. 간접비 자동계산 — 총공사비(도급액)</h2>
<div class="note">
  <b>적용 요율 = 03 전기설비 「원가」 시트 제비율.</b> 발주처 지정 요율이 확정되면 <code>tools/calc_overhead.py</code>의 <code>ELECTRIC_RATES</code>만 교체하면 결과가 갱신됩니다. 직접공사비 합계(03 제외)에 일괄 적용한 <b>개략 추정</b>이며, 한전수탁비 등 전기 전용 고정액은 미포함입니다.
</div>
<div class="kpis">
  <div class="kpi"><div class="lab">직접공사비</div><div class="val">{eok(cost["direct"])}억</div><div class="sub">{won(cost["direct"])}원</div></div>
  <div class="kpi"><div class="lab">간접비·일반관리비·이윤</div><div class="val">+{eok(indirect)}억</div><div class="sub">공급가액 − 직접공사비</div></div>
  <div class="kpi"><div class="lab">공급가액</div><div class="val">{eok(cost["supply"])}억</div><div class="sub">{won(cost["supply"])}원</div></div>
  <div class="kpi"><div class="lab">부가가치세</div><div class="val">{eok(cost["vat"])}억</div><div class="sub">{won(cost["vat"])}원</div></div>
  <div class="kpi"><div class="lab">도급액(총공사비)</div><div class="val">약 {eok(cost["dogup"])}억</div><div class="sub">{won(cost["dogup"])}원 · 직접비의 {mult:.3f}배</div></div>
</div>

<h3 style="margin:14px 0 6px;font-size:14.5px;color:#243b58">산출 내역</h3>
<table>
  <thead><tr><th style="width:40px">단계</th><th style="width:210px">항목</th><th style="width:140px">금액(원)</th><th>산식</th></tr></thead>
  <tbody>
{cost_rows_html}
  </tbody>
</table>
<div class="note">
  <b>주의.</b> ① 요율은 03 전기 원가 시트와 동일 수치입니다. ② 미매칭 누락 단가를 보정하면 직접공사비가 늘어 도급액도 비례 증가합니다. ③ 04·05·06 원본 제경비·VAT <code>식(1)</code> 행과 이중계상 여부는 별도 확인이 필요합니다. ④ 상세는 <code>총괄표.xlsx</code> 「원가계산서」·「요율비교」 시트 참조.
</div>

<h2>5. 원본 · 단가 출처</h2>
<table>
  <thead><tr><th style="width:34px">No</th><th style="width:260px">원본 파일</th><th>단가 출처 · 비고</th></tr></thead>
  <tbody>
{rows7}
    <tr style="opacity:.6"><td class="c">03</td><td class="l"><code>03_…전기설비(지구외).xlsx</code></td><td class="l"><b>제외</b> — 02와 품목 209개 100% 동일 복사본, 값 셀 전부 <code>#REF!</code>. 이중계상 방지를 위해 합계·도급액에서 제외</td></tr>
  </tbody>
</table>

<div class="note">
  <b>유의사항</b>
  <ol>
    <li><b>토목(01·04·05·06)</b> — 표준시장단가2026·시장시공가격·표준일위대가2026을 품명+규격 유사도(임계 0.56)로 매칭. 01은 토목·조경 XLS 분리.</li>
    <li><b>경계석류(화강석)</b> — 표준 DB에 화강석 경계석 단가가 없어 콘크리트 「경계블록」으로 환산 매칭 → 재료비 과소 계상 가능, 정밀 견적 시 화강석 시세 보정.</li>
    <li><b>금액 합계</b>는 매칭+검토+확정 건만 반영(미매칭·미산출 0). 세부는 각 <code>내역서작업/*_표준단가산출.xlsx</code> 통합내역 시트 참조.</li>
  </ol>
</div>

<footer>
  화성 청원지구 공내역서 표준단가 산출 총괄표 — 직접공사비(제경비·VAT 미포함) 기준. {GEN_DATE} 자동 생성(총괄표.xlsx). 끝.
</footer>

</div>
</body>
</html>
"""


def main() -> None:
    html = render()
    OUT.write_text(html, encoding="utf-8")
    print(f"생성 완료 → {OUT}  ({len(html):,} bytes)")


if __name__ == "__main__":
    main()
