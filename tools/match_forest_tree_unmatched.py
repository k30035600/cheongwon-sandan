#!/usr/bin/env python3
"""01 조경 미매칭(주 단위) × forestinfo 조경수 관측시세 시뮬레이션 — 산출 미수정.

출력: 05_내역서/조경수_미매칭점검.xlsx
"""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
ROOT = TOOLS.parent
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
CSV_PATH = BASE / "일위대가DB" / "조경수_관측시세.csv"
OUT = BASE / "조경수_미매칭점검.xlsx"

sys.path.insert(0, str(TOOLS))
import apply_standard_prices as asp  # noqa: E402

LABELS = {"01j": "01 조경"}
JO_XLSX = "01_화성 청원지구 조경_표준단가산출.xlsx"
THRESH = asp.THRESHOLD
REVIEW = asp.REVIEW_THRESHOLD

SPECIES_ALIAS = {
    "스트로브잣": "스트로브잣나무",
    "영산홍": "철쭉 영산홍",
    "산철쭉": "철쭉 산철쭉",
    "백철쭉": "철쭉 백철쭉",
    "자산홍": "철쭉 자산홍",
}

HDR = PatternFill("solid", fgColor="D9E1F2")
OKF = PatternFill("solid", fgColor="E2EFDA")
RVF = PatternFill("solid", fgColor="FFF2CC")
NMF = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)


def clean_tree_name(raw: str) -> str:
    s = re.sub(r"^\s*[가-힣]\)\.\s*", "", raw.strip())
    s = re.sub(r"^\s*\d+\)\.\s*", "", s)
    return SPECIES_ALIAS.get(s, s)


def norm_spec(s: str) -> str:
    s = s.strip().replace("×", "X").replace("Ⅹ", "X").replace("x", "X")
    return re.sub(r"\s+", "", s).upper()


def parse_dims(spec: str) -> dict[str, float]:
    s = norm_spec(spec)
    dims: dict[str, float] = {}
    for m in re.finditer(r"([HRBW])([\d.]+)", s):
        dims[m.group(1)] = float(m.group(2))
    if not dims and re.match(r"^R[\d.]+$", s):
        dims["R"] = float(s[1:])
    return dims


def spec_distance(item_spec: str, row_spec: str) -> float:
    """0=일치에 가까움, 클수록 불일치."""
    a, b = parse_dims(item_spec), parse_dims(norm_spec(row_spec))
    if not a or not b:
        return 1.0 if norm_spec(item_spec) != norm_spec(row_spec) else 0.0
    keys = set(a) | set(b)
    err = 0.0
    for k in keys:
        if k not in a or k not in b:
            err += 0.35
        else:
            err += abs(a[k] - b[k]) / max(a[k], b[k], 0.1)
    return err / max(len(keys), 1)


def newest_xlsx() -> Path | None:
    for d in (WORK, BASE):
        p = d / JO_XLSX
        if p.exists():
            return p
    cands = list(BASE.glob("01*조경*_표준단가산출.xlsx")) + list(WORK.glob("01*조경*_표준단가산출.xlsx"))
    return max(cands, key=lambda p: p.stat().st_mtime) if cands else None


def load_unmatched_trees() -> list[dict]:
    p = newest_xlsx()
    if not p:
        return []
    items = []
    wb = load_workbook(p, read_only=True, data_only=True)
    for sname in ("미매칭", "미산출"):
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [("" if h is None else str(h).strip()) for h in rows[0]]
        idx = {h: i for i, h in enumerate(hdr)}

        def col(r, *names):
            for n in names:
                for k, j in idx.items():
                    if n in k and j < len(r):
                        return r[j]
            return None

        for r in rows[1:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            unit = str(col(r, "단위") or "").strip()
            if unit != "주":
                continue
            name = str(col(r, "공종명", "명칭") or "").strip()
            spec = str(col(r, "규격") or "").strip()
            qty = col(r, "수량")
            try:
                qty = float(qty) if qty is not None else 0
            except (TypeError, ValueError):
                qty = 0
            items.append({
                "ledger": LABELS["01j"],
                "row": col(r, "행"),
                "section": str(col(r, "공종") or ""),
                "name_raw": name,
                "name": clean_tree_name(name),
                "spec": spec,
                "unit": unit,
                "qty": qty,
            })
    wb.close()
    return items


def load_prices() -> list[dict]:
    if not CSV_PATH.exists():
        return []
    prices = []
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        for rec in csv.DictReader(f):
            mid = int(float(rec["price_mid"] or 0))
            if mid <= 0:
                continue
            species = rec["species"]
            spec = rec["spec_raw"]
            prices.append({
                "code": f"{species}|{rec['spec_norm']}",
                "name": species,
                "spec": spec,
                "unit": "주",
                "mat": mid, "lab": 0.0, "exp": 0.0, "total": mid,
                "date": f"forestinfo {rec['year']}",
                "_src": f"{species} {spec} ({rec['year']})",
                "_year": int(rec["year"]),
                "_spec_norm": rec["spec_norm"],
            })
    return asp.precompute(prices)


def match_tree(item: dict, prices: list[dict]) -> tuple[dict | None, float, str]:
    """수종명 + 규격 직접 매칭 후 asp.find_best_match 폴백."""
    iname = item["name"]
    ispec = item["spec"]
    best, best_score, best_note = None, -1.0, ""

    species_rows = [p for p in prices if iname in p["name"] or p["name"] in iname
                    or SequenceMatcher(None, asp.norm_text(iname), asp.norm_text(p["name"])).ratio() > 0.72]
    if not species_rows:
        species_rows = prices

    for p in species_rows:
        name_r = SequenceMatcher(None, asp.norm_text(iname), asp.norm_text(p["name"])).ratio()
        if iname in p["name"] or p["name"] in iname:
            name_r = max(name_r, 0.92)
        dist = spec_distance(ispec, p["spec"])
        if norm_spec(ispec) == p.get("_spec_norm") or norm_spec(ispec) == norm_spec(p["spec"]):
            dist = 0.0
        spec_r = max(0.0, 1.0 - min(dist, 1.0))
        score = 0.35 + name_r * 0.4 + spec_r * 0.25
        if score > best_score:
            best, best_score = p, score
            best_note = f"수종 {name_r:.2f} 규격거리 {dist:.2f}"

    if best_score >= THRESH:
        return best, best_score, best_note

    fb, fb_score, _, _ = asp.find_best_match(
        {"name": iname, "spec": ispec, "unit": "주"}, prices
    )
    if fb_score > best_score:
        return fb, fb_score, "키워드매칭"
    return best if best_score >= 0 else (None, best_score, best_note)


def main():
    if not CSV_PATH.exists():
        print("[중단] 조경수_관측시세.csv 없음 — fetch_forest_tree_prices.py 먼저 실행")
        sys.exit(1)
    prices = load_prices()
    items = load_unmatched_trees()
    print(f"조경수 단가 {len(prices):,}건 / 01·주 미매칭 {len(items):,}건")

    WORK.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "점검결과"
    cols = ["내역서", "행", "공종", "명칭", "수종", "규격", "단위", "수량",
            "점수", "상태", "매칭규격", "단가(중간)", "단가(하한~상한)", "산출금액", "비고"]
    ws.append(["forestinfo 조경수 관측시세 — 01 미매칭(주) 해소 가능성(기존 산출 미반영)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([
        "출처: 산림조합중앙회·임산물유통정보시스템 → 가격정보 → 나무 가격정보 → 조경수 관측시세 "
        "(참고가·고시 아님)"
    ])
    ws.append([f"단가원: {CSV_PATH.name} {len(prices):,}건 / 2022년 우선 / 임계 {REVIEW}·{THRESH}"])
    ws.append(cols)
    hdr_row = 4
    for c in range(1, len(cols) + 1):
        ws.cell(row=hdr_row, column=c).fill = HDR
        ws.cell(row=hdr_row, column=c).font = BOLD

    n_ok = n_rv = n_no = 0
    add_sum = 0.0
    buck = defaultdict(lambda: {"ok": 0, "rv": 0, "no": 0})

    for it in items:
        price, score, note = match_tree(it, prices)
        if not price or score < THRESH:
            n_no += 1
            buck["no"]["no"] += 1
            ws.append([it["ledger"], it["row"], it["section"], it["name_raw"], it["name"],
                       it["spec"], it["unit"], it["qty"],
                       round(score, 3) if score >= 0 else None, "미매칭",
                       "", "", "", "", note])
            fill = NMF
        else:
            status = "매칭" if score >= REVIEW else "검토"
            amt = round(it["qty"] * price["total"])
            if status == "매칭":
                n_ok += 1
                buck["no"]["ok"] += 1
                add_sum += amt
            else:
                n_rv += 1
                buck["no"]["rv"] += 1
            lo_hi = ""
            ws.append([it["ledger"], it["row"], it["section"], it["name_raw"], it["name"],
                       it["spec"], it["unit"], it["qty"], round(score, 3), status,
                       price["spec"], price["total"], lo_hi, amt, note])
            fill = OKF if status == "매칭" else RVF
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).fill = fill

    ws2 = wb.create_sheet("요약")
    ws2.append(["구분", "건수"])
    ws2.append(["01·주 미매칭 대상", len(items)])
    ws2.append(["조경수 단가 풀", len(prices)])
    ws2.append(["신규 매칭(높음)", n_ok])
    ws2.append(["검토", n_rv])
    ws2.append(["여전히 미매칭", n_no])
    ws2.append(["해소 가능 합계", n_ok + n_rv])
    ws2.append(["매칭 산출금액(추정)", round(add_sum)])
    ws2.append([])
    ws2.append(["참고", "산림조합중앙회·forestinfo.or.kr 조경수 관측시세. 한국농촌경제원 임업관측 참고가. 2022년 이후 미갱신."])

    wb.save(OUT)
    print(f"매칭 {n_ok} / 검토 {n_rv} / 미매칭 {n_no} (해소 {n_ok + n_rv})")
    print(f"매칭 산출금액(추정): {add_sum:,.0f}원")
    print(f"저장: {OUT}")


if __name__ == "__main__":
    main()
