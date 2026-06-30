#!/usr/bin/env python3
"""공종별 물량 집계 — 통합내역 「공종」× 단위."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from naeyeok_gongjong import resolve_work

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

WORK = Path(__file__).resolve().parents[1] / "05_내역서" / "내역서작업"

FILES = [
    ("05", "05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx"),
    ("04", "04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx"),
    ("01", "01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx"),
    ("06", "06 지구단위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx"),
]

# Ⅱ절 토공 작업량(절+성) SSOT
TOGONG_WORK = {"05": 9839, "04": 29609, "01": 122589, "06": 252965}

# 공종명 → (내역코드 → 통합내역 공종번호)
CAT_SEC: dict[str, dict[str, int]] = {
    "1. 토공": {"05": 1, "04": 1, "01": 1, "06": 1},
    "2. 우수공": {"05": 2, "04": 2, "01": 2},
    "3. 오수공": {"01": 3},
    "4. 상수공": {"04": 3, "01": 4},
    "5. 포장공": {"05": 3, "04": 4, "01": 5},
    "6. 구조물공": {"05": 4, "04": 5, "01": 6},
    "7. 부대공": {"05": 5, "04": 6, "01": 7},
}

CROSS_ROWS = [
    ("1. 토공", "m3", "작업량(절+성)"),
    ("2. 우수공", "m", "관로 연장"),
    ("3. 오수공", "m", "관로 연장"),
    ("4. 상수공", "m", "관로 연장"),
    ("5. 포장공", "m2", "포장 면적"),
    ("6. 구조물공", "m3", "콘크리트·체적"),
    ("7. 부대공", "items", "항목 수"),
]

UNIT_LABEL = {"m3": "㎥", "m2": "㎡", "m": "m", "ton": "ton", "items": "건"}

# 공종별 대표 공정(02 전기 제외 · 공사공정표 SSOT)
CROSS_PERIOD = {
    "1. 토공": "M2~M11",
    "2. 우수공": "M3~M10",
    "3. 오수공": "M8~M10",
    "4. 상수공": "M5~M11",
    "5. 포장공": "M4~M13",
    "6. 구조물공": "M3~M11",
    "7. 부대공": "M5~M13",
}

TOGONG_PERIOD = {
    "05": ("M2~M3", "단지외 1순위 · 2개월"),
    "04": ("M3~M5", "회전교차로 중(M3~) 병행 · 3개월"),
    "01": ("M6~M9", "진입로 통행 개시(M6) 후 · 4개월"),
    "06": ("M6~M11", "최대 물량 · 6개월(토목과 M6~ 병행)"),
}

CONTRACT_GONGJONG = [
    {
        "label": "회전교차로",
        "code": "05",
        "gongjong": "1. 토공 · 2. 우수공 · 3. 포장공 · 4. 구조물공 · 5. 부대공",
        "period": "M2~M5",
    },
    {
        "label": "진입도로",
        "code": "04",
        "gongjong": "1. 토공 · 2. 우수공 · 3. 상수공 · 4. 포장공 · 5. 구조물공 · 6. 부대공",
        "period": "M3~M9",
        "note": "① 중 M3~ · 토공 M3~M5",
    },
    {
        "label": "토목",
        "code": "01",
        "gongjong": (
            "1. 토공 · 2. 우수공 · 3. 오수공 · 4. 상수공 · 5. 포장공 · "
            "6. 구조물공 · 7. 부대공 · 8. 하수처리장 · 9. 지하저류조"
        ),
        "period": "M6~M13",
        "note": "토공 M6~M9",
    },
    {
        "label": "지구단위",
        "code": "06",
        "gongjong": (
            "1. 토공 · 2. A1 · 3. A2 · 4. A3 · 5. A4 · 6. A5 · "
            "7. B1 · 8. B2 · 9. 지원시설 · 10. 주차장"
        ),
        "period": "M6~M12",
        "note": "토공 M6~M11",
    },
    {
        "label": "조경",
        "code": "01",
        "gongjong": "10. 조경공",
        "period": "M10~M13",
    },
]


def sec_num(sec: str) -> int:
    m = re.match(r"^(\d+)", re.sub(r"\s+", "", sec))
    return int(m.group(1)) if m else 0


def norm_unit(u: str) -> str:
    u = str(u or "").strip().replace(" ", "").lower()
    if u in ("m3", "㎥", "m3"):
        return "m3"
    if u in ("m2", "㎡"):
        return "m2"
    if u in ("m", "ｍ"):
        return "m"
    if u in ("ton", "t"):
        return "ton"
    return u


def load_qty(fn: str) -> dict[int, dict]:
    wb = load_workbook(resolve_work(fn), read_only=True, data_only=True)
    ws = wb["통합내역"]
    hdr = [str(h or "").strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(hdr)}
    by_num: dict[int, dict] = defaultdict(
        lambda: {"items": 0, "m3": 0.0, "m2": 0.0, "m": 0.0, "ton": 0.0}
    )
    cur = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        sec = str(row[idx.get("공종", 1)] or "").strip()
        if sec and sec[0].isdigit():
            cur = sec_num(sec)
        name = str(row[idx.get("공종명", 3)] or "").strip()
        if not name or not cur:
            continue
        unit = norm_unit(str(row[idx.get("단위", 6)] or ""))
        try:
            q = float(row[idx.get("수량", 5)] or 0)
        except (TypeError, ValueError):
            continue
        if q <= 0:
            continue
        d = by_num[cur]
        d["items"] += 1
        if unit in d:
            d[unit] += q
    wb.close()
    return dict(by_num)


def primary_value(lab: str, cat: str, d: dict) -> float | None:
    if cat == "1. 토공":
        return float(TOGONG_WORK.get(lab, 0)) or None
    _, unit_key, _ = next(x for x in CROSS_ROWS if x[0] == cat)
    if unit_key == "items":
        return float(d.get("items", 0)) or None
    v = d.get(unit_key, 0)
    return float(v) if v > 0 else None


def fmt_qty(v: float | None, unit_key: str) -> str:
    if v is None or v <= 0:
        return "—"
    u = UNIT_LABEL.get(unit_key, unit_key)
    if unit_key == "items":
        return f"{v:,.0f}{u}"
    if unit_key == "m" and v != int(v):
        return f"{v:,.1f}{u}"
    return f"{v:,.0f}{u}"


def fmt_cell(v: float | None) -> str:
    """HTML 표용 — 단위 없이 숫자 또는 —."""
    if v is None or v <= 0:
        return "—"
    if v != int(v):
        return f"{v:,.1f}"
    return f"{v:,.0f}"


def fmt_num(v: float) -> str:
    if v != int(v):
        return f"{v:,.1f}"
    return f"{v:,.0f}"


def section_label(num: int, lab: str) -> str:
    names = {
        1: "1. 토공",
        2: "2. 우수공",
        3: "3. 오수공" if lab == "01" else "3. 상수공" if lab == "04" else "3. 포장공",
        4: "4. 상수공" if lab in ("01", "04") else "4. 포장공",
        5: "5. 포장공" if lab in ("01", "04", "05") else "5. 구조물공",
        6: "6. 구조물공" if lab in ("01", "04") else "6. 부대공",
        7: "7. 부대공",
        8: "8. 하수처리장",
        9: "9. 지하저류조",
    }
    if lab == "05":
        return {1: "1. 토공", 2: "2. 우수공", 3: "3. 포장공", 4: "4. 구조물공", 5: "5. 부대공"}.get(num, "")
    if lab == "04":
        return {
            1: "1. 토공", 2: "2. 우수공", 3: "3. 상수공", 4: "4. 포장공",
            5: "5. 구조물공", 6: "6. 부대공",
        }.get(num, "")
    return names.get(num, f"{num}.")


def detail_units(d: dict) -> str:
    parts = []
    for k, u in [("m3", "㎥"), ("m2", "㎡"), ("m", "m"), ("ton", "ton")]:
        if d.get(k, 0) > 0:
            v = d[k]
            s = f"{v:,.1f}" if k == "m" and v != int(v) else f"{v:,.0f}"
            parts.append(f"{s}{u}")
    if d.get("items"):
        parts.append(f"{d['items']:,.0f}건")
    return " · ".join(parts) if parts else "—"


def collect_gongjong_summary() -> dict:
    """종합보고서·공사공정표용 공정별 집계(02 전기 제외)."""
    from calc_togong_schedule import agg_togong

    all_data: dict[str, dict[int, dict]] = {}
    for lab, _, fn in FILES:
        all_data[lab] = load_qty(fn)

    cross: list[dict] = []
    for cat, unit_key, note in CROSS_ROWS:
        row: dict = {"cat": cat, "unit_key": unit_key, "note": note, "period": CROSS_PERIOD.get(cat, "—")}
        total = 0.0
        has = False
        for lab in ("05", "04", "01", "06"):
            sec = CAT_SEC.get(cat, {}).get(lab)
            if sec is None:
                row[lab] = None
                continue
            d = all_data[lab].get(sec, {})
            v = primary_value(lab, cat, d)
            row[lab] = v
            if v:
                total += v
                has = True
        row["total"] = total if has else None
        cross.append(row)

    togong_phases: list[dict] = []
    total_work = 0.0
    total_haul = 0.0
    for lab, label, fn in FILES:
        tg = agg_togong(fn)
        code = lab
        period, note = TOGONG_PERIOD[code]
        total_work += tg["work"]
        total_haul += tg["haul"]
        togong_phases.append({
            "label": label.split(" ", 1)[1] if " " in label else label,
            "code": code,
            "cut": tg["cut"],
            "fill": tg["fill"],
            "work": tg["work"],
            "haul": tg["haul"],
            "period": period,
            "note": note,
        })
    for p in togong_phases:
        p["pct"] = p["work"] / total_work * 100 if total_work else 0.0

    return {
        "meta": {
            "start": "2026. 10. 1.",
            "end": "2027. 11. 30.",
            "months": 14,
            "togong_total": total_work,
            "haul_total": total_haul,
        },
        "cross": cross,
        "togong_phases": togong_phases,
        "contracts": CONTRACT_GONGJONG,
    }


def main() -> None:
    all_data: dict[str, dict[int, dict]] = {}
    for lab, _, fn in FILES:
        all_data[lab] = load_qty(fn)

    print("=== 3-1 교차표 (대표 물량) ===")
    for cat, unit_key, note in CROSS_ROWS:
        parts = [cat, note]
        total = 0.0
        for lab in ("05", "04", "01", "06"):
            sec = CAT_SEC.get(cat, {}).get(lab)
            if sec is None:
                parts.append("—")
                continue
            d = all_data[lab].get(sec, {})
            v = primary_value(lab, cat, d)
            if v:
                total += v
            parts.append(fmt_qty(v, unit_key))
        parts.append(fmt_qty(total if total else None, unit_key))
        print("\t".join(parts))

    print("\n=== 01 상세 ===")
    for num in range(1, 10):
        d = all_data["01"].get(num, {})
        if not d.get("items"):
            continue
        cat = next((c for c, m in CAT_SEC.items() if m.get("01") == num), f"{num}.")
        _, unit_key, _ = next(((a, b, c) for a, b, c in CROSS_ROWS if a == cat), (cat, "items", ""))
        if num >= 8:
            unit_key = "m3"
        pv = primary_value("01", cat if num <= 7 else "1. 토공", d) if num == 1 else (
            d.get("m3") if num >= 8 else primary_value("01", cat, d)
        )
        if num == 8:
            pv = d.get("m3")
        elif num == 9:
            pv = d.get("m3")
        elif num <= 7:
            pv = primary_value("01", cat, d)
        print(f"{cat if num<=7 else ('8. 하수처리장' if num==8 else '9. 지하저류조')}\t{fmt_qty(pv, unit_key if num<=7 else 'm3')}\t{detail_units(d)}")

    for lab in ("05", "04"):
        print(f"\n=== {lab} ===")
        sec_names = {
            "05": {1: "1. 토공", 2: "2. 우수공", 3: "3. 포장공", 4: "4. 구조물공", 5: "5. 부대공"},
            "04": {1: "1. 토공", 2: "2. 우수공", 3: "3. 상수공", 4: "4. 포장공", 5: "5. 구조물공", 6: "6. 부대공"},
        }[lab]
        for num, name in sec_names.items():
            d = all_data[lab].get(num, {})
            if not d.get("items"):
                continue
            cat = next(c for c, m in CAT_SEC.items() if m.get(lab) == num)
            _, unit_key, _ = next(x for x in CROSS_ROWS if x[0] == cat)
            pv = primary_value(lab, cat, d)
            print(f"{name}\t{fmt_qty(pv, unit_key)}\t{d['items']}건\t{detail_units(d)}")

    print("\n=== 06 ===")
    d1 = all_data["06"].get(1, {})
    print(f"1. 토공\t{fmt_qty(TOGONG_WORK['06'], 'm3')}\t{detail_units(d1)}")
    a_total = sum(all_data["06"].get(n, {}).get("items", 0) for n in range(2, 9))
    print(f"2~8 A·B\t{a_total}건\t(주·㎡·m 블록별)")


if __name__ == "__main__":
    main()
