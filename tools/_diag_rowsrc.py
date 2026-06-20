#!/usr/bin/env python3
import sys
from pathlib import Path
from collections import Counter
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(__file__).resolve().parents[1] / "05_내역서"
WORK = BASE / "내역서작업"

# 검토_전체: 파일(내역서)·공종 분포 + 행 범위
wb = openpyxl.load_workbook(BASE / "검토_전체.xlsx", read_only=True, data_only=True)
ws = wb.active
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=3))][:14]
print("검토_전체 헤더(앞):", [str(h)[:8] for h in hdr])
files = Counter()
rows_by_file = {}
for r in ws.iter_rows(min_row=4, values_only=True):
    f = r[0]
    if f is None:
        continue
    files[str(f)] += 1
    rows_by_file.setdefault(str(f), []).append(r[1])
wb.close()
print("\n검토_전체 내역서(1열) 분포:")
for k, v in files.items():
    rng = rows_by_file[k]
    print(f"  {k}: {v}건  행범위 {min(rng)}~{max(rng)}")

# 통합내역 행 범위 (01 토목)
wb = openpyxl.load_workbook(WORK / "01_화성 청원지구 토목_표준단가산출.xlsx",
                            read_only=True, data_only=True)
w = wb["통합내역"]
h = [c.value for c in next(w.iter_rows(min_row=1, max_row=1))]
ri = h.index("행")
rows = [r[ri] for r in w.iter_rows(min_row=2, values_only=True) if r[ri] is not None]
print(f"\n01 토목 통합내역 행범위 {min(rows)}~{max(rows)} · {len(rows)}건")
wb.close()
