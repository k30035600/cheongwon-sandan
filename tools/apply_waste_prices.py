#!/usr/bin/env python3
"""건설폐기물 공내역서 — 동일 파일 경비(협회단가) 시트 연동 산출."""
from __future__ import annotations

import sys
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 공내역서·산출물 모두 05_내역서 하위
WORK_DIR = BASE / "내역서작업"
SRC_DIR = BASE / "공내역서"
SRC = SRC_DIR / "07_화성 청원지구 건설폐기물처리.XLS"
OUT = WORK_DIR / "07_화성 청원지구 건설폐기물처리_표준단가산출.xlsx"
OUT_MD = OUT.with_name(OUT.stem + "_요약.md")

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
MONEY_FMT = "#,##0"


def load_expense_prices(wb: xlrd.Book) -> dict[str, dict]:
    sh = wb.sheet_by_name("경비")
    prices: dict[str, dict] = {}
    for r in range(2, sh.nrows):
        code = ""
        for c in range(sh.ncols):
            v = str(sh.cell_value(r, c)).strip()
            if v.startswith("G"):
                code = v
                break
        if not code:
            continue
        prices[code] = {
            "name": str(sh.cell_value(r, 0)).strip(),
            "spec": str(sh.cell_value(r, 1)).strip(),
            "unit": str(sh.cell_value(r, 2)).strip(),
            "mat": float(sh.cell_value(r, 3) or 0),
            "lab": float(sh.cell_value(r, 4) or 0),
            "exp": float(sh.cell_value(r, 5) or 0),
            "total": float(sh.cell_value(r, 6) or 0),
            "note": str(sh.cell_value(r, 7)).strip(),
        }
    return prices


def load_items(wb: xlrd.Book) -> list[dict]:
    sh = wb.sheet_by_name("내역서")
    items: list[dict] = []
    section = ""
    for r in range(3, sh.nrows):
        name = str(sh.cell_value(r, 0)).strip()
        spec = str(sh.cell_value(r, 1)).strip()
        qty = sh.cell_value(r, 2) if sh.cell_type(r, 2) == xlrd.XL_CELL_NUMBER else None
        unit = str(sh.cell_value(r, 3)).strip()
        code = ""
        for c in range(sh.ncols):
            v = str(sh.cell_value(r, c)).strip()
            if v.startswith("G"):
                code = v
                break
        if qty == 1 and unit == "식" and name:
            section = name
            continue
        if not qty or unit == "식" or qty <= 0:
            continue
        items.append(
            {
                "row": r + 1,
                "section": section,
                "name": name,
                "spec": spec,
                "qty": float(qty),
                "unit": unit,
                "code": code,
            }
        )
    return items


def build_results(items: list[dict], prices: dict[str, dict]):
    results = []
    totals = {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0}
    for item in items:
        out = {**item, "status": "미산출", "source": ""}
        p = prices.get(item["code"])
        if p:
            mat_a = item["qty"] * p["mat"]
            lab_a = item["qty"] * p["lab"]
            exp_a = item["qty"] * p["exp"]
            sum_a = item["qty"] * (p["total"] or p["mat"] + p["lab"] + p["exp"])
            out.update(
                {
                    "status": "매칭",
                    "source": "경비시트(26.협회단가)",
                    "price_name": p["name"],
                    "price_spec": p["spec"],
                    "price_note": p["note"],
                    "mat_u": p["mat"],
                    "lab_u": p["lab"],
                    "exp_u": p["exp"],
                    "tot_u": p["total"],
                    "mat_a": mat_a,
                    "lab_a": lab_a,
                    "exp_a": exp_a,
                    "sum_a": sum_a,
                }
            )
            totals["mat"] += mat_a
            totals["lab"] += lab_a
            totals["exp"] += exp_a
            totals["sum"] += sum_a
        results.append(out)
    return results, totals


def write_xlsx(results, totals, prices: dict):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "안내"
    ws0.append(["건설폐기물처리 — 단가 산출"])
    ws0.append([])
    ws0.append(["원본", SRC.name])
    ws0.append(["단가 출처", "동일 파일 `경비` 시트 (26. 협회단가)"])
    ws0.append(["연동 코드", "G00000004(운반), G00000005(처리)"])
    ws0["A1"].font = Font(bold=True, size=14)

    ws = wb.create_sheet("합계요약")
    ws.append(["구분", "수량(Ton)", "재료비", "노무비", "경비", "합계"])
    for c in range(1, 7):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = HEADER_FILL
    for r in results:
        ws.append([r["name"], r["qty"], r.get("mat_a", 0), r.get("lab_a", 0), r.get("exp_a", 0), r.get("sum_a", 0)])
    ws.append([])
    ws.append(["★ 합계", sum(r["qty"] for r in results), totals["mat"], totals["lab"], totals["exp"], totals["sum"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT

    headers = [
        "행", "공종", "코드", "명칭", "규격", "단위", "수량", "상태", "단가출처",
        "단가품명", "단가규격", "비고",
        "재료단가", "노무단가", "경비단가", "합계단가",
        "재료금액", "노무금액", "경비금액", "합계금액",
    ]
    ws1 = wb.create_sheet("통합내역")
    ws1.append(headers)
    for c in range(1, len(headers) + 1):
        ws1.cell(1, c).font = Font(bold=True)
        ws1.cell(1, c).fill = HEADER_FILL
    for r in results:
        ws1.append([
            r["row"], r["section"], r["code"], r["name"], r["spec"], r["unit"], r["qty"],
            r["status"], r.get("source", ""), r.get("price_name", ""), r.get("price_spec", ""),
            r.get("price_note", ""),
            r.get("mat_u", ""), r.get("lab_u", ""), r.get("exp_u", ""), r.get("tot_u", ""),
            r.get("mat_a", ""), r.get("lab_a", ""), r.get("exp_a", ""), r.get("sum_a", ""),
        ])
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=13, max_col=20):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT

    ws2 = wb.create_sheet("경비단가표")
    ws2.append(["코드", "명칭", "규격", "단위", "재료비", "노무비", "경비", "합계", "비고"])
    for c in range(1, 10):
        ws2.cell(1, c).font = Font(bold=True)
        ws2.cell(1, c).fill = HEADER_FILL
    for code, p in sorted(prices.items()):
        ws2.append([code, p["name"], p["spec"], p["unit"], p["mat"], p["lab"], p["exp"], p["total"], p["note"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            w = min(48, max(len(str(c.value or "")) for c in col) + 2)
            sheet.column_dimensions[col[0].column_letter].width = w
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        return alt


def write_md(results, totals):
    lines = [
        "# 건설폐기물처리 — 단가 산출 요약",
        "",
        f"- **원본**: `{SRC.name}`",
        "- **단가 출처**: 동일 파일 `경비` 시트 (26. 협회단가)",
        "",
        "## 재료비 / 노무비 / 경비 / 합계",
        "",
        "| 항목 | 수량 | 재료비 | 노무비 | 경비 | 합계 |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for r in results:
        lines.append(
            f"| {r['name'].strip()} | {r['qty']:,.0f} | "
            f"{r.get('mat_a', 0):,.0f} | {r.get('lab_a', 0):,.0f} | "
            f"{r.get('exp_a', 0):,.0f} | {r.get('sum_a', 0):,.0f} |"
        )
    lines.extend([
        f"| **합계** | **{sum(r['qty'] for r in results):,.0f}** | "
        f"**{totals['mat']:,.0f}** | **{totals['lab']:,.0f}** | "
        f"**{totals['exp']:,.0f}** | **{totals['sum']:,.0f}** |",
        "",
        "## 단가 (Ton당)",
        "",
        "- **폐기물 운반비** (G00000004): 24톤·30km 이내 — 경비 **16,450**원/Ton",
        "- **폐콘크리트 처리비** (G00000005): 경비 **31,142**원/Ton",
        "",
        "※ 본 공사는 **경비(협회단가) 100%** 구조(재료·노무 0).",
        "",
        "끝.",
    ])
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main():
    wb = xlrd.open_workbook(str(SRC))
    prices = load_expense_prices(wb)
    items = load_items(wb)
    results, totals = build_results(items, prices)
    saved = write_xlsx(results, totals, prices)
    write_md(results, totals)
    print(f"항목 {len(items)} / 매칭 {sum(1 for r in results if r['status']=='매칭')}")
    print(f"재료 {totals['mat']:,.0f}  노무 {totals['lab']:,.0f}  경비 {totals['exp']:,.0f}  합계 {totals['sum']:,.0f}")
    print(f"저장: {saved}")


if __name__ == "__main__":
    main()
