#!/usr/bin/env python3
"""내역서 통합내역 「1. 토공」 물량 집계 → 공정 기간 배분."""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from naeyeok_gongjong import resolve_work

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

WORK = Path(__file__).resolve().parents[1] / "05_내역서" / "내역서작업"

FILES = [
    ("05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx"),
    ("01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx"),
    ("06 지구단위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx"),
]


def classify(name: str, spec: str, unit: str) -> str:
    u = (unit or "").replace(" ", "")
    t = ((name or "") + (spec or "")).replace(" ", "")
    if u in ("m3", "㎥", "M3"):
        if any(x in t for x in ("흙운반", "임목운반", "반출", "유용", "무대", "덤프", "야적")):
            return "운반_m3"
        if "운반" in t and "불도" not in t:
            return "운반_m3"
        if any(x in t for x in ("성토", "흙쌓", "쌓기", "되메")):
            return "성토_m3"
        if any(x in t for x in ("절토", "흙깎", "흙깍", "깎기", "굴착", "터파", "발파", "땅깍", "암", "리핑")):
            return "절토_m3"
        if "토사" in t and "쌓" not in t:
            return "절토_m3"
        return "기타_m3"
    if u in ("㎡", "m2", "M2"):
        return "면적_m2"
    return f"기타_{u or '?'}"


def agg_togong(xlsx: str) -> dict:
    wb = load_workbook(resolve_work(xlsx), read_only=True, data_only=True)
    ws = wb["통합내역"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c or "").strip() for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    sec_i = idx.get("공종", 1)
    name_i = idx.get("공종명", idx.get("품명", 3))
    spec_i = idx.get("규격", 4)
    qty_i = idx.get("수량", 5)
    unit_i = idx.get("단위", 6)

    by_cat: dict[str, float] = {}
    for row in rows[1:]:
        if not row:
            continue
        sec = str(row[sec_i] or "").strip() if len(row) > sec_i else ""
        if re.match(r"^2\.", sec.replace(" ", "")):
            break
        if not re.match(r"^1\.\s*토", sec.replace(" ", "")):
            continue
        name = str(row[name_i] or "").strip() if len(row) > name_i else ""
        spec = str(row[spec_i] or "").strip() if len(row) > spec_i else ""
        qty = row[qty_i] if len(row) > qty_i else None
        unit = str(row[unit_i] or "").strip() if len(row) > unit_i else ""
        try:
            q = float(qty)
        except (TypeError, ValueError):
            continue
        if q <= 0:
            continue
        cat = classify(name, spec, unit)
        by_cat[cat] = by_cat.get(cat, 0) + q

    cut = by_cat.get("절토_m3", 0)
    fill = by_cat.get("성토_m3", 0)
    haul = by_cat.get("운반_m3", 0)
    work = cut + fill
    total_m3 = sum(v for k, v in by_cat.items() if k.endswith("_m3"))
    return {
        "by_cat": by_cat,
        "cut": cut,
        "fill": fill,
        "haul": haul,
        "work": work,
        "total_m3": total_m3,
        "area": by_cat.get("면적_m2", 0),
    }


def allocate_months(work_m3: dict[str, float], pool: int, *, min_m: int = 1) -> dict[str, int]:
    total = sum(work_m3.values())
    if total <= 0:
        return {k: min_m for k in work_m3}
    raw = {k: max(min_m, round(pool * v / total)) for k, v in work_m3.items()}
    # adjust sum to pool
    diff = pool - sum(raw.values())
    keys = sorted(work_m3, key=lambda k: work_m3[k], reverse=True)
    i = 0
    while diff != 0 and keys:
        k = keys[i % len(keys)]
        if diff > 0:
            raw[k] += 1
            diff -= 1
        elif raw[k] > min_m:
            raw[k] -= 1
            diff += 1
        i += 1
    return raw


def main() -> None:
    results: dict[str, dict] = {}
    for label, fn in FILES:
        results[label] = agg_togong(fn)

    total_work = sum(v["work"] for v in results.values())
    print("=== 토공 물량 (절토+성토 m3) ===")
    for label, v in results.items():
        pct = v["work"] / total_work * 100 if total_work else 0
        print(
            f"{label}: 절토 {v['cut']:,.0f} + 성토 {v['fill']:,.0f} "
            f"= {v['work']:,.0f} m3 ({pct:.1f}%) | 운반 {v['haul']:,.0f}"
        )
    print(f"합계 작업량: {total_work:,.0f} m3\n")

    # 토공 전용 가용 월 (14개월 중 M2~M10 = 9 months for earthwork before mostly paving)
    pool = 9
    work_only = {k: v["work"] for k, v in results.items()}
    months = allocate_months(work_only, pool, min_m=1)
    print(f"=== 토공 기간 배분 (가용 {pool}개월, 물량 비례) ===")
    for label in results:
        m = months[label]
        w = work_only[label]
        print(f"  {label}: {m}개월 (물량 {w:,.0f} m3)")


if __name__ == "__main__":
    main()
