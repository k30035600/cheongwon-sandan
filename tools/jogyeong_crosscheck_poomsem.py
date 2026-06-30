"""조경 확인 3건 — 표준품셈·일위대가 산정 (01 토목 내역 · 조경 품셈 적용).

원내역: 01_화성 청원지구 토목.XLS 「1. 토공」·수목관리 386주(상수리 B=27~31 381 + 소나무 B=29 5).
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
NOIM_CSV = ROOT / "05_내역서" / "일위대가DB" / "_공통" / "시중노임_2026.csv"
ILWAE_XLSX = ROOT / "05_내역서" / "내역서작업" / "_공통" / "미매칭_일위대가산출.xlsx"
JO_ILDAE_CSV = ROOT / "05_내역서" / "일위대가DB" / "조경" / "조경표준일위대가_2024.csv"

# 2026. 1. 1. 시중노임 (build_nojng_wages.py SSOT)
DEFAULT_NOIM = {"조경공": 235_204, "보통인부": 172_068}

# 수목 100주당 1회 기계살포 — 대형목 (R 12~20, H 4~6) · B=27~31cm 이식 교목
SPRAY_CLASS = "대형목"
SPRAY_JG_MH = 0.25
SPRAY_BO_MH = 0.25

# 지주목 1조(1개소) — 사각 지주목(가로수·R 8~12)
STAKE_TYPE = "사각 지주목(가로지지대)"
STAKE_JG_MH = 0.15
STAKE_BO_MH = 0.15

# 경비·재료 (100주당 살포 / 1개소 지주 — 검증·조정 가능)
MACHINE_WON_PER_100 = 20_000  # 동력살포기 손료+유류 (100주당)
MAT_BUNGHAe_PER_100 = 25_000  # 살균제·전착제 등 (100주당)
MAT_CHUNGHAe_PER_100 = 22_000  # 살충제 등 (100주당)
MAT_STAKE_PER_EA = 18_000  # 사각 지주목 세트(목재·볼트·패드·말뚝)


@dataclass
class LineItem:
    key: str
    name: str
    spec: str
    unit: str
    qty: float
    tree_count: int
    poomsem_class: str
    jg_mh: float
    bo_mh: float
    noim_jg: int
    noim_bo: int
    labor: int
    machine: int
    material: int
    unit_price: int
    amount: int
    basis: str
    ilwae_row: int | None = None


def load_noim() -> dict[str, int]:
    out = dict(DEFAULT_NOIM)
    if not NOIM_CSV.exists():
        return out
    with NOIM_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            name = (row.get("직종명") or "").strip()
            try:
                out[name] = int(str(row.get("2026.1.1") or "0").replace(",", ""))
            except ValueError:
                pass
    return out


def _round_won(x: float) -> int:
    return int(round(x))


def calc_spray_unit(
    *,
    tree_count: int,
    kind: str,
    noim: dict[str, int],
    jg_mh: float = SPRAY_JG_MH,
    bo_mh: float = SPRAY_BO_MH,
) -> tuple[int, int, int, int, str]:
    """1회 = 관리 수목 전체 기계살포 1회."""
    factor = tree_count / 100.0
    nj, nb = noim["조경공"], noim["보통인부"]
    labor = _round_won(factor * (jg_mh * nj + bo_mh * nb))
    machine = _round_won(factor * MACHINE_WON_PER_100)
    mat_rate = MAT_BUNGHAe_PER_100 if kind == "병해" else MAT_CHUNGHAe_PER_100
    material = _round_won(factor * mat_rate)
    unit = labor + machine + material
    basis = (
        f"2026 표준품셈 [조경] 유지관리 기계살포 · {SPRAY_CLASS} "
        f"({jg_mh}조+{bo_mh}보/100주) × {tree_count}주/100 "
        f"= 노무 {labor:,} + 기계 {machine:,} + 약제 {material:,} "
        f"(시중노임 조경공 {nj:,}·보통인부 {nb:,})"
    )
    return labor, machine, material, unit, basis


def calc_stake_unit(*, noim: dict[str, int]) -> tuple[int, int, int, str]:
    nj, nb = noim["조경공"], noim["보통인부"]
    labor = _round_won(STAKE_JG_MH * nj + STAKE_BO_MH * nb)
    material = MAT_STAKE_PER_EA
    unit = labor + material
    basis = (
        f"2026 표준품셈 [조경] 지주목 설치 · {STAKE_TYPE} "
        f"({STAKE_JG_MH}조+{STAKE_BO_MH}보/1조) = 노무 {labor:,} + 재료 {material:,} "
        f"(조경일위2024 강관지주 18,539원/m는 가설·해체 참고 · 본건은 사각 지주목 세트)"
    )
    return labor, material, unit, basis


def build_line_items() -> list[LineItem]:
    """01 토목 원내역 수량 기준."""
    noim = load_noim()
    trees = 386
    items: list[LineItem] = []

    for key, name, qty, kind, ilwae_row in (
        ("bung", "가). 교목 병해방제", 18, "병해", 108),
        ("chung", "나). 교목 충해방제", 9, "충해", 109),
    ):
        lab, mac, mat, unit, basis = calc_spray_unit(
            tree_count=trees, kind=kind, noim=noim,
        )
        items.append(
            LineItem(
                key=key,
                name=name,
                spec=f"수목관리 {trees}주 · {SPRAY_CLASS} · 1회=전체 살포",
                unit="회",
                qty=qty,
                tree_count=trees,
                poomsem_class=SPRAY_CLASS,
                jg_mh=SPRAY_JG_MH,
                bo_mh=SPRAY_BO_MH,
                noim_jg=noim["조경공"],
                noim_bo=noim["보통인부"],
                labor=lab,
                machine=mac,
                material=mat,
                unit_price=unit,
                amount=unit * int(qty),
                basis=basis,
                ilwae_row=ilwae_row,
            )
        )

    lab, mat, unit, basis = calc_stake_unit(noim=noim)
    items.append(
        LineItem(
            key="stake",
            name="다). 가로지지대",
            spec="B=20cm · 사각 지주목 1조/개소",
            unit="개소",
            qty=386,
            tree_count=trees,
            poomsem_class=STAKE_TYPE,
            jg_mh=STAKE_JG_MH,
            bo_mh=STAKE_BO_MH,
            noim_jg=noim["조경공"],
            noim_bo=noim["보통인부"],
            labor=lab,
            machine=0,
            material=mat,
            unit_price=unit,
            amount=unit * 386,
            basis=basis,
            ilwae_row=22,
        )
    )
    return items


def sync_ilwidae_xlsx(items: list[LineItem] | None = None) -> Path | None:
    """미매칭_일위대가산출 — 조경 확인 3행 제시단가·근거 갱신."""
    if not ILWAE_XLSX.exists():
        return None
    from openpyxl import load_workbook

    items = items or build_line_items()
    wb = load_workbook(ILWAE_XLSX)
    ws = wb["일위대가산출"]
    for it in items:
        if not it.ilwae_row:
            continue
        r = it.ilwae_row
        route = "C.품셈 조경방제" if it.key in ("bung", "chung") else "C.품셈 지주목"
        ws.cell(r, 1).value = route
        ws.cell(r, 13).value = it.material if it.key != "stake" else it.material
        ws.cell(r, 14).value = it.labor
        ws.cell(r, 15).value = it.machine if it.machine else None
        tot = it.unit_price
        ws.cell(r, 16).value = tot
        ws.cell(r, 17).value = tot
        ws.cell(r, 18).value = it.amount
        ws.cell(r, 21).value = it.basis
    wb.save(ILWAE_XLSX)
    return ILWAE_XLSX


def render_jogyeong_html(items: list[LineItem] | None = None) -> str:
    items = items or build_line_items()
    total = sum(it.amount for it in items)

    def row(it: LineItem) -> str:
        mac = f"{it.machine:,}" if it.machine else "—"
        return (
            f"<tr><td class='l2'>{it.name}</td><td class='l2'>{it.spec}</td>"
            f"<td class='c'>{it.qty:g}{it.unit}</td>"
            f"<td class='c'>{it.poomsem_class}</td>"
            f"<td class='num'>{it.jg_mh:g}+{it.bo_mh:g}</td>"
            f"<td class='num'>{it.labor:,}</td>"
            f"<td class='num'>{mac}</td>"
            f"<td class='num'>{it.material:,}</td>"
            f"<td class='num'>{it.unit_price:,}</td>"
            f"<td class='num'>{it.amount:,}</td></tr>"
        )

    rows_html = "\n".join(row(it) for it in items)
    detail = "".join(
        f"<li><b>{it.name}</b> — {it.basis}</li>" for it in items
    )

    return f"""
<h2>조경 확인 대상 (미확정 집계 제외 · 품셈·일위대가 산정)</h2>
<p class="muted">
  원내역 <code>01_화성 청원지구 토목.XLS</code> 「1. 토공」·수목관리 <b>386주</b>
  (상수리 B=27~31 381 + 소나무 B=29 5). 조경·유지관리 표준품셈(기계살포·지주목) 적용.
  <b>1회</b> = 386주 전체 1회 살포. 확정 전 발주처·설계 검증 권장.
</p>
<table>
<thead><tr>
  <th>품명</th><th>규격·적용</th><th>수량</th><th>품셈 구분</th>
  <th class="num">조+보(인)</th><th class="num">노무비</th><th class="num">경비</th>
  <th class="num">재료비</th><th class="num">합계단가</th><th class="num">금액</th>
</tr></thead>
<tbody>
{rows_html}
</tbody>
<tfoot><tr class="total">
  <td colspan="9">소계 (3항목 · _ka 미반영)</td>
  <td class="num">{total:,}</td>
</tr></tfoot>
</table>
<ul class="muted" style="font-size:13px;margin-top:8px">
  <li><b>근거</b> — 국토교통부 건설공사 표준품셈 [조경] 유지관리(기계살포)·식재(지주목);
    시중노임 2026. 1. 1.; 약제·지주 세트는 물가·수종별 별도 검토.</li>
  <li><b>일위대가표</b> — <code>05_내역서/내역서작업/_공통/미매칭_일위대가산출.xlsx</code>
    「일위대가산출」 r90~91·97 (품셈 경로 C)</li>
  <li><b>조경일위 참고</b> — <code>조경표준일위대가_2024.csv</code>
    「수목류 약제살포」「강관 지주 설치 및 해체」</li>
</ul>
<blockquote style="font-size:13px">
  <p><b>산출 근거 상세</b></p>
  <ul>{detail}</ul>
</blockquote>"""


if __name__ == "__main__":
    import sys

    sys.stdout.reconfigure(encoding="utf-8")
    items = build_line_items()
    for it in items:
        print(f"{it.name} | {it.qty}{it.unit} | 단가 {it.unit_price:,} | 금액 {it.amount:,}")
    print(f"합계 {sum(it.amount for it in items):,}")
    p = sync_ilwidae_xlsx(items)
    if p:
        print(f"갱신: {p}")
