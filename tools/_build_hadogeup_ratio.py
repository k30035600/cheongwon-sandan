import sys, os
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"D:\OneDrive\Cursor\cheongwon\07_타견적"
SRC = os.path.join(BASE, "희상건설", "260707 청원지구(토목,조경,전기).xlsx")
OUT = os.path.join(BASE, "하도급원가계산", "하도급원가계산_비율(표).xlsx")

# ── 희상 원가계산서 값 추출 ──
sw = openpyxl.load_workbook(SRC, data_only=True)
wg = sw["원가계산"]
det = sw["토목,조경 내역서"]
agg = sw["집계표"]
def rc(v): return round(v) if isinstance(v,(int,float)) else 0
def v(rr): return rc(wg[f"D{rr}"].value)

# 공정별 직접공사비(재료·노무·경비·합계) — 내역서 대공종 【 】 집계
PROCS=[]
for row in det.iter_rows(values_only=True):
    nm=row[0]
    if not nm or not str(nm).strip().startswith("【"): continue
    name=str(nm).strip().strip("【】").strip()
    if "직접공사비 계" in name: continue      # 상위 합계행 제외
    if row[11] is None: continue
    PROCS.append((name, rc(row[5]), rc(row[7]), rc(row[9]), rc(row[11])))
# 전기(집계표)
for row in agg.iter_rows(values_only=True):
    if row[0] and "02. 전기설비" in str(row[0]):
        PROCS.append(("02. 전기설비", rc(row[5]), rc(row[7]), rc(row[9]), rc(row[11])))
        break

JAE=v(4); DNO=v(5); DGB=v(6); DIRECT=v(7)
GANNO=v(8); SANCHUL=v(18); SOON=v(20)
ILBAN=v(21); IYUN=v(22); SUPPLY=v(24); VAT=v(25); DOGEUP=v(26)
INDIRECT=SUPPLY-DIRECT
NOMUGYE=DNO+GANNO; GYEONGGYE=DGB+SANCHUL
GANNO_SIK=wg["C8"].value                                          # 간접노무비 산식
SANCHUL_ITEMS=[(wg[f"B{rr}"].value, wg[f"C{rr}"].value, v(rr)) for rr in range(9,18)]  # 산출경비 9개(항목,산식,금액)

# ── 워크북/스타일 ──
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "희상건설 원가계산서"
ws.sheet_view.showGridLines = False
navy="1F3A5F"; head="EEF0F4"; band1="F4F8FF"; band2="FBF4EC"; grn="E7F3EA"; blue="DDE8F5"
thin=Side(style="thin",color="D9E1EC"); border=Border(thin,thin,thin,thin)
won='#,##0'
def st(cell,v,*,bold=False,size=11,color="1C2430",fill=None,align="left",bd=True,wrap=True,fmt=None):
    c=ws[cell]; c.value=v
    c.font=Font(bold=bold,size=size,color=color,name="맑은 고딕")
    if fill:c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fmt:c.number_format=fmt
    if bd:c.border=border
    return c
COLS_L = ["A","B","C","D","E","F"]
COLS_R = ["H","I","J","K","L","M"]
for col,w in [("A",9),("B",30),("C",15),("D",15),("E",15),("F",16),("G",3),
              ("H",9),("I",30),("J",15),("K",15),("L",15),("M",16)]:
    ws.column_dimensions[col].width=w

def draw(cols, title, subtitle, *, drop_gannomu=False):
    A,B,C,D,E,F = cols
    ws.merge_cells(f"{A}1:{F}1"); st(f"{A}1",title,bold=True,size=14,color=navy,bd=False)
    ws.merge_cells(f"{A}2:{F}2"); st(f"{A}2",subtitle,size=10,color="5A6675",bd=False)
    for col,t in zip(cols,["구분","공종 · 항목 (산식)","재료비","노무비","경비","합계"]):
        st(f"{col}4",t,bold=True,fill=head,color=navy,align="center")
    state={"r":5}
    def rowline(name,jae,no,gb,tot,*,sub=False,fill=None,sik=None):
        rr=state["r"]; f0=fill if fill else (grn if sub else None)
        label=name if not sik else f"{name}  ({sik})"
        st(f"{A}{rr}","",fill=f0); st(f"{B}{rr}",label,bold=sub,fill=f0)
        for col,x in [(C,jae),(D,no),(E,gb)]:
            st(f"{col}{rr}", x if x is not None else "—", align="right", fmt=won if x is not None else None, bold=sub, fill=f0)
        st(f"{F}{rr}",tot,align="right",fmt=won,bold=True,fill=f0)
        state["r"]+=1; return rr
    def sumline(name,sik,amt,*,sub=False):
        rr=state["r"]; f0=grn if sub else None
        st(f"{A}{rr}","",fill=f0); st(f"{B}{rr}",name,bold=sub,fill=f0)
        ws.merge_cells(f"{C}{rr}:{E}{rr}")
        st(f"{C}{rr}",sik,size=9,color="5A6675",align="center",fill=f0)
        st(f"{F}{rr}",amt,align="right",fmt=won,bold=True,fill=f0)
        state["r"]+=1; return rr
    # 재계산(간접노무비 제거 반영)
    ganno = 0 if drop_gannomu else GANNO
    nomugye = DNO + ganno
    gyeonggye = DGB + SANCHUL
    soon = JAE + nomugye + gyeonggye
    ilban = round(soon*0.05)
    iyun = round((nomugye+gyeonggye+ilban)*0.07)
    supply = soon + ilban + iyun
    vat = round(supply*0.10)
    dogeup = supply + vat
    ganjub_gye = ganno + SANCHUL
    # 직접공사비(공정별)
    dir_first=state["r"]
    for name,jae,no,gb,tot in PROCS:
        rowline(name,jae,no,gb,tot)
    rowline("【 직접공사비 계 】",JAE,DNO,DGB,DIRECT,sub=True)
    dir_last=state["r"]-1
    # 간접공사비
    ind_first=state["r"]
    if not drop_gannomu:
        rowline("간접노무비",None,GANNO,None,GANNO,sik=GANNO_SIK)
    for nm,sik,amt in SANCHUL_ITEMS:
        rowline(nm,None,None,amt,amt,sik=sik)
    ganjub_label="【 간접비 계(산출경비) 】" if drop_gannomu else "【 간접비 계(간접노무비+산출경비) 】"
    rowline(ganjub_label,None,(None if drop_gannomu else GANNO),SANCHUL,ganjub_gye,sub=True)
    ind_last=state["r"]-1
    # 순공사원가
    soon_r=rowline("【 순공사원가 】",JAE,nomugye,gyeonggye,soon,sub=True,fill=blue)
    st(f"{A}{soon_r}","순공사\n원가",bold=True,fill=blue,color=navy,align="center")
    ws[f"{A}{soon_r}"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    # 총원가
    tot_first=state["r"]
    sumline("일반관리비","순공사원가 × 5%",ilban)
    sumline("이윤","(노무비계+경비계+일반관리비) × 7%",iyun)
    sumline("【 공급가액(총원가) 】","순공사원가 + 일반관리비 + 이윤",supply,sub=True)
    sumline("부가가치세","공급가액 × 10%",vat)
    sumline("【 도급액 】","공급가액 + 부가가치세",dogeup,sub=True)
    tot_last=state["r"]-1
    def merge_grp(a,b,label):
        ws.merge_cells(f"{A}{a}:{A}{b}")
        st(f"{A}{a}",label,bold=True,fill=head,color=navy,align="center")
        ws[f"{A}{a}"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    merge_grp(dir_first,dir_last,"직접\n공사비")
    merge_grp(ind_first,ind_last,"간접\n공사비")
    merge_grp(tot_first,tot_last,"총원가")
    return tot_last, dogeup

sub_txt="화성 청원지구 산업유통형 지구단위계획 · 토목·조경·전기 · 부가가치세 별도 · 직접공사비는 공정별(재료비·노무비·경비) 구분"
last_h,dg_h=draw(COLS_L,"희상건설 원가계산서 (260707 기준)",sub_txt)
last_n,dg_n=draw(COLS_R,"근일건설 원가계산서 (희상 기준·간접노무비 제외)",sub_txt,drop_gannomu=True)

# 각주
note_r=max(last_h,last_n)+1
ws.merge_cells(f"A{note_r}:M{note_r}")
st(f"A{note_r}",f"※ 좌: 희상건설(도급액 {dg_h:,}) · 우: 근일건설(희상 기준·간접노무비 제외 재계산 → 도급액 {dg_n:,}). 직접공사비=첨부 내역서 공정별 집계(01 토목·조경 / 02 전기 / 04 진입도로 / 05 회전교차로 / 06 개발행위 / 07 건설폐기물). 간접비 항목 옆 ( )는 적용 요율·산식. 일반관리비 5% · 이윤 7%는 순공사원가·(노무비계+경비계+일반관리비) 기준. 근일은 간접노무비 미계상으로 노무비계·순공사원가·공급가액·도급액이 재계산됨.",size=9,color="8A5A00",bd=False)
ws[f"A{note_r}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)

# ═══════════ 시트2 — 희상·근일 원가계산 비교 ═══════════
SRC_N = os.path.join(BASE, "근일건설(청원지구).xlsx")
nw = openpyxl.load_workbook(SRC_N, data_only=True)
ng = nw["총공사비"]

def hgv(rr):
    if rr is None: return None, None
    return wg[f"D{rr}"].value, wg[f"C{rr}"].value   # 금액, 산식
def ngv(rr):
    if rr is None: return None, None
    return ng[f"C{rr}"].value, ng[f"D{rr}"].value   # 금액, 산식

H_SUP = round(hgv(24)[0]); H_DIR = round(hgv(7)[0])
N_SUP = round(ngv(47)[0]); N_DIR = round(ngv(26)[0])
H_IND = H_SUP - H_DIR; N_IND = N_SUP - N_DIR

import re
# 회사별 base 금액(산식 대입용)
H = {"jae":round(hgv(4)[0]),"dirnomu":round(hgv(5)[0]),"gannomu":round(hgv(8)[0]),
     "direct":H_DIR,"gangang":round(hgv(11)[0]),"soon":round(hgv(20)[0]),
     "gyeongbi":round(hgv(19)[0]),"ilban":round(hgv(21)[0]),"supply":H_SUP}
H["nomugye"]=H["dirnomu"]+H["gannomu"]
N = {"jae":round(ngv(27)[0]),"dirnomu":round(ngv(28)[0]),"gannomu":round(ngv(30)[0]),
     "direct":N_DIR,"gangang":round(ngv(34)[0]),"soon":round(ngv(42)[0]),
     "gyeongbi":round(ngv(41)[0]),"ilban":round(ngv(43)[0]),"supply":N_SUP}
N["nomugye"]=round(ngv(31)[0])

def base_of(label, rv):
    t={
     "간접노무비":("직접노무비",rv["dirnomu"]),
     "산재보험료":("노무비계",rv["nomugye"]),
     "고용보험료":("노무비계",rv["nomugye"]),
     "국민건강보험료":("직접노무비",rv["dirnomu"]),
     "국민연금보험료":("직접노무비",rv["dirnomu"]),
     "노인장기요양보험료":("건강보험료",rv["gangang"]),
     "산업안전보건관리비":("재+직노",rv["jae"]+rv["dirnomu"]),
     "기타경비(공과잡비)":("재+노무비계",rv["jae"]+rv["nomugye"]),
     "건설기계대여대금지급보증":("직접공사비",rv["direct"]),
     "환경보전비":("직접공사비",rv["direct"]),
     "일반관리비":("순공사원가",rv["soon"]),
     "이윤":("노무비계+경비계+일반관리비",rv["nomugye"]+rv["gyeongbi"]+rv["ilban"]),
     "이행(준공)보증보험료":("공급가액",rv["supply"]),
     "PF·신용 리스크 프리미엄":("직접공사비+간접노무비",rv["direct"]+rv["gannomu"]),
    }
    return t.get(label)

def with_base(sikc, label, rv):
    b=base_of(label, rv)
    if not b or not sikc: return sikc or ""
    bname,bval=b
    m=re.search(r"[×xX]\s*(.+)", sikc)
    rate=m.group(1).strip() if m else sikc
    return f"{bname} {bval:,} × {rate}"

# (그룹, 항목, 희상행, 근일행, override(희상금액,근일금액,산식))
COMP = [
 ("직접공사비","재료비",4,27,None),
 ("직접공사비","직접노무비",5,28,None),
 ("직접공사비","직접경비",6,29,None),
 ("직접공사비","【 직접공사비 계 】",7,26,None),
 ("간접공사비","간접노무비",8,30,None),
 ("간접공사비","산재보험료",9,32,None),
 ("간접공사비","고용보험료",10,33,None),
 ("간접공사비","국민건강보험료",11,34,None),
 ("간접공사비","국민연금보험료",12,35,None),
 ("간접공사비","노인장기요양보험료",13,36,None),
 ("간접공사비","산업안전보건관리비",14,37,None),
 ("간접공사비","기타경비(공과잡비)",15,38,None),
 ("간접공사비","건설기계대여대금지급보증",16,39,None),
 ("간접공사비","환경보전비",17,40,None),
 ("간접공사비","이행(준공)보증보험료",None,45,None),
 ("간접공사비","PF·신용 리스크 프리미엄",None,46,None),
 ("간접공사비","일반관리비",21,43,None),
 ("간접공사비","이윤",22,44,None),
 ("간접공사비","【 간접공사비 계 】",None,None,(H_IND,N_IND,"공급가액 − 직접공사비")),
 ("합계","공급가액(총공사비)",24,47,None),
 ("합계","부가가치세",25,48,None),
 ("합계","【 도급액 】",26,49,None),
]

ws = wb.create_sheet("원가계산 비교(희상·근일)")
ws.sheet_view.showGridLines = False
for col,w in [("A",13),("B",22),("C",36),("D",17),("E",36),("F",17)]:
    ws.column_dimensions[col].width=w
ws.merge_cells("A1:F1"); st("A1","희상건설 · 근일건설 원가계산 비교 (직접공사비 · 간접공사비)",bold=True,size=14,color=navy,bd=False)
ws.merge_cells("A2:F2"); st("A2","화성 청원지구 · 부가가치세 별도(공급가액 기준) · 희상=토목·조경·전기(01·02·04~07) / 근일=전기 제외(01·04~07) → 총액은 범위 차이 포함",size=9,color="5A6675",bd=False)
ws.merge_cells("A4:A5"); st("A4","구분",bold=True,fill=head,color=navy,align="center")
ws.merge_cells("B4:B5"); st("B4","항목",bold=True,fill=head,color=navy,align="center")
ws.merge_cells("C4:D4"); st("C4","희상건설",bold=True,fill=band1,color=navy,align="center")
ws.merge_cells("E4:F4"); st("E4","근일건설",bold=True,fill=band2,color=navy,align="center")
st("C5","산식(요율)",bold=True,fill=band1,align="center"); st("D5","금액(원)",bold=True,fill=band1,align="center")
st("E5","산식(요율)",bold=True,fill=band2,align="center"); st("F5","금액(원)",bold=True,fill=band2,align="center")

r=6
group_rows={}
for grp,label,hr,nr,ov in COMP:
    is_sub = "【" in label
    fill = grn if is_sub else None
    if ov:
        ha,hf = ov[0],ov[2]; na,nf = ov[1],ov[2]
    else:
        ha,hf = hgv(hr); na,nf = ngv(nr)
        hf = with_base(hf, label, H); nf = with_base(nf, label, N)
    st(f"A{r}","",fill=fill)
    st(f"B{r}",label,bold=is_sub,fill=fill)
    st(f"C{r}",(hf or "") if ha is not None else "—",size=9,color="5A6675",fill=fill)
    st(f"D{r}",round(ha) if ha is not None else "—",align="right",fmt=won,bold=is_sub,fill=fill)
    st(f"E{r}",(nf or "") if na is not None else "—",size=9,color="5A6675",fill=fill)
    st(f"F{r}",round(na) if na is not None else "—",align="right",fmt=won,bold=is_sub,fill=fill)
    group_rows.setdefault(grp,[]).append(r)
    r+=1
# 구분 열 병합
for grp,rows_ in group_rows.items():
    a,b=rows_[0],rows_[-1]
    ws.merge_cells(f"A{a}:A{b}")
    gname={"직접공사비":"직접\n공사비","간접공사비":"간접\n공사비","합계":"합계"}.get(grp,grp)
    st(f"A{a}",gname,bold=True,fill=head,color=navy,align="center")
    ws[f"A{a}"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
# 주석
ws.merge_cells(f"A{r}:F{r}")
st(f"A{r}","※ 근일건설은 전기(02) 제외 범위이며, 일반관리비(8%)·이윤(15%)·이행보증보험료·PF 리스크 프리미엄 등 요율·항목이 희상건설과 상이함. 금액 총액 비교 시 공사 범위(전기 포함/제외) 차이에 유의.",size=9,color="8A5A00",bd=False)
ws[f"A{r}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); r+=1
diff=N_IND-H_IND
msg=(f"※ 간접공사비 근일({N_IND:,}) > 희상({H_IND:,}) · 차이 +{diff:,} 주원인 — "
     f"① 이윤율 15%(희상 7%)·일반관리비 8%(희상 5%) 요율 상향(이윤만 +{round(ngv(44)[0])-round(hgv(22)[0]):,}) · "
     f"② 근일 직접노무비 비중 {N['dirnomu']/N_DIR:.0%}(희상 {H['dirnomu']/H_DIR:.0%})로 노무비 기반 간접비(간접노무비·4대보험·기타경비) 확대(간접노무비 +{N['gannomu']-H['gannomu']:,}) · "
     f"③ 근일 전용 PF·신용 리스크 프리미엄 {round(ngv(46)[0]):,}·이행보증보험료 {round(ngv(45)[0]):,} 추가.")
ws.merge_cells(f"A{r}:F{r}")
st(f"A{r}",msg,size=9,color="8A5A00",bd=False)
ws[f"A{r}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)

try:
    wb.save(OUT); print("saved",OUT)
    print(f"[희상] 직접 {H_DIR:,} · 간접 {H_IND:,} · 공급가액 {H_SUP:,} · 도급액 {round(hgv(26)[0]):,}")
    print(f"[근일] 직접 {N_DIR:,} · 간접 {N_IND:,} · 공급가액 {N_SUP:,} · 도급액 {round(ngv(49)[0]):,}")
except PermissionError:
    print("[경고] 파일 열림 — 닫고 재실행")
