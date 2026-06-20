#!/usr/bin/env python3
"""토목(01·04·05·06) 미매칭 항목 → 수동 단가 입력표 생성.

각 미매칭 항목에 대해 단가 풀(표준시장단가2026·시장시공가격·표준일위대가2026)에서
추천 후보 1~3개와 예상금액(추천1 합계단가 × 수량)을 붙여, 사용자가 단가만 채우면
되도록 한다. 매칭 엔진은 apply_standard_prices 모듈을 그대로 재사용한다.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
import apply_standard_prices as asp  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 공내역서·내역서작업이 05_내역서 하위로 이동
OUT_DIR = BASE / "내역서작업"
OUT_XLSX = OUT_DIR / "토목_미매칭_수동단가입력표.xlsx"

SRC_FILES = [
    ("01 토목", "01_화성 청원지구 토목.XLS"),
    ("01 조경", "01_화성 청원지구 조경.XLS"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계.XLS"),
    ("05 회전교차로", "05_화성 청원로(회전교차로).XLS"),
    ("06 개발행위", "06_화성 청원지구 산업유통형 개발행위.XLS"),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="E2EFDA")
NOCAND_FILL = PatternFill("solid", fgColor="FCE4D6")
MONEY_FMT = "#,##0"


def load_pool() -> list[dict]:
    market = asp.load_market_csv(asp.MARKET_2026, "표준시장단가2026")
    if not market:
        market = asp.load_prices()
    sijang = asp.load_market_csv(asp.SIJANG_2026, "시장시공가격2026")
    ildae = asp.load_ildae_prices()
    return asp.precompute(market + sijang + ildae)


def topn(item: dict, prices: list[dict], n: int = 3):
    terms = asp.extract_search_terms(item["name"], item["spec"])
    req = asp.required_keywords(item)
    item_kw = asp.kwset(f"{item['name']} {item['spec']}")
    item["_linear"] = asp.is_linear_piece(item)
    cand = [p for p in prices if asp.unit_compatible(item["unit"], p["unit"], item["_linear"])]
    best: dict[int, tuple[float, dict, str]] = {}
    for term in terms:
        kw = item_kw | asp.kwset(term)
        for p in cand:
            s = asp.score_match(item, p, term, kw, req)
            if s <= 0:
                continue
            pid = id(p)
            if pid not in best or s > best[pid][0]:
                best[pid] = (s, p, term)
    ranked = sorted(best.values(), key=lambda x: -x[0])[:n]
    return ranked


def main() -> None:
    prices = load_pool()
    print(f"단가 풀: {len(prices):,}건")

    records: list[dict] = []
    for label, fname in SRC_FILES:
        src = BASE / "공내역서" / fname
        if not src.exists():
            src = ROOT / fname
        items, _ = asp.load_estimate(src)
        unmatched = 0
        for item in items:
            ranked = topn(item, prices, 3)
            best_score = ranked[0][0] if ranked else 0.0
            if best_score >= asp.THRESHOLD:
                continue  # 매칭/검토 → 입력표 제외
            unmatched += 1
            records.append({
                "file": label,
                "row": item["row"],
                "section": item["section"],
                "name": item["name"],
                "spec": item["spec"],
                "unit": item["unit"],
                "qty": item["qty"],
                "ranked": ranked,
            })
        print(f"{label}: 미매칭 {unmatched}건")

    # 단가 미상이라 금액 정렬 불가 → 파일별 수량 내림차순(물량 우선)
    records.sort(key=lambda r: (r["file"], -r["qty"]))
    write_xlsx(records)
    print(f"\n수동 단가 입력표: {OUT_XLSX} ({len(records)}건)")


def _clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").replace("_x000D_", "").strip()


def _cand_cells(ranked, i):
    """i번째 추천 후보의 (품명, 규격, 단위, 합계단가, 점수) — 없으면 공백."""
    if i < len(ranked):
        s, p, _term = ranked[i]
        return [_clean(p["name"]), _clean(p["spec"]), p["unit"], p["total"], round(s, 3)]
    return ["", "", "", "", ""]


def write_xlsx(records: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "수동단가입력표"

    headers = [
        "파일", "행", "공종", "품명", "규격", "단위", "수량",
        "추천1_품명", "추천1_규격", "추천1_단위", "추천1_합계단가", "추천1_점수",
        "추천2_품명", "추천2_규격", "추천2_합계단가", "추천2_점수",
        "추천3_품명", "추천3_합계단가", "추천3_점수",
        "입력_재료단가", "입력_노무단가", "입력_경비단가", "비고",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    input_cols = {20, 21, 22}
    for r in records:
        ranked = r["ranked"]
        c1 = _cand_cells(ranked, 0)
        c2 = _cand_cells(ranked, 1)
        c3 = _cand_cells(ranked, 2)
        note = "DB 후보 없음 — 단가 직접 입력 필수" if not ranked else "추천 후보 검토 후 단가 입력/수정"
        ws.append([
            r["file"], r["row"], r["section"], r["name"], r["spec"], r["unit"], r["qty"],
            c1[0], c1[1], c1[2], c1[3], c1[4],
            c2[0], c2[1], c2[3], c2[4],
            c3[0], c3[3], c3[4],
            "", "", "", note,
        ])
        ridx = ws.max_row
        for c in input_cols:
            ws.cell(ridx, c).fill = INPUT_FILL
        if not ranked:
            for c in range(1, len(headers) + 1):
                ws.cell(ridx, c).fill = NOCAND_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    money_cols = {11, 15, 18, 20, 21, 22}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(40, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[letter].width = width

    # 안내 시트
    info = wb.create_sheet("안내", 0)
    for line in [
        ["토목 미매칭 — 수동 단가 입력표"],
        [],
        ["목적", "표준 DB 자동매칭(임계 0.56)에서 제외된 항목에 단가를 직접 입력"],
        ["정렬", "파일별 → 수량 내림차순 (단가 미상이라 금액 정렬 불가, 물량 우선)"],
        ["추천 후보", "점수 0.56 미만이라 그대로 쓰면 위험. 품명·규격이 일치할 때만 합계단가 채택"],
        ["입력 방법", "추천이 맞으면 그 합계단가 참고, 아니면 재/노/경 단가 직접 입력"],
        ["입력 칸", "연두색 열: 입력_재료단가 / 입력_노무단가 / 입력_경비단가"],
        ["주황 행", "DB에 호환 후보 없음 — 단가 직접 입력 필수(82건)"],
        ["반영", "입력 후 합계 산식: 수량 × (재료+노무+경비). 총괄표 수동 합산 필요"],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=14)
    for col in info.columns:
        info.column_dimensions[col[0].column_letter].width = min(70, max(len(str(c.value or "")) for c in col) + 2)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT_XLSX)
    except PermissionError:
        alt = OUT_XLSX.with_name(OUT_XLSX.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt.name}")


if __name__ == "__main__":
    main()
