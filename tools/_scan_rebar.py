import sys, glob
from pathlib import Path
from openpyxl import load_workbook
sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
KW = ["이형철근", "철근", "형강", "H형강", "HIV", "나동선", "아스팔트", "HDPE", "스테인리스", "동파이프", "동판"]
for p in sorted(BASE.glob("*_표준단가산출.xlsx")):
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        print("skip", p.name, e); continue
    if "통합내역" not in wb.sheetnames:
        continue
    ws = wb["통합내역"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h}
    def c(r, name):
        for k, i in idx.items():
            if name in k:
                return r[i] if i < len(r) else None
        return None
    hits = []
    for r in rows[1:]:
        nm = str(c(r, "공종명") or c(r, "명칭") or "")
        sp = str(c(r, "규격") or "")
        t = nm + " " + sp
        if any(k in t for k in KW):
            hits.append((nm, sp, str(c(r,"단위") or ""), str(c(r,"상태") or ""),
                         str(c(r,"매칭품명") or ""), c(r,"재료단가"), c(r,"합계단가") or c(r,"노무단가")))
    if hits:
        print(f"\n=== {p.name[:40]} — {len(hits)}건 ===")
        for nm, sp, u, st, mp, mat, tot in hits[:25]:
            print(f"  [{st}] {nm} | {sp} ({u}) → {mp} | 재료={mat}")
    wb.close()
