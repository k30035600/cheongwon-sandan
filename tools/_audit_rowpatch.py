#!/usr/bin/env python3
"""확정단가가 통합내역의 '엉뚱한 행'에 패치됐는지 전수 검사.

검토_*_일위대가산출.xlsx 의 (파일,행,품명) 과
각 _표준단가산출.xlsx 통합내역의 (행,공종명) 을 대조해
품명이 어긋난(오패치) 행을 찾는다.
"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(__file__).resolve().parents[1] / "05_내역서"
WORK = BASE / "내역서작업"

FILE_MAP = {
    "01 토목": "01_화성 청원지구 토목_표준단가산출.xlsx",
    "01 조경": "01_화성 청원지구 조경_표준단가산출.xlsx",
    "04 진입도로": "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
    "05 회전교차로": "05_화성 청원로(회전교차로)_표준단가산출.xlsx",
    "06 개발행위": "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
}

REVIEW_FILES = [
    "검토_공종별_일위대가산출.xlsx",
    "검토_토공_일위대가산출.xlsx",
    "검토_일위대가산출.xlsx",
    "미매칭_일위대가산출.xlsx",
]


def norm(s: str) -> str:
    s = str(s or "")
    for ch in " .)0123456789가나다라마바사아자차카타파하1).":
        pass
    # 앞쪽 "1). " "가). " 등 머리표 제거
    import re
    s = re.sub(r"^[\s0-9가-힣]{0,4}[).]\s*", "", s)
    return s.replace(" ", "")


# 통합내역 캐시: file -> {행: 공종명}
cache = {}
for label, fname in FILE_MAP.items():
    p = WORK / fname
    if not p.exists():
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    w = wb["통합내역"]
    h = [c.value for c in next(w.iter_rows(min_row=1, max_row=1))]
    ri, ni, qi, ti = h.index("행"), h.index("공종명"), h.index("수량"), h.index("합계금액")
    cache[label] = {}
    for r in w.iter_rows(min_row=2, values_only=True):
        if r[ri] is not None:
            cache[label][int(r[ri])] = (r[ni], r[qi], r[ti])
    wb.close()

bad = []
checked = 0
for fn in REVIEW_FILES:
    p = BASE / fn
    if not p.exists():
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["일위대가산출"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}
    fcol = idx.get("파일")
    rcol = idx.get("행")
    pcol = idx.get("품명")
    ccol = idx.get("확정단가(입력)")
    if None in (fcol, rcol, pcol):
        wb.close()
        continue
    for row in ws.iter_rows(min_row=2, values_only=True):
        f = row[fcol]
        if f not in FILE_MAP or row[rcol] is None:
            continue
        conf = row[ccol] if ccol is not None else None
        if not conf or conf <= 0:
            continue  # 확정단가 기입된 것만(=실제 패치된 것)
        checked += 1
        rn = int(row[rcol])
        nm_review = norm(row[pcol])
        target = cache.get(f, {}).get(rn)
        if not target:
            bad.append((fn, f, rn, row[pcol], "통합내역에 행 없음", None, conf))
            continue
        nm_int = norm(target[0])
        if nm_review[:6] != nm_int[:6]:
            bad.append((fn, f, rn, row[pcol], target[0], target[1], conf))
    wb.close()

print(f"확정단가 기입 행 검사: {checked}건")
print(f"오패치(품명 불일치) 의심: {len(bad)}건\n")
for fn, f, rn, nm_r, nm_i, qty, conf in bad:
    impact = (qty or 0) * (conf or 0) if isinstance(qty, (int, float)) else 0
    print(f"  [{f}] 행{rn}")
    print(f"     검토측 품명 = {nm_r}  (확정단가 {conf:,})")
    print(f"     통합내역 품명 = {nm_i}  수량={qty}  → 잘못 반영액 ≈ {impact:,.0f}")
