# -*- coding: utf-8 -*-
"""토목조경_실행예산서.xlsx → HTML 보고서 생성.

정의:
  - 합산(2) = 토목+조경
  - 하단 「토목」(3) = 토목만(전기·조경 제외) — 하도급 내용
  - 하단 「조경」 = 조경 하도급 내용
  - 도급가(희상) = 100억 이상 ~ 300억 이하 · 희상건설 견적 제출가
팩트체크(사사오입/역산) 이슈는 보고서에 언급하지 않음.
"""
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

OUT = Path(r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
           r"\토목조경_실행예산서_보고서.html")


def R(x):
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def won(n, dec=0):
    if n is None:
        return "—"
    if dec == 0:
        return f"{int(round(n)):,}"
    return f"{n:,.{dec}f}"


def pct(n):
    return f"{n * 100:.2f}%"


def calc_indirect(jae, no, gy):
    gn = R(no * 0.197)
    nogye = no + gn
    sj = R(nogye * 0.0356)
    ey = R(nogye * 0.0115)
    gg = R(no * 0.03595)
    yg = R(no * 0.0475)
    yy = R(gg * 0.1314)
    sa = R((jae + no) * 0.026)
    gt = R((jae + nogye) * 0.062)
    d = jae + no + gy
    gm = R(d * 0.004)
    hb = R(d * 0.006)
    sub10 = gn + sj + ey + gg + yg + yy + sa + gt + gm + hb
    ilban = R((d + sub10) * 0.05)
    gyeonggye = gy + (sub10 - gn)
    iyun = R((nogye + gyeonggye + ilban) * 0.07)
    items = [
        ("간접노무비", "직접노무비×19.7%", gn),
        ("산재보험료", "노무비계×3.56%", sj),
        ("고용보험료", "노무비계×1.15%", ey),
        ("국민건강보험료", "직접노무비×3.595%", gg),
        ("국민연금보험료", "직접노무비×4.75%", yg),
        ("노인장기요양보험료", "건강보험료×13.14%", yy),
        ("산업안전보건관리비", "(재+직노)×2.60%", sa),
        ("기타경비(공과잡비)", "(재+노무비계)×6.2%", gt),
        ("건설기계대여대금지급보증", "직접공사비×0.40%", gm),
        ("환경보전비", "직접공사비×0.60%", hb),
        ("일반관리비", "순공사원가×5%", ilban),
        ("이윤", "(노무비계+경비계+일반관리비)×7%", iyun),
    ]
    it = sum(x[2] for x in items)
    return d, items, it, d + it


GONGJONG_META = [
    # (카테고리, 표시 공종명, 해당 면허, 근일, 삼아)
    ("토공", "토공(절·성토·부지조성)", "토공사업(지반조성·포장)", "○", "○ (초기)"),
    ("관로·수공", "우수·오수·상하수 관로", "상·하수도설비공사업", "○", "△ (2027 예정)"),
    ("포장", "아스팔트·콘크리트 포장", "포장공사업(지반조성·포장)", "○", "△ (면허 취득 시)"),
    ("구조물", "구조물(옹벽·RC 등)", "철근·콘크리트공사업", "○", "✗"),
    ("부대·기타", "부대·기타(휀스·경계·법면 등)", "부대(토공부대·기타)", "○", "△ (토공부대 한정)"),
    ("폐기물", "건설폐기물 처리", "폐기물(별도·건설폐기물)", "—", "—"),
]


def _classify_gongjong(name, spec):
    import re as _re
    t = f"{name} {spec}"
    if "폐기물" in t:
        return "폐기물"
    if (any(k in t for k in ["맨홀", "집수정", "받이", "수로", "암거", "저류조",
                             "하수처리", "흄관", "PVC", "이중벽", "콘크리트관",
                             "우수", "오수", "상수관", "직관"])
            or ("관" in name and any(k in t for k in ["PVC", "흄", "콘크리트관",
                                                       "이중벽", "수로", "직관"]))):
        return "관로·수공"
    if any(k in t for k in ["포장", "아스팔트", "기층", "프라임", "택코"]):
        return "포장"
    if any(k in t for k in ["옹벽", "CEP", "보강토", "거푸집", "기초",
                             "철근", "콘크리트", "구조물"]):
        return "구조물"
    if any(k in t for k in ["휀스", "메쉬", "관목", "잔디", "법면", "가드레일",
                             "경계", "표지", "도색", "차선", "규준틀", "식재"]):
        return "부대·기타"
    return "토공"


def compute_gongjong(split_by_name):
    """청원지구_실행내역_분리.xlsx 상세내역을 공종(면허)별 직접비로 집계하고
    각 블록 직접공사비에 맞춰 정합(반올림)하여 반환."""
    import re as _re
    import openpyxl
    src = Path(r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
               r"\청원지구_실행내역_분리.xlsx")
    wb = openpyxl.load_workbook(src, data_only=True)
    cats = ["토공", "관로·수공", "포장", "구조물", "부대·기타", "폐기물"]
    out = {}
    for sn in ["회전진입", "토목", "지구단위"]:
        ws = wb[sn]
        start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value and "상세 내역" in str(ws.cell(r, 1).value):
                start = r
                break
        acc = {c: 0.0 for c in cats}
        block = None
        for r in range(start + 1, ws.max_row + 1):
            c1 = str(ws.cell(r, 1).value or "").strip()
            m = _re.match(r"^(0\d)\.\s", c1)
            if m:
                block = m.group(1)
            q = ws.cell(r, 3).value
            v = ws.cell(r, 12).value
            if isinstance(q, (int, float)) and isinstance(v, (int, float)) and v:
                cat = "폐기물" if block == "07" else _classify_gongjong(
                    c1, str(ws.cell(r, 2).value or ""))
                acc[cat] += v
        # 회전진입 외 블록은 미미한 폐기물(폐기물상차)을 부대·기타로 흡수
        if sn != "회전진입":
            acc["부대·기타"] += acc.pop("폐기물", 0)
            acc["폐기물"] = 0
        # 블록 직접공사비에 맞춰 정합
        auth = split_by_name[sn][2]
        leaf_sum = sum(acc.values())
        scaled = {c: R(acc[c] * auth / leaf_sum) for c in acc}
        resid = auth - sum(scaled.values())
        scaled["토공"] += resid  # 잔차는 최대 항목(토공)에서 흡수
        out[sn] = scaled
    return out


def compute_apply_rate():
    """김보영(실행) ÷ 희상(도급) 토목 공종별 적용률 산출.
    두 통합 견적서의 대공종 요약행(1식) 직접공사비를 비교한다.
    김보영 파일은 금액열이 수식(미캐시)이므로 1식 행의 단가열(재/노/경)을 사용한다.
    """
    import re as _re
    import openpyxl
    base = Path(r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서")
    files = {
        "kb": base / "청원지구_김보영(토목,조경)_통합.xlsx",
        "hs": base / "260708 청원지구_희상(토목,조경)_통합.xlsx",
    }

    def _num(v):
        return v if isinstance(v, (int, float)) else 0

    def collect(p):
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb["견적서"]
        out = {}
        for r in range(4, ws.max_row + 1):
            c1 = str(ws.cell(r, 1).value or "").strip()
            c2 = str(ws.cell(r, 2).value or "").strip()
            if ("토목·조경" in c1 and _re.match(r"^\d{1,2}\.\s", c2)
                    and ws.cell(r, 3).value == 1):
                jae = _num(ws.cell(r, 6).value) or _num(ws.cell(r, 5).value)
                no = _num(ws.cell(r, 8).value) or _num(ws.cell(r, 7).value)
                gy = _num(ws.cell(r, 10).value) or _num(ws.cell(r, 9).value)
                tot = _num(ws.cell(r, 12).value) or (jae + no + gy)
                name = _re.sub(r"\s+", "", _re.sub(r"^\d{1,2}\.\s*", "", c2))
                if tot:
                    out[name] = tot
        return out

    kb = collect(files["kb"])
    hs = collect(files["hs"])
    order = ["토공", "우수공", "오수공", "상수공", "포장공", "구조물공",
             "부대공", "하수처리장", "지하저류조"]
    labels = {"토공": "토공", "우수공": "우수공", "오수공": "오수공",
              "상수공": "상수공", "포장공": "포장공", "구조물공": "구조물공",
              "부대공": "부대공", "하수처리장": "하수처리장",
              "지하저류조": "지하저류조"}
    rows = []
    sk = sh = 0
    for k in order:
        a, b = kb.get(k, 0), hs.get(k, 0)
        sk += a
        sh += b
        rows.append((labels[k], a, b, a / b if b else 0))
    return rows, sk, sh, (sk / sh if sh else 0), kb.get("조경공", 0), hs.get("조경공", 0)


def main():
    F6, G6, H6 = 4_245_357_794, 2_147_353_595, 4_811_196_564
    I6 = 2_697_898_148.064575
    E6 = 0.75
    D5 = 0.81995
    C5, D7 = 0.79995, 0.88745

    # 희상건설 — 토목만 (전기·조경 제외)
    F9, G9, H9 = 3_646_620_419, 1_897_107_375, 4_109_082_787
    F10, G10, H10 = 77_500_490, 57_019_245, 691_209_114
    F11, G11, H11 = 521_236_885, 193_226_975, 10_904_663

    # 하단 「토목」— 본 파일 정의상 토목만(전기·조경 제외)
    F33, G33, H33 = 3_179_386_713, 1_631_287_599, 3_543_142_087
    F50, G50, H50 = 77_500_490, 57_019_245, 691_209_114

    direct_hs = F6 + G6 + H6
    K9 = (F9 + G9 + H9) + (F10 + G10 + H10) + (F11 + G11 + H11)
    J6 = direct_hs + I6 + E6
    J6_clean = direct_hs + I6
    D9 = K9 / J6
    jae4, no4, gy4, ind4 = F6 / D9, G6 / D9, H6 / D9, I6 / D9
    design = jae4 + no4 + gy4 + ind4
    bid = design * D5
    dogup_hs = J6_clean

    ind_top = [
        ("간접노무비", "직접노무비×19.7%", 423_028_658.215),
        ("산재보험료", "노무비계×3.56%", 91_505_608.21445401),
        ("고용보험료", "노무비계×1.15%", 29_559_395.9119725),
        ("국민건강보험료", "직접노무비×3.595%", 77_197_361.74025),
        ("국민연금보험료", "직접노무비×4.75%", 101_999_295.7625),
        ("노인장기요양보험료", "건강보험료×13.14%", 10_143_733.33266885),
        ("산업안전보건관리비", "(재+직노)×2.60%", 166_210_496.114),
        ("기타경비(공과잡비)", "(재+노무비계)×6.2%", 422_575_882.92733),
        ("건설기계대여대금지급보증", "직접공사비×0.40%", 44_815_631.812),
        ("환경보전비", "직접공사비×0.60%", 67_223_447.718),
        ("일반관리비", "순공사원가×5%", 631_908_373.2374089),
        ("이윤", "(노무비계+경비계+일반관리비)×7%", 631_730_263.0789909),
    ]

    # 하단 토목 / 조경
    dt, items_t, it_t, ht = calc_indirect(F33, G33, H33)
    dj, items_j, it_j, hj = calc_indirect(F50, G50, H50)
    sum_d, sum_i, sum_h = dt + dj, it_t + it_j, ht + hj
    vat = R(sum_h * 0.1)
    dogup = sum_h + vat

    # 희상 토목만 (전기·조경 제외) — 비교용으로 동일 요율 ROUND 적용
    hs_tom_d, items_hs, hs_tom_i, hs_tom_h = calc_indirect(F9, G9, H9)
    hs_jog = F10 + G10 + H10
    hs_ele = F11 + G11 + H11

    # 4분할 실행내역 (청원지구_실행내역_분리.xlsx) — 하도급 차수별
    # 차수 순서: 회전진입 → 토목/조경 → 지구단위
    split_groups = [
        ("회전진입", "1차 · 04.진입도로+07.폐기물+05.회전교차로",
         1_794_475_779, 483_701_146, 2_505_994_618),
        ("토목", "2차 · 01. 토목(1~9공종)", 4_037_192_417, 1_008_963_933, 5_550_771_985),
        ("조경", "2차 · 01. 조경(10.조경공)", 825_728_849, 141_791_919, 1_064_272_845),
        ("지구단위", "3차 · 06. 개발행위", 2_522_148_203, 534_317_233, 3_362_111_980),
    ]
    split_by_name = {g[0]: g for g in split_groups}
    split_d = sum(x[2] for x in split_groups)
    split_i = sum(x[3] for x in split_groups)
    split_dog = sum(x[4] for x in split_groups)
    split_excl = [g for g in split_groups if g[0] != "조경"]
    excl_d = sum(x[2] for x in split_excl)
    excl_i = sum(x[3] for x in split_excl)
    excl_dog = sum(x[4] for x in split_excl)

    # 공종(면허 카테고리)별 직접공사비 — 상세내역 단가행 집계 후
    # 각 블록 직접공사비(split_by_name[*][2])에 맞춰 정합(반올림).
    gongjong = compute_gongjong(split_by_name)

    block_labels = {
        "회전진입": "회전진입", "토목": "토목", "지구단위": "지구단위",
    }
    anb_parts = []
    for bname in ["회전진입", "토목", "지구단위"]:
        cats = gongjong[bname]
        meta = [m for m in GONGJONG_META if cats.get(m[0], 0)]
        span = len(meta)
        bdir = split_by_name[bname][2]
        head = (f'<td rowspan="{span}"><strong>{block_labels[bname]}</strong>'
                f'<br><span class="muted">직접 {won(bdir)}원</span></td>')
        for i, (cat, gname, lic, kg, sa) in enumerate(meta):
            amt = cats[cat]
            first = head if i == 0 else ""
            anb_parts.append(
                f'<tr>{first}<td>{gname}</td><td>{lic}</td>'
                f'<td class="num">{won(amt)}</td>'
                f'<td class="c">{kg}</td><td class="c">{sa}</td></tr>')
    anb_rows = "\n".join(anb_parts)

    # ── 개별하도급(업종별 개별계약) 배분 — 도급공사비 안분내역서 기준 ──
    # 공종(면허)별 도급공사비 = 직접(gongjong) × 블록 도급/직접 비율(잔차는 토공 흡수).
    # 업종별 담당 수급인: 토공·부대 → 삼아, 관로·포장·구조물·폐기물 → 근일.
    ASSIGN = {
        "토공": "삼아건설", "부대·기타": "삼아건설",
        "관로·수공": "근일건설", "포장": "근일건설",
        "구조물": "근일건설", "폐기물": "근일건설",
    }
    dogub_by_block = {}
    for bname in ["회전진입", "토목", "지구단위"]:
        cats = gongjong[bname]
        bdir = split_by_name[bname][2]
        bdog = split_by_name[bname][4]  # 도급(VAT 포함)
        ratio = bdog / bdir
        scaled = {c: R(v * ratio) for c, v in cats.items() if v}
        scaled["토공"] = scaled.get("토공", 0) + (bdog - sum(scaled.values()))
        dogub_by_block[bname] = scaled

    alloc_parts = []
    geun_dogub = sam_dogub = 0
    eupjong = {}   # 업종(면허) → 도급액
    for bname in ["회전진입", "토목", "지구단위"]:
        scaled = dogub_by_block[bname]
        meta = [m for m in GONGJONG_META if scaled.get(m[0], 0)]
        span = len(meta)
        bdog = split_by_name[bname][4]
        head = (f'<td rowspan="{span}"><strong>{bname}</strong>'
                f'<br><span class="muted">도급 {won(bdog)}원</span></td>')
        for i, (cat, gname, lic, kg, sa) in enumerate(meta):
            amt = scaled[cat]
            comp = ASSIGN[cat]
            if comp == "근일건설":
                geun_dogub += amt
            else:
                sam_dogub += amt
            eupjong[lic] = eupjong.get(lic, 0) + amt
            first = head if i == 0 else ""
            gcell = f'<td class="num">{won(amt)}</td>' if comp == "근일건설" else '<td class="num">—</td>'
            scell = f'<td class="num">{won(amt)}</td>' if comp == "삼아건설" else '<td class="num">—</td>'
            alloc_parts.append(
                f'<tr>{first}<td>{gname}</td><td>{lic}</td>'
                f'<td class="num">{won(amt)}</td>{gcell}{scell}</tr>')
    alloc_rows = "\n".join(alloc_parts)
    tot_dogub = geun_dogub + sam_dogub
    geun_pct = geun_dogub / tot_dogub
    sam_pct = sam_dogub / tot_dogub

    # 업종별 집계 행 + 담당
    EUP_COMP = {
        "토공사업(지반조성·포장)": "삼아건설",
        "부대(토공부대·기타)": "삼아건설",
        "상·하수도설비공사업": "근일건설",
        "포장공사업(지반조성·포장)": "근일건설",
        "철근·콘크리트공사업": "근일건설",
        "폐기물(별도·건설폐기물)": "근일건설",
    }
    eup_order = ["토공사업(지반조성·포장)", "상·하수도설비공사업",
                 "포장공사업(지반조성·포장)", "철근·콘크리트공사업",
                 "부대(토공부대·기타)", "폐기물(별도·건설폐기물)"]
    eup_parts = []
    for lic in eup_order:
        if lic not in eupjong:
            continue
        eup_parts.append(
            f'<tr><td>{lic}</td><td class="num">{won(eupjong[lic])}</td>'
            f'<td>{EUP_COMP.get(lic, "—")}</td></tr>')
    eupjong_rows = "\n".join(eup_parts)

    # 토목 공종별 적용률 (김보영 실행 ÷ 희상 도급)
    rate_data, rate_kb, rate_hs, rate_tot, jog_kb, jog_hs = compute_apply_rate()
    rate_parts = []
    for name, a, b, pr in rate_data:
        rate_parts.append(
            f'<tr><td>{name}</td><td class="num bt">{won(a)}</td>'
            f'<td class="num hs">{won(b)}</td>'
            f'<td class="num">{pct(pr)}</td></tr>')
    rate_rows = "\n".join(rate_parts)

    rows_stack = [
        ("설계가", "100억 이상 ~ 300억 미만 · 표준품셈/일위대가", design, "역산(도급÷직접비중)"),
        ("입찰가", f"조달청 낙찰율 D열 {D5} ({pct(D5)})", bid, "설계가×0.81995"),
        ("도급가(희상)", "100억 이상 ~ 300억 이하 · 희상건설 견적 제출가",
         dogup_hs, "직접+간접(토목·조경·전기)"),
        ("잔여(입찰−도급)", "입찰가 − 도급가", bid - dogup_hs, "입찰가−도급가"),
    ]

    def diff_cell(a, b):
        d = a - b
        cls = "ok" if d == 0 else ("warn" if abs(d) < 100 else "bad")
        sign = "+" if d > 0 else ""
        return f'<td class="num {cls}">{sign}{won(d)}</td>', d

    html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 — 설계가·입찰가·희상·토목·조경 보고서</title>
<style>
  @page{{margin:15mm 12mm;}}
  body{{font-family:"Malgun Gothic","맑은 고딕",system-ui,sans-serif;line-height:1.65;
    color:#1c2430;max-width:1080px;margin:0 auto;padding:28px 20px 60px;font-size:15px;background:#f7f9fc;}}
  h1{{font-size:22px;color:#1f3a5f;border-bottom:3px solid #1f3a5f;padding-bottom:10px;}}
  h2{{font-size:18px;color:#1f3a5f;margin-top:30px;border-left:5px solid #2c5aa0;padding-left:10px;}}
  h3{{font-size:16px;color:#2c5aa0;margin-top:22px;}}
  table{{width:100%;border-collapse:collapse;background:#fff;margin:12px 0;font-size:13.5px;
    border:1px solid #d9e1ec;border-radius:8px;overflow:hidden;}}
  th,td{{padding:7px 10px;border-bottom:1px solid #e3e9f2;vertical-align:top;text-align:left;}}
  th{{background:#eaf0f8;color:#1f3a5f;text-align:center;font-weight:700;}}
  th.num,td.num{{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap;}}
  td.c{{text-align:center;white-space:nowrap;}}
  tr.total td{{font-weight:700;background:#fff2cc;}}
  tr.subtotal td{{font-weight:600;background:#f4f6fa;}}
  .ok{{color:#137333;font-weight:700;}}
  .warn{{color:#b06000;font-weight:700;}}
  .bad{{color:#b3261e;font-weight:700;}}
  .hs{{background:#ddebf7;}}
  .bt{{background:#fce4d6;}}
  blockquote{{background:#eef4fb;border:1px solid #c5d0e6;border-radius:8px;padding:10px 16px;margin:14px 0;}}
  .muted{{color:#6b7787;font-size:13px;}}
  code{{background:#eef1f6;border-radius:4px;padding:1px 5px;font-size:13px;}}
  ul{{margin:8px 0;}}
  .kpi{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin:14px 0;}}
  .kpi div{{background:#fff;border:1px solid #d9e1ec;border-radius:8px;padding:12px 14px;}}
  .kpi .l{{font-size:12px;color:#6b7787;}}
  .kpi .v{{font-size:18px;font-weight:700;color:#1f3a5f;font-variant-numeric:tabular-nums;}}
  .def{{background:#fff;border:1px solid #d9e1ec;border-radius:8px;padding:10px 14px;margin:12px 0;}}
</style></head><body>

<h1>설계가 · 입찰가 · 희상건설 · 토목 · 조경 보고서</h1>
<ul>
  <li><strong>원본(도급·하도급 비교)</strong>: <code>07_타견적/실행예산서/토목조경_실행예산서.xlsx</code></li>
  <li><strong>원본(4분할 실행내역)</strong>: <code>07_타견적/실행예산서/청원지구_실행내역_분리.xlsx</code>
    — 시트: 표지 · 대공정별집계표 · 토목 · 조경 · 회전진입 · 지구단위</li>
  <li><strong>작성일</strong>: 2026. 7. 10.</li>
</ul>

<div class="def">
<strong>용어 정의</strong>
<ol style="margin:6px 0">
  <li><strong>도급가(희상)</strong> — 희상건설의 견적 제출가이다(토목·조경·전기 포함, 공사규모 100억~300억 구간).</li>
  <li><strong>하단 토목·조경</strong> — <strong>하도급</strong> 내역이다(전기 제외). 상단 희상 도급가와 구분한다.</li>
  <li><strong>합산</strong> — 하도급 <strong>토목 + 조경</strong>의 합계이다.</li>
  <li><strong>하단 「토목」</strong> — 하도급 중 토목만(전기·조경 제외) 집계한 것으로, 회전교차로·진입도로·폐기물·지구단위를 포함한 토목 실행 집계이다.</li>
  <li><strong>4분할 실행내역</strong> — 하도급을 <strong>회전진입 → 토목/조경 → 지구단위</strong> 순으로 구분한 총액계약 차수별 내역이다
    (<code>청원지구_실행내역_분리.xlsx</code>). 07.건설폐기물은 진입도로(회전진입)에 포함한다.</li>
  <li><strong>개별하도급 (업종별 개별계약)</strong> — 하도급을 공동수급체가 아니라 <strong>공종(업종)별로 자격을 갖춘 수급인(근일건설·삼아건설)과 개별 계약</strong>으로 시행하는 방식이다(<code>도급공사비 안분내역서(개별하도급·업종별 개별계약).xlsx</code>).</li>
</ol>
</div>

<blockquote>
<strong>4분할의 취지</strong>
<ol style="margin:6px 0">
  <li><strong>초기 도급한도액의 한계</strong>를 극복하기 위함이다. 단일 계약으로 소화하기 어려운 규모이므로 구분한다.</li>
  <li><strong>총액계약 · 차수별 준공</strong>으로 보증서·선급금 발행 부담을 경감한다.</li>
  <li>초기(<strong>2026년</strong>) 공사 실적을 <strong>2027년</strong>에 반영하기 위함이다.</li>
  <li>이에 따라 <strong>공정별로 착공·준공·정산</strong>이 개별 진행된다.</li>
  <li>하도급은 <strong>개별하도급(업종별 개별계약)</strong>으로 시행한다(근일건설·삼아건설이 면허 범위 공종을 각각 개별 계약).</li>
  <li><strong>신규면허 단독으로는 진입장벽이 높다.</strong> 그 사유는 다음과 같다.
    <ul style="margin:4px 0">
      <li>신규등록 초기에는 <strong>공사실적이 0</strong>이므로 시공능력평가액(≈도급한도액)이 실질자본금·기술능력 위주로 <strong>낮게 산정</strong>된다.</li>
      <li>그 결과 <strong>초기 도급한도</strong>가 작아 대규모 공사를 <strong>단독으로 수주·시공하기 어렵다.</strong></li>
      <li>보증서(계약·선급금·하자) 발급 한도도 자본·실적에 연동되어 <strong>초기에는 제약</strong>이 크다.</li>
      <li>따라서 <strong>실적을 보유한 근일건설과 업종별로 병행(개별하도급)</strong>하여 신규면허의 초기 한계를 보완하면서 실적을 축적하는 방식이 합리적이다.</li>
    </ul>
  </li>
</ol>
</blockquote>

<blockquote>
<strong>핵심 요약</strong>
<ol style="margin:6px 0">
  <li>상단 <strong>도급가(희상)</strong>는 직접 <strong>{won(direct_hs)}</strong>원에 간접 <strong>{won(I6)}</strong>원을 더한 <strong>{won(dogup_hs)}</strong>원이다(토목·조경·전기).</li>
  <li>하단은 <strong>하도급</strong>이다. <strong>토목+조경 합산</strong> 직접 <strong>{won(sum_d)}</strong>원, 부가세 포함 하도급액 <strong>{won(dogup)}</strong>원이다.</li>
  <li>그중 <strong>토목만</strong>(전기·조경 제외) 직접 <strong>{won(dt)}</strong>원, <strong>조경</strong> 직접 <strong>{won(dj)}</strong>원이다.</li>
  <li>4분할 직접합도 <strong>{won(sum_d)}</strong>원으로 하도급 합산과 일치한다(전기 제외).</li>
  <li>개별하도급(업종별) 배분은 <strong>근일건설 {won(geun_dogub)}원({pct(geun_pct)})</strong>, <strong>삼아건설 {won(sam_dogub)}원({pct(sam_pct)})</strong>이다(도급공사비 기준, Ⅶ).</li>
</ol>
</blockquote>

<div class="kpi">
  <div><div class="l">설계가(역산)</div><div class="v">{won(design)}</div></div>
  <div><div class="l">입찰가(×{D5})</div><div class="v">{won(bid)}</div></div>
  <div><div class="l">도급가(희상) 100~300억</div><div class="v">{won(dogup_hs)}</div></div>
  <div><div class="l">하도급 합산(토+조·VAT)</div><div class="v">{won(dogup)}</div></div>
</div>

<h2>Ⅰ. 설계가 · 입찰가 · 도급가 계층</h2>
<table>
<thead><tr>
  <th>구분</th><th>공사규모·비고</th><th class="num">금액(원)</th><th>산출</th>
</tr></thead>
<tbody>
"""
    for name, note, amt, how in rows_stack:
        cls = ' class="total"' if "도급가" in name else ""
        html += (f"<tr{cls}><td>{name}</td><td>{note}</td>"
                 f"<td class=\"num\">{won(amt)}</td><td>{how}</td></tr>\n")

    html += f"""
</tbody></table>

<table>
<thead><tr>
  <th>구성</th><th class="num">설계가(역산)</th><th class="num">입찰가(×{D5})</th>
  <th class="num">도급가(희상)</th><th class="num">입찰−도급</th>
</tr></thead>
<tbody>
<tr><td>재료비</td><td class="num">{won(jae4)}</td><td class="num">{won(jae4*D5)}</td>
  <td class="num">{won(F6)}</td><td class="num">{won(jae4*D5-F6)}</td></tr>
<tr><td>노무비</td><td class="num">{won(no4)}</td><td class="num">{won(no4*D5)}</td>
  <td class="num">{won(G6)}</td><td class="num">{won(no4*D5-G6)}</td></tr>
<tr><td>경비</td><td class="num">{won(gy4)}</td><td class="num">{won(gy4*D5)}</td>
  <td class="num">{won(H6)}</td><td class="num">{won(gy4*D5-H6)}</td></tr>
<tr><td>간접공사비</td><td class="num">{won(ind4)}</td><td class="num">{won(ind4*D5)}</td>
  <td class="num">{won(I6)}</td><td class="num">{won(ind4*D5-I6)}</td></tr>
<tr class="total"><td>합계</td><td class="num">{won(design)}</td><td class="num">{won(bid)}</td>
  <td class="num">{won(dogup_hs)}</td><td class="num">{won(bid-dogup_hs)}</td></tr>
</tbody></table>
<p class="muted">조달청 열 참고: 입찰(100~300억) C5={C5}, D7={D7} (2026. 1. 30. 시행 주석).</p>

<h2>Ⅱ. 희상건설 공종 분해</h2>
<p class="muted">아래는 희상 견적 원본을 토목·조경·전기로 구분한 것이다. 이후 「토목만」 비교에서는 전기·조경을 제외한다.</p>
<table>
<thead><tr>
  <th>공종</th><th class="num">재료비</th><th class="num">노무비</th>
  <th class="num">경비</th><th class="num">직접계</th><th class="num">비중</th>
</tr></thead>
<tbody>
<tr><td>토목공사</td><td class="num">{won(F9)}</td><td class="num">{won(G9)}</td>
  <td class="num">{won(H9)}</td><td class="num">{won(hs_tom_d)}</td>
  <td class="num">{pct(hs_tom_d/K9)}</td></tr>
<tr><td>조경공사</td><td class="num">{won(F10)}</td><td class="num">{won(G10)}</td>
  <td class="num">{won(H10)}</td><td class="num">{won(hs_jog)}</td>
  <td class="num">{pct(hs_jog/K9)}</td></tr>
<tr><td>전기공사</td><td class="num">{won(F11)}</td><td class="num">{won(G11)}</td>
  <td class="num">{won(H11)}</td><td class="num">{won(hs_ele)}</td>
  <td class="num">{pct(hs_ele/K9)}</td></tr>
<tr class="total"><td>합계(=희상 직접)</td><td class="num">{won(F6)}</td><td class="num">{won(G6)}</td>
  <td class="num">{won(H6)}</td><td class="num">{won(K9)}</td>
  <td class="num">100%</td></tr>
</tbody></table>

<h3>Ⅱ-1. 희상 간접공사비 (토목·조경·전기 기준)</h3>
<table>
<thead><tr><th>항목</th><th>산식</th><th class="num">금액(원)</th></tr></thead>
<tbody>
"""
    for lab, formula, amt in ind_top:
        html += f"<tr><td>{lab}</td><td>{formula}</td><td class=\"num\">{won(amt, 2)}</td></tr>\n"
    html += f"""
<tr class="total"><td colspan="2">간접공사비 계 (=I6)</td><td class="num">{won(I6, 2)}</td></tr>
</tbody></table>

<h2>Ⅲ. 합산 — 하도급 (토목+조경)</h2>
<p class="muted">여기서 다루는 토목·조경은 모두 <strong>하도급</strong>이다. 합산은 토목 하도급과 조경 하도급의 합계이며, 전기는 제외한다.</p>
<table>
<thead><tr>
  <th>구분</th><th class="num">재료비</th><th class="num">노무비</th><th class="num">경비</th>
  <th class="num">직접공사비</th><th class="num">간접공사비</th>
  <th class="num">공급가액</th>
</tr></thead>
<tbody>
<tr><td>하도급 토목만 (전기·조경 제외)</td>
  <td class="num">{won(F33)}</td><td class="num">{won(G33)}</td><td class="num">{won(H33)}</td>
  <td class="num">{won(dt)}</td><td class="num">{won(it_t)}</td><td class="num">{won(ht)}</td></tr>
<tr><td>하도급 조경</td>
  <td class="num">{won(F50)}</td><td class="num">{won(G50)}</td><td class="num">{won(H50)}</td>
  <td class="num">{won(dj)}</td><td class="num">{won(it_j)}</td><td class="num">{won(hj)}</td></tr>
<tr class="total"><td>합산 (하도급 토목+조경)</td>
  <td class="num">{won(F33+F50)}</td><td class="num">{won(G33+G50)}</td><td class="num">{won(H33+H50)}</td>
  <td class="num">{won(sum_d)}</td><td class="num">{won(sum_i)}</td><td class="num">{won(sum_h)}</td></tr>
</tbody></table>
<table>
<thead><tr><th>항목</th><th class="num">금액(원)</th><th>비고</th></tr></thead>
<tbody>
<tr><td>공급가액 (직접+간접)</td><td class="num">{won(sum_h)}</td><td>하도급 토목+조경</td></tr>
<tr><td>부가가치세</td><td class="num">{won(vat)}</td><td>공급가액×10%</td></tr>
<tr class="total"><td>하도급액</td><td class="num">{won(dogup)}</td><td>공급가액+부가세</td></tr>
</tbody></table>

<h2>Ⅳ. 토목공사 비교 검토표 — 희상(도급) vs 하단(하도급) · 전기·조경 제외</h2>
<p class="muted">
희상「토목공사」(도급·9행)와 하단「토목」(<strong>하도급</strong>·33행)을 <strong>전기·조경 제외 토목만</strong> 기준으로 비교한다.
하단 간접은 엑셀 ROUND 수식이며, 희상 토목 간접은 동일 요율을 토목 직접비에만 적용하여 ROUND한 비교용 수치이다.
</p>

<h3>Ⅳ-1. 직접공사비 비교</h3>
<table>
<thead><tr>
  <th>항목</th>
  <th class="num hs">희상 토목(도급)</th>
  <th class="num bt">하단 토목(하도급)</th>
  <th class="num">차이(도급−하도급)</th>
  <th class="num">비율(하도급/도급)</th>
</tr></thead>
<tbody>
<tr><td>재료비</td>
  <td class="num hs">{won(F9)}</td><td class="num bt">{won(F33)}</td>
  {diff_cell(F9, F33)[0]}
  <td class="num">{pct(F33/F9)}</td></tr>
<tr><td>노무비</td>
  <td class="num hs">{won(G9)}</td><td class="num bt">{won(G33)}</td>
  {diff_cell(G9, G33)[0]}
  <td class="num">{pct(G33/G9)}</td></tr>
<tr><td>경비</td>
  <td class="num hs">{won(H9)}</td><td class="num bt">{won(H33)}</td>
  {diff_cell(H9, H33)[0]}
  <td class="num">{pct(H33/H9)}</td></tr>
<tr class="total"><td>직접공사비 계</td>
  <td class="num hs">{won(hs_tom_d)}</td><td class="num bt">{won(dt)}</td>
  {diff_cell(hs_tom_d, dt)[0]}
  <td class="num">{pct(dt/hs_tom_d)}</td></tr>
</tbody></table>

<h3>Ⅳ-2. 간접공사비 비교 (동일 요율 · ROUND)</h3>
<table>
<thead><tr>
  <th>항목</th><th>산식</th>
  <th class="num hs">희상 토목(도급)</th>
  <th class="num bt">하단 토목(하도급)</th>
  <th class="num">차이(도급−하도급)</th>
</tr></thead>
<tbody>
"""
    for (lab, formula, a), (_, _, b) in zip(items_hs, items_t):
        cell, _ = diff_cell(a, b)
        html += (f"<tr><td>{lab}</td><td>{formula}</td>"
                 f"<td class=\"num hs\">{won(a)}</td>"
                 f"<td class=\"num bt\">{won(b)}</td>{cell}</tr>\n")
    html += f"""
<tr class="total"><td colspan="2">간접공사비 계</td>
  <td class="num hs">{won(hs_tom_i)}</td><td class="num bt">{won(it_t)}</td>
  {diff_cell(hs_tom_i, it_t)[0]}</tr>
<tr class="total"><td colspan="2">공급가액 (직+간)</td>
  <td class="num hs">{won(hs_tom_h)}</td><td class="num bt">{won(ht)}</td>
  {diff_cell(hs_tom_h, ht)[0]}</tr>
</tbody></table>

<table>
<thead><tr><th>요약</th><th class="num hs">희상 토목(도급)</th>
  <th class="num bt">하단 토목(하도급)</th><th class="num">차이</th><th>비고</th></tr></thead>
<tbody>
<tr><td>직접공사비</td><td class="num hs">{won(hs_tom_d)}</td>
  <td class="num bt">{won(dt)}</td>{diff_cell(hs_tom_d, dt)[0]}
  <td>하도급/도급 = {pct(dt/hs_tom_d)}</td></tr>
<tr><td>간접공사비</td><td class="num hs">{won(hs_tom_i)}</td>
  <td class="num bt">{won(it_t)}</td>{diff_cell(hs_tom_i, it_t)[0]}
  <td>동일 요율 ROUND</td></tr>
<tr class="total"><td>공급가액</td><td class="num hs">{won(hs_tom_h)}</td>
  <td class="num bt">{won(ht)}</td>{diff_cell(hs_tom_h, ht)[0]}
  <td>전기·조경 제외</td></tr>
<tr><td>부가세</td><td class="num hs">{won(R(hs_tom_h*0.1))}</td>
  <td class="num bt">{won(R(ht*0.1))}</td>
  {diff_cell(R(hs_tom_h*0.1), R(ht*0.1))[0]}
  <td>×10%</td></tr>
<tr class="total"><td>금액(VAT포함)</td>
  <td class="num hs">{won(hs_tom_h + R(hs_tom_h*0.1))}</td>
  <td class="num bt">{won(ht + R(ht*0.1))}</td>
  {diff_cell(hs_tom_h + R(hs_tom_h*0.1), ht + R(ht*0.1))[0]}
  <td>도급 vs 하도급 · 토목만</td></tr>
</tbody></table>

<p class="muted">차이 총괄: 직접 <strong>{won(hs_tom_d - dt)}</strong>원
(재료 {won(F9-F33)} · 노무 {won(G9-G33)} · 경비 {won(H9-H33)}),
공급가액 <strong>{won(hs_tom_h - ht)}</strong>원.
하도급 토목은 희상(도급) 토목의 <strong>{pct(dt/hs_tom_d)}</strong> 수준이다.</p>

<h3>Ⅳ-3. 토목 공종별 적용률 — 김보영(실행) ÷ 희상(도급)</h3>
<p class="muted">
두 통합 견적서의 <strong>대공종 요약행(직접공사비)</strong>을 공종별로 비교한 적용률이다
(<code>청원지구_김보영(토목,조경)_통합.xlsx</code> ÷ <code>260708 청원지구_희상(토목,조경)_통합.xlsx</code>).
조경(10.조경공)은 양사 금액이 동일하여 제외한다.
</p>
<table>
<thead><tr>
  <th>공종</th><th class="num bt">김보영(실행)</th>
  <th class="num hs">희상(도급)</th><th class="num">적용률</th>
</tr></thead>
<tbody>
{rate_rows}
<tr class="total"><td>토목 소계(1~9공종)</td>
  <td class="num bt">{won(rate_kb)}</td>
  <td class="num hs">{won(rate_hs)}</td>
  <td class="num">{pct(rate_tot)}</td></tr>
<tr class="subtotal"><td>조경(참고 · 제외)</td>
  <td class="num bt">{won(jog_kb)}</td>
  <td class="num hs">{won(jog_hs)}</td>
  <td class="num">{pct(jog_kb/jog_hs) if jog_hs else '—'}</td></tr>
</tbody></table>
<p class="muted">
토목 전체 적용률은 <strong>{pct(rate_tot)}</strong>(도급 대비 실행, 할인율 약 {pct(1-rate_tot)})이며,
공종별로 <strong>81~95%</strong>의 편차가 있다(단일 균일 요율이 아니라 공종별 단가·물량 조정 결과).
직접공사비 기준이며, 간접비 산식이 동일 요율이므로 도급액 기준으로도 유사하게 유지된다.
</p>

<h2>Ⅴ. 하도급 토목 · 조경 간접비 상세</h2>
<p class="muted">아래는 모두 <strong>하도급</strong> 내역이다.</p>
<h3>Ⅴ-1. 하도급 토목만 (전기·조경 제외)</h3>
<table>
<thead><tr><th>항목</th><th>산식</th><th class="num">금액(원)</th></tr></thead>
<tbody>
<tr class="subtotal"><td colspan="2">직접공사비 계</td><td class="num">{won(dt)}</td></tr>
"""
    for lab, formula, amt in items_t:
        html += f"<tr><td>{lab}</td><td>{formula}</td><td class=\"num\">{won(amt)}</td></tr>\n"
    html += f"""
<tr class="total"><td colspan="2">간접공사비 계</td><td class="num">{won(it_t)}</td></tr>
<tr class="total"><td colspan="2">공급가액</td><td class="num">{won(ht)}</td></tr>
</tbody></table>

<h3>Ⅴ-2. 하도급 조경</h3>
<table>
<thead><tr><th>항목</th><th>산식</th><th class="num">금액(원)</th></tr></thead>
<tbody>
<tr class="subtotal"><td colspan="2">직접공사비 계</td><td class="num">{won(dj)}</td></tr>
"""
    for lab, formula, amt in items_j:
        html += f"<tr><td>{lab}</td><td>{formula}</td><td class=\"num\">{won(amt)}</td></tr>\n"
    html += f"""
<tr class="total"><td colspan="2">간접공사비 계</td><td class="num">{won(it_j)}</td></tr>
<tr class="total"><td colspan="2">공급가액</td><td class="num">{won(hj)}</td></tr>
</tbody></table>

<h2>Ⅵ. 4분할 실행내역 — 총액계약 차수별 (하도급)</h2>
<p class="muted">출처: <code>청원지구_실행내역_분리.xlsx</code>.
<strong>차수 순서는 회전진입 → 토목/조경 → 지구단위</strong>이다.
4개로 구분한 취지는 <strong>초기 도급한도액 한계</strong> 극복,
<strong>총액계약 · 차수별 준공</strong>을 통한 보증서·선급금 발행 부담 경감,
초기(<strong>2026년</strong>) 실적의 <strong>2027년</strong> 반영에 있다.
이에 따라 <strong>공정별 착공·준공·정산</strong>이 개별 진행되며,
07.건설폐기물은 진입도로(회전진입)에 포함한다.</p>

<table>
<thead><tr>
  <th>차수·공정</th><th>대상</th>
  <th class="num">직접공사비</th><th class="num">간접공사비</th>
  <th class="num">공급가액</th><th class="num">하도급액(VAT포함)</th>
  <th class="num">직접 비중</th>
</tr></thead>
<tbody>
"""
    for name, desc, d, i, dog in split_groups:
        hab = d + i
        html += (f"<tr><td><strong>{name}</strong></td><td>{desc}</td>"
                 f"<td class=\"num\">{won(d)}</td>"
                 f"<td class=\"num\">{won(i)}</td>"
                 f"<td class=\"num\">{won(hab)}</td>"
                 f"<td class=\"num\">{won(dog)}</td>"
                 f"<td class=\"num\">{pct(d/split_d)}</td></tr>\n")
    html += f"""
<tr class="total"><td colspan="2">총계 (조경 포함)</td>
  <td class="num">{won(split_d)}</td><td class="num">{won(split_i)}</td>
  <td class="num">{won(split_d+split_i)}</td><td class="num">{won(split_dog)}</td>
  <td class="num">100%</td></tr>
<tr class="subtotal"><td colspan="2">총계 (조경 제외)</td>
  <td class="num">{won(excl_d)}</td><td class="num">{won(excl_i)}</td>
  <td class="num">{won(excl_d+excl_i)}</td><td class="num">{won(excl_dog)}</td>
  <td class="num">{pct(excl_d/split_d)}</td></tr>
</tbody></table>

<table>
<thead><tr>
  <th>대조</th><th class="num">4분할 합</th><th class="num">하도급 합산(Ⅲ)</th>
  <th class="num">차이</th><th>비고</th>
</tr></thead>
<tbody>
<tr><td>직접공사비</td><td class="num">{won(split_d)}</td>
  <td class="num">{won(sum_d)}</td>{diff_cell(split_d, sum_d)[0]}
  <td>전기 제외 · 일치</td></tr>
<tr><td>간접공사비</td><td class="num">{won(split_i)}</td>
  <td class="num">{won(sum_i)}</td>{diff_cell(split_i, sum_i)[0]}
  <td>그룹별 산출 합</td></tr>
<tr><td>하도급액(VAT)</td><td class="num">{won(split_dog)}</td>
  <td class="num">{won(dogup)}</td>{diff_cell(split_dog, dogup)[0]}
  <td>4차수 합 = 합산</td></tr>
<tr><td>4분할 「토목」직접</td><td class="num">{won(split_by_name['토목'][2])}</td>
  <td class="num">—</td><td class="num">—</td>
  <td>단독 토목(1~9공종) · 2차</td></tr>
<tr><td>하도급 「토목」직접(Ⅲ)</td><td class="num">{won(dt)}</td>
  <td class="num">{won(excl_d)}</td>{diff_cell(dt, excl_d)[0]}
  <td>토목+회전진입+지구단위(=조경 제외)</td></tr>
</tbody></table>

<p class="muted">
<strong>운용 유의</strong> — 4개 공정은 각각 착공·준공·정산 단위이다.
차수별 준공으로 보증·선급 부담을 분산하며, 2026년 실적을 2027년에 반영하는 구조이다.
하단 하도급 「토목」 집계(Ⅲ·Ⅳ)는 4분할의 토목+회전진입+지구단위(조경 제외)와 일치한다.
하도급은 <strong>개별하도급(업종별 개별계약)</strong>으로 시행한다(Ⅶ).
</p>

<h2>Ⅶ. 하도급 — 개별하도급 (업종별 개별계약)</h2>
<p class="muted">
하도급은 공동수급체(분담이행)가 아니라, <strong>공종(업종)별로 자격을 갖춘 수급인과 「개별 하도급계약」</strong>을 체결하는 방식으로 시행한다.
<strong>근일건설</strong>은 보유면허(관로·포장·구조물 등) 범위를,
<strong>삼아건설</strong>은 <strong>신규면허(토공)</strong> 범위의 공종을 각각 개별 계약으로 담당한다.
배분 근거는 <code>도급공사비 안분내역서(개별하도급·업종별 개별계약).xlsx</code>이다.
</p>

<table>
<thead><tr>
  <th>수급인</th><th>담당 업종(면허)</th>
  <th class="num">도급공사비</th><th class="num">비중</th>
</tr></thead>
<tbody>
<tr><td><strong>근일건설(주)</strong><br><span class="muted">보유면허</span></td>
  <td>상·하수도설비 · 포장 · 철근·콘크리트 · (건설폐기물)</td>
  <td class="num">{won(geun_dogub)}</td>
  <td class="num">{pct(geun_pct)}</td></tr>
<tr><td><strong>삼아건설</strong><br><span class="muted">신규면허</span></td>
  <td>토공(지반조성) · 토공부대</td>
  <td class="num">{won(sam_dogub)}</td>
  <td class="num">{pct(sam_pct)}</td></tr>
<tr class="total"><td colspan="2">합계(토목만 · 도급공사비 VAT포함)</td>
  <td class="num">{won(tot_dogub)}</td><td class="num">100%</td></tr>
</tbody></table>
<p class="muted">각 수급인은 <strong>자기 면허 범위의 공종만</strong> 개별 계약으로 수급한다(무자격 시공·명의대여 방지).
삼아건설의 담당 규모는 초기 시공능력(약 5.04억원, Ⅶ-2)을 상회하므로,
초기(2026년)에는 <strong>토공 실제 시공 가능 범위로 한정</strong>하고 잔여는 근일건설이 담당하는 방향으로 계약 단계에서 조정한다.</p>

<h3>Ⅶ-1. 근일건설 · 삼아건설 — 면허별 도급한도액</h3>
<p class="muted">출처: 근일건설 <code>09_공사지명원/근일건설_지명요약</code>(평가년도 2024년·2025년 지명원) · 삼아건설 초기 산정(Ⅶ-2 근거). 단위: 백만원.</p>
<table>
<thead><tr>
  <th>업체</th><th>구분</th><th>면허(업종)</th>
  <th class="num">시공능력평가액</th><th class="num">도급한도액</th>
</tr></thead>
<tbody>
<tr><td rowspan="7"><strong>근일건설(주)</strong><br><span class="muted">보유면허</span></td>
  <td rowspan="4">지반조성·포장</td><td>토공</td>
  <td class="num">24,539</td><td class="num">24,539</td></tr>
<tr><td>포장</td><td class="num">9,035</td><td class="num">9,035</td></tr>
<tr><td>보링·그라우팅·파일</td><td class="num">6,145</td><td class="num">6,145</td></tr>
<tr class="subtotal"><td>소계</td><td class="num">39,719</td><td class="num">39,719</td></tr>
<tr><td rowspan="3">기타</td><td>철근·콘크리트공사업</td>
  <td class="num">14,430</td><td class="num">14,430</td></tr>
<tr><td>상·하수도설비공사업</td><td class="num">15,983</td><td class="num">15,983</td></tr>
<tr><td>비계·구조물해체공사업</td><td class="num">4,996</td><td class="num">4,996</td></tr>
<tr class="total"><td rowspan="2"><strong>삼아건설</strong><br><span class="muted">신규면허</span></td>
  <td>신규(초기)</td><td>토공(지반조성) 한정</td>
  <td class="num">약 504</td><td class="num">약 504</td></tr>
<tr><td>2027년(증자 후)</td><td>토공 상향 + 상하수도 추가</td>
  <td class="num">상향</td><td class="num">상향</td></tr>
</tbody></table>
<p class="muted">근일건설은 토공(24,539) · 포장(9,035) · 철근콘크리트(14,430) · 상·하수도설비(15,983) 등 다수 면허를 보유한다.
삼아건설은 신규면허로서 초기 도급한도가 <strong>토공(지반조성) 약 504백만원</strong>에 한정되며, 산정 근거는 아래 Ⅶ-2와 같다.
<strong>2027년 2.5억원 증자·상하수도공사업 추가</strong> 시 한도가 상향된다.</p>

<h3>Ⅶ-2. 삼아건설 — 신규면허 · 초기 시공능력평가 근거</h3>
<p class="muted">
삼아건설은 자본금 1.8억원과 기술자 3명을 보유한 신규면허 업체이다.
전문건설(토공·포장 등)의 등록기준(자본금 1.5억원 이상, 기술인력 업종별 2명 이상)을 충족한다.
</p>
<table>
<thead><tr>
  <th>항목</th><th>내용</th>
</tr></thead>
<tbody>
<tr><td>주체</td><td><strong>삼아건설</strong> (신규면허)</td></tr>
<tr><td>계약 방식</td><td><strong>개별하도급(업종별 개별계약)</strong> — 토공(지반조성)·토공부대 공종 담당</td></tr>
<tr><td>취득 면허</td><td>토공사업 · 포장공사업(지반조성·포장공사업) 등 <strong>신규면허</strong>
  <br><span class="muted">※ 2027년 <strong>상하수도공사업</strong> 추가 취득 예정(7,000만원 증자 후)</span></td></tr>
<tr><td>자본금(실질자본금 전제)</td><td><strong>1.8억원</strong> (등록기준 1.5억원 이상 충족 · 초기 평가 시 실질자본금으로 적용)</td></tr>
<tr><td>기술자</td><td><strong>3명</strong> — 토목 특급 1명 · 중급 1명 · 초급 1명 (등록기준 업종별 2명 이상 충족)</td></tr>
<tr><td>초기 한도 범위</td><td><strong>지반조성(토공)으로 한정</strong>한다. 보유실적 반영 전 단계이며,
  포장 등 다른 주력분야는 실적·평가 반영 후 별도 산정한다.</td></tr>
<tr><td>2027년 계획</td><td><strong>7,000만원 증자</strong> 후 <strong>상하수도공사업</strong> 면허 추가 취득 예정.
  증자로 실질자본금이 <strong>1.8억원 → 2.5억원</strong>으로 증가하여 경영평가액·초기 한도가 상향되고,
  상하수도 면허 추가로 담당 공종 범위가 확대된다.</td></tr>
<tr><td>유의</td><td>4분할·차수별 준공과 연계하여 <strong>초기 도급한도 한계를 분담·분산</strong>하는 구조의 한 축이다.
  근일(보유면허·확정 한도)과 삼아(신규면허·토공 초기 한도)를 병행한다.</td></tr>
</tbody></table>

<p class="muted">
초기 시공능력평가액 산정(전문건설 · 신규등록 · 실적 반영 전)은 다음과 같다.
기본 산식은 <code>시공능력평가액 = 공사실적평가액 + 경영평가액 + 기술능력평가액 ± 신인도평가액</code>이다.
신규등록의 경우 당해 및 그다음 연도까지 <strong>경영평점은 1</strong>로 보며,
경영평가액은 실질자본금 × 경영평점 × 80/100으로 산정한다(건설산업기본법 시행규칙 전문공사 평가방법).
본건은 초기 한도를 <strong>지반조성(토공)</strong>에 한정하여 산정한다.
</p>

<p class="muted"><strong>기술능력평가액 산정 — 토목 특급 1 · 중급 1 · 초급 1</strong></p>
<p class="muted">
기술능력평가액 = 기술능력생산액 + 퇴직공제납입금 × 10 + 최근 3년간 기술개발투자액,
기술능력생산액 = 전년도 동종업계 기술인 1인당 평균생산액 × 보유 기술인 수(등급 가중) × 30/100.
등급 가중치는 초급 1 · 중급 1.15 · 고급 1.3 · 특급 1.5를 적용한다.
</p>
<table>
<thead><tr>
  <th>등급</th><th class="num">인원</th><th class="num">가중치</th><th class="num">가중 인원</th>
</tr></thead>
<tbody>
<tr><td>특급</td><td class="num">1</td><td class="num">1.5</td><td class="num">1.50</td></tr>
<tr><td>중급</td><td class="num">1</td><td class="num">1.15</td><td class="num">1.15</td></tr>
<tr><td>초급</td><td class="num">1</td><td class="num">1.0</td><td class="num">1.00</td></tr>
<tr class="total"><td>계</td><td class="num">3</td><td class="num">—</td><td class="num">3.65</td></tr>
</tbody></table>

<table>
<thead><tr>
  <th>구성</th><th>산정 전제</th><th class="num">금액(원)</th>
</tr></thead>
<tbody>
<tr><td>공사실적평가액</td><td>보유실적 없음(신규 · 실적 반영 전)</td>
  <td class="num">0</td></tr>
<tr><td>경영평가액</td><td>실질자본금 1.8억원 × 경영평점 1 × 80%</td>
  <td class="num">144,000,000</td></tr>
<tr><td>기술능력생산액(상한 전)</td><td>1인당 평균생산액 923,000,000원 × 3.65 × 30%</td>
  <td class="num">약 1,010,685,000</td></tr>
<tr><td>기술능력생산액(상한 적용)</td><td>상한 = max(실질자본금×2, 공사실적×50%) = max(3.6억, 0)</td>
  <td class="num">360,000,000</td></tr>
<tr><td>기술능력평가액</td><td>기술능력생산액 + 퇴직공제납입금×10 + 기술개발투자액(신규 ≈ 0)</td>
  <td class="num">약 360,000,000</td></tr>
<tr><td>신인도평가액</td><td>신규·실적 없음(가감 0 전제)</td>
  <td class="num">0</td></tr>
<tr class="total"><td>시공능력평가액(토공·초기)</td>
  <td>경영평가액 + 기술능력평가액</td>
  <td class="num">약 504,000,000</td></tr>
<tr class="subtotal"><td>도급한도액(토공·초기)</td>
  <td>전문건설: 시공능력평가액 ≈ 도급한도액</td>
  <td class="num">약 504,000,000</td></tr>
</tbody></table>
<p class="muted">
※ 기술능력생산액은 <strong>실질자본금의 2배(3.6억원)</strong>를 상한으로 하며, 공사실적이 없으므로 상한이 지배한다.
1인당 평균생산액 수치와 무관하게 <strong>기술능력생산액은 3.6억원으로 제한</strong>되고,
등급 가중치를 높여도 결과는 동일하다.
따라서 초기 시공능력평가액(토공)은 <strong>약 5.04억원</strong>(경영평가액 1.44억원 + 기술능력평가액 3.6억원)으로 산정된다.
<strong>2027년 2.5억원 증자</strong> 시 상한은 5.0억원, 경영평가액은 2.0억원으로 상향된다.
포장 등 다른 분야는 실적 반영 후 별도 산정한다.
</p>

<h3>Ⅶ-3. 개별하도급 배분 — 블록·공종·업종별 (도급공사비)</h3>
<p class="muted">토목만(회전진입·토목·지구단위)을 <strong>공종과 해당 면허(업종)</strong>로 구분하여, 자격을 갖춘 수급인과 <strong>개별 하도급계약</strong>으로 배정한다.
금액은 <strong>도급공사비(직접+간접+부가세)</strong> 기준이며, 각 블록 도급액을 공종별 직접공사비 비율로 안분(반올림)한 값이다(출처: <code>도급공사비 안분내역서(개별하도급·업종별 개별계약).xlsx</code>).</p>
<table>
<thead><tr>
  <th>블록(차수)</th><th>공종</th><th>해당 면허(업종)</th>
  <th class="num">도급공사비</th>
  <th class="num">근일건설</th><th class="num">삼아건설</th>
</tr></thead>
<tbody>
{alloc_rows}
<tr class="total"><td colspan="3">합계(토목만 · 도급공사비 VAT포함)</td>
  <td class="num">{won(tot_dogub)}</td>
  <td class="num">{won(geun_dogub)}</td>
  <td class="num">{won(sam_dogub)}</td></tr>
</tbody></table>
<p class="muted">각 공종은 해당 면허를 보유한 수급인에게 개별 계약으로 배정한다.
토공(지반조성)·토공부대는 삼아건설, 상·하수도설비·포장·철근·콘크리트·건설폐기물은 근일건설이 담당한다.
공종 금액은 상세내역 집계를 각 블록 도급액에 맞춰 정합(반올림)한 값으로, 세부 물량 배분에 따라 달라질 수 있다.</p>

<h3>Ⅶ-4. 업종(면허)별 개별계약 집계 · 수급인별 합계</h3>
<table>
<thead><tr>
  <th>업종(면허)</th><th class="num">도급공사비</th><th>담당 수급인</th>
</tr></thead>
<tbody>
{eupjong_rows}
<tr class="total"><td>합계(토목만 · VAT포함)</td>
  <td class="num">{won(tot_dogub)}</td><td>—</td></tr>
</tbody></table>
<table>
<thead><tr>
  <th>수급인</th><th class="num">도급공사비</th><th class="num">비중</th>
</tr></thead>
<tbody>
<tr><td><strong>근일건설(주)</strong></td>
  <td class="num">{won(geun_dogub)}</td><td class="num">{pct(geun_pct)}</td></tr>
<tr><td><strong>삼아건설</strong></td>
  <td class="num">{won(sam_dogub)}</td><td class="num">{pct(sam_pct)}</td></tr>
<tr class="total"><td>합계(토목만 · VAT포함)</td>
  <td class="num">{won(tot_dogub)}</td><td class="num">100%</td></tr>
</tbody></table>
<p class="muted">
※ 위 배분은 <code>도급공사비 안분내역서</code>의 업종별 담당 기준을 그대로 반영한 것이다.
삼아건설 담당 규모({won(sam_dogub)}원)는 초기 시공능력(약 5.04억원, Ⅶ-2)을 상회하므로,
초기(2026년)에는 <strong>토공 실제 시공 가능 범위로 한정</strong>하고 잔여는 근일건설이 담당하는 방향으로 계약 단계에서 조정한다.
<strong>2027년 증자·상하수도공사업 면허 추가</strong> 시 삼아 담당 범위가 확대된다.
</p>

<h3>Ⅶ-5. 법령 검토 — 개별하도급(업종별 개별계약)의 적법성</h3>
<p class="muted">개별하도급(업종별 개별계약)은 공종(업종)별로 자격 있는 수급인과 각각 계약하는 방식이므로, 아래 쟁점을 사전에 검토한다.</p>

<table>
<thead><tr>
  <th>구분</th><th>근거</th><th>개별하도급(업종별 개별계약) 검토</th>
</tr></thead>
<tbody>
<tr><td>하수급인 시공자격(면허)</td>
  <td>건설산업기본법 제25조 제2항 (제16조 시공자격을 갖춘 자에게 하도급)</td>
  <td><strong>부합</strong> — 공종(업종)별로 자격을 갖춘 수급인에게만 개별 계약하므로 자격 요건에 부합한다. 삼아는 토공 면허 범위만 수급한다.</td></tr>
<tr><td>무등록 시공·명의대여</td>
  <td>건설산업기본법 제21조(명의대여 금지) · 제41조·제96조(벌칙)</td>
  <td><strong>주의</strong> — 삼아가 실제 인력·장비·현장대리인으로 자기 계약분을 직접 시공해야 한다. 명의만 제공하면 위법(형사처벌 대상).</td></tr>
<tr><td>도급한도(시공능력)</td>
  <td>시공능력평가액(전문건설)</td>
  <td><strong>주의</strong> — 삼아 담당 규모가 초기 시공능력(약 5.04억원)을 상회하므로, 초기에는 토공 실제 시공 범위로 한정하고 잔여는 근일이 담당하도록 조정한다. (도급한도액 제도는 폐지되어 초과계약 자체가 곧 위법은 아니나 명의대여 정황이 될 수 있다.)</td></tr>
<tr><td>일괄하도급 금지</td>
  <td>건설산업기본법 제29조 제1항(단서: 계획·관리·조정 + 2인 이상 업종별 분할 하도급은 예외)</td>
  <td><strong>부합</strong> — 희상이 계획·관리·조정을 수행하고 2인 이상에게 업종별로 분할 하도급하면 단서 예외에 해당한다.</td></tr>
<tr><td>재하도급 금지</td>
  <td>건설산업기본법 제29조 제3항</td>
  <td><strong>주의</strong> — 각 수급인(근일·삼아)이 자기 계약분을 다시 재하도급하지 않아야 한다. 삼아→근일 실질 이전은 명의대여·재하도급에 해당.</td></tr>
<tr><td>하도급대금·계약</td>
  <td>하도급법 제3조(서면) · 제4조(부당대금 결정 금지) · 제13조(대금지급) · 제13조의2(대금지급보증) · 제3조의4(부당특약)</td>
  <td><strong>준수 필요</strong> — 개별 계약별 서면 발급·대금지급(60일)·지급보증 등 원사업자(희상) 의무를 준수한다.</td></tr>
</tbody></table>
<p class="muted">
<strong>정리.</strong> <strong>개별하도급(업종별 개별계약)</strong>은 공종(업종)별로 자격 있는 수급인과 각각 계약하는 방식으로,
건설산업기본법 제25조 제2항(시공자격)·제29조 제1항 단서(업종별 분할 하도급)에 부합하여 상대적으로 안전하다.
다만 ① 삼아건설이 <strong>자기 면허(토공) 범위만</strong> 수급하고 <strong>실제 시공</strong>할 것,
② 삼아 담당 규모를 <strong>초기 시공능력 범위로 한정</strong>할 것(잔여는 근일 담당),
③ 희상이 <strong>계획·관리·조정을 실질 수행</strong>하고 재하도급이 발생하지 않을 것,
④ 개별 계약별 <strong>서면·대금지급·지급보증</strong> 의무를 준수할 것이 전제이다.
<strong>실제 계약 체결 전 건설 분야 법률 자문으로 최종 확인할 것을 권고한다.</strong>
</p>

<p class="muted">끝.</p>
</body></html>
"""
    OUT.write_text(html, encoding="utf-8")
    print("저장:", OUT)
    print(f"희상토목 직접 {won(hs_tom_d)} · 하단토목 {won(dt)} · 차 {won(hs_tom_d-dt)}")
    print(f"4분할 직접 {won(split_d)} · 하도급합산 {won(sum_d)} · 도급가(희상) {won(dogup_hs)}")
    print(f"개별하도급 근일 {won(geun_dogub)} ({pct(geun_pct)}) · "
          f"삼아 {won(sam_dogub)} ({pct(sam_pct)}) · 계 {won(tot_dogub)}")


if __name__ == "__main__":
    main()
