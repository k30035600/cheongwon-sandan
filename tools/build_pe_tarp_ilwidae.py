#!/usr/bin/env python3
"""PE 천막지 깔기(설치·해체 포함) — 10㎡ 기준 일위대가 산출.

토목 내역서 미매칭 「가). PE천막지」(01 토목 29행, 1,082㎡)용.
출력: 05_내역서/내역서작업/PE천막지_일위대가_10㎡기준.xlsx
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
NOIM_CSV = BASE / "일위대가DB" / "시중노임_2026.csv"
OUT = WORK / "PE천막지_일위대가_10㎡기준.xlsx"
ILWAE_OUT = BASE / "미매칭_일위대가산출.xlsx"

# 10㎡ 기준 산출 파라미터 (실무 확인·조정 가능)
BASE_M2 = 10
MAT_UNIT = 1000  # PE 천막지 ㎡당 자재단가(예시·조달·견적 확인)
MAT_WASTE = 0.05  # 자재 5% 할증
LABOR_MD = 0.04  # 10㎡당 보통인부 0.04인 (100㎡당 0.4인)
CONS_RATE = 0.05  # 주자재비 5% 소모품

# 본건 내역
PROJECT = {
    "file": "01 토목",
    "row": 29,
    "section": "1. 토     공",
    "name": "가). PE천막지",
    "spec": "표준형(200~250g급, 단기 보양·법면)",
    "unit": "㎡",
    "qty": 1082,
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FILL = PatternFill("solid", fgColor="FFF2CC")
MONEY = "#,##0"


def load_noim() -> int:
    with NOIM_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r.get("직종명") == "보통인부":
                return int(r["2026.1.1"])
    return 172068


def calc(noim: int) -> dict:
    mat_qty = round(BASE_M2 * (1 + MAT_WASTE), 2)
    mat_amt = round(mat_qty * MAT_UNIT)
    cons_amt = round(mat_amt * CONS_RATE)
    lab_amt = round(LABOR_MD * noim)
    sub_mat = mat_amt + cons_amt
    sub_lab = lab_amt
    sub_tot = sub_mat + sub_lab
    per_m2_mat = round(sub_mat / BASE_M2)
    per_m2_lab = round(sub_lab / BASE_M2)
    per_m2_tot = round(sub_tot / BASE_M2)
    return {
        "mat_qty": mat_qty,
        "mat_unit": MAT_UNIT,
        "mat_amt": mat_amt,
        "cons_amt": cons_amt,
        "lab_md": LABOR_MD,
        "noim": noim,
        "lab_amt": lab_amt,
        "sub_mat": sub_mat,
        "sub_lab": sub_lab,
        "sub_tot": sub_tot,
        "per_m2_mat": per_m2_mat,
        "per_m2_lab": per_m2_lab,
        "per_m2_tot": per_m2_tot,
        "project_amt": per_m2_tot * PROJECT["qty"],
    }


def write_xlsx(c: dict, noim: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "일위대가_10㎡"
    ws.append(["PE 천막지 깔기(설치 및 해체 포함) — 일위대가표"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"기준: {BASE_M2}㎡ 단위 산출 → 1㎡당 단가 = 소계 ÷ {BASE_M2}"])
    ws.append([f"시중노임 보통인부(2026. 1. 1.): {noim:,}원/인"])
    ws.append([])

    hdr = ["품명", "규격", "단위", "수량", "재료비 단가", "노무비 단가", "경비 단가",
           "재료비 금액", "노무비 금액", "경비 금액", "합계"]
    ws.append(hdr)
    for col in range(1, len(hdr) + 1):
        cell = ws.cell(5, col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = [
        ["[자재] PE 천막지", "표준형", "㎡", c["mat_qty"], c["mat_unit"], "", "",
         c["mat_amt"], "", "", c["mat_amt"]],
        ["[노무] 보통인부", "토목노임", "인", c["lab_md"], "", c["noim"], "",
         "", c["lab_amt"], "", c["lab_amt"]],
        ["[자재] 소모품비", "결속선·말뚝 등", "%", CONS_RATE * 100, "", "", "",
         c["cons_amt"], "", "", c["cons_amt"]],
        [f"[소계] {BASE_M2}㎡당", "", "", "", "", "", "",
         c["sub_mat"], c["sub_lab"], 0, c["sub_tot"]],
        ["[최종] 1㎡당 단가(소계÷10)", "", "㎡", 1,
         c["per_m2_mat"], c["per_m2_lab"], 0,
         c["per_m2_mat"], c["per_m2_lab"], 0, c["per_m2_tot"]],
    ]
    for r in rows:
        ws.append(r)
        if r[0].startswith("[소계]") or r[0].startswith("[최종]"):
            for col in range(1, len(hdr) + 1):
                ws.cell(ws.max_row, col).fill = SUB_FILL
                ws.cell(ws.max_row, col).font = Font(bold=True)

    for row_idx in range(6, ws.max_row + 1):
        for col in (5, 6, 7, 8, 9, 10, 11):
            v = ws.cell(row_idx, col).value
            if isinstance(v, (int, float)) and v:
                ws.cell(row_idx, col).number_format = MONEY

    ws.column_dimensions["A"].width = 28
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 12

    # 본건 반영
    pj = wb.create_sheet("본건_내역반영")
    pj.append(["청원지구 01 토목 — PE천막지 내역 반영"])
    pj["A1"].font = Font(bold=True, size=13)
    pj.append([])
    pj.append(["항목", "값"])
    for k, v in [
        ("내역서", PROJECT["file"]),
        ("행", PROJECT["row"]),
        ("공종", PROJECT["section"]),
        ("품명", PROJECT["name"]),
        ("규격", PROJECT["spec"]),
        ("단위", PROJECT["unit"]),
        ("수량(정미)", PROJECT["qty"]),
        ("1㎡당 확정단가", c["per_m2_tot"]),
        ("직접공사비(재+노)", c["project_amt"]),
    ]:
        pj.append([k, v])
    pj.append([])
    pj.append(["체크", "내용"])
    checks = [
        ("수량", "도면 정미면적만 기입 — 할증은 일위대가 자재량(10.5㎡/10㎡)에 반영"),
        ("단가", f"1㎡당 {c['per_m2_tot']:,}원 (재료 {c['per_m2_mat']:,} + 노무 {c['per_m2_lab']:,})"),
        ("품셈", "2026 표준품셈 제2장 가설공사·법면보양 유사 조항 준용(보통인부 0.04인/10㎡)"),
        ("참고", "조경표준일위2024 「비탈면보양·천막 설치해체」6,056원/㎡ — 범위·포함범위 상이 시 별도 검토"),
        ("현장", "장기보양·톤백 고정 등 추가 시 마대·모래 채우기 품목 별도 가산"),
    ]
    for chk, txt in checks:
        pj.append([chk, txt])
    pj.column_dimensions["A"].width = 14
    pj.column_dimensions["B"].width = 72
    for row_idx in range(4, pj.max_row + 1):
        if isinstance(pj.cell(row_idx, 2).value, int) and pj.cell(row_idx, 1).value in (
            "1㎡당 확정단가", "직접공사비(재+노)", "수량(정미)"
        ):
            pj.cell(row_idx, 2).number_format = MONEY

    # 산출근거
    bg = wb.create_sheet("산출근거")
    bg.append(["산출 근거 (단계별)"])
    bg["A1"].font = Font(bold=True, size=13)
    lines = [
        [],
        ["1. 자재비", f"{BASE_M2}㎡ × (1+{int(MAT_WASTE*100)}% 할증) = {c['mat_qty']}㎡ × {c['mat_unit']:,}원 = {c['mat_amt']:,}원"],
        ["2. 노무비", f"{c['lab_md']}인 × {c['noim']:,}원 = {c['lab_amt']:,}원 (10㎡당)"],
        ["3. 소모품", f"주자재 {c['mat_amt']:,}원 × {int(CONS_RATE*100)}% = {c['cons_amt']:,}원"],
        ["4. 1㎡ 환산", f"{c['sub_tot']:,}원 ÷ {BASE_M2} = {c['per_m2_tot']:,}원/㎡"],
        [],
        ["※ 자재단가", f"{MAT_UNIT:,}원/㎡는 예시값 — 조달청 시설공통자재·견적서로 교체 후 재산출"],
        ["※ 예시 대비", "보통인부 165,000원 가정 시 1㎡당 1,763원 / 2026.1.1 노임 적용 시 상단 표 기준"],
    ]
    for line in lines:
        bg.append(line)
    bg.column_dimensions["A"].width = 14
    bg.column_dimensions["B"].width = 78

    WORK.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"저장: {OUT}")
    print(f"1㎡당 {c['per_m2_tot']:,}원 · 본건 {PROJECT['qty']:,}㎡ = {c['project_amt']:,}원")


def patch_ilwidae(c: dict) -> None:
    if not ILWAE_OUT.exists():
        return
    wb = load_workbook(ILWAE_OUT)
    ws = wb["일위대가산출"]
    hdr = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(hdr, 1) if h}
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, col["품명"]).value
        if not name or "PE천막" not in str(name):
            continue
        ws.cell(row_idx, col["경로"]).value = "J.품셈 PE천막(10㎡)"
        ws.cell(row_idx, col["DB_품명"]).value = "PE 천막지 깔기"
        ws.cell(row_idx, col["DB_규격"]).value = f"{BASE_M2}㎡기준→1㎡환산"
        ws.cell(row_idx, col["DB_단위"]).value = "㎡"
        ws.cell(row_idx, col["재료단가"]).value = c["per_m2_mat"]
        ws.cell(row_idx, col["노무비"]).value = c["per_m2_lab"]
        ws.cell(row_idx, col["경비"]).value = 0
        ws.cell(row_idx, col["합계단가"]).value = c["per_m2_tot"]
        ws.cell(row_idx, col["제시단가"]).value = c["per_m2_tot"]
        ws.cell(row_idx, col["제시금액"]).value = c["project_amt"]
        ws.cell(row_idx, col["확정단가(입력)"]).value = c["per_m2_tot"]
        ws.cell(row_idx, col["확정금액"]).value = c["project_amt"]
        basis = (
            f"10㎡기준 품셈조합 — 자재 {c['mat_qty']}㎡×{c['mat_unit']:,} "
            f"+ 소모품 {c['cons_amt']:,} + 보통인부 {c['lab_md']}인×{c['noim']:,} "
            f"→ 1㎡ {c['per_m2_tot']:,}원"
        )
        ws.cell(row_idx, col["표준품셈·산출근거"]).value = basis
        for ci in (col["재료단가"], col["노무비"], col["합계단가"], col["제시단가"],
                   col["제시금액"], col["확정단가(입력)"], col["확정금액"]):
            ws.cell(row_idx, ci).number_format = MONEY
        break
    try:
        wb.save(ILWAE_OUT)
        print(f"반영: {ILWAE_OUT.name} PE천막지 행")
    except PermissionError:
        alt = ILWAE_OUT.with_name("미매칭_일위대가산출_PE반영.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")


def main() -> None:
    noim = load_noim()
    c = calc(noim)
    write_xlsx(c, noim)
    patch_ilwidae(c)


if __name__ == "__main__":
    main()
