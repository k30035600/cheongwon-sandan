#!/usr/bin/env python3
"""검토_전체 — 1. 토공 항목 품셈·일위대가 산출.

입력: 05_내역서/검토_전체.xlsx
출력: 05_내역서/검토_토공_일위대가산출.xlsx
      05_내역서/내역서작업/검토_토공_품셈산출.xlsx (유형별 산출근거)
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
REVIEW_ALL = BASE / "검토_전체.xlsx"
OUT = BASE / "검토_토공_일위대가산출.xlsx"
POOMSEM_OUT = WORK / "검토_토공_품셈산출.xlsx"
MARKET_CSV = BASE / "일위대가DB" / "표준일위대가_2026" / "표준시장단가_2026.csv"
MASTER_CSV = BASE / "일위대가DB" / "_master_일위대가.csv"

IMOK_M3 = 20_000 * 0.40  # apply_standard_prices 임목운반 위탁

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ROUTE_FILLS = {
    "① 표준단가 확정": PatternFill("solid", fgColor="E2EFDA"),
    "② 단위환산": PatternFill("solid", fgColor="DDEBF7"),
    "③ 품셈·재산출": PatternFill("solid", fgColor="FCE4D6"),
    "④ 1순위+검증": PatternFill("solid", fgColor="FFF2CC"),
}
MONEY = "#,##0"

FILE_MAP = {
    "01 토목": "01 토목",
    "01 토목·조경": "01 토목",  # 분리 이전 라벨 호환
    "04 진입도로": "04 진입도로",
    "05 회전교차로": "05 회전교차로",
}
TOK01_LABELS = ("01 토목", "01 토목·조경")


def load_market() -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not MARKET_CSV.exists():
        return out
    with MARKET_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            out[r["코드"]] = {
                "code": r["코드"],
                "name": r["품명"],
                "spec": r["규격"],
                "unit": r["단위"],
                "mat": float(r["재료비"]),
                "lab": float(r["노무비"]),
                "exp": float(r["경비"]),
                "tot": float(r["합계"]),
            }
    return out


def load_master_codes(codes: set[str]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    if not MASTER_CSV.exists():
        return out
    with MASTER_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            c = r.get("코드", "")
            if c not in codes:
                continue
            out[c] = {
                "code": c,
                "name": r["품명"],
                "spec": r["규격"],
                "unit": r["단위"],
                "mat": float(r["재료비단가"] or 0),
                "lab": float(r["노무비단가"] or 0),
                "exp": float(r["경비단가"] or 0),
                "tot": float(r["합계단가"] or 0),
            }
    return out


def clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").replace("_x000D_", "").strip()


def is_togong_section(sec: str) -> bool:
    s = clean(sec).replace(" ", "")
    return "토" in s and "공" in s


def heokkakgi_code(qty: float) -> str:
    if qty >= 10_000:
        return "CDD100200000"
    if qty >= 1_000:
        return "CDD100100000"
    return "CDD100000000"


def load_review_togong() -> list[dict]:
    wb = load_workbook(REVIEW_ALL, read_only=True, data_only=True)
    ws = wb["통합"]
    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):
        file_label = clean(ws.cell(r, 1).value)
        section = clean(ws.cell(r, 3).value)
        if not file_label or not is_togong_section(section):
            continue
        rows.append({
            "file_raw": file_label,
            "file": FILE_MAP.get(file_label, file_label),
            "row": ws.cell(r, 2).value,
            "section": section,
            "name": clean(ws.cell(r, 4).value),
            "spec": clean(ws.cell(r, 5).value),
            "unit": clean(ws.cell(r, 6).value),
            "qty": float(ws.cell(r, 7).value or 0),
            "score": ws.cell(r, 8).value,
            "cur_code": clean(ws.cell(r, 10).value),
            "cur_match": clean(ws.cell(r, 11).value),
            "cur_spec": clean(ws.cell(r, 12).value),
            "cur_tot": float(ws.cell(r, 13).value or 0),
            "cur_amt": float(ws.cell(r, 14).value or 0),
        })
    wb.close()
    return rows


def price_record(item: dict, market: dict[str, dict], master: dict[str, dict]) -> dict:
    name = item["name"]
    spec = item["spec"]
    qty = item["qty"]
    blob = name + spec

    def from_code(code: str, route: str, basis: str, *, auto: bool = True) -> dict:
        p = market.get(code) or master.get(code)
        if not p:
            return {
                "route": "③ 품셈·재산출",
                "code": code,
                "db_name": "",
                "db_spec": "",
                "db_unit": item["unit"],
                "mat": 0, "lab": 0, "exp": 0, "tot": 0,
                "basis": f"코드 {code} DB 미존재 — 수동 확인",
                "auto": False,
            }
        return {
            "route": route,
            "code": code,
            "db_name": p["name"],
            "db_spec": p["spec"],
            "db_unit": p["unit"],
            "mat": p["mat"],
            "lab": p["lab"],
            "exp": p["exp"],
            "tot": p["tot"],
            "basis": basis,
            "auto": auto,
        }

    # 벌목
    if "벌목" in name and "벌개" not in name:
        p = from_code(
            "CDA210100100", "① 표준단가 확정",
            "표준시장단가2026 CDA210100100 벌목 5m미만 · 2026 표준품셈 제3장 토공(벌개제근)",
        )
        if item["cur_code"] == "CDA210100100":
            p["route"] = "① 표준단가 확정"
        return p

    # 벌개제근
    if "벌개" in name or "제근" in name:
        return from_code(
            "CUA30200200S", "① 표준단가 확정",
            "하천설계실무요령 CUA30200200S · 표준품셈 벌개제근(뿌리뽑기)",
        )

    # 임목 운반
    if "임목" in name and "운반" in name:
        m = re.search(r"L\s*=\s*(\d+)", spec, re.I)
        dist = int(m.group(1)) if m else 0
        return {
            "route": "③ 품셈·재산출",
            "code": "IMOK-UNBAN",
            "db_name": "임목운반(위탁·파쇄장)",
            "db_spec": f"L={dist}m → 위탁 단가(품셈 外)",
            "db_unit": "㎥",
            "mat": 0,
            "lab": 0,
            "exp": IMOK_M3,
            "tot": IMOK_M3,
            "basis": (
                f"오매칭 CVB40103000S(12.5km) 제외 · "
                f"임목파쇄_단가기준 시나리오B 25톤×0.40㎥/ton=8,000원/㎥ · "
                f"거리 {dist}m(단거리) — 위탁 견적 우선"
            ),
            "auto": True,
        }

    # 토사 — 파일별 구분
    if "토사" in name and "치환" not in name:
        if item["file_raw"] in ("04 진입도로", "05 회전교차로"):
            return from_code(
                "CUA31704000S", "④ 1순위+검증",
                "하천 CUA31704000S 치환토(되메우기·교체) — 현재 후보 유지·현장 흙질 확인",
            )
        # 01 토목: 기존 DB 후보(소규모) 우선 — 수량 대규모여도 내역·매칭과 일치
        if item["file_raw"] in TOK01_LABELS and item["cur_code"].startswith("CDD100"):
            code = item["cur_code"]
        else:
            code = heokkakgi_code(qty)
        scale = {"CDD100000000": "소", "CDD100100000": "중", "CDD100200000": "대"}[code]
        note = ""
        if code == "CDD100000000" and qty >= 1_000:
            note = f" · 수량 {qty:,.0f}㎥ → 중·대규모(1,798~1,345원) 검토 가능"
        return from_code(
            code, "① 표준단가 확정",
            f"표준시장단가 {code} 흙깎기/보통토사 {scale}규모{note} · 품셈 제3장",
        )

    # 규준틀
    if "비탈규준" in blob:
        return from_code("CUA31301000S", "① 표준단가 확정", "하천 CUA31301000S · 품셈 목재규준틀(비탈)")
    if "수평규준" in blob:
        return from_code("CUA31302000S", "① 표준단가 확정", "하천 CUA31302000S · 품셈 목재규준틀(수평)")

    # 수목 이식
    if "이식" in name or ("나무" in name and "소나무" in spec):
        tot = 364_389 + 271_386  # 교목굴취 B20~29 + 교목식재 B21~29 (교육청2026)
        return {
            "route": "③ 품셈·재산출",
            "code": "5045-TRANSPLANT",
            "db_name": "교목굴취+식재(기계)",
            "db_spec": "B=29cm(21~29) · 약제·수목재료 별도",
            "db_unit": "주",
            "mat": 49_903,
            "lab": 461_392,
            "exp": 124_480,
            "tot": tot,
            "basis": (
                "오매칭 CSA11000001S(굴취 H2.0m 인력 18,396원) 제외 · "
                "충남교육청2026 교목굴취 기계 20~29cm 364,389 + "
                "교목식재 기계 18~24(21~29) 271,386 = 635,775원/주"
            ),
            "auto": True,
        }

    # 폐기물상차
    if "폐기물" in name and "상차" in name:
        p = from_code(
            "CDS101201500", "③ 품셈·재산출",
            "표준시장단가 CDS101201500 흙운반 직상차(덤프15t) — "
            "굴삭기0.7㎥ 상차 공정 대용 · 운반·처리비 별도",
        )
        p["db_name"] = "폐기물상차(굴삭0.7㎥)"
        p["db_spec"] = "직상차 15t덤프 · L=0(현장적재)"
        if "흙쌓기" in item["cur_match"] or "가물막이" in item["cur_match"]:
            p["basis"] = (
                f"오매칭 {item['cur_code']}({item['cur_match'][:20]}) 제외 · " + p["basis"]
            )
        return p

    # 가드레일 철거
    if "가드레일" in name and "철거" in name:
        inst = master.get("CTH80401010S") or market.get("CTH80401010S")
        if not inst:
            inst = {"mat": 371, "lab": 4482, "exp": 439, "tot": 5292,
                    "name": "가드레일", "spec": "지주간격 4m, 3W", "unit": "m"}
        tot = round(inst["tot"] * 0.30)
        mat = round(inst["mat"] * 0.30)
        lab = round(inst["lab"] * 0.30)
        exp = tot - mat - lab
        return {
            "route": "③ 품셈·재산출",
            "code": "CTH804-DEMO30",
            "db_name": "가드레일 철거",
            "db_spec": "설치품 CTH80401010S × 30%",
            "db_unit": "m",
            "mat": mat, "lab": lab, "exp": exp, "tot": tot,
            "basis": (
                f"내역 규격「설치품의 30%」· CTH80401010S 설치 {inst['tot']:,.0f}원/m × 30% "
                f"= {tot:,}원/m · 설치 단가 대비 철거 관행"
            ),
            "auto": True,
        }

    return {
        "route": "③ 품셈·재산출",
        "code": item["cur_code"],
        "db_name": item["cur_match"],
        "db_spec": item["cur_spec"],
        "db_unit": item["unit"],
        "mat": 0, "lab": 0, "exp": 0,
        "tot": item["cur_tot"],
        "basis": "유형 미분류 — 수동 검토",
        "auto": False,
    }


def write_ilwidae_workbook(items: list[dict], priced: list[dict]) -> None:
    wb = Workbook()
    route_counts: dict[str, int] = {}
    for p in priced:
        route_counts[p["route"]] = route_counts.get(p["route"], 0) + 1

    sm = wb.active
    sm.title = "경로별요약"
    sm.append(["검토_전체 — 1. 토공 품셈·일위대가 산출"])
    sm.append([])
    sm.append(["경로", "건수", "처리"])
    desc = {
        "① 표준단가 확정": "표준시장단가·하천DB — 현재 매칭과 일치 또는 규모 보정",
        "② 단위환산": "단위 환산 후 적용",
        "③ 품셈·재산출": "오매칭·거리·규격 불일치 — 품셈·DB 재조합",
        "④ 1순위+검증": "현 후보 유지·현장 확인",
    }
    for route in ("① 표준단가 확정", "② 단위환산", "③ 품셈·재산출", "④ 1순위+검증"):
        if route_counts.get(route):
            sm.append([route, route_counts[route], desc[route]])
    sm.append([])
    sm.append(["합계", len(priced), ""])
    sm.append(["자동 확정", sum(1 for p in priced if p["auto"]), "확정단가 열 기입됨"])
    sm["A1"].font = Font(bold=True, size=13)

    ws = wb.create_sheet("일위대가산출")
    headers = [
        "경로", "파일", "행", "공종", "품명", "규격", "단위", "수량", "매칭점수",
        "현재_코드", "현재_매칭품명", "현재_합계단가",
        "DB_코드", "DB_품명", "DB_규격", "DB_단위",
        "재료단가", "노무단가", "경비단가", "합계단가", "합계금액",
        "확정단가(입력)", "확정금액", "표준품셈·산출근거",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for item, p in zip(items, priced):
        amt = round(p["tot"] * item["qty"]) if p["tot"] and item["qty"] else None
        row = [
            p["route"], item["file"], item["row"], item["section"],
            item["name"], item["spec"], item["unit"], item["qty"], item["score"],
            item["cur_code"], item["cur_match"], item["cur_tot"],
            p["code"], p["db_name"], p["db_spec"], p["db_unit"],
            p["mat"] or None, p["lab"] or None, p["exp"] or None, p["tot"] or None, amt,
            p["tot"] if p["auto"] else "",
            amt if p["auto"] else "",
            p["basis"],
        ]
        ws.append(row)
        ridx = ws.max_row
        fill = ROUTE_FILLS.get(p["route"], ROUTE_FILLS["③ 품셈·재산출"])
        for c in range(1, len(headers) + 1):
            ws.cell(ridx, c).fill = fill
            ws.cell(ridx, c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(17, 24):
            if isinstance(ws.cell(ridx, c).value, (int, float)):
                ws.cell(ridx, c).number_format = MONEY

    ws.freeze_panes = "A2"
    widths = [14, 11, 5, 12, 22, 24, 5, 8, 7, 12, 18, 10,
              12, 18, 20, 5, 9, 9, 9, 10, 12, 11, 12, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    info = wb.create_sheet("안내")
    for line in [
        ["검토_전체 — 1. 토공 17건"],
        [],
        ["색상", "연두=① 확정 / 노랑=④ 검증 / 주황=③ 재산출"],
        ["반영", "python -X utf8 tools/apply_confirmed_prices.py"],
        ["품셈 상세", "내역서작업/검토_토공_품셈산출.xlsx"],
        ["주의", "수목이식·임목운반·폐기물상차는 현장 견적·발주처 확인 권장"],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=14)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(f"저장: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")


def write_poomsem_detail(items: list[dict], priced: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "유형별요약"
    ws.append(["검토_토공 — 품셈·일위대가 산출근거"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["유형", "건수", "단가(대표)", "근거"])

    type_groups: dict[str, list[tuple[dict, dict]]] = {}
    for item, p in zip(items, priced):
        key = p["code"].split("-")[0]
        type_groups.setdefault(key, []).append((item, p))

    for key, group in type_groups.items():
        p0 = group[0][1]
        ws.append([p0["db_name"][:20], len(group), p0["tot"], p0["basis"][:60]])

    # 상세 시트 — 재산출 4유형
    templates = [
        ("임목운반", lambda n: "임목" in n and "운반" in n),
        ("수목이식", lambda n: "이식" in n or ("나무" in n and "소나무" in n)),
        ("폐기물상차", lambda n: "폐기물" in n),
        ("가드레일철거", lambda n: "가드레일" in n and "철거" in n),
    ]
    for title, pred in templates:
        matched = [(it, pr) for it, pr in zip(items, priced) if pred(it["name"])]
        if not matched:
            continue
        sh = wb.create_sheet(title)
        sh.append([f"{title} — 품셈·일위대가"])
        sh["A1"].font = Font(bold=True, size=12)
        sh.append([])
        sh.append(["파일", "행", "품명", "규격", "수량", "단위", "합계단가", "합계금액", "산출근거"])
        for c in range(1, 10):
            sh.cell(3, c).font = Font(bold=True)
            sh.cell(3, c).fill = HEADER_FILL
        for it, pr in matched:
            amt = round(pr["tot"] * it["qty"])
            sh.append([
                it["file"], it["row"], it["name"], it["spec"],
                it["qty"], it["unit"], pr["tot"], amt, pr["basis"],
            ])
        for r in range(4, sh.max_row + 1):
            for c in (7, 8):
                if isinstance(sh.cell(r, c).value, (int, float)):
                    sh.cell(r, c).number_format = MONEY
        sh.column_dimensions["A"].width = 12
        sh.column_dimensions["I"].width = 70

    WORK.mkdir(parents=True, exist_ok=True)
    wb.save(POOMSEM_OUT)
    print(f"저장: {POOMSEM_OUT}")


def main() -> None:
    if not REVIEW_ALL.exists():
        print(f"없음: {REVIEW_ALL}")
        sys.exit(1)

    market = load_market()
    codes = {
        "CDA210100100", "CDD100000000", "CDD100100000", "CDD100200000",
        "CUA30200200S", "CUA31301000S", "CUA31302000S", "CUA31704000S",
        "CTH80401010S", "CDS101201500",
    }
    master = load_master_codes(codes)
    print(f"표준시장단가 {len(market):,} · 마스터 {len(master)}건")

    items = load_review_togong()
    print(f"토공 검토 {len(items)}건")
    priced = [price_record(it, market, master) for it in items]

    for route in sorted({p["route"] for p in priced}):
        n = sum(1 for p in priced if p["route"] == route)
        print(f"  {route}: {n}건")

    write_ilwidae_workbook(items, priced)
    write_poomsem_detail(items, priced)


if __name__ == "__main__":
    main()
