#!/usr/bin/env python3
"""토공 흙운반·반출 물량 집계 및 25톤 덤프 적재량 추출."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

WORK = Path(__file__).resolve().parents[1] / "05_내역서" / "내역서작업"

FILES = [
    ("05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx"),
    ("01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx"),
    ("06 지구단위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx"),
]


def haul_cat(name: str, spec: str) -> str:
    t = ((name or "") + (spec or "")).replace(" ", "")
    if "야적" in t or "외부" in t or "반출" in t:
        return "장외반출(야적·외부)"
    if "지구내" in t or ("유용" in t and "잔토" not in t):
        return "장내유용"
    if "구조물" in t or "잔토" in t:
        return "구조물잔토"
    if "무대" in t:
        return "무대(1·2차)"
    if "도저" in t or "불도" in t:
        return "도저운반(무대)"
    if "임목" in t:
        return "임목운반"
    if "운반" in t:
        return "기타운반"
    return "기타"


def dump_ton(spec: str) -> int | None:
    t = (spec or "").replace(" ", "")
    m = re.search(r"덤프(\d+)(?:톤|t|T)?", t, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)톤", t)
    if m and "덤프" in t:
        return int(m.group(1))
    return None


def load_haul_rows() -> list[dict]:
    rows: list[dict] = []
    for label, fn in FILES:
        wb = load_workbook(WORK / fn, read_only=True, data_only=True)
        ws = wb["통합내역"]
        data = list(ws.iter_rows(values_only=True))
        wb.close()
        hdr = [str(c or "").strip() for c in data[0]]
        idx = {h: i for i, h in enumerate(hdr)}
        for row in data[1:]:
            sec = str(row[idx["공종"]] or "").strip()
            if re.match(r"^2\.", sec.replace(" ", "")):
                break
            if not re.match(r"^1\.\s*토", sec.replace(" ", "")):
                continue
            name = str(row[idx["공종명"]] or "")
            spec = str(row[idx.get("규격", 4)] or "")
            unit = str(row[idx["단위"]] or "").replace(" ", "")
            if unit not in ("m3", "M3", "㎥"):
                continue
            try:
                q = float(row[idx["수량"]])
            except (TypeError, ValueError):
                continue
            if q <= 0:
                continue
            t = (name + spec).replace(" ", "")
            if not any(x in t for x in ("운반", "반출", "유용", "무대", "야적")):
                continue
            rows.append(
                {
                    "file": label,
                    "cat": haul_cat(name, spec),
                    "name": name.strip(),
                    "spec": spec.strip(),
                    "qty": q,
                    "dump_t": dump_ton(spec),
                }
            )
    return rows


def load_01_tok_cut_fill_haul() -> dict:
    """01 토목 — 토공_물량비교표와 동일 분류(상세)."""
    wb = load_workbook(WORK / "토목" / "01_화성 청원지구 토목_표준단가산출.xlsx", read_only=True, data_only=True)
    ws = wb["통합내역"]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c or "").strip() for c in data[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    cut = fill = 0.0
    haul_detail: dict[str, float] = defaultdict(float)
    for row in data[1:]:
        sec = str(row[idx["공종"]] or "").strip()
        if re.match(r"^2\.", sec.replace(" ", "")):
            break
        if not re.match(r"^1\.\s*토", sec.replace(" ", "")):
            continue
        name = str(row[idx["공종명"]] or "")
        spec = str(row[idx.get("규격", 4)] or "")
        unit = str(row[idx["단위"]] or "").replace(" ", "")
        try:
            q = float(row[idx["수량"]])
        except (TypeError, ValueError):
            continue
        if q <= 0 or unit not in ("m3", "M3", "㎥"):
            continue
        t = (name + spec).replace(" ", "")
        if any(x in t for x in ("운반", "반출", "유용", "무대", "야적")):
            if "외부" in t or "야적" in t or "반출" in t:
                haul_detail["외부반출"] += q
            elif "지구내" in t or "유용" in t:
                haul_detail["지구내 유용"] += q
            elif "도저" in t or ("불도" in t and "운반" in t):
                haul_detail["도저 운반(무대)"] += q
            elif "구조물" in t or "잔토" in t:
                haul_detail["구조물 잔토"] += q
            elif "무대" in t:
                haul_detail["무대(1·2차)"] += q
            else:
                haul_detail["기타 운반"] += q
        elif any(x in t for x in ("흙깎", "흙깍", "절토", "터파", "땅깍", "발파", "암")) and "운반" not in t:
            if any(x in t for x in ("흙쌓", "성토", "되메")):
                fill += q
            else:
                cut += q
        elif any(x in t for x in ("흙쌓", "성토", "되메")):
            fill += q
    return {"cut": cut, "fill": fill, "haul": dict(haul_detail)}


def subcat(name: str, spec: str) -> str:
    t = ((name or "") + (spec or "")).replace(" ", "")
    if "야적" in t:
        return "장외(야적장)"
    if "진입도로" in t:
        return "장외(진입도로)"
    if "지구내" in t and "구조물" not in t:
        return "장내(지구내유용)"
    if "구조물" in t and "유용" in t:
        return "장내(구조물유용)"
    if "지하저류" in t:
        return "장내(지하저류조)"
    if "무대" in t:
        return "무대(1·2차)"
    if "도저" in t or ("불도" in t and "L=" in (spec or "")):
        return "도저운반(무대)"
    if "임목" in t:
        return "임목운반"
    return "기타운반"


def main() -> None:
    haul_rows = load_haul_rows()
    by_cat: dict[str, float] = defaultdict(float)
    by_file: dict[str, float] = defaultdict(float)
    by_dump: dict[str, float] = defaultdict(float)

    for h in haul_rows:
        by_cat[h["cat"]] += h["qty"]
        by_file[h["file"]] += h["qty"]
        key = f"{h['dump_t']}톤" if h["dump_t"] else "덤프미표기/기타"
        by_dump[key] += h["qty"]

    total_haul = sum(by_cat.values())

    print("=" * 60)
    print("1. 25톤 덤프 적재량 (내역서·단가 기준)")
    print("=" * 60)
    print("  · 임목운반(위탁): apply_standard_prices.py — 25톤 × 0.40㎥/ton = 10㎥/회")
    print("    (시나리오B 20,000원/ton → 8,000원/㎥)")
    print("  · 06·04 토공: 규격에 「덤프24톤」 명시 → 표준시장단가 「24톤 덤프」 품셈 적용")
    print("  · 일반 토사(흐트러진상): 24~25톤 덤프 ≈ 12~16㎥/회 (토질·수분에 따라)")
    print("       — 표준 품셈 관행: 24톤 = 14~15㎥, 25톤 = 15~16㎥ (보통토사)")
    print("  · 본 내역서는 **운반 수량을 ㎥(흐트러진 토량)** 로 산출, 1회 적재㎥은 규격·품셈에 따름")
    print()

    print("=" * 60)
    print("2-A. 세부 용도별 — 4개 내역 합계")
    print("=" * 60)
    sub_tot: dict[str, float] = defaultdict(float)
    sub_by_file: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for h in haul_rows:
        sc = subcat(h["name"], h["spec"])
        sub_tot[sc] += h["qty"]
        sub_by_file[h["file"]][sc] += h["qty"]
    st = sum(sub_tot.values())
    for k, v in sorted(sub_tot.items(), key=lambda x: -x[1]):
        print(f"  {k:22s} {v:>12,.0f} ㎥  ({v/st*100:.1f}%)")
    print(f"  {'합계':22s} {st:>12,.0f} ㎥")
    print()
    for label, _fn in FILES:
        sub = sub_by_file[label]
        if not sub:
            continue
        ts = sum(sub.values())
        print(f"  [{label}] 소계 {ts:,.0f} ㎥")
        for k, v in sorted(sub.items(), key=lambda x: -x[1]):
            print(f"      {k}: {v:,.0f}")
    print()

    print("=" * 60)
    print("2-B. 흙운반 물량 — 제출 4개 내역 합계(1. 토공)")
    print("=" * 60)
    for k, v in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"  {k:20s} {v:>12,.0f} ㎥  ({v/total_haul*100:.1f}%)")
    print(f"  {'합계(운반)':20s} {total_haul:>12,.0f} ㎥")
    print()
    for k, v in by_file.items():
        print(f"    └ {k}: {v:,.0f} ㎥")
    print()

    print("  [덤프 톤수별 운반 ㎥ — 규격 표기 있는 항목]")
    for k, v in sorted(by_dump.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v:,.0f} ㎥")
    print()

    print("=" * 60)
    print("3. 01 토목 상세 (토공_물량비교표 분류)")
    print("=" * 60)
    d = load_01_tok_cut_fill_haul()
    print(f"  절토(흙깎기 등): {d['cut']:,.0f} ㎥")
    print(f"  성토(흙쌓기 등): {d['fill']:,.0f} ㎥")
    print(f"  절토−성토 잉여: {d['cut']-d['fill']:,.0f} ㎥")
    print("  흙운반 용도별:")
    ht = sum(d["haul"].values())
    for k, v in sorted(d["haul"].items(), key=lambda x: -x[1]):
        print(f"    {k:18s} {v:>10,.0f} ㎥")
    print(f"    {'운반 계':18s} {ht:>10,.0f} ㎥")
    print()

    # trip estimate
    print("=" * 60)
    print("4. 25톤(≈15㎥/회) 기준 왕복 횟수 추정")
    print("=" * 60)
    cap = 15.0
    for label, vol in sorted(by_cat.items(), key=lambda x: -x[1]):
        trips = vol / cap
        print(f"  {label}: {vol:,.0f} ㎥ ÷ {cap:.0f} ㎥/회 ≈ {trips:,.0f} 회")
    print(f"  전체: {total_haul:,.0f} ㎥ ≈ {total_haul/cap:,.0f} 회")
    print()

    print("=" * 60)
    print("5. 25톤 덤프 명시 내역 (원문)")
    print("=" * 60)
    for h in haul_rows:
        s = h["spec"] + h["name"]
        if h["dump_t"] or ("25" in s and "덤프" in s.replace(" ", "")):
            print(f"  [{h['file']}] {h['qty']:,.0f} ㎥ | {h['cat']}")
            print(f"    {h['name']}")
            print(f"    {h['spec']}")


if __name__ == "__main__":
    main()
