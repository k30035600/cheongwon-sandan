import sys, os
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_DIR = r"D:\OneDrive\Cursor\cheongwon\00_부동산개발"
XLSX = os.path.join(OUT_DIR, "청원지구_사업수지분석표(추정).xlsx")
HTML = os.path.join(OUT_DIR, "청원지구_사업수지분석표(추정).html")
PY = 3.305785   # 1평 = 3.305785㎡

# ── 실데이터 ─────────────────────────────
LAND     = 17_411_053_500     # 토지조서 Ⅰ 취득가액 합계(36필지, 면세)
ACQ_AREA = 83_083.07          # 취득총면적
SUPPLY   = 13_901_806_101     # 희상 260707 공급가액(부가세 별도)
DIRECT   = 11_203_907_953     # 직접공사비 총계(안분 기준)
MIP_M2   = round(LAND/ACQ_AREA)       # 매입단가 원/㎡(참고)
MIP      = round(LAND*PY/ACQ_AREA)    # 매입단가 원/평

# 조성공사비 공종(직접공사비) — 공급가액 안분 기준 (희상 260707 내역서 01절)
TOMOK_DIRECT    = 4_622_358_410   # 01 토목 (조경 제외)
JOGYEONG_DIRECT =   825_728_849   # 01 조경 (10.조경공)
SECT = [("토목 (01)",TOMOK_DIRECT),("조경 (01)",JOGYEONG_DIRECT),
        ("진입도로 (04)",921_547_807),("회전교차로 (05)",1_216_108_387),
        ("개발행위 (06)",2_856_374_055),("건설폐기물처리 (07)",36_421_922),
        ("전기설비 (02)",725_368_523)]
_r = SUPPLY/DIRECT
ALLOC = [(n, round(d*_r)) for n,d in SECT]
ALLOC[-1] = (ALLOC[-1][0], SUPPLY - sum(a for _,a in ALLOC[:-1]))  # 잔차 보정

# 유상 획지(분양) — 토지조서 Ⅲ (주차장용지 C-2는 근생매각으로 유상 편입)
SALE = [("가구 A-1","공업 C29 기타기계",4650),("가구 A-2","공업 C24 1차금속",8617),
        ("가구 A-3","공업 C24 1차금속",7803),("가구 A-4","공업 C24 1차금속",4996),
        ("가구 A-5","공업 C24 1차금속",3425),("가구 B-1","공업 C25 금속가공",8897),
        ("가구 B-2","공업 C25 금속가공",8448),("가구 C-1","지원시설용지",1139),
        ("가구 C-2","주차장용지",745)]
# 획지별 분양 시나리오 — 단가(원/평)·입주(분양) 예정(비고)
SALE_SCEN = [
    ("가구 A-1","공업 C29 기타기계",4650,3_300_000,"협의 중"),
    ("가구 A-2","공업 C24 1차금속",8617,3_300_000,"협의 중"),
    ("가구 A-3","공업 C24 1차금속",7803,3_300_000,"협의 중"),
    ("가구 A-4","공업 C24 1차금속",4996,3_300_000,"협의 중"),
    ("가구 A-5","공업 C24 1차금속",3425,3_300_000,"협의 중"),
    ("가구 B-1","공업 C25 금속가공",8897,3_300_000,"협의 중"),
    ("가구 B-2","공업 C25 금속가공",8448,3_300_000,"협의 중"),
    ("가구 C-1","지원시설용지",1139,3_300_000,"협의 중"),
    ("가구 C-2","주차장용지",745,3_300_000,"협의 중"),
]
# 공공귀속(무상) — 토지조서 Ⅳ·Ⅴ·Ⅵ (오수처리시설은 공공기반시설 → 무상귀속 유지)
PUBLIC = [("도로(구역내 6노선)","도로",11062),("소공원(1·2)","공원",5746),
          ("경관녹지1(산178·국유림)","녹지",4054),("경관녹지2","녹지",10405),
          ("전기공급설비","송전탑",164),("오수처리시설(D-1)","오수처리",466)]

SALE_M2   = sum(a for *_,a in SALE)      # 48,720
PUBLIC_M2 = sum(a for *_,a in PUBLIC)    # 31,897
ZONE_M2   = 80_617                       # 지구단위 구역면적(토지조서)
ZONE_PYE  = ZONE_M2 / PY
SALE_PYE  = SALE_M2 / PY

# ── 가정(기본값) ─────────────────────────────
PRICE_PY = 3_300_000          # 분양단가 원/평
EQUITY   = 24_000_000_000     # 자기자본 240억
IDLE_DEF = 15_400             # 유휴부지(구역 외 인접 11필지) ㎡
IDLE_COST= 3_602_446_415      # 유휴부지 취득원가(구역 외 11필지, 개발부담금 base 제외용)
TOJI_YY  = 3_000_000_000      # 토지용역비 30억
GUKYU    = 0                  # 산178 국유림 — 무상매입 후 무상귀속(용지비 0)
GUKYU_M2 = 4_054              # 경관녹지1 편입(산178 일원) — 공공귀속에 포함
BUDAE    = 0.046
# 대주단 인정 상한 — 설계·인허가·감리 항목별 분리(개발행위 기준)
DESIGN_PY  = 30_000            # 설계용역 원/평(구역면적 기준)
PERMIT_PY  = 20_000            # 인허가용역 원/평(개발행위 포함·합계 5만/평 상한)
GAMRI_RATE = 0.035             # 감리비 3.5%(공급가액 기준·대주단 인정 상한)
GWANRI   = 0.030
PANMAE   = 0.020
JOSU     = 0.015
MORTGAGE_CAP_RATIO = 1.20       # 근저당 채권최고액 = 실채권 × 120%
PF_MORTGAGE_MAX = 3_884_000_000  # 근저당 채권최고액(120%) 합계 — 토지조서 등기
PF_REFI = round(PF_MORTGAGE_MAX / MORTGAGE_CAP_RATIO)  # 실채권 — 선순위 확보용 기존채무 대환(차입요소)
PF_PRIORITY = 12_000_000_000    # 우선수익권(신탁) — 기표 전·수익권증서 인수 불요(참고)
PF_LOAN  = 30_000_000_000     # PF 총대출(요청금액) 300억 — 입력 가정
PF_RATE  = 0.08
PF_YEARS = 2.0
YEBI     = 0.030
# 제부담금(표준경비)
SANJI    = 57_325            # 산지전용면적(구역내 임야−국유지·존치)
SANLIM   = 8_250            # 대체산림 단가(준보전 8,190+공시지가 0.1%)
NONGJI   = 4_300            # 농지전용면적(구역내 전·답)
NONGJIR  = 12_000          # 농지 단가(공시지가 20%)
BUDAM    = SANJI*SANLIM + NONGJI*NONGJIR   # 제부담금 계
# 개발부담금 — 기부채납(무상귀속) 원가 흡수로 과세표준 0 가정
DEVR     = 0.0
JISANG   = 0
TARGET   = 0.10
# Ⅳ 분양단가 민감도 — 300만원/평부터 20만원 단위
SCENARIO_PRICES = list(range(3_000_000, 4_000_001, 200_000))

FOOTNOTE = (
    "※ 부가가치세 별도.\n"
    f"공공귀속({PUBLIC_M2:,}㎡)은 무상귀속 수입 0.\n"
    f"산178 국유림(경관녹지1 {GUKYU_M2:,}㎡)은 무상매입 후 무상귀속(용지비 0).\n"
    f"유휴부지({IDLE_DEF:,}㎡)는 비분양·원가흡수(수입 0), 취득원가 {IDLE_COST:,}은 용지비·개발부담금 공제.\n"
    "주차장용지(745㎡)는 근생 매각 유상.\n"
    "제부담금 표준경비.\n"
    "개발부담금 0(기부채납·무상귀속 원가 흡수로 과세표준 미달 가정).\n"
    f"설계·인허가 {DESIGN_PY+PERMIT_PY:,}원/평(구역 {ZONE_M2:,}㎡) · 감리 {GAMRI_RATE:.1%}(공급가액).\n"
    f"근저당 채권최고액(120%) {PF_MORTGAGE_MAX:,} · 실채권(100%) {PF_REFI:,} — 선순위 확보용 대환(PF 총대출 포함).\n"
    f"금융이자 연 {PF_RATE:.0%} 가정이나, 실제 기표는 공사 기성률에 따라 분할 실행되므로 실효금리는 인하될 수 있음.\n"
    f"우선수익권 {PF_PRIORITY:,} 기표 전 신탁·인수 불요.\n"
    f"PF 총대출(요청금액) {PF_LOAN:,} — 입력 가정.\n"
    "국유지·정상지가상승분·PF조건은 실확정치로 갱신.\n"
    "작성 2026. 7. 8.\n"
    f"목표이익률 {TARGET:.0%}\n"
    f"자기자본 {EQUITY:,}\n"
    f"공사비 공급가액 {SUPPLY:,}"
)

LAND_ZONE = LAND - IDLE_COST     # 사업구역(36필지 중 구역 내) 토지매입비
LAND_ZONE_BD = round(LAND_ZONE * BUDAE)
IDLE_BD = round(IDLE_COST * BUDAE)
ZONE_YONGJI = LAND_ZONE + TOJI_YY + LAND_ZONE_BD   # 사업구역 용지비 소계
OUT_YONGJI = IDLE_COST + IDLE_BD                   # 사업구역 외 용지비 소계

def soft_costs():
    """설계·인허가·감리(대주단 인정 상한)."""
    seolbi = round(ZONE_PYE * DESIGN_PY)
    inheo  = round(ZONE_PYE * PERMIT_PY)
    gamri  = round(SUPPLY * GAMRI_RATE)
    return seolbi, inheo, gamri

def rev_from_lots(lots):
    return sum(m2/PY * price for _,_,m2,price,_ in lots)

def pnl_lots(lots):
    b = pnl(PRICE_PY)
    rev = rev_from_lots(lots)
    panmae = rev * PANMAE
    delta_pan = panmae - b['panmae']
    profit = b['profit'] + (rev - b['revenue']) - delta_pan
    out = dict(b)
    out.update(revenue=rev, rev_sale=rev, panmae=panmae, profit=profit,
               m_rev=(profit/rev if rev else 0),
               m_cost=(profit/b['total'] if b['total'] else 0),
               roe=(profit/EQUITY if EQUITY else 0))
    return out

def pf_total(yongji, supply, seolbi, inheo, gamri):
    """PF 총대출 = 사업비 차입소요 − 자기자본."""
    return max(0, yongji + supply + seolbi + inheo + gamri - EQUITY)

def pnl(price_py):
    rev_sale = SALE_PYE*price_py
    revenue  = rev_sale
    yongji  = LAND + TOJI_YY + GUKYU + LAND*BUDAE
    seolbi, inheo, gamri = soft_costs()
    constr  = SUPPLY
    soft    = seolbi + inheo + gamri
    gwanri  = (yongji+constr+soft+BUDAM)*GWANRI
    panmae  = rev_sale*PANMAE
    other   = yongji+constr+soft+BUDAM+gwanri+panmae   # 이자·수수료 전 비용
    loan    = PF_LOAN
    josu    = loan*JOSU
    interest= loan*PF_RATE*PF_YEARS
    sub     = other+josu+interest
    yebi    = sub*YEBI
    tot_bef = sub+yebi
    gaebal  = 0
    total   = tot_bef + gaebal
    profit  = revenue-total
    fix     = yongji+constr+soft+BUDAM+gwanri+josu+interest
    return dict(rev_sale=rev_sale,revenue=revenue,yongji=yongji,
                constr=constr,seolbi=seolbi,inheo=inheo,gamri=gamri,soft=soft,
                inheo_total=soft,budam=BUDAM,gwanri=gwanri,panmae=panmae,
                pf_refi=PF_REFI,pf_loan=loan,josu=josu,
                interest=interest,sub=sub,yebi=yebi,tot_bef=tot_bef,gaebal=gaebal,
                total=total,profit=profit,fix=fix,
                m_rev=(profit/revenue if revenue else 0),m_cost=(profit/total if total else 0),
                eq_ratio=(EQUITY/total if total else 0),roe=(profit/EQUITY if EQUITY else 0))

def _price_for(t):
    """목표이익률 t를 만족하는 분양단가(원/평) — PF 총대출 고정 가정."""
    a = 1 + YEBI
    k = JOSU + PF_RATE*PF_YEARS
    yongji = LAND + TOJI_YY + GUKYU + LAND*BUDAE
    seolbi, inheo, gamri = soft_costs()
    soft = seolbi + inheo + gamri
    constr = SUPPLY
    gwanri = (yongji+constr+soft+BUDAM)*GWANRI
    C0 = yongji+constr+soft+BUDAM+gwanri            # 판매비 제외(단가 독립)
    R = a*(C0 + k*PF_LOAN) / ((1-t) - a*PANMAE)
    return R/SALE_PYE

_b = pnl(PRICE_PY); FIXC=_b['fix']
def be_py():
    return _price_for(0.0)
def tgt_py(t):
    return _price_for(t)

# ═══════════════ 엑셀 ═══════════════
wb = openpyxl.Workbook(); ws = wb.active; ws.title="사업수지(추정)"
ws.sheet_view.showGridLines=False
navy="1F3A5F"; ylw="FFF6CC"; gray="EEF0F4"; grn="E7F3EA"
thin=Side(style="thin",color="D9E1EC"); border=Border(thin,thin,thin,thin)
won='#,##0'; pct='0.0%'; num='#,##0'; yr='0.0'
for col,wd in [("A",2),("B",38),("C",12),("D",11),("E",13),("F",17),("G",30)]:
    ws.column_dimensions[col].width=wd

def st(cell,val,*,bold=False,size=10,color="1C2430",fill=None,align="left",fmt=None,bd=True):
    c=ws[cell]; c.value=val
    c.font=Font(bold=bold,size=size,color=color,name="맑은 고딕")
    if fill:c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    if fmt:c.number_format=fmt
    if bd:c.border=border
    return c

def sec(r,text):
    ws.merge_cells(f"B{r}:G{r}")
    st(f"B{r}",text,bold=True,size=11,color="FFFFFF",fill=navy,bd=False)

st("B2","화성 청원지구 — 사업수지분석표(추정) (산업용지 조성·분양)",bold=True,size=14,color=navy,bd=False)
ws.merge_cells("B2:G2")
st("B3","시행 진명개발(주) · 구역 80,617㎡ · 분양단가 원/평 · 부가가치세 별도 · 공사비 안분=희상 260707 공급가액",size=9,color="5A6675",bd=False)
ws.merge_cells("B3:G3")

# ── 입력 가정 ──
sec(5,"◼ 입력 가정  (노란 셀만 수정)")
inp=[("분양단가 (원/평)",PRICE_PY,won,"유상 획지 기본단가(획지별 개별조정 가능)"),
 ("자기자본",EQUITY,won,"지정값 240억"),
 ("토지조서 총매입비",LAND,won,f"토지조서 36필지 취득가액 합(면세) · 사업구역 {LAND_ZONE:,} + 사업구역 외 {IDLE_COST:,}"),
 ("사업구역 외 취득원가",IDLE_COST,won,"구역 외 인접 11필지(유휴부지)·비분양·용지비·개발부담금 공제"),
 ("유휴부지면적 (㎡)",IDLE_DEF,num,"구역 외 인접 11필지 · 비분양·원가흡수(수입 0)"),
 ("매입단가 (원/평)",MIP,won,f"참고용 평균단가 · 토지매입비 {LAND:,} ÷ {ACQ_AREA/PY:,.0f}평 (≈{MIP_M2:,}원/㎡)"),
 ("토지용역비",TOJI_YY,won,"토지용역 — 지주작업·매입용역 (지정 30억)"),
 ("국유지(산178) 매입비",GUKYU,won,"무상매입 후 무상귀속(경관녹지1) → 용지비 0"),
 ("취득부대비율",BUDAE,pct,"취득세·등기 등 × 토지매입비"),
 ("구역면적 (㎡)",ZONE_M2,num,"지구단위 구역 80,617㎡ · 설계·인허가 산정 기준"),
 ("설계용역 단가 (원/평)",DESIGN_PY,won,"대주단 인정 상한 3만/평"),
 ("인허가용역 단가 (원/평)",PERMIT_PY,won,"개발행위 포함·상한 2만/평(합계 5만/평)"),
 ("감리비율",GAMRI_RATE,pct,"× 공사 공급가액 · 대주단 인정 상한 3.5%"),
 ("산지전용면적 (㎡)",SANJI,num,"대체산림 부과면적(구역내 임야−국유·존치)"),
 ("대체산림 단가 (원/㎡)",SANLIM,won,"준보전산지 8,190+공시지가 0.1% (2026 고시)"),
 ("농지전용면적 (㎡)",NONGJI,num,"농지보전 부과면적(구역내 전·답)"),
 ("농지 단가 (원/㎡)",NONGJIR,won,"공시지가 20%(진흥지역 밖)·상한 5만"),
 ("일반관리비율(시행)",GWANRI,pct,"× (용지+조성+설계·인허가·감리+부담금)"),
 ("판매비율",PANMAE,pct,"× 분양수입"),
 ("조달수수료율",JOSU,pct,"× PF 총대출액"),
 ("PF 총대출(요청금액)",PF_LOAN,won,"입력 가정 300억"),
 ("근저당 채권최고액(120%)",PF_MORTGAGE_MAX,won,"토지조서 공동담보 합 · 등기상 한도"),
 ("PF 금리(연)",PF_RATE,pct,"책임준공 PF"),
 ("PF 기간(년)",PF_YEARS,yr,"가중 소요기간"),
 ("예비비율",YEBI,pct,"× 비용 소계")]
r=6; CELL={}
keys=["price","equity","landtotal","idlecost","idle","mip","tojiyy","gukyu","budae","zone","designpy","permitpy","gamri",
      "sanji","sanlim","nongji","nongjir","gwanri","panmae","josu","pfloan","pfmortgage","pfrate","pfyr","yebi"]
for (lab,val,fmt,note),k in zip(inp,keys):
    st(f"B{r}",lab); st(f"C{r}",val,fill=ylw,align="right",fmt=fmt,bold=True)
    ws.merge_cells(f"D{r}:G{r}"); st(f"D{r}",note,size=9,color="5A6675")
    CELL[k]=f"C{r}"; r+=1
st(f"B{r}","사업구역 토지매입비(파생)",bold=True); st(f"C{r}",f"={CELL['landtotal']}-{CELL['idlecost']}",align="right",fmt=won,bold=True,fill=gray)
ws.merge_cells(f"D{r}:G{r}"); st(f"D{r}","토지조서 총매입비 − 사업구역 외 취득원가(구역 내 25필지·유상분양 대상)",size=9,color="5A6675"); CELL["landzone"]=f"C{r}"; r+=1
st(f"B{r}","실채권(100%·대환)",bold=True); st(f"C{r}",f"=ROUND({CELL['pfmortgage']}/1.2,0)",align="right",fmt=won,bold=True,fill=gray)
ws.merge_cells(f"D{r}:G{r}"); st(f"D{r}","선순위(근저당) 확보용 기존채무 대환 · 차입요소 · 채권최고액÷120%",size=9,color="5A6675"); CELL["pfrefi"]=f"C{r}"; r+=1
st(f"B{r}","우선수익권(신탁·기표전)",size=9,color="5A6675"); st(f"C{r}",PF_PRIORITY,align="right",fmt=won,size=9,color="5A6675")
ws.merge_cells(f"D{r}:G{r}"); st(f"D{r}","수익권증서 120억 · 아직 기표 전 신탁금액 · PF 인수 불요(참고)",size=9,color="5A6675",bd=False); r+=1

# ── Ⅰ. 수입 ──
r+=1; sec(r,"Ⅰ. 수입 (분양 · 비분양 · 공공귀속)"); r+=1
hdr=r
for col,t,al in [("B","구분",""),("C","면적(㎡)","right"),("D","면적(평)","right"),("E","단가(원/평)","right"),("F","금액(원)","right"),("G","비고","")]:
    st(f"{col}{r}",t,bold=True,fill=gray,align="center")
r+=1
st(f"B{r}","분양수입 (유상 획지)",bold=True,fill=gray); ws.merge_cells(f"C{r}:G{r}"); r+=1
sale_first=r
for name,use,m2 in SALE:
    st(f"B{r}",f"  {name} · {use}"); st(f"C{r}",m2,align="right",fmt=num)
    st(f"D{r}",f"=C{r}/{PY}",align="right",fmt=num); st(f"E{r}",f"={CELL['price']}",align="right",fmt=won)
    st(f"F{r}",f"=D{r}*E{r}",align="right",fmt=won); st(f"G{r}","",size=9,color="5A6675")
    r+=1
sale_last=r-1
st(f"B{r}","  유상 소계",bold=True,fill=gray); st(f"C{r}",f"=SUM(C{sale_first}:C{sale_last})",align="right",fmt=num,bold=True,fill=gray)
st(f"D{r}",f"=SUM(D{sale_first}:D{sale_last})",align="right",fmt=num,bold=True,fill=gray); st(f"E{r}","",fill=gray)
st(f"F{r}",f"=SUM(F{sale_first}:F{sale_last})",align="right",fmt=won,bold=True,fill=gray); st(f"G{r}","공업 46,836 + 지원 1,139 + 주차 745",size=9,color="5A6675",fill=gray)
SALE_TOT=f"F{r}"; SALE_PYC=f"D{r}"; r+=1
# 비분양(유휴부지)
st(f"B{r}","비분양·원가흡수 (구역 외 유휴부지)",bold=True,fill=gray); ws.merge_cells(f"C{r}:G{r}"); r+=1
st(f"B{r}","  구역 외 인접 11필지"); st(f"C{r}",f"={CELL['idle']}",align="right",fmt=num)
st(f"D{r}",f"=C{r}/{PY}",align="right",fmt=num); st(f"E{r}","—",align="center")
st(f"F{r}",0,align="right",fmt=won); st(f"G{r}","비분양 · 취득원가 용지비 흡수 · 개발부담금 공제",size=9,color="5A6675"); r+=1
st(f"B{r}","  비분양 소계",bold=True,fill=gray); st(f"C{r}",f"={CELL['idle']}",align="right",fmt=num,bold=True,fill=gray)
st(f"D{r}",f"=C{r}/{PY}",align="right",fmt=num,bold=True,fill=gray); st(f"E{r}","",fill=gray)
st(f"F{r}",0,align="right",fmt=won,bold=True,fill=gray); st(f"G{r}",f"취득원가 {IDLE_COST:,}원(Ⅱ 용지비)",size=9,color="5A6675",fill=gray); r+=1
# 공공귀속
st(f"B{r}","공공귀속 (무상귀속·수입 0)",bold=True,fill=gray); ws.merge_cells(f"C{r}:G{r}"); r+=1
pub_first=r
for name,kind,m2 in PUBLIC:
    st(f"B{r}",f"  {name}"); st(f"C{r}",m2,align="right",fmt=num)
    st(f"D{r}",f"=C{r}/{PY}",align="right",fmt=num); st(f"E{r}","—",align="center")
    st(f"F{r}",0,align="right",fmt=won); st(f"G{r}",f"{kind} · 국유림 무상매입 후 무상귀속" if "산178" in name else f"{kind} · 준공 후 관리청 무상귀속",size=9,color="5A6675")
    r+=1
pub_last=r-1
st(f"B{r}","  공공귀속 소계",bold=True,fill=gray); st(f"C{r}",f"=SUM(C{pub_first}:C{pub_last})",align="right",fmt=num,bold=True,fill=gray)
st(f"D{r}",f"=SUM(D{pub_first}:D{pub_last})",align="right",fmt=num,bold=True,fill=gray); st(f"E{r}","",fill=gray)
st(f"F{r}",0,align="right",fmt=won,bold=True,fill=gray); st(f"G{r}","국토계획법 제65조 무상귀속(원가는 분양원가에 흡수)",size=9,color="5A6675",fill=gray); r+=1
st(f"B{r}","총수입",bold=True); ws.merge_cells(f"C{r}:E{r}")
st(f"F{r}",f"={SALE_TOT}",align="right",fmt=won,bold=True,fill=grn); st(f"G{r}","분양수입만(비분양·공공귀속 수입 0)",size=9,color="5A6675")
TOTREV=f"F{r}"; r+=1

# ── Ⅱ. 비용 ──
r+=1; sec(r,"Ⅱ. 비용 (부가가치세 별도)"); r+=1
def cost(label,formula,*,bold=False,note=""):
    global r
    fill=gray if bold else None
    st(f"B{r}",label,bold=bold,fill=fill); ws.merge_cells(f"C{r}:E{r}")
    if fill: st(f"C{r}","",fill=fill)
    if formula is None: st(f"F{r}","",fill=fill)
    else: st(f"F{r}",formula,align="right",fmt=won,bold=bold,fill=fill)
    st(f"G{r}",note,size=9,color="5A6675",fill=fill); rr=f"F{r}"; r+=1; return rr
cost("1) 용지비",None,bold=True)
cost("   사업구역",None,bold=True,note="지구단위 구역 내(80,617㎡)")
c_z_land=cost("     토지매입비",f"={CELL['landzone']}",note="구역 내(25필지)·유상분양 대상")
c_z_ty=cost("     토지용역비",f"={CELL['tojiyy']}",note="토지용역 — 지주작업·매입용역 (30억)")
c_z_bd=cost("     취득제세·부대비",f"={c_z_land}*{CELL['budae']}",note=f"× {BUDAE:.1%}")
c_zone_sub=cost("   사업구역 소계",f"={c_z_land}+{c_z_ty}+{c_z_bd}",bold=True)
cost("   사업구역 외",None,bold=True,note="구역 외 인접 11필지(유휴부지)")
c_out_land=cost("     유휴부지 취득원가",f"={CELL['idlecost']}",note="비분양·개발부담금 공제")
c_out_bd=cost("     취득제세·부대비",f"={c_out_land}*{CELL['budae']}",note=f"× {BUDAE:.1%}")
c_out_sub=cost("   사업구역 외 소계",f"={c_out_land}+{c_out_bd}",bold=True)
c_gk=cost("   국유지(산178)",f"={CELL['gukyu']}",note="무상매입 후 무상귀속(경관녹지1)")
YONGJI=cost("   용지비 계",f"={c_zone_sub}+{c_out_sub}+{c_gk}",bold=True,
           note=f"토지 총매입비 {LAND:,}(사업구역 {LAND_ZONE:,}+사업구역 외 {IDLE_COST:,}) · 용지비 사업구역 {ZONE_YONGJI:,} + 사업구역 외 {OUT_YONGJI:,} + 국유지 0")
cost("2) 조성공사비",None,bold=True)
sect_cells=[]
for name,amt in ALLOC:
    sect_cells.append(cost(f"   {name}",amt))
supply_cell=cost("   공사 공급가액 계",f"={'+'.join(sect_cells)}",bold=True,note="희상 260707 공급가액(부가세 별도)")
CONSTR=cost("   조성공사비 계",f"={supply_cell}",bold=True)
ZONEPYC=f"={CELL['zone']}/{PY}"
cost("3) 인허가용역비",None,bold=True)
c_gm=cost("   감리비",f"={supply_cell}*{CELL['gamri']}",note="공급가액 × 비율(대주단 상한)")
c_sb=cost("   설계용역비",f"={ZONEPYC}*{CELL['designpy']}",note="구역면적(평)×단가")
c_ih=cost("   인허가용역비",f"={ZONEPYC}*{CELL['permitpy']}",note="개발행위 포함·구역면적(평)×단가")
SOFT=f"{c_gm}+{c_sb}+{c_ih}"
c_soft=cost("   인허가용역비 계",f"={SOFT}",bold=True,note="감리+설계+인허가")
cost("4) 제부담금",None,bold=True)
c_sl=cost("   대체산림자원조성비",f"={CELL['sanji']}*{CELL['sanlim']}",note="준보전산지 표준단가")
c_nj=cost("   농지보전부담금",f"={CELL['nongji']}*{CELL['nongjir']}",note="공시지가 20%(진흥 밖)")
c_bdm=cost("   제부담금 계",f"={c_sl}+{c_nj}",bold=True)
c_gw=cost("5) 일반관리비",f"=({YONGJI}+{CONSTR}+{SOFT}+{c_bdm})*{CELL['gwanri']}",bold=True)
c_pm=cost("6) 판매비",f"={SALE_TOT}*{CELL['panmae']}",bold=True,note="분양수입 기준")
PFLOAN=f"={CELL['pfloan']}"
c_js=cost("7) 조달수수료",f"={PFLOAN}*{CELL['josu']}",bold=True,note="PF 총대출(요청금액) × 수수료율")
c_it=cost("8) 금융이자(PF)",f"={PFLOAN}*{CELL['pfrate']}*{CELL['pfyr']}",bold=True,
         note="PF 총대출(요청금액)×금리×기간 · 기성률에 따라 실효금리 인하 가능")
SUBC=cost("비용 소계 (예비비 전)",f"={YONGJI}+{CONSTR}+{SOFT}+{c_bdm}+{c_gw}+{c_pm}+{c_js}+{c_it}",bold=True)
c_yb=cost("9) 예비비",f"={SUBC}*{CELL['yebi']}")
TOTB=cost("비용 계",f"={SUBC}+{c_yb}",bold=True)
c_gb=cost("10) 개발부담금",0,bold=True,note="0 — 기부채납·무상귀속 원가 흡수로 과세표준 미달")
TOTAL=cost("총비용",f"={TOTB}+{c_gb}",bold=True)

# ── Ⅲ. 재무·손익 ──
r+=1; sec(r,"Ⅲ. 재무구조·손익"); r+=1
def fin(label,formula,*,fmt=won,bold=False,fill=None,note="",size=10):
    global r
    st(f"B{r}",label,bold=bold,size=size); ws.merge_cells(f"C{r}:E{r}")
    st(f"F{r}",formula,align="right",fmt=fmt,bold=bold,fill=fill,size=size)
    st(f"G{r}",note,size=9,color="5A6675"); rr=f"F{r}"; r+=1; return rr
fin("총사업비",f"={TOTAL}",bold=True)
fin("자기자본",f"={CELL['equity']}",bold=True,note="지정 240억")
fin("자기자본비율",f"={CELL['equity']}/{TOTAL}",fmt=pct,bold=True,fill=grn,note="자기자본 ÷ 총사업비")
fin("타인자본(차입 소요)",f"={TOTAL}-{CELL['equity']}",bold=True,note="총사업비 − 자기자본")
fin("근저당 실채권(대환)",f"={CELL['pfrefi']}",note="선순위(근저당) 확보용 기존채무 대환")
fin("PF 총대출(요청금액)",f"={CELL['pfloan']}",bold=True,fill=grn,note="입력 가정 300억 · 타인자본+근저당 실채권")
fin("사업이익",f"={TOTREV}-{TOTAL}",bold=True,fill=grn,size=11,note="총수입 − 총비용")
fin("이익률 (수입 대비)",f"=({TOTREV}-{TOTAL})/{TOTREV}",fmt=pct,bold=True)
fin("이익률 (원가 대비)",f"=({TOTREV}-{TOTAL})/{TOTAL}",fmt=pct,bold=True)
fin("자기자본이익률(ROE)",f"=({TOTREV}-{TOTAL})/{CELL['equity']}",fmt=pct,bold=True,note="사업이익 ÷ 자기자본")
FIXCELL=fin("고정비 (판매·예비 전)",f"={YONGJI}+{CONSTR}+{SOFT}+{c_bdm}+{c_gw}+{c_js}+{c_it}",note="시나리오 계산용")
A_EXP=f"(1+{CELL['yebi']})"
K_EXP=f"({CELL['josu']}+{CELL['pfrate']}*{CELL['pfyr']})"
C0_EXP=f"({YONGJI}+{CONSTR}+{SOFT}+{c_bdm}+{c_gw})"
be_f=f"={A_EXP}*({C0_EXP}+{K_EXP}*{CELL['pfloan']})/((1-0)-{A_EXP}*{CELL['panmae']})/{SALE_PYC}"
fin("손익분기 분양단가 (원/평)",be_f,bold=True,note="이익 0 · PF 300억 고정")
BE_CELL=f"F{r-1}"
t10_f=f"={A_EXP}*({C0_EXP}+{K_EXP}*{CELL['pfloan']})/((1-0.1)-{A_EXP}*{CELL['panmae']})/{SALE_PYC}"
fin("목표이익률 10% 분양단가 (원/평)",t10_f,bold=True,fill=grn,note="수입 대비 10% · PF 300억 고정")

# ── Ⅳ. 시나리오 ──
r+=1; sec(r,"Ⅳ. 분양단가 민감도 (원/평)"); r+=1
for col,t in [("B","분양단가(원/평)"),("C","총수입"),("F","사업이익"),("G","이익률(수입)")]:
    st(f"{col}{r}",t,bold=True,fill=gray,align="center")
ws.merge_cells(f"C{r}:E{r}"); r+=1
for v in SCENARIO_PRICES:
    is_base = abs(v - PRICE_PY) < 1
    st(f"B{r}",v,align="right",fmt=won,fill=grn if is_base else None,bold=is_base)
    st(f"C{r}",f"={SALE_PYC}*B{r}",align="right",fmt=won); ws.merge_cells(f"C{r}:E{r}")
    other_r=f"({C0_EXP}+C{r}*{CELL['panmae']})"
    total_r=f"({A_EXP}*({other_r}+{CELL['pfloan']}*{K_EXP}))"
    st(f"F{r}",f"=C{r}-{total_r}",align="right",fmt=won,fill=grn if is_base else None,bold=is_base)
    st(f"G{r}",f"=IF(C{r}=0,0,F{r}/C{r})",align="right",fmt=pct,fill=grn if is_base else None,bold=is_base)
    r+=1
# ── Ⅴ. 획지별 분양 시나리오 ──
r+=1; sec(r,"Ⅴ. 획지별 분양 시나리오"); r+=1
for col,t,al in [("B","구분","center"),("C","면적(㎡)","center"),("D","면적(평)","center"),("E","단가(원/평)","center"),("F","금액(원)","center"),("G","비고","center")]:
    st(f"{col}{r}",t,bold=True,fill=gray,align=al)
r+=1
scen_first=r
SCEN_PY_CELLS=[]
for name,use,m2,price,note in SALE_SCEN:
    st(f"B{r}",f"{name} · {use}"); st(f"C{r}",m2,align="right",fmt=num)
    st(f"D{r}",f"=C{r}/{PY}",align="right",fmt=num)
    st(f"E{r}",price,fill=ylw,align="right",fmt=won,bold=True)
    st(f"F{r}",f"=D{r}*E{r}",align="right",fmt=won)
    st(f"G{r}",note,size=9,color="5A6675")
    SCEN_PY_CELLS.append(f"E{r}"); r+=1
scen_last=r-1
st(f"B{r}","유상 소계",bold=True,fill=gray)
st(f"C{r}",f"=SUM(C{scen_first}:C{scen_last})",align="right",fmt=num,bold=True,fill=gray)
st(f"D{r}",f"=SUM(D{scen_first}:D{scen_last})",align="right",fmt=num,bold=True,fill=gray)
st(f"E{r}","",fill=gray)
SCEN_TOT=f"F{r}"; st(f"F{r}",f"=SUM(F{scen_first}:F{scen_last})",align="right",fmt=won,bold=True,fill=gray)
st(f"G{r}","획지별 단가·입주(분양) 예정",size=9,color="5A6675",fill=gray); r+=1
tb_scen=f"(({FIXCELL}+{SCEN_TOT}*{CELL['panmae']})*(1+{CELL['yebi']}))"
st(f"B{r}","시나리오 사업이익",bold=True); ws.merge_cells(f"C{r}:E{r}")
st(f"F{r}",f"={SCEN_TOT}-{tb_scen}",align="right",fmt=won,bold=True,fill=grn)
st(f"G{r}",f"=IF({SCEN_TOT}=0,0,F{r}/{SCEN_TOT})",align="right",fmt=pct,bold=True); r+=1
r+=1; ws.merge_cells(f"B{r}:G{r}")
st(f"B{r}",FOOTNOTE,size=9,color="8A5A00",bd=False)
ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb.calculation.fullCalcOnLoad=True

# ═══════════════ HTML ═══════════════
def f(n): return f"{round(n):,}"
b=pnl(PRICE_PY); be=be_py(); t10=tgt_py(TARGET)
budam_sanlim=SANJI*SANLIM; budam_nongji=NONGJI*NONGJIR
sale_rows=""
for name,use,m2 in SALE:
    pye=m2/PY; rev=pye*PRICE_PY
    sale_rows+=f'<tr><td class="l">{name} · {use}</td><td class="r">{f(m2)}</td><td class="r">{f(pye)}</td><td class="r">{f(PRICE_PY)}</td><td class="r">{f(rev)}</td><td class="note"></td></tr>\n'
pub_rows=""
for name,kind,m2 in PUBLIC:
    note=f"{kind}·국유림 무상매입 후 무상귀속" if "산178" in name else f"{kind}·무상귀속"
    pub_rows+=f'<tr><td class="l">{name}</td><td class="r">{f(m2)}</td><td class="r">{f(m2/PY)}</td><td class="c">—</td><td class="r">0</td><td class="note">{note}</td></tr>\n'
sect_rows=""
for name,amt in ALLOC:
    sect_rows+=f'<tr><td class="l">　{name}</td><td class="r">{f(amt)}</td><td class="note">직접공사비 안분</td></tr>\n'
scen_rows=""
for p in SCENARIO_PRICES:
    d=pnl(p)
    lbl=" (기본)" if abs(p-PRICE_PY)<1 else ""
    cls=' class="hi"' if lbl else ''
    scen_rows+=f'<tr{cls}><td class="r">{f(p)}{lbl}</td><td class="r">{f(d["revenue"])}</td><td class="r">{f(d["profit"])}</td><td class="r">{d["m_rev"]:.1%}</td></tr>\n'
scen_lot_rows=""
scen_rev=0
for name,use,m2,price,note in SALE_SCEN:
    pye=m2/PY; rev=pye*price; scen_rev+=rev
    scen_lot_rows+=f'<tr><td class="l">{name} · {use}</td><td class="r">{f(m2)}</td><td class="r">{f(pye)}</td><td class="r">{f(price)}</td><td class="r">{f(rev)}</td><td class="note">{note}</td></tr>\n'
bs=pnl_lots(SALE_SCEN)

html=f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"><title>청원지구 — 사업수지분석표(추정)</title>
<style>
@page{{size:A4 portrait;margin:0;}}
*{{box-sizing:border-box;}}
body{{font-family:"Malgun Gothic","맑은 고딕",sans-serif;margin:0;background:#8a92a0;color:#1c2430;line-height:1.5;padding:16px 0;font-size:13px;}}
.wrap{{width:210mm;min-height:297mm;margin:0 auto;padding:16mm 14mm;background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.25);}}
h1{{font-size:19px;color:#1f3a5f;margin:0 0 4px;border-bottom:3px solid #1f3a5f;padding-bottom:8px;}}
.meta{{color:#5a6675;font-size:12px;margin-bottom:12px;}}
h2{{font-size:14px;color:#fff;background:#1f3a5f;padding:5px 10px;margin:16px 0 6px;border-radius:3px;}}
table{{width:100%;border-collapse:collapse;font-size:12px;margin:4px 0;}}
th,td{{border:1px solid #d9e1ec;padding:4px 8px;}}
th{{background:#eef0f4;color:#1f3a5f;text-align:center;}}
td.l{{text-align:left;}} td.r{{text-align:right;}} td.c{{text-align:center;}}
tr.sum td{{background:#eef0f4;font-weight:700;}} tr.hi td{{background:#e7f3ea;font-weight:700;}}
td.note{{color:#5a6675;font-size:10.5px;}}
.small{{color:#8a5a00;font-size:11px;margin-top:8px;}}
@media print{{body{{background:#fff;padding:0;}}.wrap{{width:auto;min-height:0;margin:0;padding:15mm 14mm;box-shadow:none;}}}}
</style></head><body><div class="wrap">
<h1>화성 청원지구 — 사업수지분석표(추정)</h1>
<div class="meta">산업용지 조성·분양 · 시행 진명개발(주) · 구역 80,617㎡ · 자기자본 240억 · PF 300억 · <b>부가가치세 별도</b> · <b>추정치</b><br>
분양단가 {f(PRICE_PY)}원/평 기준 · 공사비=희상 260707 공급가액 {f(SUPPLY)} · 매입단가 {f(MIP)}원/평</div>

<h2>Ⅰ. 수입 (분양 · 비분양 · 공공귀속)</h2>
<table>
<colgroup><col style="width:30%"><col style="width:12%"><col style="width:11%"><col style="width:13%"><col style="width:16%"><col style="width:18%"></colgroup>
<tr><th>구분</th><th>면적(㎡)</th><th>면적(평)</th><th>단가(원/평)</th><th>금액(원)</th><th>비고</th></tr>
<tr class="sum"><td class="l" colspan="6">분양수입 (유상 획지)</td></tr>
{sale_rows}<tr class="sum"><td class="l">유상 소계</td><td class="r">{f(SALE_M2)}</td><td class="r">{f(SALE_PYE)}</td><td class="c">—</td><td class="r">{f(b['rev_sale'])}</td><td class="note">공업 46,836 + 지원 1,139 + 주차 745</td></tr>
<tr class="sum"><td class="l" colspan="6">비분양·원가흡수 (구역 외 유휴부지)</td></tr>
<tr><td class="l">구역 외 인접 11필지</td><td class="r">{f(IDLE_DEF)}</td><td class="r">{f(IDLE_DEF/PY)}</td><td class="c">—</td><td class="r">0</td><td class="note">취득원가 {f(IDLE_COST)}원 · 개발부담금 공제</td></tr>
<tr class="sum"><td class="l" colspan="6">공공귀속 (무상귀속 · 수입 0)</td></tr>
{pub_rows}<tr class="sum"><td class="l">공공귀속 소계</td><td class="r">{f(PUBLIC_M2)}</td><td class="r">{f(PUBLIC_M2/PY)}</td><td class="c">—</td><td class="r">0</td><td class="note">국토계획법 제65조 무상귀속</td></tr>
<tr class="hi"><td class="l">총수입</td><td class="c" colspan="3"></td><td class="r">{f(b['revenue'])}</td><td class="note">분양수입만</td></tr>
</table>

<h2>Ⅱ. 비용 (부가가치세 별도)</h2>
<table>
<tr><th class="l">항목</th><th>금액(원)</th><th class="l">비고</th></tr>
<tr class="sum"><td class="l">1) 용지비</td><td class="r">{f(b['yongji'])}</td><td class="note">토지 총매입비 {f(LAND)}(사업구역 {f(LAND_ZONE)}+사업구역 외 {f(IDLE_COST)}) + 토지용역·부대비 · 국유지 0</td></tr>
<tr><td class="l">　토지조서 총매입비</td><td class="r">{f(LAND)}</td><td class="note">36필지 취득가액 · 사업구역 {f(LAND_ZONE)} + 사업구역 외 {f(IDLE_COST)}</td></tr>
<tr><td class="l">　사업구역</td><td class="r">{f(ZONE_YONGJI)}</td><td class="note">지구단위 구역 내(80,617㎡)</td></tr>
<tr><td class="l">　　　토지매입비</td><td class="r">{f(LAND_ZONE)}</td><td class="note">구역 내·유상분양 대상</td></tr>
<tr><td class="l">　　　토지용역비</td><td class="r">{f(TOJI_YY)}</td><td class="note">토지용역 — 지주작업·매입용역 (30억)</td></tr>
<tr><td class="l">　　　취득제세·부대비</td><td class="r">{f(LAND_ZONE_BD)}</td><td class="note">× {BUDAE:.1%}</td></tr>
<tr><td class="l">　사업구역 외</td><td class="r">{f(OUT_YONGJI)}</td><td class="note">구역 외 인접 11필지(유휴부지)</td></tr>
<tr><td class="l">　　　유휴부지 취득원가</td><td class="r">{f(IDLE_COST)}</td><td class="note">비분양·개발부담금 공제</td></tr>
<tr><td class="l">　　　취득제세·부대비</td><td class="r">{f(IDLE_BD)}</td><td class="note">× {BUDAE:.1%}</td></tr>
<tr class="sum"><td class="l">2) 조성공사비</td><td class="r">{f(b['constr'])}</td><td class="note">공급가액 {f(SUPPLY)}(부가세 별도)</td></tr>
{sect_rows}<tr class="sum"><td class="l">3) 인허가용역비</td><td class="r">{f(b['soft'])}</td><td class="note">감리+설계+인허가</td></tr>
<tr><td class="l">　감리비</td><td class="r">{f(b['gamri'])}</td><td class="note">공급가액 × {GAMRI_RATE:.1%}(대주단 상한)</td></tr>
<tr><td class="l">　설계용역비</td><td class="r">{f(b['seolbi'])}</td><td class="note">구역 {f(ZONE_PYE)}평 × {f(DESIGN_PY)}원/평</td></tr>
<tr><td class="l">　인허가용역비</td><td class="r">{f(b['inheo'])}</td><td class="note">구역 {f(ZONE_PYE)}평 × {f(PERMIT_PY)}원/평(개발행위)</td></tr>
<tr class="sum"><td class="l">4) 제부담금</td><td class="r">{f(b['budam'])}</td><td class="note">표준경비</td></tr>
<tr><td class="l">　대체산림자원조성비</td><td class="r">{f(budam_sanlim)}</td><td class="note">산지 {f(SANJI)}㎡ × {f(SANLIM)}원(준보전산지)</td></tr>
<tr><td class="l">　농지보전부담금</td><td class="r">{f(budam_nongji)}</td><td class="note">농지 {f(NONGJI)}㎡ × {f(NONGJIR)}원(공시지가20%)</td></tr>
<tr><td class="l">5) 일반관리비</td><td class="r">{f(b['gwanri'])}</td><td class="note">(용지+조성+인허가용역+부담금) × {GWANRI:.0%}</td></tr>
<tr><td class="l">6) 판매비</td><td class="r">{f(b['panmae'])}</td><td class="note">분양수입 × {PANMAE:.0%}</td></tr>
<tr><td class="l">7) 조달수수료</td><td class="r">{f(b['josu'])}</td><td class="note">PF 총대출(요청금액) {f(b['pf_loan'])} × {JOSU:.1%}</td></tr>
<tr><td class="l">8) 금융이자(PF)</td><td class="r">{f(b['interest'])}</td><td class="note">PF 총대출(요청금액) {f(b['pf_loan'])} × {PF_RATE:.0%} × {PF_YEARS}년 · 기성률에 따라 실효금리 인하 가능</td></tr>
<tr><td class="l">9) 예비비</td><td class="r">{f(b['yebi'])}</td><td class="note">비용소계 × {YEBI:.0%}</td></tr>
<tr><td class="l">10) 개발부담금</td><td class="r">0</td><td class="note">기부채납·무상귀속 원가 흡수로 과세표준 미달</td></tr>
<tr class="sum"><td class="l">총비용</td><td class="r">{f(b['total'])}</td><td class="note"></td></tr>
</table>

<h2>Ⅲ. 재무구조·손익</h2>
<table>
<tr><td class="l">총사업비</td><td class="r">{f(b['total'])}</td><td class="note"></td></tr>
<tr><td class="l">자기자본</td><td class="r">{f(EQUITY)}</td><td class="note">지정 240억</td></tr>
<tr class="hi"><td class="l">자기자본비율</td><td class="r">{b['eq_ratio']:.1%}</td><td class="note">자기자본 ÷ 총사업비</td></tr>
<tr><td class="l">타인자본(차입 소요)</td><td class="r">{f(b['total']-EQUITY)}</td><td class="note">총사업비 − 자기자본</td></tr>
<tr><td class="l">근저당 실채권(대환)</td><td class="r">{f(b['pf_refi'])}</td><td class="note">선순위(근저당) 확보용 기존채무 대환</td></tr>
<tr class="hi"><td class="l">PF 총대출(요청금액)</td><td class="r">{f(b['pf_loan'])}</td><td class="note">입력 가정 300억 · 타인자본+근저당 실채권</td></tr>
<tr class="hi"><td class="l">사업이익</td><td class="r">{f(b['profit'])}</td><td class="note">총수입 − 총비용</td></tr>
<tr><td class="l">이익률 (수입 대비)</td><td class="r">{b['m_rev']:.1%}</td><td class="note"></td></tr>
<tr><td class="l">이익률 (원가 대비)</td><td class="r">{b['m_cost']:.1%}</td><td class="note"></td></tr>
<tr><td class="l">자기자본이익률(ROE)</td><td class="r">{b['roe']:.1%}</td><td class="note">사업이익 ÷ 자기자본</td></tr>
<tr><td class="l">손익분기 분양단가</td><td class="r">{f(be)} 원/평</td><td class="note">이익 0</td></tr>
<tr class="hi"><td class="l">목표이익률 10% 분양단가</td><td class="r">{f(t10)} 원/평</td><td class="note">수입 대비 10%</td></tr>
</table>

<h2>Ⅳ. 분양단가 민감도 (원/평)</h2>
<p class="meta" style="margin:0 0 6px">300만원/평부터 20만원 단위 · 손익분기 {f(be)}원/평</p>
<table>
<tr><th class="r">분양단가(원/평)</th><th class="r">총수입</th><th class="r">사업이익</th><th class="r">이익률(수입)</th></tr>
{scen_rows}</table>

<h2>Ⅴ. 획지별 분양 시나리오</h2>
<table>
<colgroup><col style="width:30%"><col style="width:12%"><col style="width:11%"><col style="width:13%"><col style="width:16%"><col style="width:18%"></colgroup>
<tr><th>구분</th><th>면적(㎡)</th><th>면적(평)</th><th>단가(원/평)</th><th>금액(원)</th><th>비고</th></tr>
{scen_lot_rows}<tr class="sum"><td class="l">유상 소계</td><td class="r">{f(SALE_M2)}</td><td class="r">{f(SALE_PYE)}</td><td class="c">—</td><td class="r">{f(scen_rev)}</td><td class="note">획지별 단가·입주(분양) 예정</td></tr>
<tr class="hi"><td class="l">시나리오 사업이익</td><td class="c" colspan="3"></td><td class="r">{f(bs['profit'])}</td><td class="note">이익률(수입) {bs['m_rev']:.1%}</td></tr>
</table>
<div class="small">{FOOTNOTE.replace(chr(10), "<br>")}</div>
</div></body></html>"""
with open(HTML,"w",encoding="utf-8") as fp: fp.write(html)
print("HTML saved")
try:
    wb.save(XLSX); print("XLSX saved")
except PermissionError:
    alt=XLSX.replace(".xlsx","_new.xlsx"); wb.save(alt)
    print(f"[경고] 원본 열림 → {os.path.basename(alt)} 로 저장")
print(f"[{f(PRICE_PY)}원/평] 총수입 {f(b['revenue'])}(분양만) / 총비용 {f(b['total'])} / 이익 {f(b['profit'])} ({b['m_rev']:.1%})")
print(f"자기자본비율 {b['eq_ratio']:.1%} · ROE {b['roe']:.1%} · 손익분기 {f(be)}원/평 · 목표10% {f(t10)}원/평")
print(f"제부담금 {f(b['budam'])} · 개발부담금 {f(b['gaebal'])}")
print(f"PF 총대출 {f(b['pf_loan'])} · 이자 {f(b['interest'])}")
