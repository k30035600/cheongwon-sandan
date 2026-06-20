#!/usr/bin/env python3
"""확정단가 반영 후 검증(읽기 전용):
  1) 일위확정 패치된 통합내역 행의 품명이 확정 품명과 일치하는지(오패치 0건)
  2) 파일별 최대 합계금액 상위 행(이상치 점검)
"""
from __future__ import annotations
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"

FILE_MAP = {
    "01 토목": "01_화성 청원지구 토목_표준단가산출.xlsx",
    "01 조경": "01_화성 청원지구 조경_표준단가산출.xlsx",
    "04 진입도로": "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
    "05 회전교차로": "05_화성 청원로(회전교차로)_표준단가산출.xlsx",
    "06 개발행위": "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
}
IL_FILES = [
    BASE / "미매칭_일위대가산출.xlsx",
    BASE / "검토_일위대가산출.xlsx",
    BASE / "검토_토공_일위대가산출.xlsx",
    BASE / "검토_공종별_일위대가산출.xlsx",
]


def norm(s) -> str:
    s = str(s or "")
    s = re.sub(r"^[\s0-9A-Za-z가-힣]{0,4}[).]\s*", "", s)
    s = re.sub(r"[\s.·,()/]+", "", s)
    return s


def load_override_names() -> dict[str, set[str]]:
    """파일별 확정 품명 정규화 집합."""
    out: dict[str, set[str]] = {k: set() for k in FILE_MAP}
    for p in IL_FILES:
        if not p.exists():
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb["일위대가산출"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c or "").strip() for c in rows[0]]
        ci_file = hdr.index("파일") if "파일" in hdr else None
        ci_name = hdr.index("품명") if "품명" in hdr else None
        ci_conf = hdr.index("확정단가(입력)") if "확정단가(입력)" in hdr else None
        ci_sug = hdr.index("제시단가") if "제시단가" in hdr else None
        for r in rows[1:]:
            if ci_file is None or ci_name is None:
                break
            fl = str(r[ci_file]).strip() if ci_file < len(r) and r[ci_file] else ""
            if fl not in out:
                continue
            has = False
            if ci_conf is not None and ci_conf < len(r) and r[ci_conf] not in (None, "", 0):
                has = True
            if ci_sug is not None and ci_sug < len(r) and r[ci_sug] not in (None, "", 0):
                has = True
            if has and r[ci_name]:
                out[fl].add(norm(r[ci_name]))
        wb.close()
    return out


def main():
    ov_names = load_override_names()
    total_bad = 0
    grand = 0
    for label, fname in FILE_MAP.items():
        p = WORK / fname
        if not p.exists():
            print(f"[{label}] 파일 없음")
            continue
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb["통합내역"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c or "").strip() for c in rows[0]]
        ci = {h: i for i, h in enumerate(hdr)}
        bad = []
        patched = []
        file_sum = 0
        recs = []
        for r in rows[1:]:
            if not r or r[ci["행"]] is None:
                continue
            name = r[ci["공종명"]]
            code = str(r[ci.get("단가코드", -1)] or "")
            amt = r[ci.get("합계금액", -1)]
            try:
                amt = float(amt or 0)
            except (TypeError, ValueError):
                amt = 0
            file_sum += amt
            recs.append((amt, r[ci["행"]], str(name or ""), str(r[ci["규격"]] or ""),
                         r[ci.get("수량", -1)], str(r[ci.get("단가코드", -1)] or "")))
            if code == "일위확정":
                patched.append(name)
                if norm(name) not in ov_names.get(label, set()):
                    bad.append((r[ci["행"]], str(name or "")[:24]))
        grand += file_sum
        print(f"[{label}] 일위확정 패치 {len(patched)}건 / 오패치 {len(bad)}건 / 합계 {file_sum:,.0f}")
        for row_no, nm in bad[:10]:
            print(f"    ✗ 행{row_no} 「{nm}」 — 확정 품명에 없음")
        total_bad += len(bad)
        # 상위 합계금액 3건
        top = sorted(recs, reverse=True)[:3]
        for amt, row_no, nm, spec, qty, code in top:
            print(f"    ▲ 행{row_no} {amt:,.0f}원  「{nm[:22]}」 {spec[:14]} ×{qty} [{code}]")
        wb.close()
    print("=" * 60)
    print(f"총 오패치 {total_bad}건 / 직접공사비(01토목+조경+04+05+06) {grand:,.0f}")


if __name__ == "__main__":
    main()
