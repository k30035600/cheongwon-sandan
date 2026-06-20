#!/usr/bin/env python3
"""01~07 공내역서 표준단가 산출 결과 — 내역서작업 총괄표 생성."""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from calc_overhead import (  # noqa: E402
    CIVIL_JOGYEONG_RATES,
    CIVIL_TOMOK_RATES,
    compute_cost_statement_civil,
    compute_cost_statement_electric,
)
from compare_cost_rates import write_rate_compare_sheet  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 내역서작업이 05_내역서 하위로 이동
WORK_DIR = BASE / "내역서작업"
OUT_DIR = BASE
OUT_MD = OUT_DIR / "총괄표.md"
OUT_XLSX = OUT_DIR / "총괄표.xlsx"
OUT_GROUPED_XLSX = OUT_DIR / "총괄표_공종별.xlsx"

INTEGRATED_HEADERS = [
    "행", "공종", "공종명", "규격", "수량", "단위", "상태", "매칭점수",
    "단가코드", "매칭품명", "매칭규격",
    "재료단가", "노무단가", "경비단가", "합계단가",
    "재료금액", "노무금액", "경비금액", "합계금액", "비고",
]

# 공종별 총괄표 — 전기 / 조경 / 토목(5개 내역서)
GROUPS: list[dict] = [
    {
        "prefix": "전기",
        "title": "전기설비",
        "scope_note": "02 전기설비 (03 지구외·02 중복 제외)",
        "row_names": {"전기설비"},
    },
    {
        "prefix": "조경",
        "title": "조경",
        "scope_note": "01 조경",
        "row_names": {"조경"},
    },
    {
        "prefix": "토목",
        "title": "토목·도로·개발행위·폐기물",
        "scope_note": "01 토목 + 04 진입도로 + 05 회전교차로 + 06 개발행위 + 07 건설폐기물",
        "row_names": {
            "토목", "진입도로", "화성 청원로(회전교차로)", "개발행위", "건설폐기물처리",
        },
    },
]

# 통합(토목+조경+전기) — 공종별 파일 맨 앞 2시트
WHOLE_GROUP: dict = {
    "prefix": "전체",
    "title": "토목·조경·전기 통합",
    "scope_note": "01~07 (03 전기 중복 제외) — 토목+조경+전기설비",
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
MONEY_FMT = "#,##0"
PCT_FMT = "0.00%"

def load_money_totals(xlsx_name: str) -> dict[str, int]:
    """합계요약 시트 ★ 매칭합계(토목 등) 또는 ★ 합계(전기) 행에서 재·노·경·합계 읽기."""
    cands = [d / xlsx_name for d in (BASE, WORK_DIR)]
    cands = [p for p in cands if p.exists()]
    if not cands:
        return {"mat": 0, "lab": 0, "exp": 0, "sum": 0}
    path = max(cands, key=lambda p: p.stat().st_mtime)
    wb = load_workbook(path, read_only=True, data_only=True)
    if "합계요약" not in wb.sheetnames:
        wb.close()
        return {"mat": 0, "lab": 0, "exp": 0, "sum": 0}
    ws = wb["합계요약"]
    total_labels = ("★ 매칭합계", "★ 합계", "★매칭합계")
    for row in ws.iter_rows(values_only=True):
        label = str(row[0] or "").strip() if row else ""
        if label in total_labels:
            wb.close()
            vals = list(row) + [0] * 8
            return {
                "mat": int(vals[4] or 0),
                "lab": int(vals[5] or 0),
                "exp": int(vals[6] or 0),
                "sum": int(vals[7] or 0),
            }
    wb.close()
    return {"mat": 0, "lab": 0, "exp": 0, "sum": 0}


# 금액 SSOT: 내역서작업/*_표준단가산출.xlsx 합계요약 (재생성 후 자동 반영)
ROWS = [
    {
        "no": "01",
        "name": "토목",
        "src": "01_화성 청원지구 토목.XLS",
        "xlsx": "01_화성 청원지구 토목_표준단가산출.xlsx",
        "price_src": "표준시장단가2026 + 시장시공가격 + 표준일위대가2026 + 물가정보2026(철근·형강)",
        "note": "물가정보 이형철근 952천원/톤·H형강 등 반영. 임목 3건 위탁 시장가 수동 반영(+9,399만)",
    },
    {
        "no": "01",
        "name": "조경",
        "src": "01_화성 청원지구 조경.XLS",
        "xlsx": "01_화성 청원지구 조경_표준단가산출.xlsx",
        "price_src": "표준시장단가2026 + 시장시공가격 + 표준일위대가2026 + 조경일위2024(폴백)",
        "note": "조경 수목 자재가 표준단가 폐지 → 조경일위·forestinfo 별도 보완",
        "exclude_unmatched_agg": True,
    },
    {
        "no": "02",
        "name": "전기설비",
        "src": "02_화성 청원지구 전기설비.xlsx",
        "xlsx": "02_화성 청원지구 전기설비_표준단가산출.xlsx",
        "price_src": "파일 내장 일위대가(품셈) + 합산자재 + 단가조사(중기) + 표준시장단가",
        "note": "209/209건 산출 — 크레인 10톤 6일·1식 380만(일위확정), 크레인트럭 20ton 단가조사",
    },
    {
        "no": "04",
        "name": "진입도로",
        "src": "04_화성 청원지구 진입도로 실시설계.XLS",
        "xlsx": "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx",
        "price_src": "표준시장단가2026 + 시장시공가격 + 표준일위대가2026 + 물가정보2026(철근·형강)",
        "note": "제경비·VAT `식(1)` 별도, 물가정보 철근·형강 단가 반영(미매칭 38→36)",
    },
    {
        "no": "05",
        "name": "화성 청원로(회전교차로)",
        "src": "05_화성 청원로(회전교차로).XLS",
        "xlsx": "05_화성 청원로(회전교차로)_표준단가산출.xlsx",
        "price_src": "표준시장단가2026 + 시장시공가격 + 표준일위대가2026 + 물가정보2026(철근·형강)",
        "note": "제경비·VAT `식(1)` 별도, 물가정보 철근·형강 단가 반영(미매칭 40→34)",
    },
    {
        "no": "06",
        "name": "개발행위",
        "src": "06_화성 청원지구 산업유통형 개발행위.XLS",
        "xlsx": "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx",
        "price_src": "표준시장단가2026 + 시장시공가격 + 표준일위대가2026(조달청·국도·하천·항만·교육청조경)",
        "note": "미매칭 26",
    },
    {
        "no": "07",
        "name": "건설폐기물처리",
        "src": "07_화성 청원지구 건설폐기물처리.XLS",
        "xlsx": "07_화성 청원지구 건설폐기물처리_표준단가산출.xlsx",
        "price_src": "동일 파일 경비 시트 (26. 협회단가)",
        "mat": 0,
        "lab": 0,
        "exp": 31_458_312,
        "sum": 31_458_312,
        "note": "경비 100% (재·노 0)",
    },
]


def load_status_counts(xlsx_name: str) -> dict[str, int]:
    """통합내역 시트 상태별 건수 — 매칭(매칭+원본) / 검토 / 미매칭(미매칭+미산출).

    산출물은 05_내역서(BASE)와 내역서작업(OUT_DIR) 양쪽에 존재할 수 있어,
    두 위치의 동일 파일 중 **가장 최근 수정본**을 읽는다(폴더 재편 이후 stale 방지).
    """
    cands = [d / xlsx_name for d in (BASE, WORK_DIR)]
    cands = [p for p in cands if p.exists()]
    if not cands:
        for d in (BASE, WORK_DIR):
            cands += list(d.glob(f"*{xlsx_name[:6]}*표준단가산출.xlsx"))
    if not cands:
        return {"matched": 0, "review": 0, "unmatched": 0, "total": 0}
    path = max(cands, key=lambda p: p.stat().st_mtime)
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "통합내역" if "통합내역" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    status_idx = next((i for i, h in enumerate(hdr) if h and "상태" in str(h)), 7)
    c: Counter[str] = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        st = str(row[status_idx] or "").strip()
        if st:
            c[st] += 1
    wb.close()
    matched = c.get("매칭", 0) + c.get("원본", 0)
    review = c.get("검토", 0)
    unmatched = c.get("미매칭", 0) + c.get("미산출", 0)
    total = matched + review + unmatched
    return {"matched": matched, "review": review, "unmatched": unmatched, "total": total}


def resolve_xlsx_path(xlsx_name: str) -> Path | None:
    """내역서작업·루트 중 최신 *_표준단가산출.xlsx 경로."""
    cands = [d / xlsx_name for d in (BASE, WORK_DIR)]
    cands = [p for p in cands if p.exists()]
    if not cands:
        for d in (BASE, WORK_DIR):
            cands += list(d.glob(f"*{xlsx_name[:6]}*표준단가산출.xlsx"))
    if not cands:
        return None
    return max(cands, key=lambda p: p.stat().st_mtime)


def load_integrated_lines(xlsx_name: str) -> list[list]:
    """통합내역 시트 전 행(헤더 제외) — 공종별 내역서 시트용."""
    path = resolve_xlsx_path(xlsx_name)
    if not path:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "통합내역" if "통합내역" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    lines: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not any(row):
            continue
        vals = list(row)
        if len(vals) < len(INTEGRATED_HEADERS):
            vals.extend([None] * (len(INTEGRATED_HEADERS) - len(vals)))
        lines.append(vals[: len(INTEGRATED_HEADERS)])
    wb.close()
    return lines


def group_totals(sub_rows: list[dict]) -> dict:
    """공종 그룹 합계 — 조경 exclude_unmatched_agg 규칙 유지."""
    unmatched_all = sum(r["unmatched"] for r in sub_rows)
    unmatched_excl = sum(
        r["unmatched"] for r in sub_rows if not r.get("exclude_unmatched_agg")
    )
    return {
        "matched": sum(r["matched"] for r in sub_rows),
        "review": sum(r["review"] for r in sub_rows),
        "unmatched": unmatched_excl,
        "unmatched_all": unmatched_all,
        "unmatched_excluded": unmatched_all - unmatched_excl,
        "total": sum(r["total"] for r in sub_rows),
        "mat": sum(r["mat"] for r in sub_rows),
        "lab": sum(r["lab"] for r in sub_rows),
        "exp": sum(r["exp"] for r in sub_rows),
        "sum": sum(r["sum"] for r in sub_rows),
    }


def enrich_rows() -> list[dict]:
    rows = []
    for base in ROWS:
        r = {**base, **load_status_counts(base["xlsx"])}
        loaded = load_money_totals(base["xlsx"])
        if loaded["sum"] > 0:
            r.update(loaded)
        rows.append(r)
    return rows


def fmt_money(n: float) -> str:
    return f"{n:,.0f}"


def fmt_oku(n: float) -> str:
    return f"{n / 100_000_000:.2f}억"


def aggregate(rows: list[dict]) -> dict:
    unmatched_all = sum(r["unmatched"] for r in rows)
    unmatched_excl = sum(
        r["unmatched"] for r in rows if not r.get("exclude_unmatched_agg")
    )
    return {
        "matched": sum(r["matched"] for r in rows),
        "review": sum(r["review"] for r in rows),
        "unmatched": unmatched_excl,
        "unmatched_all": unmatched_all,
        "unmatched_excluded": unmatched_all - unmatched_excl,
        "total": sum(r["total"] for r in rows),
        "mat": sum(r["mat"] for r in rows),
        "lab": sum(r["lab"] for r in rows),
        "exp": sum(r["exp"] for r in rows),
        "sum": sum(r["sum"] for r in rows),
    }


def pct(part: float, whole: float) -> str:
    if whole == 0:
        return "0.0%"
    return f"{part / whole * 100:.1f}%"


def write_md(rows: list[dict], t: dict) -> None:
    mat_oku = t["mat"] / 100_000_000
    lab_oku = t["lab"] / 100_000_000
    exp_oku = t["exp"] / 100_000_000
    sum_oku = t["sum"] / 100_000_000
    priced = t["matched"] + t["review"]
    lines = [
        "# 화성 청원지구 공내역서 — 표준단가 산출 총괄표",
        "",
        "- **작성일**: 2026. 6. 19.",
        "- **범위**: 01·02·04·05·06·07 공내역서 표준단가·협회단가 산출 결과 통합",
        "- **금액 성격**: **직접공사비**(재료·노무·경비) 추정 — **제경비·부가가치세·원단위 절사 미포함**",
        "- **합계 산식**: 01(토목+조경) + 02 + 04 + 05 + 06 + 07 (**03 전기(지구외)는 02와 동일 내역서·#REF!로 제외**)",
        "",
        "### 건수 구분 정의",
        "",
        "| 구분 | 의미 |",
        "|---|---|",
        "| **매칭** | 단가 매칭 확정(전기 **원본** 기입 포함) |",
        "| **미매칭** | 단가 미연결(**미매칭**·**미산출**) — **합계 행은 01 조경 제외** |",
        "| **전체** | 수량 내역 전 건수 (매칭 + 검토 + 미매칭) |",
        "",
        f"> ① **직접공사비 합계** — **약 {sum_oku:.1f}억** ({fmt_money(t['sum'])}원)",
        f"> ② **재료비** {mat_oku:.1f}억 · **노무비** {lab_oku:.1f}억 · **경비** {exp_oku:.1f}억",
        f"> ③ **매칭** {t['matched']} / **미매칭** {t['unmatched']} "
        f"(조경 {t['unmatched_excluded']}건 제외·전체 미매칭 {t['unmatched_all']}) / **전체** {t['total']}건 "
        f"(매칭률 **{pct(t['matched'], t['total'])}**)",
        f"> ④ **검토** {t['review']}건 — 금액 합계에 포함(매칭+검토 = {priced}건)",
        "",
        "## 1. 내역서별 직접공사비",
        "",
        "| No | 구분 | 매칭 | 검토 | 미매칭 | 전체 | 재료비 | 노무비 | 경비 | 합계 | 합계(억) |",
        "|---:|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for r in rows:
        lines.append(
            f"| {r['no']} | {r['name']} | {r['matched']} | {r['review']} | {r['unmatched']} | {r['total']} | "
            f"{fmt_money(r['mat'])} | {fmt_money(r['lab'])} | {fmt_money(r['exp'])} | "
            f"{fmt_money(r['sum'])} | {fmt_oku(r['sum'])} |"
        )
    lines.extend([
        f"| **합계** | **01~07** | **{t['matched']}** | **{t['review']}** | **{t['unmatched']}** ※ | **{t['total']}** | "
        f"**{fmt_money(t['mat'])}** | **{fmt_money(t['lab'])}** | **{fmt_money(t['exp'])}** | "
        f"**{fmt_money(t['sum'])}** | **{fmt_oku(t['sum'])}** |",
        "",
        f"※ 합계 **미매칭** — 01 조경 제외({t['unmatched_excluded']}건). 전체 미매칭 {t['unmatched_all']}건.",
        "",
        "## 2. 재료·노무·경비 구성 (01~07)",
        "",
        "| 구분 | 금액 | 비율 |",
        "|---|---:|---:|",
        f"| 재료비 | {fmt_money(t['mat'])} | {pct(t['mat'], t['sum'])} |",
        f"| 노무비 | {fmt_money(t['lab'])} | {pct(t['lab'], t['sum'])} |",
        f"| 경비 | {fmt_money(t['exp'])} | {pct(t['exp'], t['sum'])} |",
        f"| **합계** | **{fmt_money(t['sum'])}** | **100%** |",
        "",
        "## 3. 매칭·검토·미매칭 현황",
        "",
        "| No | 구분 | 매칭 | 검토 | 미매칭 | 전체 | 매칭률 |",
        "|---:|---|---:|---:|---:|---:|---:|",
    ])
    for r in rows:
        lines.append(
            f"| {r['no']} | {r['name']} | {r['matched']} | {r['review']} | {r['unmatched']} | "
            f"{r['total']} | {pct(r['matched'], r['total'])} |"
        )
    lines.extend([
        f"| **합계** | **01~07** | **{t['matched']}** | **{t['review']}** | **{t['unmatched']}** ※ | "
        f"**{t['total']}** | **{pct(t['matched'], t['total'])}** |",
        "",
        "※ 합계 **미매칭** — 01 조경 제외.",
        "",
        "## 4. 원본·단가 출처",
        "",
        "| No | 원본 파일 | 단가 출처 | 비고 |",
        "|---:|---|---|---|",
    ])
    for r in rows:
        lines.append(f"| {r['no']} | `{r['src']}` | {r['price_src']} | {r['note']} |")
    lines.extend([
        "",
        "## 5. 유의사항",
        "",
        "1. **전기(02)** — 파일 내장 **일위대가(품셈)·합산자재·단가조사(중기)** 단가를 우선 적용해 **208/209건(99.5%)** 산출. 크레인트럭 20ton은 단가조사(재료+노무+경비=94,126/hr)로 산출, 미산출 1건은 크레인 10톤(원본 단위 누락). 02 파일이 지구내+지구외를 모두 포함.",
        "2. **토목(01·04·05·06)** — 원본 XLS에 품셈 시트가 없으나, **표준시장단가2026(6,347)·시장시공가격(575)·표준일위대가2026(10,765, 조달청·국도·하천·항만·충남교육청 조경)** 총 17,687건을 **품명+규격 유사도(임계 0.56)**로 매칭해 재료·노무·경비 산출. **01은 2026-06-19 토목·조경 XLS 분리**(토목 493건·조경 83건). (조경 수목 자재가는 2021년 고시 폐지로 표준 단가 없음 → 별도 시세 입력 필요)",
        "2-1. **경계석류(화강석)** — 표준 DB에 화강석 경계석 단가가 없어, **`m`·`ea` 행 모두 콘크리트 「경계블록 설치(m)」 표준단가로 환산 매칭**(1ea = 개당 길이 m, 통상 1m). 화강석 재료비가 콘크리트 기준이므로 **재료비 과소 계상 가능** — 정밀 견적 시 화강석 시세 별도 보정 필요.",
        "3. **04·05·06** 상단 제경비·부가세 `식(1)` 행은 직접노무비 확정 후 별도 계산.",
        "4. **03 전기(지구외) 제외** — 03은 02와 품목 209개가 100% 동일한 복사본이고 값 셀이 전부 #REF!(깨진 수식)이므로 이중계상 방지를 위해 합계에서 제외(사용자 지시, 2026-06-18). 02 파일이 지구내+지구외 전체를 포함.",
        "5. **금액 합계**는 **매칭 + 검토 + 원본** 건만 반영. **미매칭·미산출** 건은 금액 0.",
        "6. **조경수 보완 점검** — forestinfo 조경수 관측시세(산림조합중앙회·임산물유통정보시스템, 참고가·고시 아님)로 01·주 미매칭 31건을 별도 시뮬레이션한 결과 **매칭 30·검토 1·미매칭 0**. 결과는 루트 `조경수_미매칭점검.xlsx`에 두며, 표준단가산출 금액에는 자동 합산하지 않는다.",
        "7. **민간 플랫폼 참고** — 그린나우·트리디비·트리4989는 실거래·직거래 단가 확인용 보조 단가원으로 정리(`일위대가DB/_외부원본/조경수_민간단가플랫폼.md`). 로그인·유료·공개 API 부재로 자동 수집 대상이 아니다.",
        "8. 세부 내역은 각 `내역서작업/*_표준단가산출.xlsx` 통합내역 시트 참조.",
        "9. **폴더 역할** — `내역서작업\\`에는 최종 `*_표준단가산출.xlsx`·`*_요약.md`만 두고, 루트 `05_내역서\\`에는 총괄표·검토·매칭·미매칭 파일을 둔다.",
        "",
        "끝.",
    ])
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def _style_cost_row(ws, row: int, ncol: int, *, bold: bool = False, fill=None, money_cols=(3, 4)):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill
        if c in money_cols and isinstance(cell.value, (int, float)):
            cell.number_format = MONEY_FMT
        if c == 5 and isinstance(cell.value, float):
            cell.number_format = PCT_FMT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _compute_cost(t: dict, engine: str) -> dict:
    if engine == "tomok":
        return compute_cost_statement_civil(t["mat"], t["lab"], t["exp"], CIVIL_TOMOK_RATES)
    if engine == "jogyeong":
        return compute_cost_statement_civil(
            t["mat"], t["lab"], t["exp"], CIVIL_JOGYEONG_RATES,
            rate_source="토목공사 간접공사비 기준(2026.4.13)·조경",
        )
    return compute_cost_statement_electric(t["mat"], t["lab"], t["exp"])


def write_cost_sheet(
    wb: Workbook,
    t: dict,
    *,
    sheet_name: str = "원가계산서",
    title: str = "화성 청원지구 — 원가계산서",
    basis_note: str | None = None,
    engine: str = "electric",
) -> dict:
    """원가계산서 시트 — 직접공사비 + 제비율 → 도급액. engine: electric|tomok|jogyeong."""
    cs = _compute_cost(t, engine)
    ws = wb.create_sheet(sheet_name)
    if basis_note is None:
        basis_note = (
            "총괄표 직접공사비(01~07, 03 전기 중복 제외) + "
            "03_화성 청원지구 전기설비.xlsx 「원가」 시트 제비율"
        )

    ws.append([title])
    ws.append(["작성일", "2026. 6. 19."])
    ws.append(["기준", basis_note])
    ws.append(["직접공사비", cs["direct"], cs["direct"] / 100_000_000, "재료+노무+경비"])
    ws.append(["도급액(총공사비)", cs["contract"], cs["contract"] / 100_000_000, f"직접비의 {cs['multiplier']:.3f}배"])
    ws["A1"].font = Font(bold=True, size=14)
    for r in (4, 5):
        ws.cell(r, 2).number_format = MONEY_FMT
        ws.cell(r, 3).number_format = "0.00"

    ws.append([])
    ws.append(["【적용 요율】"])
    ws.append(["항목", "요율", "적용 기준(대상액)"])
    rate_hdr = ws.max_row
    for c in range(1, 4):
        ws.cell(rate_hdr, c).font = Font(bold=True)
        ws.cell(rate_hdr, c).fill = HEADER_FILL
    for label, rate, basis in cs["rate_rows"]:
        ws.append([label, rate, basis])
        rr = ws.max_row
        ws.cell(rr, 2).number_format = PCT_FMT

    ws.append([])
    ws.append(["【산출 내역】"])
    ws.append(["단계", "항목", "금액(원)", "금액(억)", "산식"])
    calc_hdr = ws.max_row
    for c in range(1, 6):
        ws.cell(calc_hdr, c).font = Font(bold=True)
        ws.cell(calc_hdr, c).fill = HEADER_FILL

    for item in cs["rows"]:
        name = ("  " * item["indent"]) + item["name"]
        ws.append([item["step"], name, item["amount"], item["amount"] / 100_000_000, item["formula"]])
        rr = ws.max_row
        fill = TOTAL_FILL if item.get("total") else (SUBTOTAL_FILL if item.get("bold") else None)
        _style_cost_row(ws, rr, 5, bold=item.get("bold"), fill=fill)

    ws.append([])
    note_row = ws.max_row + 1
    if engine == "tomok":
        note = (
            "※ 토목공사 간접공사비 적용기준(2026.4.13, 한국표준품셈정보원·현행) 적용 — "
            "공사기간 13~36개월·직접공사비 50억-300억·종합건설업 기준"
            "(간접노무비 19.7%·기타경비 6.9%, 추정가격 50-100억 → 일반관리비 6.5%·이윤 12%). "
            "건설기계대여대금 지급보증(법정부담금) 0.4%×직접공사비 신설 반영. "
            "산업안전보건관리비는 대상액(재+직노) 구간별 율·기초액 자동 적용. "
            "환경보전비는 토목 도로(0.9%) 기준(공사종류별 상이). "
            "04·05·06 원본 제경비·VAT 식(1) 행은 별도(이중계상 주의)."
        )
    elif engine == "jogyeong":
        note = (
            "※ 토목공사 간접공사비 적용기준(2026.4.13)의 조경 요율 적용 — "
            "공사기간 13~36개월 기준(간접노무비 19.1%·기타경비 6.3%), "
            "추정가격 5억 미만 → 일반관리비 8%·이윤 15% 자동 적용. "
            "건설기계대여대금 지급보증 0.18%×직접공사비 반영. "
            "산업안전보건관리비는 대상액 구간 적용, 환경보전비 0.5%(보수). "
            "조경 수목 자재가 미반영분은 직접공사비 보정 시 도급액 동시 증가."
        )
    else:
        note = (
            "※ 03 전기 원가 시트 제비율(간접노무비 16.7%·이윤 15%·일반관리비 6.5% 등)을 "
            "직접공사비에 적용한 개략 산출. "
            "한전수탁비 등 전기 전용 고정액은 제외(02 전기 내역에 포함 여부 별도 확인). "
            "미매칭 단가 보정 시 직접공사비·도급액 동시 증가."
        )
    ws.cell(note_row, 1, note)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 44
    return cs


def write_xlsx(rows: list[dict], t: dict) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "총괄"
    ws.append(["화성 청원지구 공내역서 — 표준단가 산출 총괄표"])
    ws.append(["작성일", "2026. 6. 19."])
    ws.append(["합계 산식", "01+02+04+05+06+07 (03 전기 중복 제외)"])
    ws.append(["건수", "매칭 / 미매칭(01 조경 제외) / 전체 (검토는 별도 열)"])
    ws.append([])
    headers = [
        "No", "구분", "원본 파일", "단가 출처",
        "매칭", "검토", "미매칭", "전체", "매칭률",
        "재료비", "노무비", "경비", "합계", "합계(억)", "비고",
    ]
    ws.append(headers)
    hdr_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr_row, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for r in rows:
        ws.append([
            r["no"], r["name"], r["src"], r["price_src"],
            r["matched"], r["review"], r["unmatched"], r["total"],
            r["matched"] / r["total"] if r["total"] else 0,
            r["mat"], r["lab"], r["exp"], r["sum"], r["sum"] / 100_000_000,
            r["note"],
        ])
    ws.append([])
    total_row = ws.max_row + 1
    ws.append([
        "", "합계(01~07)", "", "",
        t["matched"], t["review"], t["unmatched"], t["total"],
        t["matched"] / t["total"] if t["total"] else 0,
        t["mat"], t["lab"], t["exp"], t["sum"], t["sum"] / 100_000_000,
        f"미매칭 합계에서 01 조경 {t['unmatched_excluded']}건 제외",
    ])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(total_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    ws2 = wb.create_sheet("구성비")
    ws2.append(["구분", "금액", "비율"])
    for c in range(1, 4):
        ws2.cell(1, c).font = Font(bold=True)
        ws2.cell(1, c).fill = HEADER_FILL
    for label, key in [("재료비", "mat"), ("노무비", "lab"), ("경비", "exp")]:
        ws2.append([label, t[key], t[key] / t["sum"] if t["sum"] else 0])
    ws2.append(["합계", t["sum"], 1])

    write_cost_sheet(wb, t)
    write_rate_compare_sheet(wb)

    for sheet in wb.worksheets:
        start = hdr_row + 1 if sheet.title == "총괄" else 2
        for row in sheet.iter_rows(min_row=start, max_row=sheet.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.column in (10, 11, 12, 13):
                    cell.number_format = MONEY_FMT
                if isinstance(cell.value, float) and cell.column in (9, 14):
                    cell.number_format = "0.0%" if cell.column == 9 else "0.00"
        for col in sheet.columns:
            w = min(52, max(len(str(c.value or "")) for c in col) + 2)
            sheet.column_dimensions[col[0].column_letter].width = w
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws["A1"].font = Font(bold=True, size=14)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT_XLSX)
        return OUT_XLSX
    except PermissionError:
        alt = OUT_XLSX.with_name(OUT_XLSX.stem + "_업데이트.xlsx")
        wb.save(alt)
        return alt


def _style_summary_sheet(ws, hdr_row: int, money_cols: tuple[int, ...]) -> None:
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in money_cols:
                cell.number_format = MONEY_FMT
            if isinstance(cell.value, float) and cell.column in (9, 14):
                cell.number_format = "0.0%" if cell.column == 9 else "0.00"
    for col in ws.columns:
        w = min(52, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[col[0].column_letter].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_group_summary_sheet(
    wb: Workbook,
    sheet_name: str,
    group: dict,
    sub_rows: list[dict],
    gt: dict,
) -> int:
    """공종별 총괄 시트 — 내역서별 직접공사비·건수."""
    ws = wb.create_sheet(sheet_name)
    ws.append([f"화성 청원지구 — {group['title']} 총괄표"])
    ws.append(["작성일", "2026. 6. 19."])
    ws.append(["범위", group["scope_note"]])
    ws.append(["건수", "매칭 / 미매칭(조경 제외 시 해당) / 전체"])
    ws.append([])
    headers = [
        "No", "구분", "원본 파일", "단가 출처",
        "매칭", "검토", "미매칭", "전체", "매칭률",
        "재료비", "노무비", "경비", "합계", "합계(억)", "비고",
    ]
    ws.append(headers)
    hdr_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr_row, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for r in sub_rows:
        ws.append([
            r["no"], r["name"], r["src"], r["price_src"],
            r["matched"], r["review"], r["unmatched"], r["total"],
            r["matched"] / r["total"] if r["total"] else 0,
            r["mat"], r["lab"], r["exp"], r["sum"], r["sum"] / 100_000_000,
            r["note"],
        ])
    ws.append([])
    total_row = ws.max_row + 1
    note = ""
    if gt["unmatched_excluded"]:
        note = f"미매칭 합계에서 01 조경 {gt['unmatched_excluded']}건 제외"
    ws.append([
        "", f"합계({group['prefix']})", "", "",
        gt["matched"], gt["review"], gt["unmatched"], gt["total"],
        gt["matched"] / gt["total"] if gt["total"] else 0,
        gt["mat"], gt["lab"], gt["exp"], gt["sum"], gt["sum"] / 100_000_000,
        note,
    ])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(total_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
    ws["A1"].font = Font(bold=True, size=14)
    _style_summary_sheet(ws, hdr_row, (10, 11, 12, 13))
    return hdr_row


def write_group_items_sheet(wb: Workbook, sheet_name: str, sub_rows: list[dict]) -> None:
    """공종별 통합내역 — 내역서작업 표준단가산출 통합내역 병합."""
    ws = wb.create_sheet(sheet_name)
    headers = ["내역서"] + INTEGRATED_HEADERS
    ws.append(headers)
    hdr_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr_row, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for sr in sub_rows:
        for line in load_integrated_lines(sr["xlsx"]):
            ws.append([sr["name"]] + line)
    money_cols = tuple(range(18, 22))  # 재료·노무·경비·합계금액 (+내역서 열 1칸 시프트)
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in money_cols:
                cell.number_format = MONEY_FMT
            if isinstance(cell.value, (int, float)) and cell.column in (13, 14, 15, 16):
                cell.number_format = MONEY_FMT
    for col in ws.columns:
        w = min(40, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[col[0].column_letter].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_total_cost_summary(
    wb: Workbook, sheet_name: str, group_cs: list[tuple[str, dict, dict]], t: dict
) -> None:
    """전체 원가계산서 — 공종별(전기·조경·토목) 도급액 합산."""
    ws = wb.create_sheet(sheet_name)
    ws.append(["화성 청원지구 — 토목·조경·전기 통합 원가계산서 (공종별 합산)"])
    ws.append(["작성일", "2026. 6. 19."])
    ws.append([
        "기준",
        "공종별로 상이한 제비율 적용 후 합산 — "
        "전기: 03 전기 원가 요율 / 토목·조경: 토목공사 간접공사비 기준(2026.4.13, 13~36개월). "
        "전기는 토목 간접공사비 기준 적용제외 대상이므로 별도 요율 유지.",
    ])
    ws.append([])

    headers = ["공종", "재료비", "직접노무비", "직접경비", "직접공사비", "도급액", "도급액(억)", "배율", "적용요율"]
    ws.append(headers)
    hdr_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr_row, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    label_map = {"전기": "전기설비(02)", "조경": "조경(01)", "토목": "토목·도로·개발·폐기물"}
    src_map = {"전기": "03 전기 원가 요율", "조경": "토목기준 조경요율", "토목": "토목기준 토목요율"}
    tot_direct = tot_contract = 0
    for prefix, gt, cs in group_cs:
        ws.append([
            label_map.get(prefix, prefix),
            cs["mat"], cs["lab"], cs["exp"], cs["direct"],
            cs["contract"], cs["contract"] / 100_000_000,
            cs["multiplier"], src_map.get(prefix, ""),
        ])
        tot_direct += cs["direct"]
        tot_contract += cs["contract"]

    total_row = ws.max_row + 1
    ws.append([
        "합계(전체)", "", "", "", tot_direct,
        tot_contract, tot_contract / 100_000_000,
        tot_contract / tot_direct if tot_direct else 0, "공종별 합산",
    ])
    for c in range(1, len(headers) + 1):
        cell = ws.cell(total_row, c)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column in (2, 3, 4, 5, 6):
                cell.number_format = MONEY_FMT
            if isinstance(cell.value, float) and cell.column in (7, 8):
                cell.number_format = "0.00"
    ws["A1"].font = Font(bold=True, size=14)
    for col in ws.columns:
        w = min(46, max(len(str(c.value or "")) for c in col) + 2)
        ws.column_dimensions[col[0].column_letter].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_grouped_workbook(rows: list[dict], t: dict) -> Path:
    """전체(총괄·원가) + 전기·조경·토목 3그룹 × (총괄·원가·내역) = 11시트."""
    wb = Workbook()
    wb.remove(wb.active)

    engine_map = {"전기": "electric", "조경": "jogyeong", "토목": "tomok"}
    basis_map = {
        "electric": "03_화성 청원지구 전기설비.xlsx 「원가」 시트 제비율",
        "tomok": "토목공사 간접공사비 적용기준(2026.4.13)·토목 요율(공사기간 13~36개월)",
        "jogyeong": "토목공사 간접공사비 적용기준(2026.4.13)·조경 요율(공사기간 13~36개월)",
    }

    # 전체 = 공종별(전기·조경·토목) 도급액 합산
    write_group_summary_sheet(wb, "전체_총괄", WHOLE_GROUP, rows, t)
    group_cs: list[tuple[str, dict, dict]] = []  # (prefix, gt, cost statement)
    for group in GROUPS:
        sub_rows = [r for r in rows if r["name"] in group["row_names"]]
        gt = group_totals(sub_rows)
        engine = engine_map[group["prefix"]]
        group_cs.append((group["prefix"], gt, _compute_cost(gt, engine)))
    write_total_cost_summary(wb, "전체_원가계산서", group_cs, t)

    for group in GROUPS:
        sub_rows = [r for r in rows if r["name"] in group["row_names"]]
        gt = group_totals(sub_rows)
        prefix = group["prefix"]
        engine = engine_map[prefix]
        write_group_summary_sheet(wb, f"{prefix}_총괄", group, sub_rows, gt)
        write_cost_sheet(
            wb,
            gt,
            sheet_name=f"{prefix}_원가계산서",
            title=f"화성 청원지구 — {group['title']} 원가계산서",
            basis_note=f"{group['scope_note']} 직접공사비 + {basis_map[engine]}",
            engine=engine,
        )
        write_group_items_sheet(wb, f"{prefix}_내역서", sub_rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT_GROUPED_XLSX)
        return OUT_GROUPED_XLSX
    except PermissionError:
        alt = OUT_GROUPED_XLSX.with_name(OUT_GROUPED_XLSX.stem + "_업데이트.xlsx")
        wb.save(alt)
        return alt


def main() -> None:
    rows = enrich_rows()
    t = aggregate(rows)
    write_md(rows, t)
    saved = write_xlsx(rows, t)
    grouped = write_grouped_workbook(rows, t)
    print(f"총괄표 저장: {OUT_MD}")
    print(f"총괄표 저장: {saved}")
    print(f"공종별 총괄표 저장: {grouped}")
    print(
        f"매칭 {t['matched']} / 미매칭 {t['unmatched']} "
        f"(조경 {t['unmatched_excluded']}건 제외·전체 {t['unmatched_all']}) / 전체 {t['total']} "
        f"(검토 {t['review']})"
    )
    print(f"직접공사비 합계(01~07): {fmt_money(t['sum'])} ({fmt_oku(t['sum'])})")
    gsum = sum(group_totals([r for r in rows if r["name"] in g["row_names"]])["sum"] for g in GROUPS)
    print(f"공종별 합계 검증: {fmt_money(gsum)} (통합 {fmt_money(t['sum'])})")


if __name__ == "__main__":
    main()
