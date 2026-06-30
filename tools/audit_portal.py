#!/usr/bin/env python3
"""청원지구_포털.html 기준 전수조사(팩트체크).

출력: 경로 존재 · 탭 정의 · JS 집계 일치 · _ka SSOT · README 대조
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_report_html import (  # noqa: E402
    _detail_hdr_row,
    load_ka,
)
from naeyeok_gongjong import classify_ka_mihwakjeong, detail_row_has_sum, is_ka_pending_excluded  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
PORTAL = ROOT / "청원지구_포털.html"
KA = ROOT / "08_제출내역서" / "청원지구_단가통합(전기제외)_ka.xlsx"
PENDING_HTML = ROOT / "05_내역서" / "내역서작업" / "_공통" / "청원지구_미확정_ka.html"
REPORT_HTML = ROOT / "05_내역서" / "내역서작업" / "_공통" / "청원지구_종합보고서.html"
PORTAL_STATS = ROOT / "05_내역서" / "내역서작업" / "_공통" / "portal_stats.js"
PORTAL_KA_STATS = ROOT / "05_내역서" / "내역서작업" / "_공통" / "portal_ka_stats.js"

DETAIL_SHEETS = ("회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물")

issues: list[str] = []
oks: list[str] = []
warns: list[str] = []


def issue(msg: str) -> None:
    issues.append(msg)


def ok(msg: str) -> None:
    oks.append(msg)


def warn(msg: str) -> None:
    warns.append(msg)


def section(title: str) -> None:
    print(f"\n{'=' * 68}\n■ {title}\n{'=' * 68}")


def extract_paths(html: str) -> set[str]:
    paths: set[str] = set()
    for m in re.finditer(r'enc\("([^"]+)"\)', html):
        paths.add(m.group(1).split("?")[0].split("#")[0])
    for m in re.finditer(r'B \+ "(/[^"]+)"', html):
        paths.add("05_내역서" + m.group(1))
    for m in re.finditer(r'main:\s*"([^"]+)"', html):
        p = m.group(1)
        if not p.startswith("http"):
            paths.add(p.split("?")[0])
    for m in re.finditer(r'src="([^"]+)"', html):
        p = m.group(1)
        if not p.startswith("http"):
            paths.add(p.split("?")[0])
    for m in re.finditer(r'\[\s*"([^"]+\.(?:html|md|xlsx|XLS))"', html):
        paths.add(m.group(1))
    paths.add(KA.relative_to(ROOT).as_posix())
    return paths


def parse_js_obj(path: Path) -> dict:
    txt = path.read_text(encoding="utf-8")
    m = re.search(r"=\s*(\{[\s\S]*?\});", txt)
    if not m:
        return {}
    raw = m.group(1)
    raw = re.sub(r"(\w+)\s*:", r'"\1":', raw)
    return json.loads(raw)


def count_pending_live() -> dict:
    wb = load_workbook(KA, data_only=True)
    totals = {"rv": 0, "um": 0, "all": 0}
    for sn in DETAIL_SHEETS:
        ws = wb[sn]
        hdr = _detail_hdr_row(ws)
        if hdr is None:
            continue
        for r in range(hdr + 1, ws.max_row + 1):
            if not ws.cell(r, 3).value:
                continue
            name = ws.cell(r, 4).value
            if not name or str(name).startswith("★"):
                continue
            section_v = ws.cell(r, 2).value
            if is_ka_pending_excluded(section_v, name):
                continue
            has_sum = detail_row_has_sum(
                mat_u=ws.cell(r, 8).value,
                lab_u=ws.cell(r, 10).value,
                exp_u=ws.cell(r, 12).value,
                mat_amt=ws.cell(r, 9).value,
                lab_amt=ws.cell(r, 11).value,
                exp_amt=ws.cell(r, 13).value,
                unit_sum=ws.cell(r, 14).value,
                total_amt=ws.cell(r, 15).value,
            )
            cat = classify_ka_mihwakjeong(ws.cell(r, 16).value, has_sum=has_sum)
            if cat == "검토":
                totals["rv"] += 1
                totals["all"] += 1
            elif cat == "미매칭":
                totals["um"] += 1
                totals["all"] += 1
    wb.close()
    return totals


def main() -> int:
    html = PORTAL.read_text(encoding="utf-8")

    section("1. 포털 참조 파일 존재")
    paths = extract_paths(html)
    missing = [p for p in sorted(paths) if not (ROOT / p).exists()]
    print(f"추출 경로 {len(paths)}건 · 존재 {len(paths) - len(missing)} · 누락 {len(missing)}")
    for p in missing:
        print(f"  ✗ {p}")
        issue(f"누락 파일: {p}")
    if not missing:
        ok(f"포털 참조 로컬 파일 {len(paths)}건 전부 존재")

    section("2. 필수 vendor·스크립트")
    for p in ("assets/vendor/xlsx.full.min.js", "assets/vendor/marked.min.js"):
        full = ROOT / p
        if full.exists():
            ok(p)
        else:
            print(f"  ✗ {p}")
            issue(f"vendor 누락 — xlsx/md 미리보기 불가: {p}")

    for p in (PORTAL_STATS, PORTAL_KA_STATS):
        if p.exists():
            ok(p.name)
        else:
            issue(f"집계 JS 없음: {p.relative_to(ROOT)}")

    section("3. 탭·드롭다운 ID ↔ TABS 정의")
    tab_ids = set(re.findall(r'data-tab="([^"]+)"', html))
    # 동적 id 패턴
    for prefix in ("gong-", "jimy-"):
        tab_ids.update({f"{prefix}{x}" for x in ("01t", "01j", "02", "03", "04", "05", "06", "07", "geunil", "heesang")})
    tab_ids.update(
        {
            "schedule", "schedule-togong", "um-ka", "report",
            "cost-hw", "cost-jin", "cost-tok", "cost-jo", "cost-jigu", "cost-waste",
            "gong", "jimy",
        }
    )
    # TABS keys from source patterns
    tabs_defined = set(re.findall(r'TABS\["([^"]+)"\]', html))
    tabs_defined.update(re.findall(r'data-tab="' + r'([^"]+)"', html))
    # static tabs in TABS = { ... }
    for m in re.finditer(r"^\s+(\w[\w-]*):\s*\{", html, re.M):
        key = m.group(1)
        if key not in ("label", "refTag", "main", "refs", "placeholder", "dropGroup", "group", "navId", "items"):
            if key in tab_ids or key.startswith(("cost-", "gong", "schedule", "um", "report", "jimy", "jinmyeong", "toji", "env", "plan", "permit")):
                tabs_defined.add(key)
    orphan_nav = sorted(tab_ids - tabs_defined)
    if orphan_nav:
        for tid in orphan_nav:
            warn(f"nav data-tab '{tid}' — TABS[...] 정의 미확인(동적 생성 가능)")
    else:
        ok("주요 nav tab id 전부 TABS에 대응")

    section("4. _ka SSOT · portal_ka_stats.js")
    if not KA.exists():
        issue("_ka.xlsx 없음")
    else:
        wb = load_workbook(KA, data_only=True)
        ws = wb["총공사비"]
        direct = ws.cell(12, 7).value  # 합계 직접공사비
        contract_row = None
        for r in range(1, 30):
            if ws.cell(r, 1).value == "도급액":
                contract_row = r
                break
        contract_sum = ws.cell(contract_row, 8).value if contract_row else None
        wb.close()
        ka_js = parse_js_obj(PORTAL_KA_STATS) if PORTAL_KA_STATS.exists() else {}
        live_pending = count_pending_live()
        print(f"  _ka 총공사비 합계 직접(G12): {direct:,}" if isinstance(direct, (int, float)) else f"  직접: {direct}")
        print(f"  _ka 도급 합계열: {contract_sum:,}" if isinstance(contract_sum, (int, float)) else contract_sum)
        print(f"  portal_ka_stats directWon: {ka_js.get('directWon')}")
        print(f"  portal_ka_stats contractWon: {ka_js.get('contractWon')}")
        print(f"  live 미확정: 검토 {live_pending['rv']} / 미매칭 {live_pending['um']} = {live_pending['all']}")
        print(f"  portal_ka_stats pending: rv={ka_js.get('review')} um={ka_js.get('unmatched')} all={ka_js.get('pendingTotal')}")

        if isinstance(direct, (int, float)) and ka_js.get("directWon"):
            diff = abs(int(direct) - int(ka_js["directWon"]))
            if diff <= 2:
                ok(f"_ka 직접공사비 ↔ portal_ka_stats ({int(direct):,}원)")
            else:
                issue(f"_ka 직접공사비 불일치: xlsx {int(direct):,} vs JS {ka_js['directWon']:,} (차 {diff:,})")
        if live_pending["all"] == ka_js.get("pendingTotal"):
            ok(f"미확정 {live_pending['all']}건 — portal_ka_stats 일치")
        else:
            issue(
                f"미확정 건수 불일치: live {live_pending['all']} vs portal_ka_stats {ka_js.get('pendingTotal')}"
            )

        m20 = re.search(r"미매칭\s*<strong>(\d+)</strong>", PENDING_HTML.read_text(encoding="utf-8"))
        html_um = int(m20.group(1)) if m20 else None
        if html_um == live_pending["all"]:
            ok(f"미확정 HTML 본문 {html_um}건 — live 일치")
        else:
            issue(f"미확정 HTML {html_um} vs live {live_pending['all']}")

        portal_claim = re.search(r"검토·미매칭\s*(\d+)건", html)
        if portal_claim:
            claimed = int(portal_claim.group(1))
            if claimed != live_pending["all"]:
                issue(f"포털 종합보고 refs '검토·미매칭 {claimed}건' ≠ live {live_pending['all']}")
            else:
                ok(f"포털 refs 미확정 {claimed}건 문구 일치")

    section("5. 총괄표 portal_stats.js")
    ps = parse_js_obj(PORTAL_STATS) if PORTAL_STATS.exists() else {}
    md_common = ROOT / "05_내역서" / "내역서작업" / "_공통" / "총괄표.md"
    md_root = ROOT / "05_내역서" / "총괄표.md"
    for label, md in [("_공통/총괄표.md", md_common), ("05_내역서/총괄표.md", md_root)]:
        if md.exists() and ps.get("directWon"):
            m = re.search(r"([\d,]+)\s*원", md.read_text(encoding="utf-8")[:3000])
            if m:
                md_won = int(m.group(1).replace(",", ""))
                if md_won == ps["directWon"]:
                    ok(f"{label} 직접공사비 ↔ portal_stats ({md_won:,}원)")
                else:
                    warn(f"{label} 직접 {md_won:,} vs portal_stats {ps['directWon']:,}")

    readme = ROOT / "README.md"
    if readme.exists() and ps.get("directWon"):
        rm = re.search(r"([\d,]+)원\s*\(\s*[\d.]+\s*억\s*\)", readme.read_text(encoding="utf-8"))
        if rm:
            rw = int(rm.group(1).replace(",", ""))
            if rw != ps["directWon"]:
                warn(f"README 직접공사비 {rw:,} ≠ portal_stats {ps['directWon']:,} (README 갱신 필요)")
            else:
                ok("README ↔ portal_stats 직접공사비")

    section("6. 종합보고서 HTML ↔ _ka")
    if REPORT_HTML.exists() and KA.exists():
        rep = REPORT_HTML.read_text(encoding="utf-8")
        ka_js = parse_js_obj(PORTAL_KA_STATS)
        for key, pat in [
            ("direct", r"제출 최종.*?([\d,]+)\s*원"),
            ("dogeup", r"도급.*?([\d,]+)\s*원"),
        ]:
            m = re.search(pat, rep[:8000])
            if m and ka_js.get(f"{key if key != 'dogeup' else 'contract'}Won"):
                rep_won = int(m.group(1).replace(",", ""))
                js_key = "directWon" if key == "direct" else "contractWon"
                if abs(rep_won - ka_js[js_key]) > 1000:
                    warn(f"종합보고서 {key} {rep_won:,} vs portal_ka {ka_js[js_key]:,}")
                else:
                    ok(f"종합보고서 {key} 금액 ↔ _ka JS")

    section("7. _ka 시트 · 공종 HTML")
    cost_html = [
        ("cost-hw", "회전교차로", "회전교차로_직접간접비.html"),
        ("cost-jin", "진입도로", "진입도로_직접간접비.html"),
        ("cost-tok", "토목", "토목_직접간접비.html"),
        ("cost-jo", "조경", "조경_직접간접비.html"),
        ("cost-jigu", "토목", "지구단위_직접간접비.html"),
        ("cost-waste", "폐기물", "폐기물_직접간접비.html"),
    ]
    if KA.exists():
        wb = load_workbook(KA, read_only=True)
        sheets = set(wb.sheetnames)
        wb.close()
        for _, gj, fname in cost_html:
            p = ROOT / "05_내역서" / "내역서작업" / gj / fname
            if p.exists():
                ok(f"{fname}")
            else:
                issue(f"공종 HTML 없음: {p.relative_to(ROOT)}")
        for sn in ("총공사비", "회전교차로", "진입도로", "토목(조경)", "지구단위", "폐기물"):
            if sn in sheets:
                ok(f"_ka 시트 「{sn}」")
            else:
                issue(f"_ka 시트 누락: {sn}")

    section("8. deploy 버전 · 초기 iframe")
    dep_comment = re.search(r"deploy-sync:\s*main@([0-9a-f]+)", html)
    dep_js = re.search(r'DEPLOY_V\s*=\s*"([^"]+)"', html)
    if dep_comment and dep_js and dep_comment.group(1) == dep_js.group(1):
        ok(f"DEPLOY_V = {dep_js.group(1)} (주석 일치)")
    elif dep_comment and dep_js:
        warn(f"deploy-sync {dep_comment.group(1)} ≠ DEPLOY_V {dep_js.group(1)}")

    iframe_src = re.search(r'id="main-frame"[^>]+src="([^"]+)"', html)
    if iframe_src:
        init = iframe_src.group(1)
        if (ROOT / init).exists():
            ok(f"초기 iframe: {init}")
        else:
            issue(f"초기 iframe 누락: {init}")

    section("9. 경로·SSOT 주의")
    # 총괄표.md 두 위치
    if md_common.exists() and md_root.exists():
        warn("총괄표.md가 05_내역서/ 와 _공통/ 두 곳에 존재 — 포털은 _공통/ 참조")
    # 조경 표준단가산출 중복
    jo_dup = [
        ROOT / "05_내역서" / "내역서작업" / "01_화성 청원지구 조경_표준단가산출.xlsx",
        ROOT / "05_내역서" / "내역서작업" / "조경" / "01_화성 청원지구 조경_표준단가산출.xlsx",
    ]
    if all(p.exists() for p in jo_dup):
        warn("조경 표준단가산출 xlsx — 내역서작업/ 루트·조경/ 하위 중복 (빌더는 조경/ 사용)")

    section("10. 요약")
    print(f"  ✓ OK   {len(oks)}건")
    print(f"  △ 주의 {len(warns)}건")
    print(f"  ✗ 이슈 {len(issues)}건")
    if warns:
        print("\n[주의]")
        for w in warns:
            print(f"  △ {w}")
    if issues:
        print("\n[이슈 — 조치 필요]")
        for i in issues:
            print(f"  ✗ {i}")
        return 1
    print("\n전수조사: 치명 이슈 없음" if not issues else "")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
