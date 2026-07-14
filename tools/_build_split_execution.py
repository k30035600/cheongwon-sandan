# -*- coding: utf-8 -*-
"""실행예산 통합견적서 → 3그룹 분리 실행내역서.

그룹:
  1) 토목조경    : 01. 토목·조경
  2) 회전진입    : 04. 진입도로 + 07. 건설폐기물(진입도로 공정 포함) + 05. 회전교차로
  3) 지구단위    : 06. 개발행위

각 그룹 시트에 직접공사비 집계 + 간접공사비(요율·산식 포함, 엑셀 수식) + 부가세 + 검산.
추가 시트: 표지, 대공정별 집계표.
간접비 요율은 원본과 동일(전체 재현 시 21.69억 일치 검증됨).
"""
import sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

SRC = (r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
       r"\청원지구_김보영(토목,조경)_통합.xlsx")
OUT = (r"D:\OneDrive\Cursor\cheongwon\07_타견적\실행예산서"
       r"\청원지구_실행내역_분리.xlsx")

# 대공정 집계부 공종행(견적서 상단) : 공종행 리스트, 소계행
# 01은 토목(01T: 1~9공종)과 조경(01J: 10.조경공)으로 분리
AGG = {
    "01T": (list(range(9, 18)), None),   # 1.토공 ~ 9.지하저류조
    "01J": ([18], None),                 # 10. 조경공
    "04": (list(range(20, 26)), 26),
    "07": ([44], 45),
    "05": (list(range(27, 32)), 32),
    "06": (list(range(33, 43)), 43),
}
# 상세부 행 구간 (01T 토목 63~778, 01J 조경 779~888)
DET = {
    "01T": (63, 778),
    "01J": (779, 888),
    "04": (889, 1141),
    "05": (1142, 1365),
    "06": (1366, 1525),
    "07": (1526, 1528),
}
LABELS = {"01T": "01. 토목(1~9공종)", "01J": "01. 조경(10.조경공)",
          "04": "04. 진입도로", "05": "05. 회전교차로",
          "06": "06. 개발행위(지구단위)", "07": "07. 건설폐기물처리"}

GROUPS = [
    ("토목", "토목 실행내역서", ["01T"]),
    ("조경", "조경 실행내역서", ["01J"]),
    ("회전진입", "진입도로·건설폐기물·회전교차로 실행내역서", ["04", "07", "05"]),
    ("지구단위", "개발행위(지구단위) 실행내역서", ["06"]),
]

# 간접비 항목(라벨, 산식표기, 요율) — 순서 고정
IND = [
    ("간접노무비",             "직접노무비 × 19.7%",            0.197),
    ("산재보험료",             "노무비계 × 3.56%",              0.0356),
    ("고용보험료",             "노무비계 × 1.15%",              0.0115),
    ("국민건강보험료",         "직접노무비 × 3.595%",           0.03595),
    ("국민연금보험료",         "직접노무비 × 4.75%",            0.0475),
    ("노인장기요양보험료",     "건강보험료 × 13.14%",           0.1314),
    ("산업안전보건관리비",     "(재료비+직접노무비) × 2.60%",   0.026),
    ("기타경비(공과잡비)",     "(재료비+노무비계) × 6.2%",      0.062),
    ("건설기계대여대금지급보증", "직접공사비 × 0.40%",           0.004),
    ("환경보전비",             "직접공사비 × 0.60%",            0.006),
    ("일반관리비",             "순공사원가 × 5%",               0.05),
    ("이윤",                   "(노무비계+경비계+일반관리비) × 7%", 0.07),
]

navy = "1F3864"
head = PatternFill("solid", fgColor="1F3864")
band = PatternFill("solid", fgColor="D9E1F2")
sfill = PatternFill("solid", fgColor="FFF2CC")
gfill = PatternFill("solid", fgColor="E2EFDA")
redf = PatternFill("solid", fgColor="FCE4EC")
thin = Side(style="thin", color="BFBFBF")
med = Side(style="medium", color="808080")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
FN = "맑은 고딕"
NF = "#,##0"
COLS = 13


def R(x):
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def calc_indirect(jae, no, gyeong, direct):
    """원본과 동일 요율로 간접비 산출(원 단위 반올림)."""
    gn = R(no * 0.197)
    nogye = no + gn
    sj = R(nogye * 0.0356)
    gy = R(nogye * 0.0115)
    gg = R(no * 0.03595)
    yg = R(no * 0.0475)
    yy = R(gg * 0.1314)
    sa = R((jae + no) * 0.026)
    gt = R((jae + nogye) * 0.062)
    gm = R(direct * 0.004)
    hb = R(direct * 0.006)
    sub10 = gn + sj + gy + gg + yg + yy + sa + gt + gm + hb
    sun = direct + sub10
    ilban = R(sun * 0.05)
    gyeonggye = gyeong + (sub10 - gn)
    iyun = R((nogye + gyeonggye + ilban) * 0.07)
    vals = [gn, sj, gy, gg, yg, yy, sa, gt, gm, hb, ilban, iyun]
    return vals, sum(vals)


def st(ws, r, c, v, *, bold=False, color="000000", size=10, fill=None,
       align="left", nf=None, border=True, wrap=False):
    cell = ws.cell(r, c)
    cell.value = v
    cell.font = Font(name=FN, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if nf:
        cell.number_format = nf
    if border:
        cell.border = bd
    return cell


def col_widths(ws):
    for i, w in enumerate([30, 24, 8, 6, 12, 14, 12, 14, 12, 14, 12, 15, 12]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def two_row_header(ws, r):
    hdr1 = ["공종/품명", "규격", "수량", "단위",
            "재료비", "", "노무비", "", "경비", "", "합계", "", "비고"]
    hdr2 = ["", "", "", "", "단가", "금액", "단가", "금액",
            "단가", "금액", "단가", "금액", ""]
    for i, h in enumerate(hdr1):
        st(ws, r, i + 1, h, bold=True, color="FFFFFF", fill=head, align="center")
    for i, h in enumerate(hdr2):
        if h:
            st(ws, r + 1, i + 1, h, bold=True, color="FFFFFF", fill=head, align="center")
        else:
            ws.cell(r + 1, i + 1).fill = head
            ws.cell(r + 1, i + 1).border = bd
    for cc in (1, 2, 3, 4, 13):
        ws.merge_cells(start_row=r, start_column=cc, end_row=r + 1, end_column=cc)
    for base in (5, 7, 9, 11):
        ws.merge_cells(start_row=r, start_column=base, end_row=r, end_column=base + 1)
    return r + 2


# ---------------------------------------------------------------------------
def build_group_sheet(wb_out, ws_src, name, subtitle, keys, gdata):
    ws = wb_out.create_sheet(name)
    col_widths(ws)

    ws.merge_cells(f"A1:M1")
    st(ws, 1, 1, f"화성 청원지구 산업유통형 지구단위계획 — {subtitle}",
       bold=True, color=navy, size=14, align="center", border=False)
    ws.merge_cells(f"A2:M2")
    note = " + ".join(LABELS[k] for k in keys)
    st(ws, 2, 1, f"대상 공정 : {note}", color="595959", size=9,
       align="center", border=False)

    r = two_row_header(ws, 4)

    # 직접공사비 집계부(대공정 소계행)
    subtotal_rows = []
    for k in keys:
        jae = no = gy = tot = 0
        # 원본이 수식화되어 금액(F/H/J/L)이 캐시 없음 → 단가(E/G/I) 사용
        # 집계 공종행은 1식(수량=1)이라 단가=금액
        for cr in AGG[k][0]:
            jae += ws_src.cell(cr, 5).value or 0
            no += ws_src.cell(cr, 7).value or 0
            gy += ws_src.cell(cr, 9).value or 0
            tot += (ws_src.cell(cr, 5).value or 0) + (ws_src.cell(cr, 7).value or 0) \
                + (ws_src.cell(cr, 9).value or 0)
        st(ws, r, 1, LABELS[k], bold=True, color=navy, fill=band)
        st(ws, r, 2, "1식", align="center")
        st(ws, r, 3, 1, align="center")
        st(ws, r, 4, "식", align="center")
        # 수량 1식 → 단가=금액
        st(ws, r, 5, jae, nf=NF, align="right")
        st(ws, r, 6, jae, nf=NF, align="right")
        st(ws, r, 7, no, nf=NF, align="right")
        st(ws, r, 8, no, nf=NF, align="right")
        st(ws, r, 9, gy, nf=NF, align="right")
        st(ws, r, 10, gy, nf=NF, align="right")
        st(ws, r, 11, tot, nf=NF, align="right")
        st(ws, r, 12, tot, nf=NF, align="right", bold=True)
        subtotal_rows.append(r)
        r += 1

    # Ⅰ. 직접공사비 계
    dr = r
    st(ws, dr, 1, "Ⅰ. 직접공사비 계", bold=True, color="FFFFFF", fill=head)
    for base in (6, 8, 10, 12):
        cl = get_column_letter(base)
        refs = "+".join(f"{cl}{sr}" for sr in subtotal_rows)
        st(ws, dr, base, f"={refs}", bold=True, color="FFFFFF", fill=head,
           nf=NF, align="right")
    r += 1

    # 간접공사비 (수식)
    st(ws, r, 1, "Ⅱ. 간접공사비", bold=True, color=navy, fill=gfill)
    st(ws, r, 2, "산식", bold=True, color=navy, fill=gfill, align="center")
    r += 1
    ind_first = r
    ind_rows = {}
    for label, formula_txt, rate in IND:
        st(ws, r, 1, "　" + label, align="left")
        st(ws, r, 2, formula_txt, color="595959", size=9, align="left")
        ind_rows[label] = r
        r += 1
    ind_last = r - 1

    F, H, J, L = f"F{dr}", f"H{dr}", f"J{dr}", f"L{dr}"
    rn = ind_rows
    gn = f"J{rn['간접노무비']}"
    nogye = f"({H}+{gn})"          # 노무비계
    # 각 항목 J(경비) 수식, L=J
    fx = {
        "간접노무비":            f"=ROUND({H}*0.197,0)",
        "산재보험료":            f"=ROUND({nogye}*0.0356,0)",
        "고용보험료":            f"=ROUND({nogye}*0.0115,0)",
        "국민건강보험료":        f"=ROUND({H}*0.03595,0)",
        "국민연금보험료":        f"=ROUND({H}*0.0475,0)",
        "노인장기요양보험료":    f"=ROUND(J{rn['국민건강보험료']}*0.1314,0)",
        "산업안전보건관리비":    f"=ROUND(({F}+{H})*0.026,0)",
        "기타경비(공과잡비)":    f"=ROUND(({F}+{nogye})*0.062,0)",
        "건설기계대여대금지급보증": f"=ROUND({L}*0.004,0)",
        "환경보전비":            f"=ROUND({L}*0.006,0)",
    }
    sub10_ref = f"SUM(J{rn['간접노무비']}:J{rn['환경보전비']})"
    fx["일반관리비"] = f"=ROUND(({L}+{sub10_ref})*0.05,0)"
    # 경비계 = 직접경비 + (간접10 - 간접노무비)
    gyeonggye = f"({J}+{sub10_ref}-{gn})"
    fx["이윤"] = f"=ROUND(({nogye}+{gyeonggye}+J{rn['일반관리비']})*0.07,0)"
    for label, _t, _r in IND:
        rr = rn[label]
        st(ws, rr, 10, fx[label], nf=NF, align="right")
        st(ws, rr, 12, f"=J{rr}", nf=NF, align="right")

    # Ⅱ. 간접공사비 계
    r = ind_last + 1
    ir = r
    st(ws, ir, 1, "Ⅱ. 간접공사비 계", bold=True, color="FFFFFF", fill=head)
    st(ws, ir, 10, f"=SUM(J{ind_first}:J{ind_last})", bold=True,
       color="FFFFFF", fill=head, nf=NF, align="right")
    st(ws, ir, 12, f"=SUM(L{ind_first}:L{ind_last})", bold=True,
       color="FFFFFF", fill=head, nf=NF, align="right")
    r += 1

    # 합 계
    hr = r
    st(ws, hr, 1, "합 계 (직접+간접)", bold=True, fill=sfill)
    st(ws, hr, 6, f"={F}", nf=NF, align="right", bold=True, fill=sfill)
    st(ws, hr, 8, f"={H}", nf=NF, align="right", bold=True, fill=sfill)
    st(ws, hr, 10, f"={J}+J{ir}", nf=NF, align="right", bold=True, fill=sfill)
    st(ws, hr, 12, f"={L}+L{ir}", nf=NF, align="right", bold=True, fill=sfill)
    r += 1

    # 부가가치세
    vr = r
    st(ws, vr, 1, "부가가치세", bold=True)
    st(ws, vr, 2, "합계 × 10%", color="595959", size=9)
    st(ws, vr, 10, f"=ROUND(L{hr}*0.1,0)", nf=NF, align="right")
    st(ws, vr, 12, f"=J{vr}", nf=NF, align="right")
    r += 1

    # 공사비 계(도급액)
    tr = r
    st(ws, tr, 1, "공사비 계 (도급액)", bold=True, color="FFFFFF", fill=head)
    st(ws, tr, 12, f"=L{hr}+L{vr}", bold=True, color="FFFFFF", fill=head,
       nf=NF, align="right")
    r += 2

    # ── 검산 ──
    st(ws, r, 1, "【 검산 】 파이썬 산출값과 엑셀 수식 대사", bold=True,
       color="C00000", size=11, border=False)
    r += 1
    jae = sum((ws_src.cell(cr, 5).value or 0)
              for k in keys for cr in AGG[k][0])
    no = sum((ws_src.cell(cr, 7).value or 0)
             for k in keys for cr in AGG[k][0])
    gy = sum((ws_src.cell(cr, 9).value or 0)
             for k in keys for cr in AGG[k][0])
    direct = jae + no + gy
    ivals, itot = calc_indirect(jae, no, gy, direct)
    gdata.update(dict(jae=jae, no=no, gy=gy, direct=direct,
                      ivals=ivals, itot=itot))
    chk = [
        ("직접공사비 계", direct, f"=L{dr}"),
        ("간접공사비 계", itot, f"=L{ir}"),
        ("합계(직접+간접)", direct + itot, f"=L{hr}"),
        ("부가가치세", R((direct + itot) * 0.1), f"=L{vr}"),
        ("공사비 계", (direct + itot) + R((direct + itot) * 0.1), f"=L{tr}"),
    ]
    st(ws, r, 1, "항목", bold=True, fill=band, align="center")
    st(ws, r, 2, "파이썬 검산값", bold=True, fill=band, align="center")
    st(ws, r, 6, "엑셀 수식결과", bold=True, fill=band, align="center")
    st(ws, r, 10, "차이", bold=True, fill=band, align="center")
    r += 1
    for lab, pv, exf in chk:
        st(ws, r, 1, lab)
        st(ws, r, 2, pv, nf=NF, align="right")
        st(ws, r, 6, exf, nf=NF, align="right")
        st(ws, r, 10, f"=F{r}-B{r}", nf=NF, align="right")
        r += 1
    r += 1

    # ── 상세내역 ──
    st(ws, r, 1, "【 상세 내역 】", bold=True, color=navy, size=12, border=False)
    r += 1
    for k in keys:
        st(ws, r, 1, LABELS[k], bold=True, color=navy, fill=band)
        r += 1
        lo, hi = DET[k]
        for cr in range(lo, hi + 1):
            for c in range(1, COLS + 1):
                sc = ws_src.cell(cr, c)
                dc = ws.cell(r, c)
                dc.value = sc.value
                if sc.number_format and sc.number_format != "General":
                    dc.number_format = sc.number_format
                dc.font = Font(name=FN, size=9)
                dc.alignment = Alignment(vertical="center", wrap_text=True)
                dc.border = bd
            # 원본 단가(E/G/I)는 상수로 복사됨 → 금액을 단가×수량 수식으로 재생성
            # (원본 금액 셀은 수식이라 data_only에서 None으로 복사됨)
            q = ws_src.cell(cr, 3).value
            if isinstance(q, (int, float)) and q > 0:
                for up_c, amt_c in [(5, 6), (7, 8), (9, 10)]:
                    mc = ws.cell(r, amt_c)
                    mc.value = f"={get_column_letter(up_c)}{r}*C{r}"
                    mc.number_format = NF
                    mc.alignment = Alignment(horizontal="right", vertical="center")
                kc = ws.cell(r, 11)
                kc.value = f"=E{r}+G{r}+I{r}"
                kc.number_format = NF
                kc.alignment = Alignment(horizontal="right", vertical="center")
                lc = ws.cell(r, 12)
                lc.value = f"=F{r}+H{r}+J{r}"
                lc.number_format = NF
                lc.alignment = Alignment(horizontal="right", vertical="center")
            r += 1
        r += 1

    ws.freeze_panes = "A6"
    return dr, ir, hr, vr, tr


# ---------------------------------------------------------------------------
def build_summary_sheet(wb_out, gstats):
    ws = wb_out.create_sheet("대공정별집계표", 1)
    col_widths(ws)
    ws.merge_cells("A1:M1")
    st(ws, 1, 1, "화성 청원지구 — 대공정별 집계표 (실행)",
       bold=True, color=navy, size=15, align="center", border=False)
    ws.merge_cells("A2:M2")
    st(ws, 2, 1, "07. 건설폐기물은 진입도로(04) 공정에 포함",
       color="595959", size=9, align="center", border=False)

    r = 4
    hdrs = ["구분", "대공정", "재료비", "노무비", "경비", "직접공사비",
            "간접공사비", "합계(직+간)", "부가세", "도급액(합계+VAT)"]
    cmap = [1, 2, 3, 4, 5, 6, 8, 9, 11, 12]
    for h, c in zip(hdrs, cmap):
        st(ws, r, c, h, bold=True, color="FFFFFF", fill=head, align="center")
    r += 1

    # 대공정 단위(직접비만) + 그룹 소계(간접·합계·부가세·도급)
    src = gstats["_perproc"]
    tj = tn = tg = td = ti = 0
    for gi, (gname, keys) in enumerate([("토목", ["01T"]),
                                        ("조경", ["01J"]),
                                        ("회전진입", ["04", "07", "05"]),
                                        ("지구단위", ["06"])]):
        gs = gstats[gname]
        first = r
        for k in keys:
            p = src[k]
            st(ws, r, 1, gname if k == keys[0] else "")
            st(ws, r, 2, LABELS[k])
            st(ws, r, 3, p["jae"], nf=NF, align="right")
            st(ws, r, 4, p["no"], nf=NF, align="right")
            st(ws, r, 5, p["gy"], nf=NF, align="right")
            st(ws, r, 6, p["tot"], nf=NF, align="right")
            r += 1
        # 그룹 소계
        st(ws, r, 1, "", fill=gfill)
        st(ws, r, 2, f"[{gname} 소계]", bold=True, fill=gfill)
        st(ws, r, 3, f"=SUM(C{first}:C{r-1})", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 4, f"=SUM(D{first}:D{r-1})", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 5, f"=SUM(E{first}:E{r-1})", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 6, f"=SUM(F{first}:F{r-1})", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 8, gs["itot"], nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 9, f"=F{r}+H{r}", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 11, f"=ROUND(I{r}*0.1,0)", nf=NF, align="right", bold=True, fill=gfill)
        st(ws, r, 12, f"=I{r}+K{r}", nf=NF, align="right", bold=True, fill=gfill)
        tj += gs["jae"]; tn += gs["no"]; tg += gs["gy"]; td += gs["direct"]; ti += gs["itot"]
        r += 1

    # 총계 (조경 포함) — 4그룹 전체
    st(ws, r, 1, "", fill=sfill)
    st(ws, r, 2, "총   계 (조경 포함)", bold=True, color="C00000", fill=sfill)
    st(ws, r, 3, tj, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 4, tn, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 5, tg, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 6, td, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 8, ti, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 9, f"=F{r}+H{r}", nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 11, f"=ROUND(I{r}*0.1,0)", nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 12, f"=I{r}+K{r}", nf=NF, align="right", bold=True, fill=sfill)
    grand = r
    r += 1

    # 총계 (조경 제외) — 조경 그룹 차감
    gj = gstats["조경"]
    st(ws, r, 1, "", fill=band)
    st(ws, r, 2, "총   계 (조경 제외)", bold=True, color=navy, fill=band)
    st(ws, r, 3, tj - gj["jae"], nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 4, tn - gj["no"], nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 5, tg - gj["gy"], nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 6, td - gj["direct"], nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 8, ti - gj["itot"], nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 9, f"=F{r}+H{r}", nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 11, f"=ROUND(I{r}*0.1,0)", nf=NF, align="right", bold=True, fill=band)
    st(ws, r, 12, f"=I{r}+K{r}", nf=NF, align="right", bold=True, fill=band)
    r += 2

    # 검산: 원본 전체값 대사
    st(ws, r, 1, "【 검산 】 원본 통합견적서 대비", bold=True, color="C00000",
       size=11, border=False)
    r += 1
    ORIG = {"직접공사비": 9179545248, "간접공사비": 2168774231,
            "합계": 11348319479, "부가세": 1134831948, "공사비계": 12483151427}
    st(ws, r, 1, "항목", bold=True, fill=band, align="center")
    st(ws, r, 2, "원본값", bold=True, fill=band, align="center")
    st(ws, r, 3, "4그룹 합(조경포함)", bold=True, fill=band, align="center")
    st(ws, r, 5, "차이(반올림)", bold=True, fill=band, align="center")
    r += 1
    rows = [
        ("직접공사비", ORIG["직접공사비"], f"=F{grand}"),
        ("간접공사비", ORIG["간접공사비"], f"=H{grand}"),
        ("합계(직+간)", ORIG["합계"], f"=I{grand}"),
        ("부가가치세", ORIG["부가세"], f"=K{grand}"),
        ("공사비 계", ORIG["공사비계"], f"=L{grand}"),
    ]
    for lab, ov, exf in rows:
        st(ws, r, 1, lab)
        st(ws, r, 2, ov, nf=NF, align="right")
        st(ws, r, 3, exf, nf=NF, align="right")
        st(ws, r, 5, f"=C{r}-B{r}", nf=NF, align="right")
        r += 1
    st(ws, r, 1, "※ 간접비는 그룹별 독립 산출→반올림 오차(±수원) 발생 가능",
       color="595959", size=9, border=False)
    ws.freeze_panes = "A5"


def build_cover(wb_out, gstats):
    ws = wb_out.create_sheet("표지", 0)
    for i, w in enumerate([4, 26, 20, 18, 18, 18, 4]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:F2")
    st(ws, 2, 2, "화성 청원지구 산업유통형 지구단위계획", bold=True,
       color=navy, size=16, align="center", border=False)
    ws.merge_cells("B3:F3")
    st(ws, 3, 2, "실 행 내 역 서 (3분할)", bold=True, color="000000",
       size=20, align="center", border=False)
    ws.merge_cells("B4:F4")
    st(ws, 4, 2, "─ 도급 통합견적서 기준 실행 분리 ─", color="595959",
       size=10, align="center", border=False)

    # 그룹 요약표
    r = 7
    hdrs = ["구분", "대상 공정", "직접공사비", "간접공사비", "도급액(VAT포함)"]
    for i, h in enumerate(hdrs):
        st(ws, r, 2 + i, h, bold=True, color="FFFFFF", fill=head, align="center")
    r += 1
    mapping = [("토목", "01. 토목(1~9공종)"),
               ("조경", "01. 조경(10.조경공)"),
               ("회전진입", "04.진입도로+07.폐기물+05.회전교차로"),
               ("지구단위", "06. 개발행위")]
    for gname, desc in mapping:
        gs = gstats[gname]
        direct = gs["direct"]; itot = gs["itot"]
        hab = direct + itot
        vat = R(hab * 0.1)
        st(ws, r, 2, gname, bold=True, color=navy)
        st(ws, r, 3, desc, size=9)
        st(ws, r, 4, direct, nf=NF, align="right")
        st(ws, r, 5, itot, nf=NF, align="right")
        st(ws, r, 6, hab + vat, nf=NF, align="right", bold=True)
        r += 1
    # 합계
    td = sum(gstats[g]["direct"] for g, _ in mapping)
    ti = sum(gstats[g]["itot"] for g, _ in mapping)
    thab = td + ti
    tvat = R(thab * 0.1)
    st(ws, r, 2, "합 계", bold=True, color="C00000", fill=sfill)
    st(ws, r, 3, "3개 그룹", size=9, fill=sfill)
    st(ws, r, 4, td, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 5, ti, nf=NF, align="right", bold=True, fill=sfill)
    st(ws, r, 6, thab + tvat, nf=NF, align="right", bold=True, fill=sfill)
    r += 3

    for line in [
        f"공급가액 합계 : {thab:,} 원",
        f"부  가  세    : {tvat:,} 원",
        f"도  급  액    : {thab + tvat:,} 원",
    ]:
        ws.merge_cells(f"B{r}:F{r}")
        st(ws, r, 2, line, bold=True, size=11, border=False)
        r += 1
    r += 2
    ws.merge_cells(f"B{r}:F{r}")
    st(ws, r, 2, "작성일 : 2026. 7. 9.", size=10, align="right", border=False)
    r += 1
    ws.merge_cells(f"B{r}:F{r}")
    st(ws, r, 2, "※ 07. 건설폐기물은 진입도로(04) 공정에 포함",
       color="595959", size=9, align="right", border=False)


def main():
    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    ws_src = wb_src["견적서"]

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    gstats = {"_perproc": {}}
    # 대공정별 직접비(집계표용) — 원본 수식화로 금액 없음 → 단가(수량1)로
    for k, (rows, subr) in AGG.items():
        jae = sum((ws_src.cell(cr, 5).value or 0) for cr in rows)
        no = sum((ws_src.cell(cr, 7).value or 0) for cr in rows)
        gy = sum((ws_src.cell(cr, 9).value or 0) for cr in rows)
        gstats["_perproc"][k] = dict(jae=jae, no=no, gy=gy, tot=jae + no + gy)

    for name, subtitle, keys in GROUPS:
        gd = {}
        build_group_sheet(wb_out, ws_src, name, subtitle, keys, gd)
        gstats[name] = gd
        print(f"【{name}】 직접 {gd['direct']:,} · 간접 {gd['itot']:,} · "
              f"도급 {gd['direct'] + gd['itot'] + R((gd['direct']+gd['itot'])*0.1):,}")

    build_summary_sheet(wb_out, gstats)
    build_cover(wb_out, gstats)

    # 시트 순서: 표지, 대공정별집계표, 4그룹
    order = ["표지", "대공정별집계표", "토목", "조경", "회전진입", "지구단위"]
    wb_out._sheets.sort(key=lambda s: order.index(s.title))

    # 전체 검산 요약
    grp = ("토목", "조경", "회전진입", "지구단위")
    td = sum(gstats[g]["direct"] for g in grp)
    ti = sum(gstats[g]["itot"] for g in grp)
    print(f"\n[검산] 4그룹 직접합 {td:,} (원본 9,179,545,248 / 차 {td-9179545248})")
    print(f"[검산] 4그룹 간접합 {ti:,} (원본 2,168,774,231 / 차 {ti-2168774231})")

    wb_out.save(OUT)
    print("\n저장:", OUT)


if __name__ == "__main__":
    main()
