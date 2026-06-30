#!/usr/bin/env python3
"""제출 _ka.xlsx → 청원지구_종합보고서.html · portal_ka_stats.js 자동 생성."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
TOOLS = Path(__file__).resolve().parent
if str(TOOLS) not in sys.path:
    sys.path.insert(0, str(TOOLS))

from naeyeok_gongjong import (  # noqa: E402
    classify_ka_mihwakjeong,
    detail_row_has_sum,
    is_ka_gayeonghyeon,
    is_ka_status_excluded,
    is_ka_pending_excluded,
)
from report_gongjong_qty import collect_gongjong_summary, fmt_cell, fmt_num  # noqa: E402

KA_FILE = "08_제출내역서/청원지구_단가통합(전기제외)_ka.xlsx"
KA = ROOT / "08_제출내역서" / "청원지구_단가통합(전기제외)_ka.xlsx"
SCHEDULE_HTML = ROOT / "08_제출내역서" / "청원지구_공사공정표.html"
OUT = ROOT / "05_내역서" / "내역서작업" / "_공통" / "청원지구_종합보고서.html"
KA_STATS_JS = ROOT / "05_내역서" / "내역서작업" / "_공통" / "portal_ka_stats.js"
GEN_DATE = "2026. 6. 29."

# 계약차수(제출 통합본 시트 구분)
CONTRACT_PHASES = [
    {"sheet": "회전교차로", "star": "★ 회전교차로 합계", "label": "회전교차로", "no": "05", "dogeup_cols": (2,)},
    {"sheet": "진입도로", "star": "★ 진입도로 합계", "label": "진입도로", "no": "04", "dogeup_cols": (3,)},
    {"sheet": "토목(조경)", "star": "★ 토목(조경) 합계", "label": "토목·조경", "no": "01", "dogeup_cols": (4, 5)},
    {"sheet": "지구단위", "star": "★ 지구단위 합계", "label": "지구단위", "no": "06", "dogeup_cols": (6,)},
]
WASTE = {"sheet": "폐기물", "star": "★ 폐기물 합계", "label": "건설폐기물", "no": "07", "dogeup_cols": (7,)}

# 토목(조경) 시트 — 계약차수별 요약 분리
TJ_SPLIT = [
    {"star": "★ 01 토목 소계", "label": "토목", "file": "01 토목", "dogeup_col": 4},
    {"star": "★ 01 조경 소계", "label": "조경", "file": "01 조경", "dogeup_col": 5},
]


def won(v) -> str:
    return f"{int(round(float(v or 0))):,}"


def eok(v) -> str:
    return f"{float(v or 0) / 1e8:,.2f}"


def pct(part: float, whole: float) -> str:
    return f"{part / whole * 100:.1f}" if whole else "0.0"


def _num(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _find_star_row(ws, star_label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if v and str(v).strip() == star_label:
            return r
    return None


def _read_star(ws, row: int) -> dict:
    matched = int(_num(ws.cell(row, 4).value))
    total = int(_num(ws.cell(row, 5).value))
    mat = _num(ws.cell(row, 7).value)
    lab = _num(ws.cell(row, 8).value)
    exp = _num(ws.cell(row, 9).value)
    direct = _num(ws.cell(row, 10).value) or (mat + lab + exp)
    rate = matched / total if total else 0.0
    return {
        "matched": matched,
        "total": total,
        "rate": rate,
        "mat": mat,
        "lab": lab,
        "exp": exp,
        "direct": direct,
    }


def _read_sections(ws, hdr_row: int = 3) -> list[dict]:
    """상단 공종(차수) 요약만 — ★ 소계 이전, 내역 일괄(◆) 제외."""
    out: list[dict] = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        marker = ws.cell(r, 1).value
        if marker and str(marker).startswith("◆"):
            break
        name = ws.cell(r, 3).value
        if not name:
            if out:
                break
            continue
        name = str(name).strip()
        if name.startswith("★"):
            if not ws.cell(r, 1).value and "합계" in name:
                break
            continue
        no = ws.cell(r, 1).value
        if not no:
            continue
        matched = ws.cell(r, 4).value
        total = ws.cell(r, 5).value
        mat = _num(ws.cell(r, 7).value)
        lab = _num(ws.cell(r, 8).value)
        exp = _num(ws.cell(r, 9).value)
        direct = _num(ws.cell(r, 10).value) or (mat + lab + exp)
        if direct == 0 and mat == lab == exp == 0:
            if not matched and total in (0, None, 1):
                continue
        out.append({
            "no": str(no).strip(),
            "file": str(ws.cell(r, 2).value or "").strip(),
            "name": name,
            "matched": int(_num(matched)),
            "total": int(_num(total)),
            "mat": mat,
            "lab": lab,
            "exp": exp,
            "direct": direct,
        })
    return out


def _split_tj_summary(
    ws,
    cost_ws,
    base: dict,
    *,
    dogeup_row: int = 21,
) -> list[dict]:
    """토목(조경) 시트 — ★ 소계·총공사비 열(01 토목/01 조경) 기준 분리."""
    parts: list[dict] = []
    all_sections = base.get("sections") or []
    for sub in TJ_SPLIT:
        row = _find_star_row(ws, sub["star"])
        if row is None:
            raise ValueError(f"토목(조경) — '{sub['star']}' 행 없음")
        star = _read_star(ws, row)
        dogeup = _num(cost_ws.cell(dogeup_row, sub["dogeup_col"]).value)
        direct = star["direct"]
        parts.append({
            "sheet": base["sheet"],
            "label": sub["label"],
            "no": "01",
            **star,
            "dogeup": dogeup,
            "mult": dogeup / direct if direct else 0.0,
            "sections": [s for s in all_sections if s.get("file") == sub["file"]],
        })
    return parts


def _build_summary_phases(phases: list[dict], wb, cost_ws, *, dogeup_row: int = 21) -> list[dict]:
    """계약차수별 요약표용 — 토목/조경 분리 행."""
    out: list[dict] = []
    for p in phases:
        if p["sheet"] == "토목(조경)":
            out.extend(_split_tj_summary(wb[p["sheet"]], cost_ws, p, dogeup_row=dogeup_row))
        else:
            out.append(p)
    return out


def _detail_hdr_row(ws) -> int | None:
    for r in range(1, min(ws.max_row + 1, 800)):
        if str(ws.cell(r, 4).value or "").strip() == "공종명":
            return r
    return None


def _classify_status(st, *, has_sum: bool = False) -> str | None:
    """매칭(빈칸)·원본 제외. 검토·미매칭 중 합계 있으면 가영현."""
    if is_ka_status_excluded(st):
        return None
    s = str(st or "").strip()
    if is_ka_gayeonghyeon(s, has_sum=has_sum):
        return "gyh"
    if s in ("미매칭", "미산출"):
        return "um"
    if s == "검토":
        return "rv"
    return None


def _count_sheet_pending(ws, *, file_filter: str | None = None) -> dict:
    """◆ 내역 일괄 — 미확정(검토·미매칭만, 가영현 제외)."""
    hdr = _detail_hdr_row(ws)
    blank = {"rv": 0, "um": 0, "all": 0}
    if hdr is None:
        return blank
    counts = {"rv": 0, "um": 0, "all": 0}
    for r in range(hdr + 1, ws.max_row + 1):
        if not ws.cell(r, 3).value:
            continue
        naeyeok = str(ws.cell(r, 1).value or "").strip()
        if file_filter and naeyeok != file_filter:
            continue
        section = ws.cell(r, 2).value
        name = ws.cell(r, 4).value
        if not name or str(name).startswith("★"):
            continue
        if is_ka_pending_excluded(section, name):
            continue
        has_sum = detail_row_has_sum(
            mat_u=ws.cell(r, 8).value,
            lab_u=ws.cell(r, 10).value,
            exp_u=ws.cell(r, 12).value,
            mat_amt=ws.cell(r, 9).value,
            lab_amt=ws.cell(r, 11).value,
            exp_amt=ws.cell(r, 13).value,
            unit_sum=ws.cell(r, 14).value,
            total_amt=ws.cell(r, 15).value,
        )
        cat = classify_ka_mihwakjeong(ws.cell(r, 16).value, has_sum=has_sum)
        if cat == "검토":
            counts["rv"] += 1
        elif cat == "미매칭":
            counts["um"] += 1
        else:
            continue
        counts["all"] += 1
    return counts


def _count_sheet_status(ws, *, file_filter: str | None = None) -> dict:
    """◆ 내역 일괄 — 매칭 제외 · 검토·미매칭·가영현 집계."""
    hdr = _detail_hdr_row(ws)
    blank = {"rv": 0, "um": 0, "gyh": 0, "all": 0}
    if hdr is None:
        return blank
    counts = {"rv": 0, "um": 0, "gyh": 0, "all": 0}
    for r in range(hdr + 1, ws.max_row + 1):
        if not ws.cell(r, 3).value:
            continue
        naeyeok = str(ws.cell(r, 1).value or "").strip()
        if file_filter and naeyeok != file_filter:
            continue
        section = ws.cell(r, 2).value
        name = ws.cell(r, 4).value
        if not name or str(name).startswith("★"):
            continue
        if is_ka_pending_excluded(section, name):
            continue
        has_sum = detail_row_has_sum(
            mat_u=ws.cell(r, 8).value,
            lab_u=ws.cell(r, 10).value,
            exp_u=ws.cell(r, 12).value,
            mat_amt=ws.cell(r, 9).value,
            lab_amt=ws.cell(r, 11).value,
            exp_amt=ws.cell(r, 13).value,
            unit_sum=ws.cell(r, 14).value,
            total_amt=ws.cell(r, 15).value,
        )
        cat = _classify_status(ws.cell(r, 16).value, has_sum=has_sum)
        if not cat:
            continue
        counts[cat] += 1
        counts["all"] += 1
    return counts


def _attach_item_status(summary_phases: list[dict], wb) -> tuple[dict, dict]:
    """summary_phases 각 행 + 전체 합계 — _ka 내역 「상태」(매칭 제외) · 미확정(가영현 제외)."""
    tot = {"rv": 0, "um": 0, "gyh": 0, "all": 0}
    pending = {"rv": 0, "um": 0, "all": 0}
    for p in summary_phases:
        ws = wb[p["sheet"]]
        ff = f"01 {p['label']}" if p["sheet"] == "토목(조경)" else None
        st = _count_sheet_status(ws, file_filter=ff)
        pd = _count_sheet_pending(ws, file_filter=ff)
        p["item_status"] = st
        p["item_pending"] = pd
        for k in ("rv", "um", "gyh", "all"):
            tot[k] += st[k]
        for k in ("rv", "um", "all"):
            pending[k] += pd[k]
    return tot, pending


def load_ka() -> dict:
    wb = load_workbook(KA, data_only=True)
    cost_ws = wb["총공사비"]
    dogeup_row = 21
    phases: list[dict] = []

    for cfg in CONTRACT_PHASES + [WASTE]:
        ws = wb[cfg["sheet"]]
        row = _find_star_row(ws, cfg["star"])
        if row is None:
            raise ValueError(f"{cfg['sheet']} — '{cfg['star']}' 행 없음")
        star = _read_star(ws, row)
        dogeup = sum(_num(cost_ws.cell(dogeup_row, c).value) for c in cfg["dogeup_cols"])
        mult = dogeup / star["direct"] if star["direct"] else 0.0
        sections = _read_sections(ws)
        phases.append({
            **cfg,
            **star,
            "dogeup": dogeup,
            "mult": mult,
            "sections": sections,
        })

    summary_phases = _build_summary_phases(phases, wb, cost_ws, dogeup_row=dogeup_row)
    contract = phases[:4]
    waste = phases[4]
    tot = {
        "mat": sum(p["mat"] for p in phases),
        "lab": sum(p["lab"] for p in phases),
        "exp": sum(p["exp"] for p in phases),
        "direct": sum(p["direct"] for p in phases),
        "dogeup": sum(p["dogeup"] for p in phases),
        "matched": sum(p["matched"] for p in phases),
        "total": sum(p["total"] for p in phases),
    }
    tot["rate"] = tot["matched"] / tot["total"] if tot["total"] else 0.0
    tot["mult"] = tot["dogeup"] / tot["direct"] if tot["direct"] else 0.0
    contract_only = {
        "direct": sum(p["direct"] for p in contract),
        "dogeup": sum(p["dogeup"] for p in contract),
    }
    status_tot, pending_tot = _attach_item_status(summary_phases, wb)
    wb.close()
    return {
        "phases": phases,
        "summary_phases": summary_phases,
        "contract": contract,
        "waste": waste,
        "tot": tot,
        "contract_only": contract_only,
        "status_tot": status_tot,
        "pending_tot": pending_tot,
    }


def write_portal_ka_stats(ka: dict) -> Path:
    """포털 종합보고 탭 — _ka SSOT 수치."""
    kt = ka["tot"]
    co = ka["contract_only"]
    st = ka["status_tot"]
    pd = ka["pending_tot"]
    payload = {
        "file": KA_FILE,
        "sumDesc": (
            f"★ 제출 최종 · 직접 {won(kt['direct'])}원 · "
            f"도급 {won(kt['dogeup'])}원 · "
            f"미확정 검토 {pd['rv']} / 미매칭 {pd['um']}"
        ),
        "directWon": int(round(kt["direct"])),
        "directOku": eok(kt["direct"]),
        "contractWon": int(round(kt["dogeup"])),
        "contractOku": eok(kt["dogeup"]),
        "contract4Oku": eok(co["dogeup"]),
        "contractMult": round(kt["mult"], 3),
        "review": pd["rv"],
        "unmatched": pd["um"],
        "gayeonghyeon": st["gyh"],
        "pendingTotal": pd["all"],
        "matOku": eok(kt["mat"]),
        "labOku": eok(kt["lab"]),
        "expOku": eok(kt["exp"]),
    }
    KA_STATS_JS.write_text(
        "// auto-generated by tools/build_report_html.py — do not edit\n"
        f"window.PORTAL_KA_STATS = {json.dumps(payload, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )
    return KA_STATS_JS


def render_ka_status_rows(phases: list[dict]) -> str:
    lines = []
    for p in phases:
        st = p.get("item_status") or {}
        lines.append(
            f'  <tr><td class="c">{p["no"]}</td><td class="l"><b>{p["label"]}</b></td>'
            f'<td class="r">{st.get("rv", 0)}</td><td class="r">{st.get("um", 0)}</td>'
            f'<td class="r">{st.get("gyh", 0)}</td><td class="r">{st.get("all", 0)}</td></tr>'
        )
    return "\n".join(lines)


def render_phase_summary_rows(phases: list[dict]) -> str:
    lines = []
    for p in phases:
        lines.append(
            f'  <tr><td class="l"><b>{p["label"]}</b></td><td class="c">{p["no"]}</td>'
            f'<td class="r">{won(p["mat"])}</td><td class="r">{won(p["lab"])}</td>'
            f'<td class="r">{won(p["exp"])}</td><td class="r">{won(p["direct"])}</td>'
            f'<td class="r">{won(p["dogeup"])}</td>'
            f'<td class="r">{p["matched"]}/{p["total"]}</td>'
            f'<td class="r">{p["rate"] * 100:.1f}%</td></tr>'
        )
    return "\n".join(lines)


def render_section_blocks(summary_phases: list[dict]) -> str:
    """계약차수별 공종(차수) 상세 — 토목·조경 분리."""
    blocks: list[str] = []
    for p in summary_phases:
        if p.get("no") == "07":
            continue
        if not p["sections"]:
            continue
        rows = "\n".join(
            f'    <tr><td class="l">{s["name"]}</td><td class="c">{s["matched"]}/{s["total"]}</td>'
            f'<td class="r">{won(s["mat"])}</td><td class="r">{won(s["lab"])}</td>'
            f'<td class="r">{won(s["exp"])}</td><td class="r">{won(s["direct"])}</td></tr>'
            for s in p["sections"]
        )
        blocks.append(
            f'<h3>{p["label"]} — 공종(차수)별 직접공사비</h3>'
            f'<table><tr><th>공종</th><th>매칭/전체</th><th>재료비</th><th>노무비</th>'
            f'<th>경비</th><th>직접합계</th></tr>{rows}'
            f'<tr style="font-weight:700;background:#eef2f9"><td class="l">★ {p["label"]} 소계</td>'
            f'<td class="c">{p["matched"]}/{p["total"]}</td>'
            f'<td class="r">{won(p["mat"])}</td><td class="r">{won(p["lab"])}</td>'
            f'<td class="r">{won(p["exp"])}</td><td class="r">{won(p["direct"])}</td></tr></table>'
        )
    return "\n".join(blocks)


def render_gongjong_section(gj: dict) -> str:
    """Ⅳ. 공정별 집계 (02 전기 제외)."""
    meta = gj["meta"]
    sched_link = "../../08_제출내역서/청원지구_공사공정표.html"

    cross_rows = "\n".join(
        f'  <tr><td class="l"><b>{r["cat"]}</b></td><td class="c">{r["note"]}</td>'
        f'<td class="r">{fmt_cell(r.get("05"))}</td><td class="r">{fmt_cell(r.get("04"))}</td>'
        f'<td class="r">{fmt_cell(r.get("01"))}</td><td class="r">{fmt_cell(r.get("06"))}</td>'
        f'<td class="r"><b>{fmt_cell(r.get("total"))}</b></td><td class="c">{r["period"]}</td></tr>'
        for r in gj["cross"]
    )

    tg = gj["togong_phases"]
    tg_sum_cut = sum(p["cut"] for p in tg)
    tg_sum_fill = sum(p["fill"] for p in tg)
    tg_sum_work = meta["togong_total"]
    tg_sum_haul = meta["haul_total"]

    togong_rows = "\n".join(
        f'  <tr><td class="c"><b>{p["label"]}</b></td><td class="c">{p["code"]}</td>'
        f'<td class="r">{fmt_num(p["cut"])}</td><td class="r">{fmt_num(p["fill"])}</td>'
        f'<td class="r"><b>{fmt_num(p["work"])}</b></td><td class="c">{p["pct"]:.1f}%</td>'
        f'<td class="r">{fmt_num(p["haul"])}</td><td class="c">{p["period"]}</td>'
        f'<td class="l">{p["note"]}</td></tr>'
        for p in tg
    )

    contract_rows = "\n".join(
        f'  <tr><td class="c"><b>{c["label"]}</b></td><td class="c">{c["code"]}</td>'
        f'<td class="l">{c["gongjong"]}</td><td class="c">{c["period"]}'
        + (f'<br><span class="muted">{c["note"]}</span>' if c.get("note") else "")
        + "</td></tr>"
        for c in gj["contracts"]
    )

    return f"""
<h2>Ⅲ. 공정별 집계 (02 전기 제외)</h2>
<p class="muted">출처: 각 내역서 <code>*_표준단가산출.xlsx</code> 「통합내역」 수량 합계.
  착공 <b>{meta["start"]}</b> · 준공 <b>{meta["end"]}</b> · 총공기 <b>{meta["months"]}개월</b>.
  상세 Gantt·월별 일정: <a href="{sched_link}">청원지구_공사공정표.html</a>.</p>

<div class="kpis">
  <div class="kpi"><div class="lab">착공예정</div><div class="val">{meta["start"]}</div><div class="sub">M1</div></div>
  <div class="kpi"><div class="lab">준공목표</div><div class="val">{meta["end"]}</div><div class="sub">M14</div></div>
  <div class="kpi"><div class="lab">토공 작업량</div><div class="val">{fmt_num(tg_sum_work)}㎥</div><div class="sub">절+성 합계</div></div>
  <div class="kpi"><div class="lab">흙운반</div><div class="val">{fmt_num(tg_sum_haul)}㎥</div><div class="sub">건설폐기물 연동</div></div>
</div>

<h3>1. 토공 물량·기간 (구분별)</h3>
<table>
  <tr>
    <th>구분</th><th>코드</th><th>절토(㎥)</th><th>성토(㎥)</th><th>작업량(㎥)</th><th>비율</th><th>운반(㎥)</th><th>토공 기간</th><th>산정</th>
  </tr>
{togong_rows}
  <tr style="font-weight:700;background:#fff7ed">
    <td class="c" colspan="2">합계</td>
    <td class="r">{fmt_num(tg_sum_cut)}</td><td class="r">{fmt_num(tg_sum_fill)}</td>
    <td class="r">{fmt_num(tg_sum_work)}</td><td class="c">100%</td>
    <td class="r">{fmt_num(tg_sum_haul)}</td><td class="c">—</td>
    <td class="l">{meta["months"]}개월 총공기 내 배치</td>
  </tr>
</table>

<h3>2. 공종별 물량 — 구분별 교차표</h3>
<table>
  <tr>
    <th>공종</th><th>대표 단위</th><th>회전교차로</th><th>진입도로</th><th>토목</th><th>지구단위</th><th>합계</th><th>공정</th>
  </tr>
{cross_rows}
</table>
<p class="muted">「합계」는 동일 단위만 합산. 1. 토공 대표값은 절토+성토 작업량(㎥). 집계: <code>tools/report_gongjong_qty.py</code>.</p>

<h3>3. 구분별 공종·공정 기간</h3>
<table>
  <tr><th style="width:96px">구분</th><th style="width:48px">코드</th><th>공종 (합계요약)</th><th style="width:120px">공정 기간</th></tr>
{contract_rows}
  <tr><td class="c"><b>공통</b></td><td class="c">07</td><td class="l">건설폐기물 · 가설·준공</td><td class="c">M1~M14</td></tr>
</table>

<div class="box ok">
  <b>물량 → 공정 연계.</b>
  단지외(회전·진입) 토공 <b>{fmt_num(tg[0]["work"] + tg[1]["work"])}㎥</b>를 M2~M5에 선행 ·
  M6 진입로 1차 포장 통행 후 단지내 <b>{fmt_num(tg[2]["work"] + tg[3]["work"])}㎥</b> 집중 ·
  운반 <b>{fmt_num(tg_sum_haul)}㎥</b>는 건설폐기물과 전 기간 연동 ·
  최대 부하 M6~M9(토목·지구단위 병행).
</div>
"""


def render() -> str:
    ka = load_ka()
    gj = collect_gongjong_summary()
    phase_rows = render_phase_summary_rows(ka["summary_phases"])
    section_blocks = render_section_blocks(ka["summary_phases"])
    gongjong_section = render_gongjong_section(gj)
    ka_status_rows = render_ka_status_rows(ka["summary_phases"])
    kt = ka["tot"]
    co = ka["contract_only"]
    st = ka["status_tot"]
    pd = ka["pending_tot"]
    sp = {p["label"]: p for p in ka["summary_phases"]}
    mat_pct = pct(kt["mat"], kt["direct"])
    lab_pct = pct(kt["lab"], kt["direct"])
    exp_pct = pct(kt["exp"], kt["direct"])

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 — 종합보고서</title>
<style>
  :root{{--navy:#1f3a5f;--blue:#2c5aa0;--line:#d9e1ec;--bg:#f4f6fa;--ok:#1a7f37;}}
  *{{box-sizing:border-box;}}
  body{{font-family:"Malgun Gothic","맑은 고딕",system-ui,sans-serif;margin:0;
       background:var(--bg);color:#1c2430;line-height:1.7;font-size:15px;}}
  .wrap{{max-width:980px;margin:0 auto;padding:30px 22px 60px;}}
  header.doc{{border-bottom:3px solid var(--navy);padding-bottom:16px;margin-bottom:10px;}}
  header.doc h1{{font-size:23px;color:var(--navy);margin:0 0 6px;letter-spacing:-.4px;}}
  .meta{{color:#5a6675;font-size:13px;}}
  h2{{font-size:18px;color:var(--navy);margin:30px 0 10px;padding-left:10px;border-left:5px solid var(--blue);}}
  h3{{font-size:15.5px;color:#2a3a4d;margin:18px 0 6px;}}
  table{{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line);
        border-radius:8px;overflow:hidden;font-size:13px;margin:8px 0;}}
  th,td{{padding:8px 10px;border:1px solid var(--line);vertical-align:middle;}}
  th{{background:#eaf0f8;color:var(--navy);text-align:center;font-weight:700;}}
  td.l{{text-align:left;}} td.r{{text-align:right;white-space:nowrap;}} td.c{{text-align:center;white-space:nowrap;}}
  .kpis{{display:flex;flex-wrap:wrap;gap:12px;margin:16px 0;}}
  .kpi{{flex:1 1 150px;background:#fff;border:1px solid var(--line);border-radius:10px;padding:13px 15px;}}
  .kpi .lab{{font-size:12px;color:#6b7785;}} .kpi .val{{font-size:20px;font-weight:800;color:var(--navy);margin-top:2px;}}
  .kpi .sub{{font-size:11.5px;color:#8a94a3;margin-top:2px;}}
  .box{{border-radius:8px;padding:13px 18px;margin:14px 0;border:1px solid var(--line);background:#fff;}}
  .box.ok{{border-left:5px solid var(--ok);background:#f2fbf4;}}
  .box.hl{{border-left:5px solid var(--blue);background:#eef4fb;}}
  .box ol{{margin:6px 0;padding-left:20px;}} .box li{{margin:4px 0;}}
  code{{background:#eef2f7;padding:1px 6px;border-radius:4px;font-size:12.5px;}}
  .muted{{color:#6b7785;font-size:12.5px;}}
  a{{color:var(--blue);}}
  footer{{margin-top:34px;font-size:12px;color:#8a94a3;border-top:1px solid var(--line);padding-top:12px;}}
</style>
</head>
<body>
<div class="wrap">
<header class="doc">
  <h1>화성 청원지구 — 종합보고서</h1>
  <div class="meta">산업유통형 지구단위계획 · 마도면 청원리 산175-2번지 일원 80,617㎡ · 작성 {GEN_DATE}</div>
</header>

<div class="box hl">
  본 종합보고의 <b>최종 SSOT</b>는
  <code>{KA_FILE}</code>이다.
  금액·도급·미확정 건수는 해당 통합본 「총공사비」·구분 시트 ★합계·「◆ 내역 일괄」 상태 기준이며,
  02 전기설비는 제출 범위에서 제외한다.
  공정·물량은 <code>{KA_FILE}</code>·
  <code>08_제출내역서/청원지구_공사공정표.html</code>·표준단가산출 「통합내역」을 교차 참조한다.
</div>

<div class="kpis">
  <div class="kpi"><div class="lab">제출 최종 · 직접공사비</div><div class="val" style="color:var(--ok)">{won(kt["direct"])}원</div><div class="sub">재·노·경 합계</div></div>
  <div class="kpi"><div class="lab">제출 최종 · 도급액</div><div class="val">{won(kt["dogeup"])}원</div><div class="sub">직접의 {kt["mult"]:.3f}배</div></div>
  <div class="kpi"><div class="lab">4대 계약 도급</div><div class="val">{won(co["dogeup"])}원</div><div class="sub">회전·진입·토목·조경·지구</div></div>
  <div class="kpi"><div class="lab">미확정 (_ka)</div><div class="val">{pd["rv"]} / {pd["um"]}</div><div class="sub">검토 / 미매칭 · {pd["all"]}건 (가영현·품질관리·조경확인 3건 제외)</div></div>
</div>

<h2>Ⅰ. 사업 개요</h2>
<table>
  <tr><th style="width:130px">항목</th><th>내용</th></tr>
  <tr><td class="c">사업명</td><td class="l">화성 청원지구 산업유통형 지구단위계획</td></tr>
  <tr><td class="c">위치·면적</td><td class="l">화성시 마도면 청원리 산175-2번지 일원 · 80,617㎡</td></tr>
  <tr><td class="c">시행자</td><td class="l">진명개발(주) (법인등록 134811-0724593 · 설립 2023. 1. 10.)</td></tr>
  <tr><td class="c">인허가</td><td class="l">화성시 고시 제2026-76호 · 전략환경영향평가 HG20250176(한강유역환경청 본협의 완료)</td></tr>
  <tr><td class="c">공내역서</td><td class="l">01 토목·조경 / 02 전기 / 04 진입도로 / 05 회전교차로 / 06 개발행위 / 07 폐기물 (03 전기 지구외는 02 중복·제외)</td></tr>
  <tr><td class="c">제출 SSOT</td><td class="l"><code>{KA_FILE}</code> · 02 전기 제외 · 본 보고서 최종 집계 기준</td></tr>
</table>

<h2>Ⅱ. 제출 최종 집계 (_ka SSOT)</h2>
<p class="muted"><code>{KA_FILE}</code> · 02 전기 제외 · 간접비·도급액은 「총공사비」 시트 내역서별 산출.</p>

<div class="kpis">
  <div class="kpi"><div class="lab">4대 계약 직접공사비</div><div class="val" style="color:var(--ok)">{won(ka["contract_only"]["direct"])}원</div><div class="sub">01·04~06</div></div>
  <div class="kpi"><div class="lab">4대 계약 도급액</div><div class="val">{won(ka["contract_only"]["dogeup"])}원</div><div class="sub">회전·진입·토목·조경·지구</div></div>
  <div class="kpi"><div class="lab">전체(07 포함) 직접</div><div class="val">{won(kt["direct"])}원</div><div class="sub">07 폐기물 포함</div></div>
  <div class="kpi"><div class="lab">전체 도급액</div><div class="val">{won(kt["dogeup"])}원</div><div class="sub">직접의 {kt["mult"]:.3f}배</div></div>
</div>

<h3>1. 계약차수별 요약</h3>
<table>
  <tr>
    <th>계약차수</th><th>No</th><th>재료비</th><th>노무비</th><th>경비</th>
    <th>직접공사비</th><th>도급액</th><th>매칭/전체</th><th>매칭률</th>
  </tr>
{phase_rows}
  <tr style="font-weight:700;background:#fff2cc">
    <td class="l">★ 합계 (01·04~07)</td><td class="c">—</td>
    <td class="r">{won(kt["mat"])}</td><td class="r">{won(kt["lab"])}</td><td class="r">{won(kt["exp"])}</td>
    <td class="r">{won(kt["direct"])}</td>
    <td class="r">{won(kt["dogeup"])}</td>
    <td class="r">{kt["matched"]}/{kt["total"]}</td><td class="r">{kt["rate"] * 100:.1f}%</td>
  </tr>
</table>
<p class="muted">01 토목·01 조경은 동일 시트(토목(조경)) 내 ★소계·「총공사비」 01 토목/01 조경 열 기준 분리.
  전체 직접공사비 구성(재·노·경): {mat_pct}% / {lab_pct}% / {exp_pct}%.</p>

<h3>2. 공종(차수)별 직접공사비</h3>
{section_blocks}
{gongjong_section}

<h2>Ⅳ. 검토·미매칭·가영현 (_ka 기준 · 매칭 제외)</h2>
<p class="muted"><code>{KA_FILE}</code> 「◆ 내역 일괄」 — <b>매칭 확정은 상태 빈칸</b>(집계 제외).
  검토·미매칭 중 <b>합계단가·합계금액(또는 단가·금액)이 있는 행</b>은 「가영현」.
  산업안전보건관리비·안전관리비는 간접비 별도 적용으로 제외.</p>
<table>
  <tr><th style="width:34px">No</th><th>구분</th><th>검토</th><th>미매칭</th><th>가영현</th><th>소계</th></tr>
{ka_status_rows}
  <tr style="font-weight:700;background:#eef2f9"><td class="c">—</td><td class="l">★ 합계 (01·04~07)</td>
    <td class="r">{st["rv"]}</td><td class="r">{st["um"]}</td><td class="r">{st["gyh"]}</td>
    <td class="r">{st["all"]}</td></tr>
</table>
<p class="muted">미매칭 {st["um"]}건 · 검토 {st["rv"]}건 · 가영현 {st["gyh"]}건(합계 입력·1차 검토).
  포털 「미확정 (_ka)」는 검토·미매칭 {pd["all"]}건만 표시(가영현·품질관리비·조경 확인 3건 제외).</p>

<h2>Ⅴ. 결론</h2>
<div class="box ok">
  <ol>
    <li>제출 최종 통합본(<code>_ka</code>) 기준 <b>직접공사비 {won(kt["direct"])}원</b> ·
        <b>도급액 {won(kt["dogeup"])}원</b>(직접의 {kt["mult"]:.3f}배)으로 정리하였다.</li>
    <li>4대 계약차수 도급 — 회전교차로 {won(sp["회전교차로"]["dogeup"])}원 · 진입도로 {won(sp["진입도로"]["dogeup"])}원 ·
        토목 {won(sp["토목"]["dogeup"])}원 · 조경 {won(sp["조경"]["dogeup"])}원 · 지구단위 {won(sp["지구단위"]["dogeup"])}원
        (합 {won(co["dogeup"])}원).</li>
    <li>건설폐기물(07) 포함 전체 도급 <b>{won(kt["dogeup"])}원</b> ·
        _ka 미확정 {pd["all"]}건(검토 {pd["rv"]} · 미매칭 {pd["um"]}; 가영현 {st["gyh"]}건 별도).</li>
    <li>공정별 집계(02 전기 제외) 토공 작업량 <b>{fmt_num(gj["meta"]["togong_total"])}㎥</b> ·
        총공기 <b>{gj["meta"]["months"]}개월</b>({gj["meta"]["start"]}~{gj["meta"]["end"]}).</li>
  </ol>
</div>

<h2>부록 A. 토공 물량 검증 (인허가도면 ↔ 공내역)</h2>
<p>인허가도면 <code>04_인허가도면/00. 토공 합본</code>(45쪽)은 이미지형 PDF로 독립 「토공 수량집계표」가 없다.
공내역서 「1. 토공」 집계로 <b>토량 수지</b>를 역산·검증하였다. 상세: <code>토목/토공_물량비교표.html</code>.</p>
<table>
  <tr><th style="width:240px">구분</th><th>산출(㎥)</th><th>근거</th></tr>
  <tr><td class="l">① 절토 계(소할 제외)</td><td class="r">89,087</td><td class="l">토사+리핑+연암+보통암</td></tr>
  <tr><td class="l">② 성토 계</td><td class="r">26,538</td><td class="l">흙쌓기 전체</td></tr>
  <tr><td class="l">③ 절토 − 성토</td><td class="r">62,549</td><td class="l">사토 대상</td></tr>
  <tr><td class="l">④ 외부반출 / 절토</td><td class="r">68.9%</td><td class="l">반출 비율</td></tr>
</table>

<footer>화성 청원지구 종합보고서 · {GEN_DATE} · SSOT: {KA_FILE} · 끝.</footer>
</div>
</body>
</html>"""


def main() -> None:
    ka = load_ka()
    html = render()
    OUT.write_text(html, encoding="utf-8")
    stats_js = write_portal_ka_stats(ka)
    kt = ka["tot"]
    st = ka["status_tot"]
    pd = ka["pending_tot"]
    print(f"  {OUT.relative_to(ROOT)}")
    print(f"  {stats_js.relative_to(ROOT)}")
    print(f"  _ka 직접 {won(kt['direct'])}원 · 도급 {won(kt['dogeup'])}원")
    print(f"  _ka 미확정 {pd['all']} — 검토 {pd['rv']} / 미매칭 {pd['um']} (가영현 {st['gyh']}건 제외)")


if __name__ == "__main__":
    print("=== 종합보고서 HTML ===")
    main()
    from build_ka_cost_html import main as build_ka_cost_html  # noqa: E402

    print("=== _ka 공종별 HTML ===")
    build_ka_cost_html()
