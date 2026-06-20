#!/usr/bin/env python3
"""조달청 시설공통자재 가격정보(토목·건축·기계설비·전기,정보통신) 4개 XLSX → 통합 CSV.

한국표준품셈정보원에서 받은 「시설공통자재 가격정보」(조달청 자재단가, 재료비)를
단일 마스터 CSV로 통합한다. 표준단가산출 매칭의 재료비 단가원으로 사용.

출력: 05_내역서/일위대가DB/조달청시설자재_2026.csv
컬럼: 분야,자재구분,물품분류번호,물품식별번호,품명,규격,단위,가격,게시일자
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
SRC_DIR = BASE / "한국표준품셈정보원"
OUT_CSV = BASE / "일위대가DB" / "조달청시설자재_2026.csv"

FILES = {
    "토목": "시설공통자재(토목).xlsx",
    "건축": "시설공통자재(건축).xlsx",
    "기계설비": "시설공통자재(기계설비).xlsx",
    "전기정보통신": "시설공통자재(전기,정보통신).xlsx",
}

OUT_HEADERS = [
    "분야", "자재구분", "물품분류번호", "물품식별번호",
    "품명", "규격", "단위", "가격", "게시일자",
]


def parse_file(div: str, path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header: list[str] | None = None
    col: dict[str, int] = {}
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if row[0] == "NO":
            header = [str(c).strip() if c is not None else "" for c in row]
            col = {h: i for i, h in enumerate(header)}
            continue
        if header is None or not isinstance(row[0], int):
            continue

        def get(key: str):
            idx = col.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        name = (str(get("품명") or "")).strip()
        if not name:
            continue
        try:
            price = float(str(get("가격") or "0").replace(",", ""))
        except ValueError:
            continue
        if price <= 0:
            continue
        posted = get("게시일자")
        posted = str(posted)[:10] if posted else ""
        rows.append({
            "분야": div,
            "자재구분": (str(get("자재구분") or "")).strip(),
            "물품분류번호": (str(get("물품분류번호") or "")).strip(),
            "물품식별번호": (str(get("물품식별번호") or "")).strip(),
            "품명": name,
            "규격": (str(get("규격") or "")).strip(),
            "단위": (str(get("단위") or "")).strip(),
            "가격": int(price),
            "게시일자": posted,
        })
    wb.close()
    return rows


def main() -> None:
    all_rows: list[dict] = []
    for div, fn in FILES.items():
        path = SRC_DIR / fn
        if not path.exists():
            print(f"[건너뜀] 없음: {path}")
            continue
        rows = parse_file(div, path)
        all_rows.extend(rows)
        print(f"{div}: {len(rows)}건")

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_HEADERS)
        w.writeheader()
        w.writerows(all_rows)
    print(f"통합 저장: {OUT_CSV} (총 {len(all_rows)}건)")


if __name__ == "__main__":
    main()
