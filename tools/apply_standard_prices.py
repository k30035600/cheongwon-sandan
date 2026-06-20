#!/usr/bin/env python3
"""공내역서 수량 × 조달청 표준시장단가(공공데이터)로 재료비/노무비/경비 산출."""
from __future__ import annotations

import argparse
import csv
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 공내역서·내역서작업·일위대가DB·산출물 모두 05_내역서 하위
WORK_DIR = BASE / "내역서작업"
SRC_DIR = BASE / "공내역서"
SRC_XLS = SRC_DIR / "01_화성 청원지구 토목.XLS"
PRICE_CSV = ROOT / "tools" / "_poomsem_cache" / "표준시장단가.csv"
_STD2026 = BASE / "일위대가DB" / "표준일위대가_2026"
ILDAE_MASTER = _STD2026 / "표준일위대가_2026_마스터.csv"
MARKET_2026 = _STD2026 / "표준시장단가_2026.csv"      # 최신(2026 하반기) 표준시장단가
SIJANG_2026 = _STD2026 / "시장시공가격_2026.csv"      # 시장시공가격
LANDSCAPE_ILDAE = BASE / "일위대가DB" / "조경표준일위대가_2024.csv"  # 서울시설공단 조경 표준설계 일위대가(2024 하반기 노임)
MULGA_2026 = BASE / "일위대가DB" / "물가정보_2026.csv"  # 물가정보 변동품목(이형철근·형강·전선 등 재료비)
JOJADANG_2026 = BASE / "일위대가DB" / "조달청시설자재_2026.csv"  # 조달청 시설공통자재(토목·건축·기계·전기 재료비)
OUT_XLSX = WORK_DIR / "01_화성 청원지구 토목_표준단가산출.xlsx"
OUT_MD = WORK_DIR / "01_화성 청원지구 토목_표준단가산출_요약.md"

# 임목파쇄 — 위탁 시장가(ton→㎥ 0.40 환산) · 경비 100% · 01_…토목.XLS 전용
# 근거: 05_내역서/내역서작업/임목파쇄_단가기준_체크리스트.md 시나리오 B
IMOK_OUTSOURCE_TON_PER_M3 = 0.40
IMOK_OUTSOURCE_RULES: list[tuple[str, str, float, str]] = [
    # (품명 접두, 규격 키워드, ㎥당 합계단가, 단가명)
    ("가). 임목 운반", "", 20_000 * IMOK_OUTSOURCE_TON_PER_M3, "임목운반(위탁·25톤)"),
    ("나). 임목파쇄", "가지", (82_500 + 20_000) * IMOK_OUTSOURCE_TON_PER_M3, "임목파쇄 가지·잡목(위탁처리+운반)"),
    ("다). 임목파쇄", "뿌리", (95_000 + 20_000) * IMOK_OUTSOURCE_TON_PER_M3, "임목파쇄 뿌리(위탁처리+운반·상단)"),
]

THRESHOLD = 0.56        # 미매칭 컷오프(이상이면 매칭/검토). near-miss 구제 위해 0.62→0.56
REVIEW_THRESHOLD = 0.75  # 이상이면 '매칭(높음)', 미만이면 '검토'

UNIT_MAP = {
    "㎡": "m2",
    "m2": "m2",
    "M2": "m2",
    "a": "m2",
    "100㎡": "100m2",
    "㎥": "m3",
    "m3": "m3",
    "M3": "m3",
    "m": "m",
    "M": "m",
    "ｍ": "m",
    "개": "개",
    "ea": "개",
    "EA": "개",
    "개소": "개소",
    "본": "본",
    "주": "주",
    "대": "대",
    "Ton": "ton",
    "TON": "ton",
    "ton": "ton",
    "set": "set",
    "매": "매",
    "회": "회",
    "조": "조",
    "kg": "kg",
    "경간": "경간",
    "P.S": "PS",
    "sum": "sum",
    "개월": "개월",
    "D/M": "D/M",
    "공/㎥": "공/m3",
}

# 항목명에 포함되면 단가 품명에도 반드시 포함되어야 하는 키워드
REQUIRED_IN_PRICE = [
    "되메우기",
    "터파기",
    "흙깎기",
    "흙깍기",
    "흙쌓기",
    "흙운반",
    "아스팔트",
    "텍코팅",
    "프라임",
    "잔디붙임",
    "원심력",
    "주철관",
    "pvc",
    "스테인리스",
    "경계석",
    "차선도색",
    "혼합골재",
    "보조기층",
    "CCTV",
]

# 필수키워드 동의어 — 항목 키워드가 DB에서 다음 대체어로 충족되어도 통과
# (화강석 경계석은 표준 DB에 없어 콘크리트 경계블록 설치 단가를 대용)
REQUIRED_SYNONYMS = {
    "경계석": ("경계석", "경계블록", "경계블럭"),
}

SPEC_KEYWORDS = [
    "소규모",
    "중규모",
    "대규모",
    "보통토사",
    "풍화암",
    "연암",
    "보통암",
    "줄떼",
    "평떼",
    "소켓식",
    "kp",
    "d150",
    "d200",
    "d250",
    "d300",
    "d400",
    "d500",
    "d600",
    "d800",
    "d1000",
    "d1200",
    "d1350",
    "d1500",
    "표층",
    "중간층",
    "기층",
    "텍코팅",
    "프라임",
    "화강석",
    "페인트",
    "굴삭기",
    "불도저",
    "백호",
    "덤프",
    "로우더",
    "진동",
    "발파",
    "리핑",
    "연속",
    "병행",
    "반복",
    "이형",
    "플랜지",
    "슬립",
    "나사",
    "절단",
    "이중벽",
    "고강성",
    "스테인리스",
    "주철",
    "원심력",
    "pvc",
    "sts",
    "u형",
    "u型",
    "box",
    "맨홀",
    "유량계",
    "제수변",
    "공기밸브",
    "경계석",
    "차선",
    "잔디",
    "법면",
    "임목",
    "pe천막",
    "천막",
    "터파기",
    "되메우기",
    "흙운반",
    "아스팔트",
    "혼합골재",
    "보조기층",
    "강관",
    "흄관",
    "비계",
    "시스템",
    "5m",
    "8m",
    "12m",
    "20m",
    "30m",
    "type1",
    "type2",
    "type3",
    "type4",
    "type5",
    "기계",
    "인력",
    "육상",
    "신설",
    "직선",
    "곡선",
    "경사",
    "낮춤",
    "디스트리뷰터",
    "대형장비",
    "소형장비",
]

GENERIC_NAMES = {
    "법면보호공",
    "하수관",
    "하수관천공및접합",
    "원심력철근콘크리트관접합및부설",
    "원심력철근콘크리트관접합및부설소켓식",
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
UNMATCH_FILL = PatternFill("solid", fgColor="FCE4D6")
MONEY_FMT = "#,##0"
THIN = Side(style="thin", color="BBBBBB")


def norm_unit(unit: str) -> str:
    return UNIT_MAP.get((unit or "").strip().replace(" ", ""), (unit or "").strip().replace(" ", ""))


def norm_text(text: str) -> str:
    s = re.sub(r"\s+", "", text or "")
    s = s.replace("흙깍기", "흙깎기").replace("땅깍기", "흙깎기")
    s = re.sub(r"[\(\),/\-\+\.·]", "", s)
    return s.lower()


def strip_prefix(name: str) -> str:
    s = name.strip()
    while True:
        ns = re.sub(r"^[ \t]*([가-힣]\)|\d+\))\.\s*", "", s)
        ns = re.sub(r"^[ \t]*\d+\.\s*", "", ns)
        if ns == s:
            break
        s = ns
    return s.strip()


def kwset(text: str) -> set[str]:
    s = (text or "").lower()
    return {k for k in SPEC_KEYWORDS if k in s}


def required_keywords(item: dict) -> list[str]:
    blob = norm_text(f"{item['name']} {item['spec']}")
    found = []
    for kw in REQUIRED_IN_PRICE:
        if kw.lower() in blob or norm_text(kw) in blob:
            found.append(kw.lower())
    return found


def extract_search_terms(name: str, spec: str) -> list[str]:
    n = strip_prefix(name)
    terms: list[str] = []
    parens = re.findall(r"\(([^)]+)\)", f"{n} {spec or ''}")
    base = re.sub(r"\([^)]*\)", "", n).strip()
    if spec:
        terms.append(spec.strip())
    for p in parens:
        terms.append(p.strip())
    if base:
        terms.append(base)
    extra: list[str] = []
    for t in terms:
        for part in re.split(r"[,/及및·]", t):
            part = part.strip()
            if len(part) >= 2:
                extra.append(part)
    terms.extend(extra)
    ordered: list[str] = []
    for t in terms:
        t = t.replace("흙깍기", "흙깎기").replace("땅깍기", "흙깎기")
        if t not in ordered:
            ordered.append(t)
    base_norm = norm_text(base)
    if base_norm in {norm_text(x) for x in GENERIC_NAMES} or "법면보호" in base:
        ordered = [t for t in ordered if norm_text(t) != base_norm] + [base]
    return ordered


def load_prices() -> list[dict]:
    if not PRICE_CSV.exists():
        raise FileNotFoundError(f"단가 DB 없음: {PRICE_CSV}")
    rows = list(csv.reader(PRICE_CSV.read_text("cp949").splitlines()))
    prices: list[dict] = []
    for row in rows[1:]:
        if len(row) < 9:
            continue
        try:
            prices.append(
                {
                    "code": row[1].strip(),
                    "name": row[2].strip(),
                    "spec": row[3].strip(),
                    "unit": row[4].strip().replace(" ", ""),
                    "mat": float(row[5] or 0),
                    "lab": float(row[6] or 0),
                    "exp": float(row[7] or 0),
                    "total": float(row[8] or 0),
                    "date": row[0].strip(),
                }
            )
        except ValueError:
            continue
    return prices


def load_ildae_prices() -> list[dict]:
    """일위대가DB/표준일위대가_2026/표준일위대가_2026_마스터.csv 를 단가 후보로 적재.

    조달청·국도·하천·항만 표준일위대가(재료비/노무비/경비/합계 단가)를 매칭 풀에 추가한다.
    컬럼: 출처,시트,호표,코드,품명,규격,단위,재료비단가,노무비단가,경비단가,합계단가,비고
    """
    if not ILDAE_MASTER.exists():
        return []
    prices: list[dict] = []
    with ILDAE_MASTER.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if len(row) < 11 or row[0] == "출처" or not row[0].startswith("표준일위대가"):
                continue
            try:
                mat = float(row[7] or 0)
                lab = float(row[8] or 0)
                exp = float(row[9] or 0)
                tot = float(row[10] or 0) or (mat + lab + exp)
            except ValueError:
                continue
            name = (row[4] or "").strip()
            if not name:
                continue
            prices.append(
                {
                    "code": (row[3] or "").strip(),
                    "name": name,
                    "spec": (row[5] or "").strip(),
                    "unit": (row[6] or "").strip().replace(" ", ""),
                    "mat": mat,
                    "lab": lab,
                    "exp": exp,
                    "total": tot,
                    "date": row[0].replace("표준일위대가:", "표준ILD:"),
                }
            )
    return prices


def load_landscape_ildae() -> list[dict]:
    """서울시설공단 조경 표준설계 일위대가(조경표준일위대가_2024.csv)를 단가 후보로 적재.

    수목식재·이식·전정·잔디·조경 포장·돌공사 등 조경 공종 일위대가(재료/노무/경비/합계 단가).
    표준 DB에 없는 조경공·식재류 미매칭 보정용. 컬럼: 구분,품명,규격,단위,재료비,노무비,경비,합계
    """
    if not LANDSCAPE_ILDAE.exists():
        return []
    prices: list[dict] = []
    with LANDSCAPE_ILDAE.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            name = (row.get("품명") or "").strip()
            if not name:
                continue
            try:
                mat = float(row.get("재료비") or 0)
                lab = float(row.get("노무비") or 0)
                exp = float(row.get("경비") or 0)
                tot = float(row.get("합계") or 0) or (mat + lab + exp)
            except ValueError:
                continue
            if tot <= 0:
                continue
            prices.append({
                "code": (row.get("구분") or "").strip(),
                "name": name,
                "spec": (row.get("규격") or "").strip(),
                "unit": (row.get("단위") or "").strip().replace(" ", ""),
                "mat": mat, "lab": lab, "exp": exp, "total": tot,
                "date": "조경일위2024",
                "_fallback": True,  # 표준DB 미매칭 시에만 보조로 사용(토목 매칭 오염 방지)
            })
    return prices


def load_mulga() -> list[dict]:
    """물가정보 변동품목(물가정보_2026.csv)을 재료비 단가 후보로 적재.

    이형철근·H형강·전선·아스팔트·HDPE 등 자재 시세(재료비). 노무·경비 없음.
    철근류가 표준DB에 자재단가가 없어 크레인 등으로 오매칭되는 문제를 바로잡는 용도.
    컬럼: 품명,규격,단위,가격,변동,비고
    """
    if not MULGA_2026.exists():
        return []
    prices: list[dict] = []
    with MULGA_2026.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            name = (row.get("품명") or "").strip()
            if not name:
                continue
            try:
                price = float(str(row.get("가격") or "0").replace(",", ""))
            except ValueError:
                continue
            if price <= 0:
                continue
            prices.append({
                "code": "물가정보",
                "name": name,
                "spec": (row.get("규격") or "").strip(),
                "unit": (row.get("단위") or "").strip().replace(" ", ""),
                "mat": price, "lab": 0.0, "exp": 0.0, "total": price,
                "date": "물가정보2026",
            })
    return prices


def load_jojadang() -> list[dict]:
    """조달청 시설공통자재 가격정보(조달청시설자재_2026.csv)를 재료비 단가 후보로 적재.

    토목·건축·기계설비·전기,정보통신 4분야 6천여 건의 자재단가(재료비). 노무·경비 없음.
    물가정보·표준시장단가에 없는 자재(파일·관류·복공판 등)의 재료비 보강용.
    컬럼: 분야,자재구분,물품분류번호,물품식별번호,품명,규격,단위,가격,게시일자
    """
    if not JOJADANG_2026.exists():
        return []
    prices: list[dict] = []
    with JOJADANG_2026.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            name = (row.get("품명") or "").strip()
            if not name:
                continue
            try:
                price = float(str(row.get("가격") or "0").replace(",", ""))
            except ValueError:
                continue
            if price <= 0:
                continue
            prices.append({
                "code": (row.get("물품식별번호") or "조달청자재").strip() or "조달청자재",
                "name": name,
                "spec": (row.get("규격") or "").strip(),
                "unit": (row.get("단위") or "").strip().replace(" ", ""),
                "mat": price, "lab": 0.0, "exp": 0.0, "total": price,
                "date": "조달청시설자재2026",
            })
    return prices


def load_market_csv(path: Path, tag: str) -> list[dict]:
    """표준시장단가_2026.csv / 시장시공가격_2026.csv 를 단가 후보로 적재.

    컬럼: 발표일자,코드,품명,규격,단위,재료비,노무비,경비,합계,공사구분,적용조건
    """
    if not path.exists():
        return []
    prices: list[dict] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if len(row) < 9 or row[0] == "발표일자":
                continue
            name = (row[2] or "").strip()
            if not name:
                continue
            try:
                mat = float(row[5] or 0)
                lab = float(row[6] or 0)
                exp = float(row[7] or 0)
                tot = float(row[8] or 0) or (mat + lab + exp)
            except ValueError:
                continue
            prices.append(
                {
                    "code": (row[1] or "").strip(),
                    "name": name,
                    "spec": (row[3] or "").strip(),
                    "unit": (row[4] or "").strip().replace(" ", ""),
                    "mat": mat,
                    "lab": lab,
                    "exp": exp,
                    "total": tot,
                    "date": tag,
                }
            )
    return prices


def precompute(prices: list[dict]) -> list[dict]:
    """매칭 내부 루프에서 반복 계산되는 정규화 필드를 미리 캐시."""
    for p in prices:
        p["_pn"] = norm_text(p["name"])
        p["_pblob"] = p["_pn"] + norm_text(p["spec"])
        p["_pk"] = kwset(f"{p['name']} {p['spec']}")
        p["_preq_blob"] = norm_text(f"{p['name']} {p['spec']}")
    return prices


# 선형(1m 규격) 부속을 개수(ea)로 계상한 항목 — m 단가 × 개당 길이로 환산
LINEAR_PIECE_KW = ("경계석", "연석", "경계블록", "경계블럭")


def is_linear_piece(item: dict) -> bool:
    blob = f"{item.get('name', '')} {item.get('spec', '')}"
    return any(k in blob for k in LINEAR_PIECE_KW)


def piece_length_m(spec: str) -> float:
    """규격에서 개당 길이(m) 추정. 'AxBx1000㎜' → 1.0. 못 찾으면 1.0(표준 1m)."""
    nums = [int(x) for x in re.findall(r"\d+", spec or "")]
    cands = [n for n in nums if 300 <= n <= 4000]
    return (cands[-1] / 1000.0) if cands else 1.0


def unit_compatible(item_unit: str, price_unit: str, linear: bool = False) -> bool:
    iu = norm_unit(item_unit)
    pu = norm_unit(price_unit)
    if iu == pu:
        return True
    if iu in {"m2", "100m2"} and pu == "m2":
        return True
    # 경계석·연석류: ea/개 계상분을 m 단가와 매칭(개당 길이로 환산)
    if linear and iu == "개" and pu == "m":
        return True
    return False


def passes_required(item: dict, price: dict, req: list[str] | None = None) -> bool:
    req = required_keywords(item) if req is None else req
    if not req:
        return True
    blob = price.get("_preq_blob") or norm_text(f"{price['name']} {price['spec']}")
    for k in req:
        alts = REQUIRED_SYNONYMS.get(k, (k,))
        if any(norm_text(a) in blob or a in blob for a in alts):
            return True
    return False


def score_match(item: dict, price: dict, term: str, item_kw: set[str] | None = None,
                req: list[str] | None = None) -> float:
    if not unit_compatible(item["unit"], price["unit"], item.get("_linear", False)):
        return -1.0
    if not passes_required(item, price, req):
        return -1.0
    tn = norm_text(term)
    pn = price.get("_pn") or norm_text(price["name"])
    blob = price.get("_pblob") or (pn + norm_text(price["spec"]))
    if not tn:
        return -1.0
    name_score = SequenceMatcher(None, tn, pn).ratio()
    if tn in pn or pn in tn:
        name_score = max(name_score, 0.9)
    if tn in blob:
        name_score = max(name_score, 0.82)
    ik = item_kw if item_kw is not None else kwset(f"{item['name']} {item['spec']} {term}")
    pk = price.get("_pk") or kwset(f"{price['name']} {price['spec']}")
    if ik and pk:
        spec_score = 0.35 + 0.65 * (len(ik & pk) / max(len(ik), 1))
    elif not item["spec"].strip() and not price["spec"].strip():
        spec_score = 0.55
    else:
        spec_score = SequenceMatcher(None, item["spec"], price["spec"]).ratio() * 0.75
    return 0.25 + name_score * 0.45 + spec_score * 0.3


def find_best_match(item: dict, prices: list[dict]) -> tuple[dict | None, float, str, list[str]]:
    terms = extract_search_terms(item["name"], item["spec"])
    item["_linear"] = is_linear_piece(item)
    candidates = [p for p in prices if unit_compatible(item["unit"], p["unit"], item["_linear"])]
    req = required_keywords(item)
    item_kw = kwset(f"{item['name']} {item['spec']}")

    def scan(pool):
        b, bs, bt = None, -1.0, ""
        for term in terms:
            kw = item_kw | kwset(term)
            for price in pool:
                s = score_match(item, price, term, kw, req)
                if s > bs:
                    b, bs, bt = price, s, term
        return b, bs, bt

    # 1차: 표준DB(비-폴백)만으로 매칭 — 토목·전기 등 본래 단가 우선
    primary = [p for p in candidates if not p.get("_fallback")]
    best, best_score, best_term = scan(primary)
    # 2차: 표준DB가 임계 미만(사실상 미매칭)일 때만 조경 일위대가 폴백 허용
    if best_score < THRESHOLD:
        fb = [p for p in candidates if p.get("_fallback")]
        if fb:
            fb_best, fb_score, fb_term = scan(fb)
            if fb_score > best_score:
                best, best_score, best_term = fb_best, fb_score, fb_term
    return best, best_score, best_term, terms


def is_section_header(name: str) -> bool:
    n = name.strip()
    if not n:
        return False
    if n.startswith("▣") or n.startswith("▶"):
        return True
    if re.match(r"^[ⅠⅡⅢⅣⅤ]", n):
        return True
    if re.match(r"^\d+\.", n):
        return True
    if "지구단위계획" in n or "진입도로" in n:
        return True
    if n.startswith("..."):
        return True
    return False


def load_estimate(src: Path | None = None) -> tuple[list[dict], list[str]]:
    path = src or SRC_XLS
    wb = xlrd.open_workbook(str(path))
    sh = wb.sheet_by_name("내역서")
    sections: list[str] = ["(미분류)"]
    items: list[dict] = []
    for r in range(3, sh.nrows):
        name = str(sh.cell_value(r, 0)).strip()
        spec = str(sh.cell_value(r, 1)).strip()
        qty = sh.cell_value(r, 2) if sh.cell_type(r, 2) == xlrd.XL_CELL_NUMBER else None
        unit = str(sh.cell_value(r, 3)).strip()
        if qty == 1 and unit == "식" and name:
            if is_section_header(name):
                sections.append(name)
            continue
        if qty and unit and unit != "식" and qty > 0:
            items.append(
                {
                    "row": r + 1,
                    "section": sections[-1],
                    "name": name,
                    "spec": spec,
                    "qty": float(qty),
                    "unit": unit,
                }
            )
    return items, sections[1:]


def qty_factor(item_unit: str, price_unit: str, item: dict | None = None) -> float:
    iu = norm_unit(item_unit)
    pu = norm_unit(price_unit)
    if iu == "100m2" and pu == "m2":
        return 0.01
    # ea/개(선형 부속) × m 단가 → 개당 길이(m)만큼 환산
    if iu == "개" and pu == "m" and item is not None and item.get("_linear"):
        return piece_length_m(item.get("spec", ""))
    return 1.0


def calc_amounts(item: dict, price: dict) -> dict:
    factor = qty_factor(item["unit"], price["unit"], item)
    qty = item["qty"] * factor
    return {
        "mat_amt": qty * price["mat"],
        "lab_amt": qty * price["lab"],
        "exp_amt": qty * price["exp"],
        "sum_amt": qty * price["total"],
    }


def imok_outsource_price(item: dict, src_name: str) -> dict | None:
    """01 토목 임목 3건 — 위탁 시장가 수동 단가."""
    if "토목" not in src_name or "조경" in src_name:
        return None
    name = item["name"].strip()
    spec = item.get("spec", "")
    for prefix, spec_kw, unit_total, label in IMOK_OUTSOURCE_RULES:
        if not name.startswith(prefix):
            continue
        if spec_kw and spec_kw not in spec:
            continue
        exp = unit_total  # 위탁 하도급 — 경비 100%
        return {
            "code": "수동위탁",
            "name": label,
            "spec": f"위탁 시장가 · {IMOK_OUTSOURCE_TON_PER_M3}ton/㎥ 환산",
            "unit": item["unit"],
            "mat": 0.0,
            "lab": 0.0,
            "exp": exp,
            "total": exp,
            "date": "임목위탁2026",
        }
    return None


def build_results(items: list[dict], prices: list[dict], src_name: str = ""):
    matched_rows = []
    unmatched_rows = []
    by_row: dict[int, dict] = {}
    totals = {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0}
    section_totals: dict[str, dict] = {}

    for item in items:
        manual = imok_outsource_price(item, src_name)
        if manual:
            price, score, term, terms = manual, 1.0, "수동위탁", ["위탁"]
        else:
            price, score, term, terms = find_best_match(item, prices)
        sec = item["section"]
        section_totals.setdefault(
            sec, {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0, "matched": 0, "items": 0}
        )
        section_totals[sec]["items"] += 1

        base = {**item, "status": "미매칭", "match_score": round(score, 3) if score >= 0 else None, "terms": ", ".join(terms[:5])}

        if not price or score < THRESHOLD:
            unmatched_rows.append({**base, "score": score})
            by_row[item["row"]] = base
            continue

        amts = calc_amounts(item, price)
        if manual:
            confidence = "수동"
            status = "매칭"
            row = {
                **item,
                **amts,
                "status": status,
                "match_score": 1.0,
                "match_term": term,
                "price_code": price["code"],
                "price_name": price["name"],
                "price_spec": price["spec"],
                "price_unit": price["unit"],
                "mat_unit": price["mat"],
                "lab_unit": price["lab"],
                "exp_unit": price["exp"],
                "total_unit": price["total"],
                "confidence": confidence,
            }
        else:
            confidence = "높음" if score >= REVIEW_THRESHOLD else "검토"
            status = "매칭" if score >= REVIEW_THRESHOLD else "검토"
            row = {
                **item,
                **amts,
                "status": status,
                "match_score": round(score, 3),
                "match_term": term,
                "price_code": price["code"],
                "price_name": price["name"],
                "price_spec": price["spec"],
                "price_unit": price["unit"],
                "mat_unit": price["mat"],
                "lab_unit": price["lab"],
                "exp_unit": price["exp"],
                "total_unit": price["total"],
                "confidence": confidence,
            }
        matched_rows.append(row)
        by_row[item["row"]] = row
        for key, val in zip(["mat", "lab", "exp", "sum"], [amts["mat_amt"], amts["lab_amt"], amts["exp_amt"], amts["sum_amt"]]):
            totals[key] += val
            section_totals[sec][key] += val
        section_totals[sec]["matched"] += 1

    review_rows = [r for r in matched_rows if r["status"] == "검토"]
    integrated = [by_row.get(it["row"], {**it, "status": "미매칭"}) for it in items]
    return matched_rows, unmatched_rows, review_rows, integrated, totals, section_totals


def style_header(ws, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit(ws, max_width=52):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max_width, max(len(str(cell.value or "")) for cell in col) + 2)
        ws.column_dimensions[letter].width = width


def apply_table(ws, money_cols: set[int] | None = None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    money_cols = money_cols or set()
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.column in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT


def write_xlsx(matched, unmatched, review, integrated, totals, section_totals, price_date: str, out_path: Path | None = None, src_name: str = ""):
    wb = Workbook()

    # --- 안내 ---
    ws0 = wb.active
    ws0.title = "안내"
    notes = [
        ["화성 청원지구 공내역서 — 표준단가 산출 결과"],
        [],
        ["원본", src_name or str(SRC_XLS.name)],
        ["단가 DB", f"조달청 표준시장단가 ({price_date})"],
        ["산출일", "자동 산출 (apply_standard_prices.py)"],
        [],
        ["시트 구성"],
        ["합계요약", "공종별 재료비·노무비·경비·합계"],
        ["통합내역", "576개 전 항목 + 단가·금액·상태(매칭/검토/미매칭)"],
        ["매칭내역", "단가 매칭 완료 항목"],
        ["검토필요", "매칭점수 0.62~0.74 — 품명·규격 수동 확인 권장"],
        ["미매칭", "DB에 대응 단가 없음 — 표준일위대가·수동 입력 필요"],
        [],
        ["주의"],
        ["1", "합계는 매칭·검토 항목만 포함. 미매칭 271건은 금액 0."],
        ["2", "표준시장단가 ≠ 표준품셈 일위대가 전체."],
        ["3", "조경·품질관리·GIS 등은 별도 단가표 필요."],
    ]
    for line in notes:
        ws0.append(line)
    ws0["A1"].font = Font(bold=True, size=14)
    autofit(ws0, 80)

    # --- 합계요약 ---
    ws = wb.create_sheet("합계요약")
    ws.append(["구분", "매칭", "전체", "매칭률", "재료비", "노무비", "경비", "합계"])
    style_header(ws, 8)
    total_items = sum(v["items"] for v in section_totals.values())
    total_matched = sum(v["matched"] for v in section_totals.values())
    for sec, st in section_totals.items():
        rate = st["matched"] / st["items"] if st["items"] else 0
        ws.append([sec, st["matched"], st["items"], rate, st["mat"], st["lab"], st["exp"], st["sum"]])
    ws.append([])
    ws.append(
        [
            "★ 매칭합계",
            total_matched,
            total_items,
            total_matched / total_items if total_items else 0,
            totals["mat"],
            totals["lab"],
            totals["exp"],
            totals["sum"],
        ]
    )
    ws.append(["※ 미매칭", len(unmatched), total_items, len(unmatched) / total_items if total_items else 0, 0, 0, 0, 0])
    apply_table(ws, {5, 6, 7, 8})
    for r in range(2, ws.max_row + 1):
        if isinstance(ws.cell(r, 4).value, float):
            ws.cell(r, 4).number_format = "0.0%"

    # --- 통합내역 ---
    ws1 = wb.create_sheet("통합내역")
    int_headers = [
        "행", "공종", "공종명", "규격", "수량", "단위", "상태", "매칭점수",
        "단가코드", "매칭품명", "매칭규격",
        "재료단가", "노무단가", "경비단가", "합계단가",
        "재료금액", "노무금액", "경비금액", "합계금액", "비고",
    ]
    ws1.append(int_headers)
    style_header(ws1, len(int_headers))
    for r in integrated:
        ws1.append(
            [
                r.get("row"),
                r.get("section"),
                r.get("name"),
                r.get("spec") or "",
                r.get("qty"),
                r.get("unit"),
                r.get("status", "미매칭"),
                r.get("match_score"),
                r.get("price_code", ""),
                r.get("price_name", ""),
                r.get("price_spec", ""),
                r.get("mat_unit", ""),
                r.get("lab_unit", ""),
                r.get("exp_unit", ""),
                r.get("total_unit", ""),
                r.get("mat_amt", ""),
                r.get("lab_amt", ""),
                r.get("exp_amt", ""),
                r.get("sum_amt", ""),
                r.get("terms", r.get("match_term", "")),
            ]
        )
    apply_table(ws1, {12, 13, 14, 15, 16, 17, 18, 19})
    for row_idx in range(2, ws1.max_row + 1):
        status = ws1.cell(row_idx, 7).value
        fill = REVIEW_FILL if status == "검토" else UNMATCH_FILL if status == "미매칭" else None
        if fill:
            for c in range(1, len(int_headers) + 1):
                ws1.cell(row_idx, c).fill = fill

    detail_headers = [
        "행", "공종", "공종명", "규격", "수량", "단위", "매칭점수", "신뢰도",
        "단가코드", "매칭품명", "매칭규격",
        "재료단가", "노무단가", "경비단가", "합계단가",
        "재료금액", "노무금액", "경비금액", "합계금액",
    ]

    def write_detail_sheet(title: str, rows: list[dict]):
        w = wb.create_sheet(title)
        w.append(detail_headers)
        style_header(w, len(detail_headers))
        for r in rows:
            w.append(
                [
                    r["row"], r["section"], r["name"], r["spec"], r["qty"], r["unit"],
                    r["match_score"], r.get("confidence", ""), r["price_code"],
                    r["price_name"], r["price_spec"],
                    r["mat_unit"], r["lab_unit"], r["exp_unit"], r["total_unit"],
                    r["mat_amt"], r["lab_amt"], r["exp_amt"], r["sum_amt"],
                ]
            )
        apply_table(w, {12, 13, 14, 15, 16, 17, 18, 19})
        if title == "검토필요":
            for row_idx in range(2, w.max_row + 1):
                for c in range(1, len(detail_headers) + 1):
                    w.cell(row_idx, c).fill = REVIEW_FILL

    write_detail_sheet("매칭내역", [r for r in matched if r["status"] == "매칭"])
    write_detail_sheet("검토필요", review)

    ws3 = wb.create_sheet("미매칭")
    ws3.append(["행", "공종", "공종명", "규격", "수량", "단위", "최고점수", "검색어"])
    style_header(ws3, 8)
    for r in unmatched:
        ws3.append([r["row"], r["section"], r["name"], r["spec"], r["qty"], r["unit"], round(r["score"], 3), r["terms"]])
    apply_table(ws3)
    for row_idx in range(2, ws3.max_row + 1):
        for c in range(1, 9):
            ws3.cell(row_idx, c).fill = UNMATCH_FILL

    for sheet in wb.worksheets:
        autofit(sheet)

    dest = out_path or OUT_XLSX
    try:
        wb.save(dest)
        saved = dest
    except PermissionError:
        alt = dest.with_name(dest.stem + "_업데이트.xlsx")
        wb.save(alt)
        saved = alt
        print(f"원본 파일 사용 중 → {alt.name} 로 저장")
    return saved


def write_md(matched, unmatched, review, totals, section_totals, price_date: str, src_name: str, out_md: Path | None = None):
    total_items = sum(v["items"] for v in section_totals.values())
    matched_ok = len([r for r in matched if r["status"] == "매칭"])
    lines = [
        "# 화성 청원지구 공내역서 — 표준단가 산출 요약",
        "",
        "## 산출 기준",
        "",
        f"- **원본**: `{src_name}`",
        f"- **단가 출처**: 조달청 표준시장단가 ({price_date})",
        "- **결과 파일**: `…_표준단가산출.xlsx` (시트: 안내·합계요약·통합내역·매칭내역·검토필요·미매칭)",
        "",
        "## 재료비 / 노무비 / 경비 합계 (매칭+검토 항목)",
        "",
        "| 구분 | 매칭 | 전체 | 재료비 | 노무비 | 경비 | 합계 |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for sec, st in section_totals.items():
        lines.append(
            f"| {sec} | {st['matched']} | {st['items']} | "
            f"{st['mat']:,.0f} | {st['lab']:,.0f} | {st['exp']:,.0f} | {st['sum']:,.0f} |"
        )
    all_matched = matched_ok + len(review)
    lines.extend(
        [
            f"| **합계** | **{all_matched}** | **{total_items}** | "
            f"**{totals['mat']:,.0f}** | **{totals['lab']:,.0f}** | "
            f"**{totals['exp']:,.0f}** | **{totals['sum']:,.0f}** |",
            "",
            f"- 매칭(높음): **{matched_ok}건** / 검토필요: **{len(review)}건** / 미매칭: **{len(unmatched)}건**",
            "",
            "끝.",
        ]
    )
    (out_md or OUT_MD).write_text("\n".join(lines), encoding="utf-8")


def parse_args():
    p = argparse.ArgumentParser(description="토목 공내역서 재료비·노무비·경비 산출")
    p.add_argument("--src", type=Path, default=SRC_XLS)
    p.add_argument("--out", type=Path, default=None)
    return p.parse_args()


def main():
    args = parse_args()
    src = args.src if args.src.is_absolute() else ROOT / args.src
    if not src.exists():
        alt = SRC_DIR / args.src.name
        if alt.exists():
            src = alt
    # 출력물은 05_내역서/내역서작업에 저장
    out = args.out or WORK_DIR / (src.stem + "_표준단가산출.xlsx")
    out_md = out.with_name(out.stem + "_요약.md")

    market = load_market_csv(MARKET_2026, "표준시장단가2026")
    if not market:  # 신규 추출본이 없으면 구 CSV로 폴백
        market = load_prices()
    sijang = load_market_csv(SIJANG_2026, "시장시공가격2026")
    ildae = load_ildae_prices()
    landscape = load_landscape_ildae()
    mulga = load_mulga()
    jojadang = load_jojadang()
    prices = precompute(market + sijang + ildae + mulga + jojadang + landscape)
    price_date = (
        f"표준시장단가2026 {len(market):,} + 시장시공가격 {len(sijang):,} "
        f"+ 표준일위대가2026 {len(ildae):,} + 물가정보2026 {len(mulga):,} "
        f"+ 조달청시설자재2026 {len(jojadang):,} + 조경일위대가2024 {len(landscape):,}건"
    )
    print(f"단가 풀: 시장단가 {len(market):,} + 시장시공 {len(sijang):,} "
          f"+ 일위대가 {len(ildae):,} + 물가정보 {len(mulga):,} + 조달청자재 {len(jojadang):,} "
          f"+ 조경일위 {len(landscape):,} = {len(prices):,}건")
    items, _ = load_estimate(src)
    matched, unmatched, review, integrated, totals, section_totals = build_results(items, prices, src.name)
    saved = write_xlsx(matched, unmatched, review, integrated, totals, section_totals, price_date, out, src.name)
    write_md(matched, unmatched, review, totals, section_totals, price_date, src.name, out_md)
    matched_ok = len([r for r in matched if r["status"] == "매칭"])
    print(f"항목 {len(items)} / 매칭 {matched_ok} / 검토 {len(review)} / 미매칭 {len(unmatched)}")
    print(f"재료비 {totals['mat']:,.0f}  노무비 {totals['lab']:,.0f}  경비 {totals['exp']:,.0f}  합계 {totals['sum']:,.0f}")
    print(f"저장: {saved}")


if __name__ == "__main__":
    main()
