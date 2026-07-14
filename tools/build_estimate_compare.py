#!/usr/bin/env python3
"""세흥건설 기준 견적 비교표 생성 → 07_타견적/견적비교.xlsx.

비교 키 = (내역서 No, 공종).  세흥건설을 기준(base)으로 하고
근일건설을 매칭한다. 금액차이 = 세흥 - 근일.
열: 공종/규격 · 수량 · 단위 · 세흥(재/노/경/합계) · 근일(재/노/경/합계)
    · 금액차이(재/노/경/합계) · 비고
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
SEHUNG = ROOT / "07_타견적" / "세흥건설(청원지구).xlsx"
GEUNIL = ROOT / "07_타견적" / "근일건설(청원지구).xlsx"
DST = ROOT / "07_타견적" / "견적비교.xlsx"

# 내역서 No → 표시명
CLASS_NAME = {
    "01": "01. 지구단위계획(세흥) ↔ 01 토목·조경(근일)",
    "04": "04. 진입도로",
    "05": "05. 회전교차로",
    "06": "06. 개발행위",
    "07": "07. 건설폐기물",
}
CLASS_ORDER = ["01", "04", "05", "06", "07"]


def n(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return 0.0
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_gong(s: str) -> str:
    """공종명 정규화: 앞 'N.' 제거, 공백 제거."""
    s = (s or "").strip()
    s = re.sub(r"^\d+\.\s*", "", s)
    return s.replace(" ", "")


def load_sehung():
    wb = load_workbook(SEHUNG, data_only=True)
    ws = wb["견적서"]
    rows = []
    for r in range(8, 42):  # 집계표 구간
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        if not c1 or not re.match(r"^\d+\.", str(c1)):
            continue  # [소계] 등 제외
        no = re.match(r"^(\d+)\.", str(c1)).group(1)
        rows.append({
            "no": no,
            "gong": str(c2).strip(),
            "key": (no, norm_gong(str(c2))),
            "qty": ws.cell(row=r, column=3).value,
            "unit": ws.cell(row=r, column=4).value,
            "mat": n(ws.cell(row=r, column=6).value),
            "lab": n(ws.cell(row=r, column=8).value),
            "exp": n(ws.cell(row=r, column=10).value),
            "sum": n(ws.cell(row=r, column=12).value),
        })
    return rows


def load_geunil():
    wb = load_workbook(GEUNIL, data_only=True)
    ws = wb["내역집계"]
    data = {}
    for r in range(4, ws.max_row + 1):
        no = ws.cell(row=r, column=2).value
        gong = ws.cell(row=r, column=4).value
        if not no or not gong:
            continue
        g = str(gong).strip()
        if g.startswith("★") or g.startswith("▣") or g.startswith("..."):
            continue
        if "산업안전보건관리비" in g:
            continue
        key = (str(no).strip(), norm_gong(g))
        # 내역집계 열: 1구분 2No 3내역서 4공종 5매칭 6전체 7매칭률 8재료비 9노무비 10경비 11합계
        data[key] = {
            "no": str(no).strip(),
            "gong": g,
            "mat": n(ws.cell(row=r, column=8).value),
            "lab": n(ws.cell(row=r, column=9).value),
            "exp": n(ws.cell(row=r, column=10).value),
            "sum": n(ws.cell(row=r, column=11).value),
        }
    return data


def main() -> None:
    sehung = load_sehung()
    geunil = load_geunil()

    wb = Workbook()
    ws = wb.active
    ws.title = "견적비교"

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment("center", "center", wrap_text=True)
    right = Alignment("right", "center")
    left = Alignment("left", "center", wrap_text=True)
    f_hdr = Font(bold=True, size=10, color="FFFFFF")
    fill_hdr = PatternFill("solid", fgColor="404E67")
    fill_se = PatternFill("solid", fgColor="DDEBF7")
    fill_ge = PatternFill("solid", fgColor="FCE4D6")
    fill_df = PatternFill("solid", fgColor="E2EFDA")
    fill_cls = PatternFill("solid", fgColor="FFF2CC")
    fill_sub = PatternFill("solid", fgColor="F2F2F2")

    # 헤더 2행
    ws.merge_cells("A1:A2"); ws["A1"] = "공종 / 규격"
    ws.merge_cells("B1:B2"); ws["B1"] = "수량"
    ws.merge_cells("C1:C2"); ws["C1"] = "단위"
    ws.merge_cells("D1:G1"); ws["D1"] = "세흥건설"
    ws.merge_cells("H1:K1"); ws["H1"] = "근일건설"
    ws.merge_cells("L1:O1"); ws["L1"] = "금액차이(세흥−근일)"
    ws.merge_cells("P1:P2"); ws["P1"] = "비고"
    subhdr = ["재료비", "노무비", "경비", "합계"]
    for i, base in enumerate((4, 8, 12)):  # D,H,L
        for j, name in enumerate(subhdr):
            ws.cell(row=2, column=base + j, value=name)
    for c in range(1, 17):
        for rr in (1, 2):
            cell = ws.cell(row=rr, column=c)
            cell.font = f_hdr; cell.fill = fill_hdr
            cell.alignment = center; cell.border = border

    def money(cell, val, fill=None):
        cell.value = round(val) if isinstance(val, (int, float)) else val
        if isinstance(val, (int, float)):
            cell.number_format = "#,##0"
        cell.alignment = right; cell.border = border
        if fill:
            cell.fill = fill

    grand = {k: [0.0, 0.0, 0.0, 0.0] for k in ("se", "ge")}
    matched_keys = set()

    for no in CLASS_ORDER:
        se_rows = [r for r in sehung if r["no"] == no]
        # 07은 세흥 집계에 없음 → 근일 단독 처리(뒤에서)
        if no == "07":
            continue
        if not se_rows:
            continue
        # 분류 헤더
        r0 = ws.max_row + 1
        ws.cell(row=r0, column=1, value=CLASS_NAME.get(no, no))
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=16)
        hc = ws.cell(row=r0, column=1)
        hc.font = Font(bold=True, size=10); hc.fill = fill_cls
        hc.alignment = left; hc.border = border

        sub = {"se": [0, 0, 0, 0], "ge": [0, 0, 0, 0]}
        for r in se_rows:
            g = geunil.get(r["key"])
            matched_keys.add(r["key"])
            rr = ws.max_row + 1
            ws.cell(row=rr, column=1, value=r["gong"]).alignment = left
            ws.cell(row=rr, column=1).border = border
            qcell = ws.cell(row=rr, column=2, value=r["qty"]); qcell.alignment = center; qcell.border = border
            ucell = ws.cell(row=rr, column=3, value=r["unit"]); ucell.alignment = center; ucell.border = border
            se_v = [r["mat"], r["lab"], r["exp"], r["sum"]]
            ge_v = [g["mat"], g["lab"], g["exp"], g["sum"]] if g else [None, None, None, None]
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), se_v[j], fill_se)
                money(ws.cell(row=rr, column=8 + j), ge_v[j] if g else "", fill_ge)
                if g:
                    money(ws.cell(row=rr, column=12 + j), se_v[j] - ge_v[j], fill_df)
                else:
                    money(ws.cell(row=rr, column=12 + j), "", fill_df)
                sub["se"][j] += se_v[j]
                if g:
                    sub["ge"][j] += ge_v[j]
            note = "" if g else "근일 미매칭"
            nc = ws.cell(row=rr, column=16, value=note); nc.alignment = left; nc.border = border
        # 분류 소계
        rr = ws.max_row + 1
        sc = ws.cell(row=rr, column=1, value=f"[ {no} 소계 ]"); sc.font = Font(bold=True); sc.alignment = right
        sc.fill = fill_sub; sc.border = border
        for c in (2, 3):
            ws.cell(row=rr, column=c).fill = fill_sub; ws.cell(row=rr, column=c).border = border
        for j in range(4):
            money(ws.cell(row=rr, column=4 + j), sub["se"][j], fill_sub)
            money(ws.cell(row=rr, column=8 + j), sub["ge"][j], fill_sub)
            money(ws.cell(row=rr, column=12 + j), sub["se"][j] - sub["ge"][j], fill_sub)
            ws.cell(row=rr, column=4 + j).font = Font(bold=True)
            ws.cell(row=rr, column=8 + j).font = Font(bold=True)
            ws.cell(row=rr, column=12 + j).font = Font(bold=True)
            grand["se"][j] += sub["se"][j]
            grand["ge"][j] += sub["ge"][j]
        ws.cell(row=rr, column=16).fill = fill_sub; ws.cell(row=rr, column=16).border = border

    # 근일 단독 (세흥 집계에 없는 항목: 07 폐기물 등)
    extra = [(k, v) for k, v in geunil.items() if k not in matched_keys]
    if extra:
        r0 = ws.max_row + 1
        ws.cell(row=r0, column=1, value="◆ 근일 단독 (세흥 집계 미포함)")
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=16)
        hc = ws.cell(row=r0, column=1); hc.font = Font(bold=True, size=10)
        hc.fill = PatternFill("solid", fgColor="F8CBAD"); hc.alignment = left; hc.border = border
        for k, g in sorted(extra):
            rr = ws.max_row + 1
            ws.cell(row=rr, column=1, value=f"{g['no']} · {g['gong']}").alignment = left
            ws.cell(row=rr, column=1).border = border
            ws.cell(row=rr, column=2).border = border
            ws.cell(row=rr, column=3).border = border
            ge_v = [g["mat"], g["lab"], g["exp"], g["sum"]]
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), "", fill_se)
                money(ws.cell(row=rr, column=8 + j), ge_v[j], fill_ge)
                money(ws.cell(row=rr, column=12 + j), -ge_v[j], fill_df)
                grand["ge"][j] += ge_v[j]
            ws.cell(row=rr, column=16, value="세흥 없음").alignment = left
            ws.cell(row=rr, column=16).border = border

    # 총계
    rr = ws.max_row + 1
    tc = ws.cell(row=rr, column=1, value="★ 총 계 (공종 집계 기준)")
    tc.font = Font(bold=True, size=11, color="FFFFFF"); tc.fill = fill_hdr
    tc.alignment = right; tc.border = border
    for c in (2, 3):
        ws.cell(row=rr, column=c).fill = fill_hdr; ws.cell(row=rr, column=c).border = border
    for j in range(4):
        for base, key in ((4, "se"), (8, "ge")):
            cell = ws.cell(row=rr, column=base + j)
            money(cell, grand[key][j], fill_hdr)
            cell.font = Font(bold=True, color="FFFFFF")
        cell = ws.cell(row=rr, column=12 + j)
        money(cell, grand["se"][j] - grand["ge"][j], fill_hdr)
        cell.font = Font(bold=True, color="FFFFFF")
    ws.cell(row=rr, column=16).fill = fill_hdr; ws.cell(row=rr, column=16).border = border

    # 주석
    rr = ws.max_row + 2
    ws.cell(row=rr, column=1,
            value="※ 비교 단위: 공종별 집계(내역서 No × 공종). 세흥 '수량 1·식'은 집계 표기. "
                  "세흥 직접공사비계에는 안전·품질관리비 425,371,600원이 별도 포함(본 표 공종 합계 제외).")
    ws.cell(row=rr, column=1).font = Font(size=9, italic=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=16)

    widths = [34, 7, 6, 13, 13, 12, 14, 13, 13, 12, 14, 13, 13, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    wb.save(DST)

    print(f"OK  {DST.name}")
    print(f"세흥 공종 {len(sehung)}행 · 근일 공종 {len(geunil)}개 · 근일단독 {len(extra)}개")
    print("총계  세흥 합계 {:,.0f} · 근일 합계 {:,.0f} · 차이 {:,.0f}".format(
        grand["se"][3], grand["ge"][3], grand["se"][3] - grand["ge"][3]))


if __name__ == "__main__":
    main()
