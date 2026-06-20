#!/usr/bin/env python3
"""저신뢰(점수<0.56) 오매칭 과대단가 교정 — 미매칭_일위대가산출.xlsx 인-place 수정.

전수조사 결과 미매칭 자동제시(점수 0.46~0.53)가 전혀 다른 품목에 매칭돼 확정으로
굳어진 과대단가를 표준품셈·조경표준일위대가 기준의 보수적 재산출치로 교체한다.
※ 재산출치는 발주처·설계자 검증 권장(근거 열에 명시).
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
UNMATCHED_IL = BASE / "미매칭_일위대가산출.xlsx"
POOMSEM_OUT = WORK / "검토_오매칭교정_품셈산출.xlsx"

# 흄관 하차(크레인 상하차) 직경별 보수적 추정(경비 기준, 검증 권장)
HUMGWAN_BY_DIA = {250: 2000, 450: 3000, 600: 5000, 700: 6000,
                  800: 8000, 900: 10000, 1000: 14000, 1200: 20000}


def humgwan_unloading(spec: str) -> tuple[float, float, float, float, str]:
    m = re.search(r"(\d{3,4})", spec.replace(" ", ""))
    dia = int(m.group(1)) if m else 0
    # 가장 가까운 직경 구간
    key = min(HUMGWAN_BY_DIA, key=lambda d: abs(d - dia)) if dia else 250
    exp = HUMGWAN_BY_DIA[key]
    basis = (f"오매칭(점수0.46) 제외 · 흄관 하차는 직경·중량 비례 → "
             f"Ø{key}mm 크레인 상하차 추정 {exp:,}원/본(장비경비, 검증 권장)")
    return 0, 0, exp, exp, basis


# 교정 규칙: (품명 포함, 규격 포함, 단위, 재료, 노무, 경비, 합계, 근거)
# spec_fn 가 있으면 규격별 동적 산출
CORRECTIONS = [
    {
        "name": "경고용테이프", "spec": "", "unit": "m",
        "mat": 100, "lab": 200, "exp": 0, "total": 300,
        "basis": ("오매칭(점수0.50, 표지판/PVC관) 제외 · 지중 매설 PE 경고테이프 200mm "
                  "자재 약100원/m + 되메우기 중 부설 노무 약200원/m = 300원/m(보수적·검증 권장)"),
    },
    {
        "name": "우수받이", "spec": "0.3", "unit": "개소",
        "mat": 150_000, "lab": 50_000, "exp": 0, "total": 200_000,
        "basis": ("오매칭(점수0.49, 스틸 방음문) 제외 · 1호 우수받이(0.3×0.4×0.9m 스틸) "
                  "강재 받이틀+스틸그레이팅 자재 약150,000 + 설치 약50,000(조경일위 PE빗물받이 "
                  "설치 50,124 준용) = 200,000원/개소(검증 권장)"),
    },
    {
        "name": "가로지지대", "spec": "", "unit": "개소",
        "mat": 50_000, "lab": 30_000, "exp": 0, "total": 80_000,
        "basis": ("오매칭(점수0.46, 차로지정표지 문형식) 제외 · 가로수 지주(B20 수목지지대) "
                  "강관지주 설치 18,539원/m(조경일위)×2주식+자재 약 80,000원/개소(보수적·검증 권장)"),
    },
    {
        "name": "흄관하차비", "spec": "", "unit": "본", "spec_fn": humgwan_unloading,
    },
    {
        "name": "점형블록", "spec": "300x300", "unit": "ea",
        "mat": 121, "lab": 883, "exp": 86, "total": 1090,
        "basis": ("오매칭(점수0.48, 고무타일 ㎡단가) 제외 · 표준시장단가 CLG802100810 "
                  "시각장애인용 점자블록 포장 12,112원/㎡ × 0.09㎡/개(300×300) = 1,090원/개(검증 권장)"),
    },
    {
        "name": "선형블록", "spec": "300x300", "unit": "ea",
        "mat": 121, "lab": 883, "exp": 86, "total": 1090,
        "basis": ("오매칭(점수0.48, 고무타일 ㎡단가) 제외 · 표준시장단가 CLG802100810 "
                  "시각장애인용 점자블록 포장 12,112원/㎡ × 0.09㎡/개(300×300) = 1,090원/개(검증 권장)"),
    },
]


def clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").strip()


def match_rule(name: str, spec: str, unit: str):
    n = clean(name)
    s = clean(spec)
    u = clean(unit)
    for rule in CORRECTIONS:
        if rule["name"] in n and (not rule["spec"] or rule["spec"] in s) and rule["unit"] == u:
            if rule.get("spec_fn"):
                mat, lab, exp, total, basis = rule["spec_fn"](s)
                return {"mat": mat, "lab": lab, "exp": exp, "total": total, "basis": basis}
            return {"mat": rule["mat"], "lab": rule["lab"], "exp": rule["exp"],
                    "total": rule["total"], "basis": rule["basis"]}
    return None


def main() -> None:
    if not UNMATCHED_IL.exists():
        print(f"없음: {UNMATCHED_IL}")
        sys.exit(1)
    wb = load_workbook(UNMATCHED_IL)
    ws = wb["일위대가산출"]
    col = {clean(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    c_name, c_spec, c_unit, c_qty = col["품명"], col["규격"], col["단위"], col["수량"]
    c_mat = col.get("재료단가") or col.get("재료비")
    c_lab = col.get("노무비") or col.get("노무단가")
    c_exp = col.get("경비") or col.get("경비단가")
    c_tot = col["합계단가"]
    c_sug = col.get("제시단가")
    c_sugamt = col.get("제시금액")
    c_conf = col["확정단가(입력)"]
    c_confamt = col.get("확정금액")
    c_basis = col.get("표준품셈·산출근거")

    log = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, c_name).value
        if not name:
            continue
        spec = ws.cell(r, c_spec).value
        unit = ws.cell(r, c_unit).value
        fix = match_rule(name, spec, unit)
        if not fix:
            continue
        qty = ws.cell(r, c_qty).value or 0
        try:
            qty = float(qty)
        except (TypeError, ValueError):
            qty = 0
        old = ws.cell(r, c_conf).value or ws.cell(r, c_tot).value
        ws.cell(r, c_mat).value = fix["mat"] or None
        ws.cell(r, c_lab).value = fix["lab"] or None
        ws.cell(r, c_exp).value = fix["exp"] or None
        ws.cell(r, c_tot).value = fix["total"]
        if c_sug:
            ws.cell(r, c_sug).value = fix["total"]
        if c_sugamt:
            ws.cell(r, c_sugamt).value = round(fix["total"] * qty)
        ws.cell(r, c_conf).value = fix["total"]
        if c_confamt:
            ws.cell(r, c_confamt).value = round(fix["total"] * qty)
        if c_basis:
            ws.cell(r, c_basis).value = fix["basis"]
        log.append((clean(ws.cell(r, col.get("파일", 2)).value), clean(name)[:22],
                    clean(spec)[:16], unit, qty, old, fix["total"], round(fix["total"] * qty)))

    wb.save(UNMATCHED_IL)
    print(f"교정 {len(log)}건 저장 → {UNMATCHED_IL.name}")

    # 품셈 재산출 근거 워크시트
    pwb = Workbook()
    pw = pwb.active
    pw.title = "오매칭교정"
    pw.append(["저신뢰 오매칭 과대단가 — 표준품셈·일위대가 재산출(검증 권장)"])
    pw["A1"].font = Font(bold=True, size=13)
    pw.append([])
    hdr = ["파일", "품명", "규격", "단위", "수량", "기존단가", "교정단가", "교정금액"]
    pw.append(hdr)
    for c in range(1, len(hdr) + 1):
        pw.cell(3, c).font = Font(bold=True)
        pw.cell(3, c).fill = PatternFill("solid", fgColor="FCE4D6")
    for row in sorted(log, key=lambda x: -(x[6] * x[4])):
        pw.append(list(row))
    for c in (6, 7, 8):
        for r in range(4, pw.max_row + 1):
            if isinstance(pw.cell(r, c).value, (int, float)):
                pw.cell(r, c).number_format = "#,##0"
    widths = [16, 22, 18, 6, 9, 14, 12, 16]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        pw.column_dimensions[get_column_letter(i)].width = w
    pwb.save(POOMSEM_OUT)
    print(f"품셈 재산출 근거 → {POOMSEM_OUT}")

    print("\n교정 내역:")
    for fl, nm, sp, u, qty, old, new, amt in log:
        try:
            oldv = f"{float(old):,.0f}"
        except (TypeError, ValueError):
            oldv = str(old)
        print(f"  [{fl}] {nm} {sp} {u} ×{qty:,.0f}: {oldv} → {new:,} (금액 {amt:,})")


if __name__ == "__main__":
    main()
