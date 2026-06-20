#!/usr/bin/env python3
"""검토필요 188건 → 일위대가 확정·환산·재산출 작업표.

각 _표준단가산출.xlsx 「검토필요」 시트에서 후보 단가·점수를 읽어
① 표준단가 확정 ② 단위환산 ③ 품셈·재매칭 ④ 수동검증 경로를 제시한다.
출력: 05_내역서/검토_일위대가산출.xlsx
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
import apply_standard_prices as asp  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
OUT = BASE / "검토_일위대가산출.xlsx"

SRC = [
    ("01 토목", "01_화성 청원지구 토목_표준단가산출.xlsx"),
    ("01 조경", "01_화성 청원지구 조경_표준단가산출.xlsx"),
    ("04 진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx"),
    ("05 회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx"),
    ("06 개발행위", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx"),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ROUTE_FILLS = {
    "① 표준단가 확정": PatternFill("solid", fgColor="E2EFDA"),
    "② 단위환산": PatternFill("solid", fgColor="DDEBF7"),
    "④ 1순위 채택+검증": PatternFill("solid", fgColor="FFF2CC"),
    "③ 품셈·재매칭": PatternFill("solid", fgColor="FCE4D6"),
}
MONEY = "#,##0"

CONVERT_UNITS = {"개소", "본", "회", "주", "ea", "EA", "Ton", "ton", "대"}
POOMSEM_KW = [
    ("운반", "표준품셈 자재운반(공통) — 거리·중량별 차량/인력운반"),
    ("접합", "표준품셈 해당 배관·관 접합 조항 — 직종 공량×시중노임"),
    ("부설", "표준품셈 해당 부설 조항 또는 표준시장단가 부설단가"),
    ("방제", "표준품셈 [조경] 병해충 방제 — 살수·살포 인부×노임+약제"),
    ("시험", "시험·검사 — 별도 단가표·발주처 기준 확인"),
]


def clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").replace("_x000D_", "").strip()


def core_name(name: str) -> str:
    s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", clean(name))
    return re.sub(r"^[.·\s]+", "", s).strip()


def name_tokens(name: str) -> list[str]:
    cn = core_name(name)
    toks = [t for t in re.split(r"[\s/()·]+", cn) if len(t) >= 2]
    return toks


def name_overlap(name: str, matched: str) -> bool:
    if not matched:
        return False
    toks = name_tokens(name)
    if not toks:
        return False
    blob = matched.replace(" ", "")
    hit = sum(1 for t in toks if t in blob or t in matched)
    return hit >= max(1, len(toks) // 2)


def conversion_note(unit: str, name: str, spec: str) -> str:
    u = clean(unit)
    if u in ("개소", "본"):
        if "접합" in name:
            return "개소→m 환산(접합부 길이 또는 1개소=규격길이)"
        if "맨홀" in name or "받이" in name:
            return "개소=1식 — 표준일위대가 '식' 또는 구성물 단위 확인"
        return "개소→m 또는 ea 환산 필요"
    if u in ("주",):
        return "수목 1주 — 조경일위·forestinfo 시세(재+노+경) 또는 표준단가 '주' 단위"
    if u in ("회",):
        return "1회=1식 시험·방제 — 품셈 회당 공량×노임"
    if u.lower() in ("ton", "t"):
        return "Ton→톤 운반 품셈(L=km) 적용"
    if u in ("EA", "ea"):
        return "EA→m 또는 1ea=규격길이(m) 환산(경계석·집수정 등)"
    if u in ("대",):
        return "1대=차량 1회분(운반 품셈)"
    return ""


def poomsem_hint(name: str) -> str:
    blob = core_name(name)
    for kw, ref in POOMSEM_KW:
        if kw in blob:
            return ref
    return "표준품셈·표준시장단가 해당 조항 재검색"


def classify_route(score: float | None, unit: str, name: str, matched: str) -> str:
    ov = name_overlap(name, matched)
    sc = float(score) if isinstance(score, (int, float)) else 0.0
    u = clean(unit)
    if any(k in name for k in ("운반", "시험", "검사")) and (sc < 0.65 or not ov):
        return "③ 품셈·재매칭"
    if sc >= 0.70 and ov:
        return "① 표준단가 확정"
    if u in CONVERT_UNITS and sc >= 0.56 and ov:
        return "② 단위환산"
    if sc >= 0.65 and ov:
        return "④ 1순위 채택+검증"
    if sc >= 0.56 and not ov:
        return "③ 품셈·재매칭"
    return "③ 품셈·재매칭"


def load_pool() -> list[dict]:
    market = asp.load_market_csv(asp.MARKET_2026, "표준시장단가2026")
    if not market:
        market = asp.load_prices()
    sijang = asp.load_market_csv(asp.SIJANG_2026, "시장시공가격2026")
    ildae = asp.load_ildae_prices()
    mulga = asp.load_mulga()
    jojadang = asp.load_jojadang()
    jogyeong = asp.load_landscape_ildae()
    return asp.precompute(market + sijang + ildae + mulga + jojadang + jogyeong)


def topn_alts(item: dict, prices: list[dict], n: int = 3) -> str:
    terms = asp.extract_search_terms(item["name"], item["spec"])
    req = asp.required_keywords(item)
    item_kw = asp.kwset(f"{item['name']} {item['spec']}")
    item["_linear"] = asp.is_linear_piece(item)
    cand = [p for p in prices if asp.unit_compatible(item["unit"], p["unit"], item["_linear"])]
    best: dict[int, tuple[float, dict, str]] = {}
    for term in terms:
        kw = item_kw | asp.kwset(term)
        for p in cand:
            s = asp.score_match(item, p, term, kw, req)
            if s <= 0:
                continue
            pid = id(p)
            if pid not in best or s > best[pid][0]:
                best[pid] = (s, p, term)
    ranked = sorted(best.values(), key=lambda x: -x[0])[:n]
    parts = []
    for sc, p, _ in ranked:
        parts.append(f"{p['name'][:18]}({sc:.2f})={p['total']:,.0f}{p['unit']}")
    return " · ".join(parts)


def basis_text(route: str, code: str, matched: str, conv: str, hint: str, alts: str) -> str:
    bits = [f"경로={route}"]
    if code:
        bits.append(f"현재후보={code} {matched[:30]}")
    if conv:
        bits.append(conv)
    if route.startswith("③"):
        bits.append(hint)
    if alts and route.startswith("③"):
        bits.append(f"대안후보: {alts[:120]}")
    elif route.startswith("④") and alts:
        bits.append(f"2·3순위: {alts[:100]}")
    return " / ".join(bits)


def load_review_records(prices: list[dict]) -> list[dict]:
    records: list[dict] = []
    for label, fname in SRC:
        fp = WORK / fname
        if not fp.exists():
            continue
        wb = load_workbook(fp, read_only=True, data_only=True)
        if "검토필요" not in wb.sheetnames:
            wb.close()
            continue
        for r in wb["검토필요"].iter_rows(min_row=2, values_only=True):
            if not r or r[2] is None:
                continue
            name = clean(r[2])
            spec = clean(r[3])
            qty = r[4] if isinstance(r[4], (int, float)) else None
            unit = clean(r[5])
            score = r[6]
            code = clean(r[8]) if len(r) > 8 else ""
            matched_name = clean(r[9]) if len(r) > 9 else ""
            matched_spec = clean(r[10]) if len(r) > 10 else ""
            mat_u = float(r[11] or 0) if len(r) > 11 and r[11] is not None else 0.0
            lab_u = float(r[12] or 0) if len(r) > 12 and r[12] is not None else 0.0
            exp_u = float(r[13] or 0) if len(r) > 13 and r[13] is not None else 0.0
            tot_u = float(r[14] or 0) if len(r) > 14 and r[14] is not None else mat_u + lab_u + exp_u

            item = {"name": name, "spec": spec, "unit": unit, "qty": qty or 0, "section": clean(r[1])}
            route = classify_route(score, unit, name, matched_name)
            conv = conversion_note(unit, name, spec) if route.startswith("②") else ""
            hint = poomsem_hint(name) if route.startswith("③") else ""
            alts = ""
            if route.startswith("③") or (route.startswith("④") and not name_overlap(name, matched_name)):
                alts = topn_alts(item, prices)

            records.append({
                "file": label,
                "section": clean(r[1]),
                "row": r[0],
                "name": name,
                "spec": spec,
                "unit": unit,
                "qty": qty,
                "score": score,
                "route": route,
                "code": code,
                "matched_name": matched_name,
                "matched_spec": matched_spec,
                "mat_u": mat_u,
                "lab_u": lab_u,
                "exp_u": exp_u,
                "tot_u": tot_u,
                "conv": conv,
                "hint": hint,
                "alts": alts,
            })
        wb.close()
    records.sort(key=lambda x: (x["route"], x["file"], -(x["score"] or 0)))
    return records


def write_workbook(records: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "일위대가산출"
    headers = [
        "경로", "파일", "행", "공종", "품명", "규격", "단위", "수량", "매칭점수",
        "현재_단가코드", "현재_매칭품명", "현재_매칭규격",
        "재료단가", "노무단가", "경비단가", "합계단가", "합계금액",
        "확정단가(입력)", "확정금액", "환산·근거 / 대안후보",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    route_counts: dict[str, int] = {}
    for rec in records:
        route = rec["route"]
        route_counts[route] = route_counts.get(route, 0) + 1
        amt = rec["tot_u"] * rec["qty"] if rec["qty"] and rec["tot_u"] else None
        basis = basis_text(route, rec["code"], rec["matched_name"], rec["conv"], rec["hint"], rec["alts"])
        row = [
            route, rec["file"], rec["row"], rec["section"], rec["name"], rec["spec"],
            rec["unit"], rec["qty"], rec["score"],
            rec["code"], rec["matched_name"], rec["matched_spec"],
            rec["mat_u"], rec["lab_u"], rec["exp_u"], rec["tot_u"], amt,
            rec["tot_u"] if route.startswith("①") else "",
            amt if route.startswith("①") else "",
            basis,
        ]
        ws.append(row)
        ridx = ws.max_row
        fill = ROUTE_FILLS.get(route, ROUTE_FILLS["③ 품셈·재매칭"])
        for c in range(1, len(headers) + 1):
            ws.cell(ridx, c).fill = fill
            ws.cell(ridx, c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in (13, 14, 15, 16, 17, 18, 19):
            if isinstance(ws.cell(ridx, c).value, (int, float)):
                ws.cell(ridx, c).number_format = MONEY

    ws.freeze_panes = "A2"
    widths = [14, 12, 6, 14, 26, 22, 6, 8, 8, 12, 22, 22, 10, 10, 10, 11, 12, 11, 12, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 요약 시트
    sm = wb.create_sheet("경로별요약", 0)
    sm.append(["검토필요 → 일위대가 산출 경로"])
    sm.append([])
    sm.append(["경로", "건수", "처리 방법"])
    sm.append(["① 표준단가 확정", route_counts.get("① 표준단가 확정", 0),
               "현재 매칭품명·합계단가 그대로 확정(점수≥0.70·품명 일치)"])
    sm.append(["② 단위환산", route_counts.get("② 단위환산", 0),
               "개소·주·EA 등 → m·식 환산 후 표준단가 적용"])
    sm.append(["④ 1순위 채택+검증", route_counts.get("④ 1순위 채택+검증", 0),
               "현재 후보 확인 후 확정단가 입력"])
    sm.append(["③ 품셈·재매칭", route_counts.get("③ 품셈·재매칭", 0),
               "오매칭 의심 — 품셈+노임 또는 대안후보 재선택"])
    sm.append([])
    sm.append(["합계", len(records), ""])
    sm["A1"].font = Font(bold=True, size=13)
    for r in (3, 4, 5, 6, 7):
        for c in range(1, 4):
            sm.cell(r, c).font = Font(bold=True) if c == 1 else Font()
    sm.column_dimensions["A"].width = 18
    sm.column_dimensions["B"].width = 8
    sm.column_dimensions["C"].width = 70

    info = wb.create_sheet("안내")
    for line in [
        ["검토 188건 — 일위대가 확정·환산·재산출"],
        [],
        ["색상", "연두=① 확정 / 하늘=② 환산 / 노랑=④ 검증 / 주황=③ 재산출"],
        ["확정단가", "「확정단가(입력)」·「확정금액」 열에 최종 단가 기입 → 표준단가산출·총괄표 반영"],
        ["현재 후보", "검토필요 시트의 매칭품명·합계단가 — 금액 합계에 이미 포함(178건)"],
        ["주의", "점수 0.70+라도 품명 불일치(예: 거푸집→케이블트레이)는 ③ 재산출"],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=14)
    info.column_dimensions["A"].width = 12
    info.column_dimensions["B"].width = 85

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(f"저장: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")


def main() -> None:
    prices = load_pool()
    print(f"단가 풀: {len(prices):,}건")
    records = load_review_records(prices)
    print(f"검토 항목: {len(records)}건")
    for route in sorted({r["route"] for r in records}):
        n = sum(1 for r in records if r["route"] == route)
        print(f"  {route}: {n}건")
    write_workbook(records)


if __name__ == "__main__":
    main()
