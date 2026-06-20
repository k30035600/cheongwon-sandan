#!/usr/bin/env python3
"""워크스페이스 내 추출 가능한 일위대가(품셈 기반)를 일위대가DB 폴더로 통합 저장.

추출 원천:
  - 02 전기(지구내): 일대목차(호표 단가표)·일위대가(상세 구성)·합산자재·단가조사
  - 07 폐기물: 경비(26.협회단가)
  - (옵션) tools/_poomsem_cache 의 조달청 표준일위대가/표준시장단가 등 외부 파일

원본 xlsx가 Excel에서 열려 잠긴 경우 tools/_poomsem_cache 의 동명 사본을 사용한다.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"  # 폴더 재편(2026-06-18): 공내역서·일위대가DB가 05_내역서 하위로 이동
CACHE = ROOT / "tools" / "_poomsem_cache"
OUT_DIR = BASE / "일위대가DB"
EXT_DIR = OUT_DIR / "_외부원본"  # 외부 표준일위대가 XLSX 투입 위치
STD_DIR = OUT_DIR / "표준일위대가_2026"  # 복사·정제한 조달청/국도 표준일위대가 정본
CODEBOOK = STD_DIR / "참고" / "표준공사코드(20260616).xlsx"  # 표준시장단가·시장시공가격 원천

# ingest_external_ildae 가 우리 산출물을 다시 읽지 않도록 제외
EXT_SKIP = {"일위대가_통합db", "_master_일위대가", "표준일위대가_2026_통합", "표준일위대가_2026_마스터"}

# 외부 표준일위대가 헤더 매핑 키워드 (레이아웃 무관 컬럼 탐지)
EXT_HEADER_KEYS = {
    "name": ["공종명", "품명", "명칭", "공 종 명", "명   칭"],
    "spec": ["규격", "규 격"],
    "unit": ["단위", "단 위"],
    "mat": ["재료비", "재 료 비"],
    "lab": ["노무비", "노 무 비"],
    "exp": ["경비", "경    비", "경 비"],
    "tot": ["합계", "합    계", "계"],
    "code": ["코드", "호표", "비고"],
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
MONEY_FMT = "#,##0.##"

MASTER_COLS = [
    "출처", "시트", "호표", "코드", "품명", "규격", "단위",
    "재료비단가", "노무비단가", "경비단가", "합계단가", "비고",
]


def open_xlsx(name: str):
    """원본 우선, 잠겨 있으면 캐시 사본 사용."""
    for path in (BASE / "공내역서" / name, ROOT / name, CACHE / name):
        if path.exists():
            try:
                return load_workbook(str(path), data_only=True), path
            except PermissionError:
                continue
    return None, None


def num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ---------- 02 전기 ----------
def extract_electric():
    wb, path = open_xlsx("02_화성 청원지구 전기설비.xlsx")
    if wb is None:
        print("02 전기 파일을 열 수 없음 (원본/캐시 모두 실패)")
        return {}, []
    src_label = "전기(02)"
    out: dict[str, list[list]] = {}
    master: list[list] = []

    # 일대목차: 호표별 일위대가 단가표
    if "일대목차" in wb.sheetnames:
        sh = wb["일대목차"]
        rows = [["번호", "코드", "호표", "품명", "규격", "단위",
                 "재료비단가", "노무비단가", "경비단가", "합계단가", "비고"]]
        for r in range(4, sh.max_row + 1):
            name = sh.cell(r, 4).value
            if not name:
                continue
            no = sh.cell(r, 1).value
            code = sh.cell(r, 2).value
            hop = sh.cell(r, 3).value
            spec = sh.cell(r, 5).value
            unit = sh.cell(r, 6).value
            mat = num(sh.cell(r, 8).value)
            lab = num(sh.cell(r, 11).value)
            exp = num(sh.cell(r, 13).value)
            tot = num(sh.cell(r, 16).value) or (mat + lab + exp)
            rows.append([no, code, hop, name, spec, unit, mat, lab, exp, tot, sh.cell(r, 17).value])
            master.append([src_label, "일대목차", hop, code, name, spec, unit, mat, lab, exp, tot, ""])
        out["전기(02)_일위대가_호표"] = rows

    # 일위대가: 호표별 구성내역(상세)
    if "일위대가" in wb.sheetnames:
        sh = wb["일위대가"]
        rows = [["번호", "구분", "코드", "품명", "규격", "단위", "수량",
                 "재료비단가", "재료비금액", "노무비단가", "노무비금액",
                 "경비단가", "경비금액", "합계금액", "비고"]]
        for r in range(4, sh.max_row + 1):
            kind = sh.cell(r, 2).value
            name = sh.cell(r, 4).value
            if not name and kind not in ("합계줄",):
                continue
            rows.append([
                sh.cell(r, 1).value, kind, sh.cell(r, 3).value, name,
                sh.cell(r, 5).value, sh.cell(r, 6).value, num(sh.cell(r, 7).value),
                num(sh.cell(r, 8).value), num(sh.cell(r, 9).value),
                num(sh.cell(r, 11).value), num(sh.cell(r, 12).value),
                num(sh.cell(r, 13).value), num(sh.cell(r, 14).value),
                num(sh.cell(r, 16).value), sh.cell(r, 17).value,
            ])
        out["전기(02)_일위대가_상세"] = rows

    # 합산자재: 자재 단가
    if "합산자재" in wb.sheetnames:
        sh = wb["합산자재"]
        rows = [["코드", "품명", "규격", "단위", "수량", "재료비단가", "합계단가", "비고"]]
        for r in range(4, sh.max_row + 1):
            name = sh.cell(r, 4).value
            if not name:
                continue
            code = sh.cell(r, 3).value
            spec = sh.cell(r, 5).value
            unit = sh.cell(r, 6).value
            mat = num(sh.cell(r, 8).value)
            tot = num(sh.cell(r, 12).value) or mat
            rows.append([code, name, spec, unit, num(sh.cell(r, 7).value), mat, tot, sh.cell(r, 13).value])
            master.append([src_label, "합산자재", "", code, name, spec, unit, mat, 0.0, 0.0, tot, "자재단가"])
        out["전기(02)_합산자재"] = rows

    # 단가조사: 자재·중기 단가조사
    if "단가조사" in wb.sheetnames:
        sh = wb["단가조사"]
        rows = [["코드", "품명", "규격", "단위", "적용단가", "거래가격", "물가정보", "물가자료"]]
        for r in range(4, sh.max_row + 1):
            name = sh.cell(r, 4).value
            if not name:
                continue
            rows.append([
                sh.cell(r, 3).value, name, sh.cell(r, 5).value, sh.cell(r, 6).value,
                num(sh.cell(r, 7).value), num(sh.cell(r, 10).value),
                num(sh.cell(r, 12).value), num(sh.cell(r, 14).value),
            ])
        out["전기(02)_단가조사"] = rows

    print(f"02 전기 추출: {path.name} → 시트 {len(out)}개")
    return out, master


# ---------- 07 폐기물 ----------
def extract_waste():
    src = BASE / "공내역서" / "07_화성 청원지구 건설폐기물처리.XLS"
    out: dict[str, list[list]] = {}
    master: list[list] = []
    if not src.exists():
        return out, master
    wb = xlrd.open_workbook(str(src))
    sh = wb.sheet_by_name("경비")
    rows = [["코드", "품명", "규격", "단위", "재료비단가", "노무비단가", "경비단가", "합계단가", "비고"]]
    for r in range(2, sh.nrows):
        name = str(sh.cell_value(r, 0)).strip()
        if not name:
            continue
        code = ""
        for c in range(sh.ncols):
            v = str(sh.cell_value(r, c)).strip()
            if v.startswith("G"):
                code = v
                break
        spec = str(sh.cell_value(r, 1)).strip()
        unit = str(sh.cell_value(r, 2)).strip()
        mat = num(sh.cell_value(r, 3))
        lab = num(sh.cell_value(r, 4))
        exp = num(sh.cell_value(r, 5))
        tot = num(sh.cell_value(r, 6)) or (mat + lab + exp)
        note = str(sh.cell_value(r, 7)).strip()
        rows.append([code, name, spec, unit, mat, lab, exp, tot, note])
        master.append(["폐기물(07)", "경비(협회단가)", "", code, name, spec, unit, mat, lab, exp, tot, note])
    out["폐기물(07)_경비_협회단가"] = rows
    print(f"07 폐기물 추출: 경비 {len(rows)-1}건")
    return out, master


def _norm_header(v) -> str:
    return str(v or "").replace(" ", "").strip()


def _detect_columns(sheet) -> tuple[int, dict] | None:
    """헤더 행을 찾아 {필드: 0-based 컬럼} 매핑 반환. 못 찾으면 None."""
    for r in range(1, min(sheet.max_row, 15) + 1):
        cells = [_norm_header(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        mapping: dict[str, int] = {}
        for field, keys in EXT_HEADER_KEYS.items():
            nkeys = [k.replace(" ", "") for k in keys]
            for ci, val in enumerate(cells):
                if val and any(val == k or k in val for k in nkeys):
                    mapping.setdefault(field, ci)
                    break
        # 품명·단위·(재료비 또는 합계) 가 있으면 유효 헤더로 간주
        if "name" in mapping and "unit" in mapping and ("mat" in mapping or "tot" in mapping):
            return r, mapping
    return None


def ingest_external_ildae():
    """일위대가DB/_외부원본/ 의 외부 표준일위대가 XLSX를 헤더 기반으로 통합.

    파일명에 '일위대가' 포함 xlsx를 대상으로, 시트별로 헤더를 탐지해
    품명·규격·단위·재료비·노무비·경비·합계를 마스터로 적재한다.
    """
    out: dict[str, list[list]] = {}
    master: list[list] = []
    if not EXT_DIR.exists():
        return out, master
    files = [
        p for p in EXT_DIR.glob("*.xlsx")
        if "일위대가" in p.name and not p.name.startswith("~$")
        and p.stem.lower() not in EXT_SKIP
    ]
    for path in files:
        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            print(f"외부 파일 열기 실패: {path.name} ({e})")
            continue
        src_label = f"외부:{path.stem[:24]}"
        file_rows = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            det = _detect_columns(sheet)
            if not det:
                continue
            hdr_row, m = det
            rows = [["호표/코드", "품명", "규격", "단위", "재료비단가", "노무비단가", "경비단가", "합계단가"]]
            for r in range(hdr_row + 1, sheet.max_row + 1):
                name = sheet.cell(r, m["name"] + 1).value
                if not name or not str(name).strip():
                    continue
                spec = sheet.cell(r, m["spec"] + 1).value if "spec" in m else ""
                unit = sheet.cell(r, m["unit"] + 1).value if "unit" in m else ""
                mat = num(sheet.cell(r, m["mat"] + 1).value) if "mat" in m else 0.0
                lab = num(sheet.cell(r, m["lab"] + 1).value) if "lab" in m else 0.0
                exp = num(sheet.cell(r, m["exp"] + 1).value) if "exp" in m else 0.0
                tot = num(sheet.cell(r, m["tot"] + 1).value) if "tot" in m else (mat + lab + exp)
                code = sheet.cell(r, m["code"] + 1).value if "code" in m else ""
                if not unit and mat == 0 and lab == 0 and exp == 0 and tot == 0:
                    continue  # 구분/소계/빈 행 스킵
                rows.append([code, name, spec, unit, mat, lab, exp, tot])
                master.append([src_label, sn[:20], code, code, name, spec, unit, mat, lab, exp, tot, path.name])
                file_rows += 1
            if len(rows) > 1:
                key = f"외부_{path.stem[:16]}_{sn[:14]}"
                out[key] = rows
        wb.close()
        print(f"외부 표준일위대가 통합: {path.name} → {file_rows}건")
    if not files:
        print(f"외부 표준일위대가 없음 — '{EXT_DIR}' 에 XLSX를 넣으면 자동 통합됩니다.")
    return out, master


def _gongjong_label(path: Path) -> str:
    """파일명에서 공종/원천 라벨 추출."""
    n = path.stem
    if "조달청2026_" in n:
        return n.split("조달청2026_", 1)[1].split("_", 1)[0].replace("공사", "")
    if "국도건설공사" in n:
        tail = n.split("_")[-1]            # 예: '1.토공'
        return "국도-" + (tail.split(".", 1)[-1] if "." in tail else tail)
    if "하천" in n:
        return "하천"
    if "항만" in n:
        return "항만"
    if "교육청" in n and "토목조경" in n:
        return "교육청-토목조경"
    if "교육청" in n:
        return "교육청-" + n.split("_")[-2] if "_" in n else "교육청"
    return n[:12]


def ingest_standard_ildae():
    """표준일위대가_2026/ 의 조달청·국도·하천·항만 표준일위대가(정본)를 통합.

    메인 시트(일위대가 / 중기단가산출 / (참고)일위대가)만 적재하고
    상세·산출식 시트는 제외한다. 헤더(품명·단위·합계)를 동적으로 탐지한다.
    """
    out: dict[str, list[list]] = {}
    master: list[list] = []
    if not STD_DIR.exists():
        return out, master
    files = sorted(
        p for p in STD_DIR.rglob("*.xlsx")
        if not p.name.startswith("~$") and "표준공사코드" not in p.name
    )
    for path in files:
        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
        except Exception as e:  # noqa: BLE001
            print(f"표준일위대가 열기 실패: {path.name} ({e})")
            continue
        label = _gongjong_label(path)
        # 내역서 프로그램 export(교육청 등)는 '일위대가목록'이 정제된 단가표.
        # 같은 파일의 상세 '일위대가' 시트(자원 구성)는 제외한다.
        if "일위대가목록" in wb.sheetnames:
            target_sheets = ["일위대가목록"]
        else:
            target_sheets = [
                sn for sn in wb.sheetnames
                if ("일위대가" in sn or "중기단가산출" in sn)
                and "상세" not in sn and "산출식" not in sn
            ]
        for sn in target_sheets:
            sheet = wb[sn]
            cmap: dict[str, int] = {}      # 헤더명 -> 0-based 컬럼
            header_found = False
            rows = [["코드", "번호", "품명", "규격", "단위",
                     "재료비단가", "노무비단가", "경비단가", "합계단가", "비목구분"]]
            cnt = 0
            for tup in sheet.iter_rows(values_only=True):
                if not header_found:
                    vals = [_norm_header(v) for v in tup]
                    if "품명" in vals and "단위" in vals and "합계" in vals:
                        for ci, v in enumerate(vals):
                            if v:
                                cmap.setdefault(v, ci)
                        header_found = True
                    continue

                def col(*names):
                    for nm in names:
                        if nm in cmap:
                            return cmap[nm]
                    return None

                def get(ci):
                    return tup[ci] if (ci is not None and ci < len(tup)) else None

                name = get(col("품명"))
                if not name or not str(name).strip():
                    continue
                code = get(col("일위대가코드", "중기단가산출코드", "코드")) or ""
                mat = num(get(col("재료비")))
                lab = num(get(col("노무비")))
                exp = num(get(col("경비")))
                tot = num(get(col("합계"))) or (mat + lab + exp)
                if not code and tot == 0:
                    continue
                spec = get(col("규격")) or ""
                unit = get(col("단위")) or ""
                no = get(col("번호")) or ""
                div = get(col("비목구분")) or ""
                nm = str(name).strip()
                rows.append([code, no, nm, spec, unit, mat, lab, exp, tot, div])
                master.append([f"표준일위대가:{label}", sn[:20], "", code, nm,
                               spec, unit, mat, lab, exp, tot, path.stem[:28]])
                cnt += 1
            if cnt:
                key = f"표준_{label}_{sn}".replace(" ", "").replace("(", "").replace(")", "")[:40]
                out[key] = rows
                print(f"표준일위대가 통합: {path.name} [{sn}] → {cnt}건")
        wb.close()
    return out, master


def ingest_market_price():
    """표준공사코드 파일의 '표준시장단가'(2026 하반기)·'시장시공가격' 시트를 CSV로 추출.

    내역서 단가 매칭 풀에 쓸 최신 시장단가(재료비/노무비/경비/합계, 세부공종코드 포함).
    시트 차원이 과대 선언(빈 행 포함)되어 있어 연속 공백 다수면 조기 종료한다.
    """
    if not CODEBOOK.exists():
        print(f"표준공사코드 파일 없음: {CODEBOOK}")
        return
    wb = load_workbook(str(CODEBOOK), data_only=True, read_only=True)
    cols = ["발표일자", "코드", "품명", "규격", "단위",
            "재료비", "노무비", "경비", "합계", "공사구분", "적용조건"]
    for sheet_name, out_name in [("표준시장단가", "표준시장단가_2026.csv"),
                                 ("시장시공가격", "시장시공가격_2026.csv")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = [cols]
        blanks = 0
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not r or not r[0]:
                blanks += 1
                if blanks > 200 and len(rows) > 1:
                    break
                continue
            blanks = 0
            # 공사구분, 발표일자, 코드, 품명, 규격, 단위, 재, 노, 경, 합, 적용조건
            rows.append([
                str(r[1] or ""), str(r[2] or ""), str(r[3] or "").strip(),
                str(r[4] or "").strip(), str(r[5] or "").strip().replace(" ", ""),
                num(r[6]), num(r[7]), num(r[8]), num(r[9]),
                str(r[0] or ""), str(r[10] or "") if len(r) > 10 else "",
            ])
        STD_DIR.mkdir(parents=True, exist_ok=True)
        path = STD_DIR / out_name
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)
        print(f"시장단가 추출: {sheet_name} → {out_name} ({len(rows)-1}건)")
    wb.close()


def write_standard_outputs(std_tables: dict[str, list[list]], std_master: list[list]):
    """표준일위대가 전용 마스터 CSV·통합 XLSX 를 STD_DIR 에 저장."""
    if not std_master:
        return None
    STD_DIR.mkdir(parents=True, exist_ok=True)
    # 마스터 CSV
    mpath = STD_DIR / "표준일위대가_2026_마스터.csv"
    with mpath.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(MASTER_COLS)
        w.writerows(std_master)
    # 통합 XLSX (공종/시트별 + _통합마스터)
    wb = Workbook()
    ws = wb.active
    ws.title = "_통합마스터"
    ws.append(MASTER_COLS)
    for row in std_master:
        ws.append(row)
    for name, rows in std_tables.items():
        base = name[:31]
        t, i = base, 1
        while t in wb.sheetnames:
            t = f"{base[:28]}_{i}"
            i += 1
        sh = wb.create_sheet(t)
        for row in rows:
            sh.append(row)
    for sheet in wb.worksheets:
        for c in range(1, sheet.max_column + 1):
            sheet.cell(1, c).font = Font(bold=True)
            sheet.cell(1, c).fill = HEADER_FILL
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = MONEY_FMT
    dest = STD_DIR / "표준일위대가_2026_통합.xlsx"
    try:
        wb.save(dest)
    except PermissionError:
        dest = dest.with_name(dest.stem + "_업데이트.xlsx")
        wb.save(dest)
    return dest


def write_csvs(tables: dict[str, list[list]]):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for name, rows in tables.items():
        path = OUT_DIR / f"{name}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)


def write_master_csv(master: list[list]):
    path = OUT_DIR / "_master_일위대가.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(MASTER_COLS)
        w.writerows(master)
    return len(master)


def write_combined_xlsx(tables: dict[str, list[list]], master: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "_통합마스터"
    ws.append(MASTER_COLS)
    for row in master:
        ws.append(row)
    for name, rows in tables.items():
        title = name.split("_", 1)[-1][:31] if "_" in name else name[:31]
        # 시트명 충돌 방지
        base = name[:31]
        t = base
        i = 1
        while t in wb.sheetnames:
            t = f"{base[:28]}_{i}"
            i += 1
        sh = wb.create_sheet(t)
        for row in rows:
            sh.append(row)

    for sheet in wb.worksheets:
        for c in range(1, sheet.max_column + 1):
            sheet.cell(1, c).font = Font(bold=True)
            sheet.cell(1, c).fill = HEADER_FILL
        sheet.freeze_panes = "A2"
        for col in sheet.columns:
            width = min(46, max(len(str(c.value or "")) for c in col) + 2)
            sheet.column_dimensions[col[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = MONEY_FMT
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    dest = OUT_DIR / "일위대가_통합DB.xlsx"
    try:
        wb.save(dest)
        return dest
    except PermissionError:
        alt = dest.with_name(dest.stem + "_업데이트.xlsx")
        wb.save(alt)
        return alt


def write_readme(tables, master_count, std_tables=None, std_count=0):
    counts = "\n".join(f"- `{n}.csv` — {len(rows)-1}건" for n, rows in tables.items())
    std_counts = ""
    if std_tables:
        std_counts = "\n".join(
            f"  - `{n}` — {len(rows)-1}건" for n, rows in std_tables.items()
        )
    text = f"""# 일위대가DB — 화성 청원지구

공내역서에서 **추출한 일위대가(품셈 기반)** 와 외부 **조달청·국도 표준일위대가(2026)** 를
통합 저장한 폴더.

## 구성 파일

- `_master_일위대가.csv` — **전체 통합 마스터 ({master_count}건)** = 내장(전기·폐기물) + 표준일위대가.
  컬럼: {', '.join(MASTER_COLS)}
- `일위대가_통합DB.xlsx` — 내장(전기·폐기물) 원천 통합(시트별) + `_통합마스터`
- `표준일위대가_2026/` — 외부 표준일위대가 **정본(복사본)** + 전용 산출물
  - `표준일위대가_2026_마스터.csv` — 표준일위대가만 통합 ({std_count}건)
  - `표준일위대가_2026_통합.xlsx` — 공종/시트별 + `_통합마스터`
  - `조달청2026_*공사_표준일위대가.xlsx` — 토목·건축·기계·전기·통신(2026 상반기)
  - `국도하천항만/` — 국도건설공사 토목(1~9)·하천(24년)·항만(20년) 단가산출서
  - `참고/표준공사코드(20260616).xlsx` — 표준공사코드·순수자원·표준시장단가·시장시공가격
{counts}

## 원천 및 범위

| 원천 | 내용 | 비고 |
|---|---|---|
| **전기(02)** | 일대목차(호표 일위대가)·상세·합산자재·단가조사 | 품셈 기반, 재/노/경 단가 |
| **폐기물(07)** | 경비(26. 협회단가) | 폐기물 운반·처리 Ton당 |
| **조달청 표준일위대가(2026)** | 토목·건축·기계·전기·통신 (업데이트 26년 상반기) | 일위대가 시트 합계단가 |
| **국도·하천·항만 단가산출서** | 국도 토목 1~9공종, 하천(24년)·항만(20년) | 중기단가산출·(참고)일위대가 |

## 토목 01·04·05·06 적용

원본 XLS에 일위대가·품셈 시트가 **없어** 자체 추출은 불가하지만,
위 **표준일위대가_2026 마스터**(특히 `표준일위대가:토목`·`국도-*`)를
**품명+규격** 유사도로 매칭해 재료비·노무비·경비를 산출한다.

## 표준일위대가 갱신 방법

1. 조달청/국토부에서 최신 표준일위대가 XLSX를 받아 `일위대가DB/_외부원본/` 하위에 둔다.
2. 정본을 `일위대가DB/표준일위대가_2026/` 로 복사한다(공종 폴더 또는 `국도하천항만/`).
3. `python -X utf8 tools/build_ildae_db.py` 실행 → 마스터·통합 XLSX 자동 재생성.
   - 메인 시트(`일위대가`·`중기단가산출`·`(참고)일위대가`)만 적재, `상세`·`산출식`은 제외.
   - 헤더(품명·단위·합계)를 자동 탐지하므로 레이아웃이 달라도 매핑된다.

> ⚠️ 조달청 NPCCS·민간 자료실·공공데이터포털은 **로그인/공동인증서**가 필요해
> 에이전트가 자동 다운로드할 수 없다. 사용자가 직접 받아 위 폴더에 넣는다.

## 단가 구조

- **표준일위대가**: 1단위(M·M2·M3·회 등) 재료비·노무비·경비·합계 → 내역서 수량 × 단가.
- **일대목차/일위대가 상세**: 호표별 단가 + 구성 자원(노무·자재·중기) 공량·단가(품셈 근거).

## 표준일위대가 시트별 건수

{std_counts}

끝.
"""
    (OUT_DIR / "README.md").write_text(text, encoding="utf-8")


def main():
    EXT_DIR.mkdir(parents=True, exist_ok=True)
    tables: dict[str, list[list]] = {}
    master: list[list] = []
    for extractor in (extract_electric, extract_waste, ingest_external_ildae):
        t, m = extractor()
        tables.update(t)
        master.extend(m)

    # 표준일위대가(조달청·국도·하천·항만) — 대용량이라 STD_DIR 별도 산출 + 글로벌 마스터 합류
    std_tables, std_master = ingest_standard_ildae()
    # 최신 2026 표준시장단가·시장시공가격 추출(매칭 풀 보강용)
    ingest_market_price()

    write_csvs(tables)
    std_xlsx = write_standard_outputs(std_tables, std_master)
    cnt = write_master_csv(master + std_master)   # 내장 + 표준 통합 마스터
    xlsx = write_combined_xlsx(tables, master)    # 내장(전기·폐기물)만 → 경량 유지
    write_readme(tables, cnt, std_tables, len(std_master))
    print(f"\n일위대가DB 저장 위치: {OUT_DIR}")
    print(f"통합 마스터(내장+표준): {cnt}건 / 내장시트 {len(tables)}개")
    print(f"내장 통합 xlsx: {xlsx.name}")
    if std_xlsx:
        print(f"표준일위대가 통합: {len(std_master)}건 / 시트 {len(std_tables)}개 → {std_xlsx.name}")


if __name__ == "__main__":
    main()
