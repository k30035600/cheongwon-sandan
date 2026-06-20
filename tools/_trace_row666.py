#!/usr/bin/env python3
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(__file__).resolve().parents[1] / "05_내역서"

# 1) 검토_전체 원본 행 666
wb = openpyxl.load_workbook(BASE / "검토_전체.xlsx", read_only=True, data_only=True)
ws = wb.active
for r in ws.iter_rows(min_row=4, values_only=True):
    if r[1] == 666:
        print("검토_전체 행666:")
        print("  공종 =", r[2])
        print("  품명 =", r[3])
        print("  규격 =", r[4])
        print("  단위 =", r[5], "/ 수량 =", r[6])
        print("  점수 =", r[7])
        print("  매칭품명 =", r[10] if len(r) > 10 else "")
        print("  합계단가 =", r[12] if len(r) > 12 else "")
        break
wb.close()

# 2) 확정 출처 추적
for fn in [
    "검토_공종별_일위대가산출.xlsx", "검토_토공_일위대가산출.xlsx",
    "검토_일위대가산출.xlsx", "미매칭_일위대가산출.xlsx",
]:
    p = BASE / fn
    if not p.exists():
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["일위대가산출"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}
    rcol = idx.get("행")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if rcol is not None and row[rcol] == 666:
            print(f"\n{fn} 행666:")
            for key in ["파일", "공종", "품명", "규격", "단위", "경로",
                        "DB_품명", "DB_단위", "합계단가", "확정단가(입력)"]:
                if key in idx:
                    print(f"  {key} = {row[idx[key]]}")
    wb.close()
