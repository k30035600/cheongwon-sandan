#!/usr/bin/env python3
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(__file__).resolve().parents[1] / "05_내역서"
WORK = BASE / "내역서작업"

# 통합내역에서 살수·세륜세차 행 찾기
wb = openpyxl.load_workbook(WORK / "01_화성 청원지구 토목_표준단가산출.xlsx",
                            read_only=True, data_only=True)
w = wb["통합내역"]
h = [c.value for c in next(w.iter_rows(min_row=1, max_row=1))]
ci = {k: h.index(k) for k in ["행", "공종", "공종명", "규격", "수량", "단위",
                              "단가코드", "매칭품명", "합계단가", "합계금액"]}
print("=== 통합내역: 살수·세륜세차·부대공 주요 행 ===")
for r in w.iter_rows(min_row=2, values_only=True):
    nm = str(r[ci["공종명"]] or "")
    if any(k in nm for k in ["살수", "세륜", "세차"]):
        print(f"  행{r[ci['행']]} [{r[ci['공종']]}] {nm[:20]} "
              f"규격={str(r[ci['규격']])[:14]} 수량={r[ci['수량']]}{r[ci['단위']]} "
              f"코드={r[ci['단가코드']]} 단가={r[ci['합계단가']]:,} 금액={r[ci['합계금액']]:,.0f}")
wb.close()

# 검토_전체에서 살수·세륜세차 행 번호
print("\n=== 검토_전체: 살수·세륜세차 행 번호 ===")
wb = openpyxl.load_workbook(BASE / "검토_전체.xlsx", read_only=True, data_only=True)
ws = wb.active
for r in ws.iter_rows(min_row=4, values_only=True):
    nm = str(r[3] or "")
    if any(k in nm for k in ["살수", "세륜", "세차"]):
        print(f"  행{r[1]} 공종={r[2]} 품명={nm[:20]} 규격={str(r[4])[:14]} "
              f"단위={r[5]} 수량={r[6]} 합계단가={r[12] if len(r)>12 else ''}")
wb.close()
