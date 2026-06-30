"""조경시설물_미매칭_20건.xlsx — 재/노/경 헤더·수식·합계행·비고(푸른조경)."""
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

XLSX = Path("05_내역서/내역서작업/조경/조경시설물/조경시설물_미매칭_20건.xlsx")
TOT_FILL = PatternFill("solid", fgColor="D9E1F2")
TOT_FONT = Font(bold=True)
MONEY = "#,##0"

# J~Q: 재단가·노단가·경단가·합단가 | 재·노·경·합계
HEADERS = {
    10: "재단가",
    11: "노단가",
    12: "경단가",
    13: "합단가",
    14: "재",
    15: "노",
    16: "경",
    17: "합계",
}


def apply(wb) -> int:
    ws = wb["미매칭20건"]
    # 합계행 제외하고 데이터 마지막 행
    last = ws.max_row
    while last > 1 and ws.cell(last, 4).value and "합계" in str(ws.cell(last, 4).value):
        last -= 1
    first = 2

    for c, h in HEADERS.items():
        ws.cell(1, c).value = h

    for r in range(first, last + 1):
        ws.cell(r, 13).value = f"=J{r}+K{r}+L{r}"
        ws.cell(r, 14).value = f"=F{r}*J{r}"
        ws.cell(r, 15).value = f"=F{r}*K{r}"
        ws.cell(r, 16).value = f"=F{r}*L{r}"
        ws.cell(r, 17).value = f"=N{r}+O{r}+P{r}"
        ws.cell(r, 19).value = "푸른조경"

    tot = last + 1
    ws.cell(tot, 4).value = "★ 합계"
    ws.cell(tot, 14).value = f"=SUM(N{first}:N{last})"
    ws.cell(tot, 15).value = f"=SUM(O{first}:O{last})"
    ws.cell(tot, 16).value = f"=SUM(P{first}:P{last})"
    ws.cell(tot, 17).value = f"=SUM(Q{first}:Q{last})"
    for c in range(1, 20):
        ws.cell(tot, c).fill = TOT_FILL
        ws.cell(tot, c).font = TOT_FONT

    for r in range(1, tot + 1):
        for c in range(10, 18):
            ws.cell(r, c).number_format = MONEY

    if "요약" in wb.sheetnames:
        ws2 = wb["요약"]
        for r in range(1, ws2.max_row + 1):
            if ws2.cell(r, 1).value == "푸른조경 합계(20건)":
                ws2.cell(r, 1).value = "합계(20건)"
                ws2.cell(r, 2).value = f"='미매칭20건'!Q{tot}"
                ws2.cell(r, 2).number_format = MONEY
            if ws2.cell(r, 1).value == "분류":
                hdr_row = r
                break
        else:
            hdr_row = None
        if hdr_row:
            r = hdr_row + 1
            while r <= ws2.max_row and ws2.cell(r, 1).value and ws2.cell(r, 1).value != "합계":
                cat = ws2.cell(r, 1).value
                ws2.cell(r, 3).value = (
                    f"=SUMIF('미매칭20건'!$H${first}:$H${last},\"{cat}\",'미매칭20건'!$Q${first}:$Q${last})"
                )
                ws2.cell(r, 3).number_format = MONEY
                r += 1
            ws2.cell(r, 1).value = "합계"
            ws2.cell(r, 2).value = f"=COUNTA('미매칭20건'!A{first}:A{last})"
            ws2.cell(r, 3).value = f"='미매칭20건'!Q{tot}"
            ws2.cell(r, 3).number_format = MONEY

    return tot


def main():
    wb = load_workbook(XLSX)
    tot = apply(wb)
    wb.save(XLSX)
    print(f"저장: {XLSX} (합계 Q{tot})")
    wb.close()


if __name__ == "__main__":
    main()
