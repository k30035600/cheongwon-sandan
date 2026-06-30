#!/usr/bin/env python3
"""공종별(회전·진입·토목·조경·지구) 차수별 직접비 + 간접비 원가 HTML 생성."""
from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from calc_overhead import (  # noqa: E402
    CIVIL_JOGYEONG_RATES,
    CIVIL_TOMOK_RATES,
    compute_cost_statement_civil,
)

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "05_내역서" / "내역서작업"
_t = date.today()
GEN_DATE = f"{_t.year}. {_t.month}. {_t.day}."

CSS = """
body{font-family:"Malgun Gothic","맑은 고딕",system-ui,sans-serif;line-height:1.65;
  color:#1c2430;max-width:1080px;margin:0 auto;padding:28px 20px 60px;font-size:15px;background:#f7f9fc;}
h1{font-size:22px;color:#1f3a5f;border-bottom:3px solid #1f3a5f;padding-bottom:10px;}
h2{font-size:18px;color:#1f3a5f;margin-top:28px;border-left:5px solid #2c5aa0;padding-left:10px;}
h3{font-size:16px;color:#2c5aa0;margin-top:22px;}
table{width:100%;border-collapse:collapse;background:#fff;margin:12px 0;font-size:14px;
  border:1px solid #d9e1ec;border-radius:8px;overflow:hidden;}
th,td{padding:8px 11px;border-bottom:1px solid #e3e9f2;vertical-align:top;text-align:left;}
th{background:#eaf0f8;color:#1f3a5f;text-align:center;font-weight:700;}
th.num,td.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap;}
td.c{text-align:center;white-space:nowrap;}
tr.total td{font-weight:700;background:#fff2cc;}
tr.subtotal td{font-weight:600;background:#f4f6fa;}
.l1{font-weight:600;}
.l2{padding-left:22px;color:#334155;}
.l3{padding-left:40px;color:#475569;font-size:13px;}
blockquote{background:#eef4fb;border:1px solid #c5d0e6;border-radius:8px;padding:10px 16px;margin:14px 0;}
.muted{color:#6b7787;font-size:13px;}
"""

GROUPS: list[dict] = [
    {
        "slug": "회전교차로",
        "title": "05 회전교차로",
        "src": "05_화성 청원로(회전교차로).XLS",
        "parts": [
            ("회전교차로", "05_화성 청원로(회전교차로)_표준단가산출.xlsx", CIVIL_TOMOK_RATES, "토목"),
        ],
    },
    {
        "slug": "진입도로",
        "title": "04 진입도로",
        "src": "04_화성 청원지구 진입도로 실시설계.XLS",
        "parts": [
            ("진입도로", "04_화성 청원지구 진입도로 실시설계_표준단가산출.xlsx", CIVIL_TOMOK_RATES, "토목"),
        ],
    },
    {
        "slug": "토목",
        "title": "01 토목",
        "src": "01_화성 청원지구 토목.XLS",
        "parts": [
            ("토목", "01_화성 청원지구 토목_표준단가산출.xlsx", CIVIL_TOMOK_RATES, "토목"),
        ],
    },
    {
        "slug": "조경",
        "title": "01 조경",
        "src": "01_화성 청원지구 조경.XLS",
        "parts": [
            ("조경", "01_화성 청원지구 조경_표준단가산출.xlsx", CIVIL_JOGYEONG_RATES, "조경"),
        ],
    },
    {
        "slug": "지구단위",
        "title": "06 산업유통형 개발행위",
        "src": "06_화성 청원지구 산업유통형 개발행위.XLS",
        "parts": [
            ("토목", "06_화성 청원지구 산업유통형 개발행위_표준단가산출.xlsx", CIVIL_TOMOK_RATES, "토목"),
        ],
    },
    {
        "slug": "폐기물",
        "title": "07 건설폐기물처리",
        "src": "07_화성 청원지구 건설폐기물처리.XLS",
        "parts": [
            ("폐기물", "07_화성 청원지구 건설폐기물처리_표준단가산출.xlsx", CIVIL_TOMOK_RATES, "토목"),
        ],
    },
]


def won(v) -> str:
    try:
        return f"{int(round(float(v or 0))):,}"
    except (TypeError, ValueError):
        return "0"


def eok(v) -> str:
    return f"{float(v or 0) / 1e8:,.2f}"


def section_level(name: str) -> int:
    n = name.strip()
    if n.startswith("..."):
        return 3
    if re.match(r"^[가나다]\)", n) or re.match(r"^\([가나다]\)", n):
        return 2
    if re.match(r"^\d+\.", n) or n.startswith("▣") or n.startswith("▶"):
        return 1
    return 1


def section_sort_key(name: str) -> tuple:
    n = name.strip()
    m = re.match(r"^(\d+)\.", n)
    if m:
        return (0, int(m.group(1)), n)
    if n.startswith("▣") or n.startswith("▶"):
        return (1, 999, n)
    return (2, 0, n)


def load_hapgye(path: Path) -> tuple[list[dict], dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["합계요약"]
    rows: list[dict] = []
    grand = {"mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0, "matched": 0, "total": 0}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        if name in ("★ 매칭합계", "※ 미매칭"):
            if name.startswith("★"):
                grand = {
                    "mat": float(ws.cell(r, 5).value or 0),
                    "lab": float(ws.cell(r, 6).value or 0),
                    "exp": float(ws.cell(r, 7).value or 0),
                    "sum": float(ws.cell(r, 8).value or 0),
                    "matched": int(ws.cell(r, 2).value or 0),
                    "total": int(ws.cell(r, 3).value or 0),
                }
            continue
        rows.append({
            "name": name,
            "level": section_level(name),
            "matched": int(ws.cell(r, 2).value or 0),
            "total": int(ws.cell(r, 3).value or 0),
            "mat": float(ws.cell(r, 5).value or 0),
            "lab": float(ws.cell(r, 6).value or 0),
            "exp": float(ws.cell(r, 7).value or 0),
            "sum": float(ws.cell(r, 8).value or 0),
        })
    rows.sort(key=lambda s: section_sort_key(s["name"]))
    wb.close()
    return rows, grand


def render_direct_table(sections: list[dict], grand: dict) -> str:
    lines = [
        "<table><thead><tr>",
        "<th>차수·공종</th><th class='num'>매칭/전체</th>",
        "<th class='num'>재료비</th><th class='num'>노무비</th><th class='num'>경비</th><th class='num'>직접비 합계</th>",
        "</tr></thead><tbody>",
    ]
    for s in sections:
        if s["sum"] == 0 and s["matched"] == 0 and s["level"] >= 2:
            continue
        cls = f"l{s['level']}"
        tr_cls = " class='subtotal'" if s["level"] == 1 else ""
        lines.append(
            f"<tr{tr_cls}><td class='{cls}'>{s['name']}</td>"
            f"<td class='num'>{s['matched']}/{s['total']}</td>"
            f"<td class='num'>{won(s['mat'])}</td><td class='num'>{won(s['lab'])}</td>"
            f"<td class='num'>{won(s['exp'])}</td><td class='num'>{won(s['sum'])}</td></tr>"
        )
    rate = grand["matched"] / grand["total"] if grand["total"] else 0
    lines.append(
        f"<tr class='total'><td>① 직접공사비 합계</td>"
        f"<td class='num'>{grand['matched']}/{grand['total']} ({rate*100:.1f}%)</td>"
        f"<td class='num'>{won(grand['mat'])}</td><td class='num'>{won(grand['lab'])}</td>"
        f"<td class='num'>{won(grand['exp'])}</td><td class='num'>{won(grand['sum'])}</td></tr>"
    )
    lines.append("</tbody></table>")
    return "\n".join(lines)


def render_cost_table(cs: dict) -> str:
    lines = [
        "<table><thead><tr><th>차수</th><th>항목</th><th class='num'>금액(원)</th><th>산식</th></tr></thead><tbody>",
    ]
    for row in cs["rows"]:
        step = row["step"]
        name = row["name"]
        amt = row["amount"]
        formula = row.get("formula") or ""
        bold = row.get("bold", False)
        total = row.get("total", False)
        tr = " class='total'" if bold and total else (" class='subtotal'" if bold else "")
        indent = "&nbsp;" * 4 if name.startswith("·") else ""
        disp = indent + name
        lines.append(
            f"<tr{tr}><td>{step or ''}</td><td>{disp}</td>"
            f"<td class='num'>{won(amt)}</td><td class='muted'>{formula or ''}</td></tr>"
        )
    lines.append("</tbody></table>")
    return "\n".join(lines)


def build_group_html(cfg: dict) -> str:
    parts_html: list[str] = []
    sum_mat = sum_lab = sum_exp = sum_direct = 0.0
    sum_dogeup = 0.0
    cost_blocks: list[str] = []

    for gj, fname, rates, rate_label in cfg["parts"]:
        path = WORK / gj / fname
        if not path.exists():
            parts_html.append(f"<p class='muted'>파일 없음: {path}</p>")
            continue
        sections, grand = load_hapgye(path)
        cs = compute_cost_statement_civil(grand["mat"], grand["lab"], grand["exp"], rates)
        sum_mat += grand["mat"]
        sum_lab += grand["lab"]
        sum_exp += grand["exp"]
        sum_direct += grand["sum"]
        sum_dogeup += cs["contract"]

        parts_html.append(f"<h2>{cfg['title']} — {gj} ({fname})</h2>")
        parts_html.append(
            f"<p class='muted'>간접비 요율: {rate_label}공사 간접공사비 적용기준(2026.4.13) · "
            f"매칭 {grand['matched']}/{grand['total']}건</p>"
        )
        parts_html.append("<h3>1. 차수별 직접공사비</h3>")
        parts_html.append(render_direct_table(sections, grand))
        parts_html.append("<h3>2. 간접비·도급액 (원가계산서)</h3>")
        parts_html.append(render_cost_table(cs))
        cost_blocks.append(
            f"<li><strong>{gj}</strong> — 직접 {won(grand['sum'])}원 · 도급 {won(cs['contract'])}원</li>"
        )

    summary = ""
    if len(cfg["parts"]) > 1:
        summary = (
            "<h2>구분 합산 (참고)</h2>"
            "<blockquote><ul>"
            + "".join(cost_blocks)
            + f"</ul><p class='muted'>직접비 합 {won(sum_direct)}원 · "
            f"도급액 합 {won(sum_dogeup)}원 — "
            "간접비 요율이 토목/조경으로 달라 단순 합산 도급은 참고치입니다.</p></blockquote>"
        )

    body = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>화성 청원지구 — {cfg['title']} 직접·간접비</title>
<style>{CSS}</style></head><body>
<h1>화성 청원지구 — {cfg['title']}</h1>
<ul>
  <li><strong>원본 공내역서</strong>: <code>{cfg['src']}</code></li>
  <li><strong>작성일</strong>: {GEN_DATE}</li>
  <li><strong>직접비</strong>: 표준단가산출 「합계요약」 차수별 재·노·경</li>
  <li><strong>간접비</strong>: 토목공사·조경공사 간접공사비 적용기준(2026.4.13) 원가계산서</li>
</ul>
{"".join(parts_html)}
{summary}
<p class="muted">끝.</p>
</body></html>"""
    return body


def main() -> None:
    for cfg in GROUPS:
        html = build_group_html(cfg)
        if cfg["slug"] == "지구단위":
            out_dir = WORK / "토목"
            out_name = "지구단위_직접간접비.html"
        else:
            out_dir = WORK / cfg["slug"]
            out_name = f"{cfg['slug']}_직접간접비.html"
        out_dir.mkdir(parents=True, exist_ok=True)
        out = out_dir / out_name
        out.write_text(html, encoding="utf-8")
        print(f"  {out.relative_to(ROOT)}")


if __name__ == "__main__":
    print("=== 공종별 직접·간접비 HTML ===")
    main()
