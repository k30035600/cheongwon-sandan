#!/usr/bin/env python3
"""세흥건설 ↔ 삼아건설(청원지구) 견적 비교 → 07_타견적/견적비교(세흥_삼아건설).xlsx.

참조: 견적비교(세흥_근일).xlsx 와 동일 3시트 구조
  1) 원가계산서 — 직접·간접비·도급액 총괄 비교
  2) 내역서     — 공종별(내역서 No × 공종) 비교
  3) 세부내역서 — 품목별 세부 비교
금액차이 = 세흥 − 삼아건설.
"""
from __future__ import annotations

import collections
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
SEHUNG = ROOT / "07_타견적" / "세흥건설(청원지구).xlsx"
SAMA = ROOT / "07_타견적" / "삼아건설(청원지구).xlsx"
DST = ROOT / "07_타견적" / "견적비교(세흥_삼아건설).xlsx"

# 삼아건설 워크북은 금액이 '수식'으로만 저장(캐시값 없음)돼 data_only 로드 시 None →
# 리프의 단가(E/G/I)·수량(C) 리터럴에서 금액을 재계산하고, 원가 카스케이드도
# build_singyu 의 법정·조달청 요율을 그대로 사용해 원가계산서와 동일하게 복원한다.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_singyu import (  # noqa: E402
    R_INDLAB, R_SANJAE, R_GOYONG, R_GUNGANG, R_YEONGUM, R_JANGGI,
    R_SANAN, R_GITA, R_GISTGYE, R_HWANGYEONG, R_ILBAN, R_IYUN,
)

CLASS_NAME = {
    "01": "01. 지구단위계획(세흥) ↔ 01 토목·조경(삼아건설)",
    "04": "04. 진입도로",
    "05": "05. 회전교차로",
    "06": "06. 개발행위",
    "07": "07. 건설폐기물",
}
CLASS_ORDER = ["01", "04", "05", "06", "07"]
CLASS_KEYWORD = {
    "지구단위계획": "01", "토목·조경": "01", "진입도로": "04",
    "회전교차로": "05", "개발행위": "06", "폐기물": "07",
}

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment("center", "center", wrap_text=True)
RIGHT = Alignment("right", "center")
LEFT = Alignment("left", "center", wrap_text=True)
F_HDR = Font(bold=True, size=10, color="FFFFFF")
FILL_HDR = PatternFill("solid", fgColor="404E67")
FILL_SE = PatternFill("solid", fgColor="DDEBF7")
FILL_SAMA = PatternFill("solid", fgColor="E2EFDA")   # 삼아건설: 연녹
FILL_DF = PatternFill("solid", fgColor="FCE4D6")
FILL_CLS = PatternFill("solid", fgColor="FFF2CC")
FILL_GNG = PatternFill("solid", fgColor="FBE5D6")
FILL_SUB = PatternFill("solid", fgColor="F2F2F2")

# 견적비교(세흥_근일).xlsx 동일 레이아웃 + B4 용지
PAPER_B4 = 12  # Excel B4(JIS)
SHEET_LAYOUT = {
    "원가계산서": {
        "widths": [12, 24, 18, 16, 18, 16, 13, 89.375],
        "freeze": "A3",
        "orientation": "portrait",
    },
    "내역서": {
        "widths": [34, 7, 6, 13, 13, 12, 14, 13, 13, 12, 14, 13, 13, 12, 14, 16],
        "freeze": "A3",
        "orientation": "landscape",
    },
    "세부내역서": {
        "widths": [66, 7.875, 6.25, 11.5, 12.75, 11.5, 12.875, 11.5,
                   13, 13, 13, 12.25, 13, 11.5, 12.25, 14.375],
        "freeze": "A540",
        "orientation": "landscape",
    },
}


def apply_sheet_layout(ws, key: str) -> None:
    """열 너비·고정창·기본행높이·B4 용지(견적비교 세흥_근일 동일)."""
    cfg = SHEET_LAYOUT[key]
    ws.sheet_format.defaultRowHeight = 16.5
    for i, w in enumerate(cfg["widths"], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = cfg["freeze"]
    ws.page_setup.paperSize = PAPER_B4
    ws.page_setup.orientation = cfg["orientation"]


def n(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def z(v):
    x = n(v)
    return 0.0 if x is None else x


def clean_lbl(s):
    """표시용 라벨 선두의 공백·점(., .., ... 등)을 제거.
    번호표(가)., 1) 등)는 유지. 예: '... 가). 중기운반비' → '가). 중기운반비'."""
    return re.sub(r"^[\s.]+", "", str(s or ""))


def norm(s: str) -> str:
    return re.sub(r"\s+", "", str(s or "")).replace("·", "").lstrip(".")


def gongnorm(s: str) -> str:
    s = re.sub(r"^\d+\.\s*", "", (s or "").strip())
    return s.replace(" ", "")


def itemcore(s):
    s = str(s or "")
    s = re.sub(r"^[.\s]*", "", s)
    s = re.sub(r"^((\d+\)|[가-하]\)|[가-하]\.)\s*)+", "", s)
    return s.replace(" ", "")


def qkey(v):
    x = n(v)
    return None if x is None else round(x)


def charbag(*parts):
    s = "".join(str(p or "") for p in parts)
    s = re.sub(r"[\s().,·\-/]", "", s)
    s = re.sub(r"^((\d+\)|[가-하]\)|[가-하]\.))+", "", s)
    return collections.Counter(s)


def bagsim(a, b):
    if not a or not b:
        return 0.0
    inter = sum((a & b).values())
    union = sum((a | b).values())
    return inter / union if union else 0.0


CHAR_TH = 0.72


def money(cell, val, fill=None, bold=False, white=False):
    if isinstance(val, (int, float)):
        cell.value = round(val)
        cell.number_format = "#,##0"
    else:
        cell.value = val if val else ""
    cell.alignment = RIGHT
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if bold or white:
        cell.font = Font(bold=bold, color="FFFFFF" if white else "000000", size=10)


# ══════════════════════════════════════════════════════════
# 1) 원가계산서
# ══════════════════════════════════════════════════════════
def load_sehung_costsheet():
    ws = load_workbook(SEHUNG, data_only=True)["견적서"]
    d = {}
    d["재료비"] = z(ws.cell(row=3, column=6).value)
    d["직접노무비"] = z(ws.cell(row=3, column=8).value)
    d["직접경비"] = z(ws.cell(row=3, column=10).value)
    for r in range(44, 61):
        label = norm(ws.cell(row=r, column=1).value)
        if label:
            d[label] = z(ws.cell(row=r, column=12).value)
    return d


def read_sama_leaves():
    """삼아건설 내역서 리프를 단가(E/G/I)×수량(C)으로 재계산해 반환.
    (금액·합계 열은 수식이라 data_only 로 읽으면 None 이므로 단가로 산출)."""
    ws = load_workbook(SAMA, data_only=True)["내역서"]
    leaves = []
    cur_no = cur_gong = cur_gong_disp = None
    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a:
            continue
        s = str(a).strip()
        msec = re.match(r"【 (\d+)\.", s)
        if msec:
            cur_no = msec.group(1)
            continue
        if s.startswith("★"):
            continue
        qty = ws.cell(r, 3).value
        # 공종 헤더: 'N. 이름' + 수량 없음
        if re.match(r"^\d+\.\s", s) and qty in (None, ""):
            cur_gong = gongnorm(s)
            cur_gong_disp = s
            continue
        if qty in (None, ""):   # sub 헤더(가./1)./가) 등)
            continue
        q = n(qty) or 0.0
        mat = round(z(ws.cell(r, 5).value) * q)
        lab = round(z(ws.cell(r, 7).value) * q)
        exp = round(z(ws.cell(r, 9).value) * q)
        spec = str(ws.cell(r, 2).value or "")
        leaves.append({
            "no": cur_no, "gong": cur_gong, "gong_disp": cur_gong_disp,
            "name": s, "spec": spec, "qty": q, "unit": ws.cell(r, 4).value,
            "mat": float(mat), "lab": float(lab), "exp": float(exp),
            "sum": float(mat + lab + exp),
            "core": itemcore(s), "qk": qkey(qty), "bag": charbag(s, spec),
        })
    return leaves


def sama_cost_cascade(mat, lab, exp):
    """리프 합(재·직노·직경)에서 원가계산서 카스케이드를 build_singyu 요율로 복원.
    직접경비(exp)에는 이윤·일반관리비 인하분 안분배당이 이미 포함돼 있음."""
    indlab = lab * R_INDLAB
    lg = lab + indlab                       # 노무비계
    sanjae = lg * R_SANJAE
    goyong = lg * R_GOYONG
    gungang = lab * R_GUNGANG
    yeongum = lab * R_YEONGUM
    janggi = gungang * R_JANGGI
    sanan = (mat + lab) * R_SANAN
    gita = (mat + lg) * R_GITA
    direct = mat + lab + exp
    gistgye = direct * R_GISTGYE
    hwan = direct * R_HWANGYEONG
    sanchul = (sanjae + goyong + gungang + yeongum + janggi
               + sanan + gita + gistgye + hwan)
    gyeongbi = exp + sanchul
    sun = mat + lg + gyeongbi
    ilban = sun * R_ILBAN
    iyun = (lg + gyeongbi + ilban) * R_IYUN
    supply = sun + ilban + iyun
    vat = supply * 0.1
    return {
        "재료비": mat, "직접노무비": lab, "직접경비": exp, "직접공사비계": direct,
        "간접노무비": indlab, "산재보험료": sanjae, "고용보험료": goyong,
        "국민건강보험료": gungang, "국민연금보험료": yeongum,
        "노인장기요양보험료": janggi, "산업안전보건관리비": sanan,
        "기타경비": gita, "건설기계대여대금지급보증": gistgye, "환경보전비": hwan,
        "일반관리비": ilban, "이윤": iyun,
        "공급가액": supply, "부가가치세": vat, "도급액": supply + vat,
    }


def load_sama_costsheet():
    ws = load_workbook(SAMA, data_only=True)["원가계산서"]
    rates = {}
    for r in range(5, ws.max_row + 1):
        item = ws.cell(r, 2).value or ws.cell(r, 1).value
        if not item:
            continue
        rates[norm(str(item))] = ws.cell(r, 3).value or ""
    leaves = read_sama_leaves()
    mat = sum(x["mat"] for x in leaves)
    lab = sum(x["lab"] for x in leaves)
    exp = sum(x["exp"] for x in leaves)
    return sama_cost_cascade(mat, lab, exp), rates


def get(d, *keys):
    for k in keys:
        if k in d:
            return d[k]
    for k in keys:
        for dk, dv in d.items():
            if k in dk:
                return dv
    return None


def build_costsheet(wb, se, sama, sama_rates):
    ws = wb.create_sheet("원가계산서")
    ws.merge_cells("A1:A2"); ws["A1"] = "구분"
    ws.merge_cells("B1:B2"); ws["B1"] = "항목"
    ws.merge_cells("C1:D1"); ws["C1"] = "세흥건설"
    ws.merge_cells("E1:F1"); ws["E1"] = "삼아건설"
    ws.merge_cells("G1:G2"); ws["G1"] = "금액차이(세흥−삼아건설)"
    ws.merge_cells("H1:H2"); ws["H1"] = "산식·비고"
    ws["C2"] = "요율"; ws["D2"] = "금액"
    ws["E2"] = "요율"; ws["F2"] = "금액"
    for c in range(1, 9):
        for rr in (1, 2):
            cell = ws.cell(row=rr, column=c)
            cell.font = F_HDR; cell.fill = FILL_HDR
            cell.alignment = CENTER; cell.border = BORDER

    # 세흥건설(청원지구).xlsx 원가는 '간접노무비'를 별도 계상하지 않는다.
    # → 직접노무비를 19.7%로 분리하지 않고 견적서 원값을 그대로 반영한다.
    se_lab = get(se, "직접노무비") or 0.0
    se_direct = get(se, "직접공사비계") or 0.0
    se_indirect = get(se, "간접공사비계") or 0.0

    sama_direct = get(sama, "직접공사비계") or 0.0
    sama_supply = get(sama, "공급가액") or 0.0
    sama_indirect = sama_supply - sama_direct if sama_supply else None
    sama_vat = get(sama, "부가가치세")
    sama_dogup = get(sama, "도급액") or get(sama, "도급액(공사비계)")

    def HR(key):
        return sama_rates.get(norm(key), "")

    rows = [
        ("직접공사비", "재료비", "", get(se, "재료비"), HR("재료비"), get(sama, "재료비"), ""),
        ("직접공사비", "직접노무비", "", se_lab, HR("직접노무비"),
         get(sama, "직접노무비"), "세흥: 간접노무비 미분리(원값). 삼아건설: 간접노무비 별도 계상"),
        ("직접공사비", "직접경비", "", get(se, "직접경비"), HR("직접경비"),
         get(sama, "직접경비"), "세흥: 안전·품질관리비 포함"),
        ("직접공사비", "【 직접공사비 계 】", "", se_direct, HR("직접공사비계"),
         sama_direct, "세흥: 견적서 직접공사비계 원값"),
        ("간접비", "간접노무비", "—", None, HR("간접노무비"),
         get(sama, "간접노무비"), "세흥: 해당 없음(미계상) · 삼아건설: 직접노무비×19.7%"),
        ("간접비", "산재보험료", "노무비×3.56%", get(se, "산재보험료"), HR("산재보험료"),
         get(sama, "산재보험료"), "2026년 산재보험료율 고시(건설업)"),
        ("간접비", "고용보험료", "노무비×1.01%", get(se, "고용보험료"), HR("고용보험료"),
         get(sama, "고용보험료"), "삼아건설: 사업주 부담 1.15%"),
        ("간접비", "국민건강보험료", "직노×3.595%", get(se, "건강보험료"), HR("국민건강보험료"),
         get(sama, "국민건강보험료"), "2026년 법정요율"),
        ("간접비", "국민연금보험료", "직노×4.75%", get(se, "연금보험료"), HR("국민연금보험료"),
         get(sama, "국민연금보험료"), "2026년 법정요율"),
        ("간접비", "노인장기요양보험료", "건보×13.14%", get(se, "노인장기요양보험료"),
         HR("노인장기요양보험료"), get(sama, "노인장기요양보험료"), "2026년 법정요율"),
        ("간접비", "퇴직공제부금", "직노×2.30%", get(se, "퇴직공제부금"), "—", None, "세흥만 계상"),
        ("간접비", "산업안전보건관리비", "(재+직노)×2.53%+기초", get(se, "산업안전보건관리비"),
         HR("산업안전보건관리비"), get(sama, "산업안전보건관리비"),
         "삼아건설: 고시 토목 50억 이상 2.60%"),
        ("간접비", "건설기계대여대금지급보증", "직접공사비×0.40%", get(se, "건설기계대여대금지급보증서"),
         HR("건설기계대여대금지급보증"), get(sama, "건설기계대여대금지급보증"), ""),
        ("간접비", "하도급대금지급보증수수료", "직접공사비×0.080%", get(se, "하도급대금지급보증"),
         "—", None, "세흥만 계상"),
        ("간접비", "환경보전비", "직접공사비×0.6~0.9%", get(se, "환경보전비"), HR("환경보전비"),
         get(sama, "환경보전비"), "삼아건설 0.6% · 세흥 0.9%(도로)"),
        ("간접비", "기타경비(공과잡비)", "—", None, HR("기타경비"),
         get(sama, "기타경비"), "삼아건설: 조달청 기타경비 6.2%"),
        ("간접비", "일반관리비", "(재+노+경)×8.0%", get(se, "일반관리비"), HR("일반관리비"),
         get(sama, "일반관리비"), "세흥 8% · 삼아건설 5.5%"),
        ("간접비", "이윤", "(노+경+일관)×15.0%", get(se, "이윤"), HR("이윤"),
         get(sama, "이윤"), "세흥 15% · 삼아건설 8%"),
        ("간접비", "이행(준공)보증보험료", "—", None, HR("이행"),
         get(sama, "이행") or 0, "삼아건설: 미적용"),
        ("간접비", "【 간접비 계 】", "", se_indirect, "",
         sama_indirect, "삼아건설=(공급가액−직접공사비)"),
        ("합계", "【 공급가액(합계) 】", "", get(se, "합계"), HR("공급가액"), sama_supply, ""),
        ("합계", "부가가치세", "10%", get(se, "부가가치세"), HR("부가가치세"), sama_vat, ""),
        ("합계", "★ 도급액(공사비 계)", "", get(se, "공사비계"), HR("도급액"), sama_dogup,
         "세흥: 13,604,000,000 반올림"),
    ]

    r = 3
    prev_grp = None
    for grp, item, se_rate, sev, sama_rate, samav, note in rows:
        is_total = item.startswith("【") or item.startswith("★")
        gc = ws.cell(row=r, column=1, value=grp if grp != prev_grp else "")
        gc.alignment = CENTER; gc.border = BORDER
        if grp != prev_grp:
            gc.font = Font(bold=True)
        prev_grp = grp
        ic = ws.cell(row=r, column=2, value=item)
        ic.alignment = LEFT; ic.border = BORDER
        fill = FILL_SUB if is_total else None
        if is_total:
            ic.font = Font(bold=True)
        src = ws.cell(row=r, column=3, value=se_rate or "")
        src.alignment = CENTER; src.border = BORDER; src.font = Font(size=9, italic=True)
        if fill:
            src.fill = fill
        money(ws.cell(row=r, column=4), sev if sev is not None else "", fill or FILL_SE, bold=is_total)
        hrc = ws.cell(row=r, column=5, value=sama_rate or "")
        hrc.alignment = CENTER; hrc.border = BORDER; hrc.font = Font(size=9, italic=True)
        if fill:
            hrc.fill = fill
        money(ws.cell(row=r, column=6), samav if samav is not None else "", fill or FILL_SAMA, bold=is_total)
        diff = (sev - samav) if (sev is not None and samav is not None) else ""
        money(ws.cell(row=r, column=7), diff, fill or FILL_DF, bold=is_total)
        nc = ws.cell(row=r, column=8, value=note)
        nc.alignment = Alignment("left", "center", wrap_text=True)
        nc.border = BORDER
        if note and len(note) > 40:
            ws.row_dimensions[r].height = 33
        if is_total:
            nc.fill = FILL_SUB; nc.font = Font(bold=True); gc.fill = FILL_SUB
        r += 1

    note_row = r + 1
    ws.cell(row=note_row, column=2,
            value="※ 두 견적의 원가 산정 방식이 달라 항목 구성이 다릅니다. "
                  "세흥 직접경비에는 안전·품질관리비가 포함되고, 삼아건설은 기타경비·법정경비를 별도 계상합니다. "
                  f"세흥건설 견적서 원가는 간접노무비를 별도 계상하지 않으며(직접노무비 {se_lab:,.0f}원 원값), "
                  "삼아건설은 조달청 토목 19.7%로 간접노무비를 별도 계상합니다. "
                  "삼아건설 내역서 단가는 세흥건설 단가를 기준으로 산출.")
    ws.cell(row=note_row, column=2).font = Font(size=9, italic=True)
    ws.cell(row=note_row, column=2).alignment = Alignment("left", "center", wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=8)
    apply_sheet_layout(ws, "원가계산서")


# ══════════════════════════════════════════════════════════
# 2) 내역서(공종별)
# ══════════════════════════════════════════════════════════
def load_sehung_gong():
    ws = load_workbook(SEHUNG, data_only=True)["견적서"]
    rows = []
    for r in range(8, 42):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        if not c1 or not re.match(r"^\d+\.", str(c1)):
            continue
        no = re.match(r"^(\d+)\.", str(c1)).group(1)
        rows.append({
            "no": no, "gong": str(c2).strip(),
            "key": (no, gongnorm(str(c2))),
            "qty": ws.cell(row=r, column=3).value,
            "unit": ws.cell(row=r, column=4).value,
            "mat": z(ws.cell(row=r, column=6).value),
            "lab": z(ws.cell(row=r, column=8).value),
            "exp": z(ws.cell(row=r, column=10).value),
            "sum": z(ws.cell(row=r, column=12).value),
        })
    return rows


def load_sama_gong():
    """공종별(내역서 No × 공종) 집계 — 리프를 (no, 공종)으로 합산."""
    data = {}
    for x in read_sama_leaves():
        key = (x["no"], x["gong"])
        d = data.setdefault(key, {
            "no": x["no"], "gong": x["gong_disp"] or x["gong"],
            "mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0,
        })
        d["mat"] += x["mat"]; d["lab"] += x["lab"]
        d["exp"] += x["exp"]; d["sum"] += x["sum"]
    return data


def build_gong_sheet(wb, sehung, sama):
    ws = wb.create_sheet("내역서")
    ws.merge_cells("A1:A2"); ws["A1"] = "공종 / 규격"
    ws.merge_cells("B1:B2"); ws["B1"] = "수량"
    ws.merge_cells("C1:C2"); ws["C1"] = "단위"
    ws.merge_cells("D1:G1"); ws["D1"] = "세흥건설"
    ws.merge_cells("H1:K1"); ws["H1"] = "삼아건설"
    ws.merge_cells("L1:O1"); ws["L1"] = "금액차이(세흥−삼아건설)"
    ws.merge_cells("P1:P2"); ws["P1"] = "비고"
    for i, base in enumerate((4, 8, 12)):
        for j, name in enumerate(("재료비", "노무비", "경비", "합계")):
            ws.cell(row=2, column=base + j, value=name)
    for c in range(1, 17):
        for rr in (1, 2):
            cell = ws.cell(row=rr, column=c)
            cell.font = F_HDR; cell.fill = FILL_HDR
            cell.alignment = CENTER; cell.border = BORDER

    grand = {"se": [0.0] * 4, "sama": [0.0] * 4}
    matched = set()

    for no in CLASS_ORDER:
        se_rows = [x for x in sehung if x["no"] == no]
        if not se_rows:
            continue
        r0 = ws.max_row + 1
        ws.cell(row=r0, column=1, value=CLASS_NAME.get(no, no))
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=16)
        hc = ws.cell(row=r0, column=1)
        hc.font = Font(bold=True, size=10); hc.fill = FILL_CLS
        hc.alignment = LEFT; hc.border = BORDER

        sub = {"se": [0.0] * 4, "sama": [0.0] * 4}
        for x in se_rows:
            h = sama.get(x["key"]); matched.add(x["key"])
            rr = ws.max_row + 1
            ws.cell(row=rr, column=1, value=clean_lbl(x["gong"])).alignment = LEFT
            ws.cell(row=rr, column=1).border = BORDER
            for c, val in ((2, x["qty"]), (3, x["unit"])):
                cc = ws.cell(row=rr, column=c, value=val)
                cc.alignment = CENTER; cc.border = BORDER
            sev = [x["mat"], x["lab"], x["exp"], x["sum"]]
            hv = [h["mat"], h["lab"], h["exp"], h["sum"]] if h else None
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), sev[j], FILL_SE)
                money(ws.cell(row=rr, column=8 + j), hv[j] if h else "", FILL_SAMA)
                money(ws.cell(row=rr, column=12 + j), (sev[j] - hv[j]) if h else "", FILL_DF)
                sub["se"][j] += sev[j]
                if h:
                    sub["sama"][j] += hv[j]
            ws.cell(row=rr, column=16, value="" if h else "삼아건설 미매칭").border = BORDER
            ws.cell(row=rr, column=16).alignment = LEFT

        if se_rows:
            rr = ws.max_row + 1
            sc = ws.cell(row=rr, column=1, value=f"[ {no} 소계 ]")
            sc.font = Font(bold=True); sc.alignment = RIGHT; sc.fill = FILL_SUB; sc.border = BORDER
            for c in (2, 3):
                ws.cell(row=rr, column=c).fill = FILL_SUB; ws.cell(row=rr, column=c).border = BORDER
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), sub["se"][j], FILL_SUB, bold=True)
                money(ws.cell(row=rr, column=8 + j), sub["sama"][j], FILL_SUB, bold=True)
                money(ws.cell(row=rr, column=12 + j), sub["se"][j] - sub["sama"][j], FILL_SUB, bold=True)
                grand["se"][j] += sub["se"][j]; grand["sama"][j] += sub["sama"][j]
            ws.cell(row=rr, column=16).fill = FILL_SUB; ws.cell(row=rr, column=16).border = BORDER

    extra = [(k, v) for k, v in sama.items() if k not in matched]
    if extra:
        r0 = ws.max_row + 1
        ws.cell(row=r0, column=1, value="◆ 삼아건설 단독 (세흥 집계 미포함)")
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=16)
        hc = ws.cell(row=r0, column=1); hc.font = Font(bold=True, size=10)
        hc.fill = PatternFill("solid", fgColor="F8CBAD"); hc.alignment = LEFT; hc.border = BORDER
        for k, h in sorted(extra):
            rr = ws.max_row + 1
            ws.cell(row=rr, column=1, value=f"{h['no']} · {clean_lbl(h['gong'])}").alignment = LEFT
            ws.cell(row=rr, column=1).border = BORDER
            ws.cell(row=rr, column=2).border = BORDER; ws.cell(row=rr, column=3).border = BORDER
            hv = [h["mat"], h["lab"], h["exp"], h["sum"]]
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), "", FILL_SE)
                money(ws.cell(row=rr, column=8 + j), hv[j], FILL_SAMA)
                money(ws.cell(row=rr, column=12 + j), -hv[j], FILL_DF)
                grand["sama"][j] += hv[j]
            ws.cell(row=rr, column=16, value="세흥 없음").border = BORDER
            ws.cell(row=rr, column=16).alignment = LEFT

    rr = ws.max_row + 1
    tc = ws.cell(row=rr, column=1, value="★ 총 계 (공종 집계)")
    tc.font = Font(bold=True, size=11, color="FFFFFF"); tc.fill = FILL_HDR
    tc.alignment = RIGHT; tc.border = BORDER
    for c in (2, 3):
        ws.cell(row=rr, column=c).fill = FILL_HDR; ws.cell(row=rr, column=c).border = BORDER
    for j in range(4):
        money(ws.cell(row=rr, column=4 + j), grand["se"][j], FILL_HDR, white=True, bold=True)
        money(ws.cell(row=rr, column=8 + j), grand["sama"][j], FILL_HDR, white=True, bold=True)
        money(ws.cell(row=rr, column=12 + j), grand["se"][j] - grand["sama"][j], FILL_HDR, white=True, bold=True)
    ws.cell(row=rr, column=16).fill = FILL_HDR; ws.cell(row=rr, column=16).border = BORDER

    rr = ws.max_row + 2
    ws.cell(row=rr, column=1,
            value="※ 공종별 집계(내역서 No × 공종) 비교. 세흥 '수량 1·식'은 집계 표기. "
                  "세흥 직접공사비계에는 안전·품질관리비 425,371,600원이 별도 포함(공종 합계 밖). "
                  "삼아건설 직접경비에는 이윤·일반관리비 인하분이 내역서에 반영됨.")
    ws.cell(row=rr, column=1).font = Font(size=9, italic=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=16)
    apply_sheet_layout(ws, "내역서")
    return grand


# ══════════════════════════════════════════════════════════
# 3) 세부내역서
# ══════════════════════════════════════════════════════════
def load_sehung_items():
    ws = load_workbook(SEHUNG, data_only=True)["견적서"]
    items = []
    cur_no = cur_gong = cur_gong_disp = None
    for r in range(61, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        q = ws.cell(row=r, column=3).value
        s_sum = ws.cell(row=r, column=12).value
        if not c1:
            continue
        c1s = str(c1).strip()
        mcls = re.match(r"^\d+\.\s*(.+)$", c1s)
        if mcls and q in (None, "") and any(k in c1s for k in CLASS_KEYWORD):
            for k, no in CLASS_KEYWORD.items():
                if k in c1s:
                    cur_no = no
            continue
        if re.match(r"^\d+\.\s", c1s) and q in (None, "") and s_sum in (None, ""):
            cur_gong = gongnorm(c1s)
            cur_gong_disp = c1s
            continue
        if q not in (None, ""):
            items.append({
                "no": cur_no, "gong": cur_gong, "gong_disp": cur_gong_disp,
                "name": c1s, "spec": (str(c2).strip() if c2 else ""),
                "qty": n(q), "unit": ws.cell(row=r, column=4).value,
                "mat": z(ws.cell(row=r, column=6).value),
                "lab": z(ws.cell(row=r, column=8).value),
                "exp": z(ws.cell(row=r, column=10).value),
                "sum": z(ws.cell(row=r, column=12).value),
                "core": itemcore(c1s), "qk": qkey(q),
            })
    return items


def load_sama_items():
    exact = {}
    by_core = collections.defaultdict(list)
    by_gong = collections.defaultdict(list)
    for rec in read_sama_leaves():
        key = (rec["no"], rec["gong"], rec["core"], rec["qk"])
        exact.setdefault(key, rec)
        by_core[(rec["no"], rec["gong"], rec["core"])].append(rec)
        by_gong[(rec["no"], rec["gong"])].append(rec)
    return exact, by_core, by_gong


def match_sama(it, exact, by_core, by_gong):
    k4 = (it["no"], it["gong"], it["core"], it["qk"])
    if k4 in exact:
        return exact[k4], ""
    lst = by_core.get((it["no"], it["gong"], it["core"]))
    if lst and len(lst) == 1:
        note = "매칭(수량상이)" if lst[0]["qty"] != it["qty"] else ""
        return lst[0], note
    cands = by_gong.get((it["no"], it["gong"]))
    if cands:
        sb = charbag(it["name"], it["spec"])
        best, best_sc = None, 0.0
        for c in cands:
            sc = bagsim(sb, c["bag"])
            if it["qk"] is not None and c["qk"] == it["qk"]:
                sc += 0.15
            if sc > best_sc:
                best_sc, best = sc, c
        if best and best_sc >= CHAR_TH:
            tag = "매칭(문자유사)" if best_sc >= 1.0 else f"매칭(문자유사 {min(best_sc, 0.99):.2f})"
            return best, tag
    return None, "삼아건설 미매칭"


def build_detail_sheet(wb, items, exact, by_core, by_gong):
    ws = wb.create_sheet("세부내역서")
    ws.merge_cells("A1:A2"); ws["A1"] = "공종 / 규격"
    ws.merge_cells("B1:B2"); ws["B1"] = "수량"
    ws.merge_cells("C1:C2"); ws["C1"] = "단위"
    ws.merge_cells("D1:G1"); ws["D1"] = "세흥건설"
    ws.merge_cells("H1:K1"); ws["H1"] = "삼아건설"
    ws.merge_cells("L1:O1"); ws["L1"] = "금액차이(세흥−삼아건설)"
    ws.merge_cells("P1:P2"); ws["P1"] = "비고"
    for i, base in enumerate((4, 8, 12)):
        for j, name in enumerate(("재료비", "노무비", "경비", "합계")):
            ws.cell(row=2, column=base + j, value=name)
    for c in range(1, 17):
        for rr in (1, 2):
            cell = ws.cell(row=rr, column=c)
            cell.font = F_HDR; cell.fill = FILL_HDR
            cell.alignment = CENTER; cell.border = BORDER

    grand = {"se": [0.0] * 4, "sama": [0.0] * 4}
    n_match = n_total = 0
    tot_lab_raw = 0.0

    by_class = collections.OrderedDict()
    for no in CLASS_ORDER:
        by_class[no] = collections.OrderedDict()
    for it in items:
        no = it["no"]
        if no not in by_class:
            by_class[no] = collections.OrderedDict()
        by_class[no].setdefault(it["gong"], {"disp": it["gong_disp"], "rows": []})
        by_class[no][it["gong"]]["rows"].append(it)

    for no, gongs in by_class.items():
        if not gongs:
            continue
        r0 = ws.max_row + 1
        ws.cell(row=r0, column=1, value=CLASS_NAME.get(no, f"{no}."))
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=16)
        hc = ws.cell(row=r0, column=1)
        hc.font = Font(bold=True, size=11); hc.fill = FILL_CLS
        hc.alignment = LEFT; hc.border = BORDER
        cls_sub = {"se": [0.0] * 4, "sama": [0.0] * 4}

        for gong, info in gongs.items():
            r1 = ws.max_row + 1
            ws.cell(row=r1, column=1, value=f"  {clean_lbl(info['disp'] or gong)}")
            ws.merge_cells(start_row=r1, start_column=1, end_row=r1, end_column=16)
            gc = ws.cell(row=r1, column=1)
            gc.font = Font(bold=True, size=10); gc.fill = FILL_GNG
            gc.alignment = LEFT; gc.border = BORDER
            gong_sub = {"se": [0.0] * 4, "sama": [0.0] * 4}

            for it in info["rows"]:
                n_total += 1
                h, note = match_sama(it, exact, by_core, by_gong)
                if h:
                    n_match += 1
                rr = ws.max_row + 1
                nm = clean_lbl(it["name"]) + (f"  · {it['spec']}" if it["spec"] else "")
                ws.cell(row=rr, column=1, value=nm).alignment = LEFT
                ws.cell(row=rr, column=1).border = BORDER
                qc = ws.cell(row=rr, column=2, value=it["qty"])
                qc.alignment = CENTER; qc.border = BORDER
                if isinstance(it["qty"], float) and it["qty"].is_integer():
                    qc.number_format = "#,##0"
                uc = ws.cell(row=rr, column=3, value=it["unit"])
                uc.alignment = CENTER; uc.border = BORDER
                # 세흥은 간접노무비를 분리하지 않으므로 노무비 원값 그대로 사용
                tot_lab_raw += it["lab"]
                se_v = [it["mat"], it["lab"], it["exp"], it["sum"]]
                hv = [h["mat"], h["lab"], h["exp"], h["sum"]] if h else None
                for j in range(4):
                    money(ws.cell(row=rr, column=4 + j), se_v[j], FILL_SE)
                    money(ws.cell(row=rr, column=8 + j), hv[j] if h else "", FILL_SAMA)
                    money(ws.cell(row=rr, column=12 + j), (se_v[j] - hv[j]) if h else "", FILL_DF)
                    gong_sub["se"][j] += se_v[j]
                    if h:
                        gong_sub["sama"][j] += hv[j]
                nc = ws.cell(row=rr, column=16, value=note)
                nc.alignment = LEFT; nc.border = BORDER

            rr = ws.max_row + 1
            sc = ws.cell(row=rr, column=1, value=f"    [ {clean_lbl(info['disp'] or gong)} 소계 ]")
            sc.font = Font(bold=True, size=9); sc.alignment = RIGHT
            sc.fill = FILL_SUB; sc.border = BORDER
            for c in (2, 3):
                ws.cell(row=rr, column=c).fill = FILL_SUB; ws.cell(row=rr, column=c).border = BORDER
            for j in range(4):
                money(ws.cell(row=rr, column=4 + j), gong_sub["se"][j], FILL_SUB, bold=True)
                money(ws.cell(row=rr, column=8 + j), gong_sub["sama"][j], FILL_SUB, bold=True)
                money(ws.cell(row=rr, column=12 + j), gong_sub["se"][j] - gong_sub["sama"][j], FILL_SUB, bold=True)
                cls_sub["se"][j] += gong_sub["se"][j]
                cls_sub["sama"][j] += gong_sub["sama"][j]
            ws.cell(row=rr, column=16).fill = FILL_SUB; ws.cell(row=rr, column=16).border = BORDER

        rr = ws.max_row + 1
        sc = ws.cell(row=rr, column=1, value=f"[ {no} 소계 ]")
        sc.font = Font(bold=True); sc.alignment = RIGHT; sc.fill = FILL_SUB; sc.border = BORDER
        for c in (2, 3):
            ws.cell(row=rr, column=c).fill = FILL_SUB; ws.cell(row=rr, column=c).border = BORDER
        for j in range(4):
            money(ws.cell(row=rr, column=4 + j), cls_sub["se"][j], FILL_SUB, bold=True)
            money(ws.cell(row=rr, column=8 + j), cls_sub["sama"][j], FILL_SUB, bold=True)
            money(ws.cell(row=rr, column=12 + j), cls_sub["se"][j] - cls_sub["sama"][j], FILL_SUB, bold=True)
            grand["se"][j] += cls_sub["se"][j]; grand["sama"][j] += cls_sub["sama"][j]
        ws.cell(row=rr, column=16).fill = FILL_SUB; ws.cell(row=rr, column=16).border = BORDER

    rr = ws.max_row + 1
    tc = ws.cell(row=rr, column=1, value="★ 총 계 (세부내역)")
    tc.font = Font(bold=True, size=11, color="FFFFFF"); tc.fill = FILL_HDR
    tc.alignment = RIGHT; tc.border = BORDER
    for c in (2, 3):
        ws.cell(row=rr, column=c).fill = FILL_HDR; ws.cell(row=rr, column=c).border = BORDER
    for j in range(4):
        money(ws.cell(row=rr, column=4 + j), grand["se"][j], FILL_HDR, white=True, bold=True)
        money(ws.cell(row=rr, column=8 + j), grand["sama"][j], FILL_HDR, white=True, bold=True)
        money(ws.cell(row=rr, column=12 + j), grand["se"][j] - grand["sama"][j], FILL_HDR, white=True, bold=True)
    ws.cell(row=rr, column=16).fill = FILL_HDR; ws.cell(row=rr, column=16).border = BORDER

    rr = ws.max_row + 2
    ws.cell(row=rr, column=1,
            value=f"※ 세흥건설은 간접노무비를 별도 계상하지 않아 노무비를 원값 그대로 표기(직접노무비 계 {tot_lab_raw:,.0f}원). "
                  f"세흥 품목 {n_total}건 중 삼아건설 매칭 {n_match}건({n_match/n_total*100:.0f}%). "
                  "삼아건설 내역서는 세흥 단가 기준 산출. 매칭 1차(공종·품명·수량)·2차(품명)·3차(문자유사).")
    ws.cell(row=rr, column=1).font = Font(size=9, italic=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=16)
    apply_sheet_layout(ws, "세부내역서")
    return grand, n_match, n_total


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    se_cost = load_sehung_costsheet()
    sama_cost, sama_rates = load_sama_costsheet()
    build_costsheet(wb, se_cost, sama_cost, sama_rates)

    se_gong = load_sehung_gong()
    sama_gong = load_sama_gong()
    ggrand = build_gong_sheet(wb, se_gong, sama_gong)

    items = load_sehung_items()
    exact, by_core, by_gong = load_sama_items()
    dgrand, n_match, n_total = build_detail_sheet(wb, items, exact, by_core, by_gong)

    wb.save(DST)
    print(f"OK  {DST.name}  시트: {wb.sheetnames}")
    print("원가계산서  세흥 도급액 {:,.0f} · 삼아건설 도급액 {:,.0f} · 차이 {:,.0f}".format(
        se_cost.get("공사비계", 0) or 0,
        get(sama_cost, "도급액") or 0,
        (se_cost.get("공사비계", 0) or 0) - (get(sama_cost, "도급액") or 0)))
    print("내역서 총계  세흥 {:,.0f} · 삼아건설 {:,.0f} · 차이 {:,.0f}".format(
        ggrand["se"][3], ggrand["sama"][3], ggrand["se"][3] - ggrand["sama"][3]))
    print(f"세부내역  세흥 {n_total}건 · 삼아건설 매칭 {n_match}건 ({n_match/n_total*100:.1f}%)")
    print("세부총계  세흥 {:,.0f} · 삼아건설 {:,.0f}".format(dgrand["se"][3], dgrand["sama"][3]))


if __name__ == "__main__":
    main()
