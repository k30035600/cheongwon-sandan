#!/usr/bin/env python3
"""10톤 기동식 크레인 — 표준품셈 기계경비 + 시중 임대 일위대가(6일·1식).

전기 02 내역 177행 「크레인 10톤」(공원 통신·양중, 수량 6·단위 공란·미산출) 반영.
출력: 05_내역서/내역서작업/크레인10톤_일위대가_6일기준.xlsx
"""
from __future__ import annotations

import csv
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
NOIM_CSV = BASE / "일위대가DB" / "시중노임_2026.csv"
OUT = WORK / "크레인10톤_일위대가_6일기준.xlsx"
ELECTRIC = WORK / "02_화성 청원지구 전기설비_표준단가산출.xlsx"

# 산출 파라미터 (실무·견적 확인 후 조정)
MACHINE_PRICE = 150_000_000  # 10톤 타이어식 크레인 예시
DEPR_PER_HR = 0.000035  # 기계가격 × 계수 = 시간당 손료(원)
HOURS_PER_DAY = 8
FUEL_L_PER_HR = 6.0
DIESEL_WON_PER_L = 1_500  # 경유 단가(예시·유류대 확인)
MISC_FUEL_RATE = 0.20
WORK_DAYS = 6

# 시중 임대 일위대가 (발주처 승인·견적 기준)
RENTAL_DAY = 600_000  # 기사 포함 일대
TRANSPORT = 200_000  # 반입·선출 1회
APPLY_METHOD = "rental"  # "rental" | "poomsem"

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FILL = PatternFill("solid", fgColor="FFF2CC")
MONEY = "#,##0"

PROJECT = {
    "file": "02 전기설비",
    "row": 177,
    "section": "1.전기공사 (지구내)::1-4.공원 통신설비 공사",
    "code": "InMastDBNonCode",
    "name": "크레인",
    "spec": "10톤",
    "orig_qty": 6,
    "orig_unit": None,
}


def load_operator_wage() -> int:
    with NOIM_CSV.open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r.get("직종명") == "건설기계운전사":
                return int(r["2026.1.1"])
    return 283_297


def calc_poomsem_day(op_wage: int) -> dict:
    machine_hr = MACHINE_PRICE * DEPR_PER_HR
    machine_day = round(machine_hr * HOURS_PER_DAY)
    fuel_main = round(FUEL_L_PER_HR * HOURS_PER_DAY * DIESEL_WON_PER_L)
    fuel_misc = round(fuel_main * MISC_FUEL_RATE)
    labor = op_wage
    total = machine_day + fuel_main + fuel_misc + labor
    return {
        "machine_day": machine_day,
        "fuel_main": fuel_main,
        "fuel_misc": fuel_misc,
        "labor": labor,
        "day_total": total,
        "days_total": total * WORK_DAYS,
    }


def calc_rental() -> dict:
    rental = RENTAL_DAY * WORK_DAYS
    total = rental + TRANSPORT
    return {
        "rental_day": RENTAL_DAY,
        "rental_total": rental,
        "transport": TRANSPORT,
        "total": total,
        "per_day_equiv": round(total / WORK_DAYS),
    }


def write_xlsx(poom: dict, rent: dict, op_wage: int) -> dict:
    apply = rent if APPLY_METHOD == "rental" else {
        "total": poom["days_total"],
        "per_day_equiv": poom["day_total"],
        "rental_total": poom["days_total"],
        "transport": 0,
    }
    wb = Workbook()

    # 1) 품셈 1일
    ws = wb.active
    ws.title = "품셈_1일8시간"
    ws.append(["10톤 기동식 크레인(타이어) — 표준품셈 기계경비 1일(8시간)"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"건설기계운전사 노임(2026. 1. 1.): {op_wage:,}원/일 · 경유 {DIESEL_WON_PER_L:,}원/L"])
    ws.append([])
    ws.append(["구분", "산식", "금액(원)", "비고"])
    rows = [
        ["A. 기계손료", f"({MACHINE_PRICE:,} × {DEPR_PER_HR}) × {HOURS_PER_DAY}h", poom["machine_day"], "시간당 손료"],
        ["B. 주연료비", f"{FUEL_L_PER_HR}L/h × {HOURS_PER_DAY}h × {DIESEL_WON_PER_L:,}원", poom["fuel_main"], "경유"],
        ["C. 잡재료비", f"주연료 × {int(MISC_FUEL_RATE*100)}%", poom["fuel_misc"], "윤활유 등"],
        ["D. 운전원", "건설기계운전사 1일", poom["labor"], "시중노임"],
        ["합계(1일)", "", poom["day_total"], ""],
        [f"합계({WORK_DAYS}일)", f"{poom['day_total']:,} × {WORK_DAYS}", poom["days_total"], "품셈 경로"],
    ]
    for r in rows:
        ws.append(r)
    for row_idx in range(4, ws.max_row + 1):
        if ws.cell(row_idx, 1).value and "합계" in str(ws.cell(row_idx, 1).value):
            for c in range(1, 5):
                ws.cell(row_idx, c).fill = SUB_FILL
                ws.cell(row_idx, c).font = Font(bold=True)
        if isinstance(ws.cell(row_idx, 3).value, int):
            ws.cell(row_idx, 3).number_format = MONEY
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    # 2) 임대 일위대가 6일
    il = wb.create_sheet("일위대가_6일1식")
    il.append(["크레인 양중 — 10톤 타이어식 · 6일 작업 · 단위 1식"])
    il["A1"].font = Font(bold=True, size=13)
    il.append([])
    il.append(["품명", "규격", "단위", "수량", "단가", "금액", "비고"])
    il.append(["크레인 임대료", "10Ton 기사 포함", "일", WORK_DAYS, rent["rental_day"],
               rent["rental_total"], "시중 견적"])
    il.append(["현장 반입·선출", "왕복", "회", 1, rent["transport"], rent["transport"], "탁송"])
    il.append(["유류대", "임대료 포함 시", "L", 0, 0, 0, "별도 정산 시 가산"])
    il.append(["소계", "", "식", 1, rent["total"], rent["total"], "VAT 별도"])
    for row_idx in range(4, il.max_row + 1):
        for c in (5, 6):
            if isinstance(il.cell(row_idx, c).value, (int, float)) and il.cell(row_idx, c).value:
                il.cell(row_idx, c).number_format = MONEY
        if il.cell(row_idx, 1).value == "소계":
            for c in range(1, 8):
                il.cell(row_idx, c).fill = SUB_FILL
                il.cell(row_idx, c).font = Font(bold=True)
    for col in "ABCDEFG":
        il.column_dimensions[col].width = 16

    # 3) 본건
    pj = wb.create_sheet("본건_전기177행")
    pj.append(["02 전기 — 177행 크레인 10톤 반영"])
    pj["A1"].font = Font(bold=True, size=13)
    pj.append([])
    method_label = "시중 임대 1식" if APPLY_METHOD == "rental" else f"품셈 {WORK_DAYS}일"
    pj.append(["항목", "값"])
    for k, v in [
        ("내역서", PROJECT["file"]),
        ("행", PROJECT["row"]),
        ("품명", f"{PROJECT['name']} {PROJECT['spec']} 양중"),
        ("원본", f"수량 {PROJECT['orig_qty']} · 단위 공란 → **식 1** 로 정정"),
        ("적용 경로", method_label),
        ("1식 합계단가", apply["total"]),
        ("일 환산(참고)", apply["per_day_equiv"]),
        ("품셈 대안(6일)", poom["days_total"]),
    ]:
        pj.append([k, v])
    pj.append([])
    pj.append(["체크", "내용"])
    for chk, txt in [
        ("견적서", "크레인 업체 2~3곳 견적 비교 첨부"),
        ("노임표", "품셈 경로 시 건설기계운전사 시중노임 공고 출력"),
        ("현장", "6일 투입 사유·양중 구간 도면·사진"),
        ("단위", "원본 단위 공란 — 내역·일위대가 모두 **식 1** 또는 **일 6** 중 하나로 통일"),
    ]:
        pj.append([chk, txt])
    pj.column_dimensions["A"].width = 14
    pj.column_dimensions["B"].width = 72
    for row_idx in range(4, 12):
        if pj.cell(row_idx, 1).value in ("1식 합계단가", "일 환산(참고)", "품셈 대안(6일)"):
            pj.cell(row_idx, 2).number_format = MONEY

    WORK.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"저장: {OUT}")
    return apply


def patch_electric(apply: dict) -> None:
    if not ELECTRIC.exists():
        print("전기 표준단가산출 없음 — 스킵")
        return
    wb = load_workbook(ELECTRIC)
    ws = wb["통합내역"]
    col = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    target = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, col["행"]).value == PROJECT["row"]:
            target = row_idx
            break
    if not target:
        wb.close()
        return

    tot = apply["total"]
    # 중기·임대 — 경비 100% 관행(전기 크레인트럭 hr 단가와 동일 계열)
    ws.cell(target, col["단위"], "식")
    ws.cell(target, col["수량"], 1)
    ws.cell(target, col["상태"], "매칭")
    ws.cell(target, col["단가출처"], "일위확정")
    ws.cell(target, col["매칭점수"], 1)
    ws.cell(target, col["참조"], "크레인10톤_6일일위대가")
    ws.cell(target, col["매칭품명"], "크레인 임대(10톤·6일·기사)")
    ws.cell(target, col["매칭규격"], f"{WORK_DAYS}일·반입포함")
    ws.cell(target, col["재료단가"], 0)
    ws.cell(target, col["노무단가"], 0)
    ws.cell(target, col["경비단가"], tot)
    ws.cell(target, col["합계단가"], tot)
    ws.cell(target, col["재료금액"], 0)
    ws.cell(target, col["노무금액"], 0)
    ws.cell(target, col["경비금액"], tot)
    ws.cell(target, col["합계금액"], tot)

    # 미산출 시트에서 제거·매칭 시트 갱신은 통합 기준으로 합계만 보정
    if "미산출" in wb.sheetnames:
        uws = wb["미산출"]
        for row_idx in range(uws.max_row, 1, -1):
            if uws.cell(row_idx, 1).value == PROJECT["row"]:
                uws.delete_rows(row_idx)

    # 합계요약 — 통합내역 공종별·★합계 재계산
    if "합계요약" in wb.sheetnames:
        from collections import defaultdict
        sec_tot: dict[str, dict] = defaultdict(lambda: {"n": 0, "ok": 0, "mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0})
        g = {"ok": 0, "all": 0, "mat": 0.0, "lab": 0.0, "exp": 0.0, "sum": 0.0}
        for row_idx in range(2, ws.max_row + 1):
            sec = str(ws.cell(row_idx, col["공종"]).value or "")
            st = ws.cell(row_idx, col["상태"]).value
            sec_tot[sec]["n"] += 1
            g["all"] += 1
            if st in ("매칭", "검토", "원본"):
                sec_tot[sec]["ok"] += 1
                g["ok"] += 1
                ma = float(ws.cell(row_idx, col["재료금액"]).value or 0)
                la = float(ws.cell(row_idx, col["노무금액"]).value or 0)
                ea = float(ws.cell(row_idx, col["경비금액"]).value or 0)
                sa = float(ws.cell(row_idx, col["합계금액"]).value or 0)
                for d, v in [("mat", ma), ("lab", la), ("exp", ea), ("sum", sa)]:
                    sec_tot[sec][d] += v
                    g[d] += v
        sws = wb["합계요약"]
        for row_idx in range(2, sws.max_row + 1):
            label = str(sws.cell(row_idx, 1).value or "")
            if label in sec_tot:
                st = sec_tot[label]
                sws.cell(row_idx, 2, st["ok"])
                sws.cell(row_idx, 3, st["n"])
                sws.cell(row_idx, 4, st["ok"] / st["n"] if st["n"] else 0)
                sws.cell(row_idx, 5, st["mat"])
                sws.cell(row_idx, 6, st["lab"])
                sws.cell(row_idx, 7, st["exp"])
                sws.cell(row_idx, 8, st["sum"])
            elif label.startswith("★"):
                sws.cell(row_idx, 2, g["ok"])
                sws.cell(row_idx, 3, g["all"])
                sws.cell(row_idx, 4, g["ok"] / g["all"] if g["all"] else 0)
                sws.cell(row_idx, 5, g["mat"])
                sws.cell(row_idx, 6, g["lab"])
                sws.cell(row_idx, 7, g["exp"])
                sws.cell(row_idx, 8, g["sum"])

    try:
        wb.save(ELECTRIC)
        print(f"반영: {ELECTRIC.name} 177행 → 식 1 @ {tot:,}원")
    except PermissionError:
        alt = ELECTRIC.with_name(ELECTRIC.stem + "_크레인반영.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")
    wb.close()


def main() -> None:
    op = load_operator_wage()
    poom = calc_poomsem_day(op)
    rent = calc_rental()
    apply = write_xlsx(poom, rent, op)
    print(f"품셈 1일 {poom['day_total']:,}원 · {WORK_DAYS}일 {poom['days_total']:,}원")
    print(f"임대 1식 {rent['total']:,}원 (일 {rent['rental_day']:,} × {WORK_DAYS} + 반입 {rent['transport']:,})")
    patch_electric(apply)
    print("총괄표 재생성…")
    subprocess.run(
        [sys.executable, "-X", "utf8", str(ROOT / "tools" / "build_consolidated_summary.py")],
        cwd=str(ROOT),
        check=True,
    )


if __name__ == "__main__":
    main()
