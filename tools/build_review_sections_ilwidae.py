#!/usr/bin/env python3
"""검토_전체 — 우수·오수·상수·포장·구조물·부대·하수처리장·지하저류조·지원·주차 품셈·일위대가.

입력: 05_내역서/검토_전체.xlsx
출력: 05_내역서/검토_공종별_일위대가산출.xlsx
      05_내역서/내역서작업/검토_공종별_품셈산출.xlsx
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
import apply_standard_prices as asp  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "05_내역서"
WORK = BASE / "내역서작업"
REVIEW_ALL = BASE / "검토_전체.xlsx"
OUT = BASE / "검토_공종별_일위대가산출.xlsx"
POOMSEM_OUT = WORK / "검토_공종별_품셈산출.xlsx"

FILE_MAP = {
    "01 토목": "01 토목",
    "01 토목·조경": "01 토목",  # 분리 이전 라벨 호환
    "04 진입도로": "04 진입도로",
    "05 회전교차로": "05 회전교차로",
    "06 개발행위": "06 개발행위",
}

SECTION_ORDER = [
    ("하수처리장", ("하수처리",)),
    ("지하저류조", ("지하저류",)),
    ("지원시설", ("지원시설",)),
    ("주차장", ("주차",)),
    ("구조물공", ("구조물",)),
    ("부대공", ("부대",)),
    ("포장공", ("포장",)),
    ("상수공", ("상수",)),
    ("우수공", ("우수",)),
    ("오수공", ("오수",)),
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
ROUTE_FILLS = {
    "① 표준단가 확정": PatternFill("solid", fgColor="E2EFDA"),
    "② 단위환산": PatternFill("solid", fgColor="DDEBF7"),
    "③ 품셈·재산출": PatternFill("solid", fgColor="FCE4D6"),
    "④ 1순위+검증": PatternFill("solid", fgColor="FFF2CC"),
}
MONEY = "#,##0"
CONVERT_UNITS = {"개소", "본", "회", "주", "ea", "EA", "Ton", "ton", "대", "개"}


def clean(v) -> str:
    return str(v or "").replace("\r", " ").replace("\n", " ").replace("_x000D_", "").strip()


def core_name(name: str) -> str:
    s = re.sub(r"^[가-힣A-Za-z\d]+\)\.?\s*", "", clean(name))
    return re.sub(r"^[.·\s]+", "", s).strip()


def name_tokens(name: str) -> list[str]:
    cn = core_name(name)
    return [t for t in re.split(r"[\s/()·]+", cn) if len(t) >= 2]


def name_overlap(name: str, matched: str) -> bool:
    if not matched:
        return False
    toks = name_tokens(name)
    if not toks:
        return False
    blob = matched.replace(" ", "")
    hit = sum(1 for t in toks if t in blob or t in matched)
    return hit >= max(1, len(toks) // 2)


def section_group(sec: str) -> str | None:
    s = re.sub(r"[\s\.]+", "", clean(sec))
    for label, kws in SECTION_ORDER:
        if any(kw in s for kw in kws):
            if label == "오수공" and "하수처리" in s:
                continue
            return label
    return None


def classify_route(score: float, unit: str, name: str, matched: str, *, forced: bool) -> str:
    ov = name_overlap(name, matched)
    if forced and ov:
        return "① 표준단가 확정"
    if sc := score:
        if sc >= 0.70 and ov:
            return "① 표준단가 확정"
        if clean(unit) in CONVERT_UNITS and sc >= 0.56 and ov:
            return "② 단위환산"
        if sc >= 0.65 and ov:
            return "④ 1순위+검증"
    return "③ 품셈·재산출"


def extra_search_terms(name: str, spec: str, unit: str) -> list[str]:
    blob = (name + " " + spec).replace(" ", "")
    u = clean(unit).lower()
    terms: list[str] = []
    if "레미콘" in name:
        terms += ["레미콘타설", "레미콘"]
    if "모래" in name and u in ("m3", "㎥"):
        terms += ["모래", "되메우기", "sand"]
    if "쇄석" in name:
        terms += ["쇄석", "쇄석깔기", "혼합골재"]
    if "잡석" in name or "석분" in name:
        terms += ["잡석", "천단잡석", "전석"]
    if "무수축" in name and ("몰탈" in name or "몰탈" in spec):
        terms += ["무수축몰탈", "무수축"]
    if "몰탈" in name and "무수축" not in name:
        terms += ["몰탈", "모르타르"]
    if "PVC" in blob.upper() or "PVC" in name.upper():
        if "절단" in name:
            terms += ["PVC", "절단", "이중벽"]
        elif "접합" in name or "이형" in name:
            terms += ["PVC", "접합", "고무링"]
        else:
            terms += ["PVC", "부설", "이중벽"]
    if "맨홀" in name:
        terms += ["맨홀", "우수"]
    if "게이트밸브" in name:
        terms += ["게이트밸브", "밸브"]
    if "교통" in name and "표지" in name:
        terms += ["교통표지판", "표지판"]
    if "볼라드" in name:
        terms += ["볼라드"]
    if "보도블록" in name:
        terms += ["보도블록"]
    if "경계석" in name:
        terms += ["경계블록", "경계석"]
    if "보강토" in name:
        terms += ["보강토옹벽"]
    if "스페이서" in name or "스페이셔" in name:
        terms += ["스페이서", "간격재", "DOWEL"]
    if "DOWEL" in name.upper():
        terms += ["DOWEL", "다owel"]
    if "문양거푸집" in name:
        terms += ["문양거푸집"]
    elif "거푸집" in name:
        terms += ["합판거푸집", "거푸집"]
    if "부직포" in name:
        terms += ["부직포"]
    if "가설방음" in name or "방음판" in name:
        terms += ["가설방음", "방음벽"]
    if "세륜" in name or "세차" in name:
        terms += ["세륜", "세차"]
    if any(k in name for k in ("사무실", "창고", "시험실", "숙소", "복지")):
        terms += ["가설사무소", "조립식"]
    if "폴리우레아" in name:
        terms += ["폴리우레아", "방수"]
    if "에폭시" in name:
        terms += ["에폭시", "방수"]
    if "점검구" in name:
        terms += ["점검구"]
    if "주철관" in name:
        terms += ["주철관", "부설"]
    if "소프트" in name and "밸브" in name:
        terms += ["공기밸브", "급속"]
    if "준설" in name:
        terms += ["준설", "하수관"]
    if "수팽창" in name:
        terms += ["수팽창"]
    if "측구" in name:
        terms += ["측구", "U형"]
    if "표시못" in name or ("분기" in name and "표" in name):
        terms += ["표시못", "분기"]
    if "집수정" in name:
        terms += ["집수정"]
    if "휀스" in name:
        terms += ["휀스", "울타리"]
    if "오수받이" in name:
        terms += ["오수받이", "물받이"]
    if "관목" in name:
        terms += ["관목", "식재"]
    if "PC" in name and "슬래브" in name:
        terms += ["PC", "슬래브"]
    if "살수" in name:
        terms += ["살수", "살수차"]
    return list(dict.fromkeys(terms))


def load_pool() -> list[dict]:
    market = asp.load_market_csv(asp.MARKET_2026, "표준시장단가2026")
    if not market:
        market = asp.load_prices()
    sijang = asp.load_market_csv(asp.SIJANG_2026, "시장시공가격2026")
    return asp.precompute(
        market + sijang + asp.load_ildae_prices() + asp.load_mulga()
        + asp.load_jojadang() + asp.load_landscape_ildae()
    )


def topn_db(item: dict, prices: list[dict], extra: list[str], n: int = 5) -> list[tuple[float, dict, str]]:
    base_terms = asp.extract_search_terms(item["name"], item["spec"])
    terms = list(dict.fromkeys(extra + base_terms))
    req = asp.required_keywords(item)
    item_kw = asp.kwset(f"{item['name']} {item['spec']}")
    item["_linear"] = asp.is_linear_piece(item)
    cand = [p for p in prices if asp.unit_compatible(item["unit"], p["unit"], item["_linear"])]
    best: dict[int, tuple[float, dict, str]] = {}
    for term in terms:
        kw = item_kw | asp.kwset(term)
        for p in cand:
            s = asp.score_match(item, p, term, kw, req)
            if s <= 0:
                continue
            pid = id(p)
            if pid not in best or s > best[pid][0]:
                best[pid] = (s, p, term)
    return sorted(best.values(), key=lambda x: -x[0])[:n]


def pack_price(p: dict, route: str, basis: str, *, auto: bool) -> dict:
    return {
        "route": route,
        "code": p.get("code", ""),
        "db_name": p.get("name", ""),
        "db_spec": p.get("spec", ""),
        "db_unit": p.get("unit", ""),
        "mat": p.get("mat", 0),
        "lab": p.get("lab", 0),
        "exp": p.get("exp", 0),
        "tot": p.get("total", 0),
        "basis": basis,
        "auto": auto,
        "match_score": None,
    }


def price_item(item: dict, prices: list[dict]) -> dict:
    name, spec = item["name"], item["spec"]
    extra = extra_search_terms(name, spec, item["unit"])
    search_item = {
        "name": name,
        "spec": spec,
        "unit": item["unit"],
        "section": item["section"],
    }
    ranked = topn_db(search_item, prices, extra, 5)

    cur_ok = (
        item["cur_tot"] > 0
        and float(item["score"] or 0) >= 0.68
        and name_overlap(name, item["cur_match"])
    )
    if cur_ok:
        route = classify_route(float(item["score"]), item["unit"], name, item["cur_match"], forced=False)
        auto = route.startswith(("①", "④", "②"))
        return {
            "route": route,
            "code": item["cur_code"],
            "db_name": item["cur_match"],
            "db_spec": item["cur_spec"],
            "db_unit": item["unit"],
            "mat": 0,
            "lab": 0,
            "exp": 0,
            "tot": item["cur_tot"],
            "basis": f"현재 매칭 {item['cur_code']} — 품명 일치·점수 {float(item['score']):.2f}",
            "auto": auto,
            "match_score": float(item["score"]),
        }

    if ranked:
        sc, p, term = ranked[0]
        alts = " · ".join(f"{x[1]['name'][:14]}({x[0]:.2f})" for x in ranked[1:3])
        route = classify_route(sc, item["unit"], name, p["name"], forced=False)
        bad = item["cur_match"] and not name_overlap(name, item["cur_match"])
        basis = f"DB {p.get('code','')} ({sc:.2f}, {term})"
        if bad:
            basis = f"오매칭 {item['cur_code']}({item['cur_match'][:18]}) → " + basis
        if alts:
            basis += f" / 대안: {alts[:80]}"
        ov = name_overlap(name, p["name"])
        auto = p["total"] > 0 and (
            route.startswith("①")
            or (route.startswith("④") and sc >= 0.55)
            or (sc >= 0.58 and ov)
            or (sc >= 0.52 and ov and bad)
        )
        return pack_price(p, route, basis, auto=auto) | {"match_score": sc}

    if item["cur_tot"] > 0:
        return {
            "route": "③ 품셈·재산출",
            "code": item["cur_code"],
            "db_name": item["cur_match"],
            "db_spec": item["cur_spec"],
            "db_unit": item["unit"],
            "mat": 0,
            "lab": 0,
            "exp": 0,
            "tot": item["cur_tot"],
            "basis": "DB 재검색 실패 — 현재 후보 유지·수동 확인",
            "auto": False,
            "match_score": float(item["score"] or 0),
        }

    return {
        "route": "③ 품셈·재산출",
        "code": "",
        "db_name": "",
        "db_spec": "",
        "db_unit": item["unit"],
        "mat": 0,
        "lab": 0,
        "exp": 0,
        "tot": 0,
        "basis": "단가 없음 — 수동 입력",
        "auto": False,
        "match_score": 0,
    }


def load_review_items() -> list[dict]:
    path = REVIEW_ALL if REVIEW_ALL.exists() else WORK / "검토_전체.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["통합"]
    rows: list[dict] = []
    for r in range(4, ws.max_row + 1):
        section = clean(ws.cell(r, 3).value)
        grp = section_group(section)
        if not grp:
            continue
        file_label = clean(ws.cell(r, 1).value)
        rows.append({
            "file_raw": file_label,
            "file": FILE_MAP.get(file_label, file_label),
            "row": ws.cell(r, 2).value,
            "section": section,
            "section_group": grp,
            "name": clean(ws.cell(r, 4).value),
            "spec": clean(ws.cell(r, 5).value),
            "unit": clean(ws.cell(r, 6).value),
            "qty": float(ws.cell(r, 7).value or 0),
            "score": ws.cell(r, 8).value,
            "cur_code": clean(ws.cell(r, 10).value),
            "cur_match": clean(ws.cell(r, 11).value),
            "cur_spec": clean(ws.cell(r, 12).value),
            "cur_tot": float(ws.cell(r, 13).value or 0),
            "cur_amt": float(ws.cell(r, 14).value or 0),
        })
    wb.close()
    order = {lbl: i for i, (lbl, _) in enumerate(SECTION_ORDER)}
    rows.sort(key=lambda x: (order.get(x["section_group"], 99), x["file"], x["row"]))
    return rows


def write_ilwidae_workbook(items: list[dict], priced: list[dict]) -> None:
    wb = Workbook()
    by_grp: dict[str, int] = defaultdict(int)
    by_route: dict[str, int] = defaultdict(int)
    for item, p in zip(items, priced):
        by_grp[item["section_group"]] += 1
        by_route[p["route"]] += 1

    sm = wb.active
    sm.title = "경로별요약"
    sm.append(["검토_전체 — 공종별 품셈·일위대가 (토공 제외)"])
    sm.append([])
    sm.append(["공종", "건수"])
    for label, _ in SECTION_ORDER:
        if by_grp.get(label):
            sm.append([label, by_grp[label]])
    sm.append(["합계", len(items)])
    sm.append([])
    sm.append(["경로", "건수"])
    for route in ("① 표준단가 확정", "② 단위환산", "③ 품셈·재산출", "④ 1순위+검증"):
        if by_route.get(route):
            sm.append([route, by_route[route]])
    sm.append(["자동 확정", sum(1 for p in priced if p["auto"]), ""])
    sm["A1"].font = Font(bold=True, size=13)

    ws = wb.create_sheet("일위대가산출")
    headers = [
        "공종", "경로", "파일", "행", "내역공종", "품명", "규격", "단위", "수량", "매칭점수",
        "현재_코드", "현재_매칭품명", "현재_합계단가",
        "DB_코드", "DB_품명", "DB_규격", "DB_단위",
        "재료단가", "노무단가", "경비단가", "합계단가", "합계금액",
        "확정단가(입력)", "확정금액", "표준품셈·산출근거",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for item, p in zip(items, priced):
        amt = round(p["tot"] * item["qty"]) if p["tot"] and item["qty"] else None
        row = [
            item["section_group"], p["route"], item["file"], item["row"], item["section"],
            item["name"], item["spec"], item["unit"], item["qty"], item["score"],
            item["cur_code"], item["cur_match"], item["cur_tot"],
            p["code"], p["db_name"], p["db_spec"], p["db_unit"],
            p["mat"] or None, p["lab"] or None, p["exp"] or None, p["tot"] or None, amt,
            p["tot"] if p["auto"] else "",
            amt if p["auto"] else "",
            p["basis"],
        ]
        ws.append(row)
        ridx = ws.max_row
        fill = ROUTE_FILLS.get(p["route"], ROUTE_FILLS["③ 품셈·재산출"])
        for c in range(1, len(headers) + 1):
            ws.cell(ridx, c).fill = fill
            ws.cell(ridx, c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(18, 25):
            if isinstance(ws.cell(ridx, c).value, (int, float)):
                ws.cell(ridx, c).number_format = MONEY

    ws.freeze_panes = "A2"
    widths = [10, 14, 11, 5, 14, 24, 20, 5, 7, 7, 12, 18, 10,
              12, 20, 18, 5, 9, 9, 9, 10, 12, 11, 12, 52]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    info = wb.create_sheet("안내")
    for line in [
        ["검토 — 우수·오수·상수·포장·구조물·부대·하수처리장·지하저류조·지원·주차"],
        [],
        ["반영", "python -X utf8 tools/apply_confirmed_prices.py"],
        ["품셈", "내역서작업/검토_공종별_품셈산출.xlsx"],
        ["주의", "개소·본·ea 환산·PVC 규격·맨홀 호칭은 현장 도면 확인"],
    ]:
        info.append(line)
    info["A1"].font = Font(bold=True, size=14)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print(f"저장: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_업데이트.xlsx")
        wb.save(alt)
        print(f"원본 사용 중 → {alt}")


def write_poomsem_workbook(items: list[dict], priced: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "공종별요약"
    ws.append(["검토_공종별 — 품셈·일위대가"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["공종", "건수", "자동확정", "재산출", "대표단가(중앙)"])

    grouped: dict[str, list[tuple[dict, dict]]] = defaultdict(list)
    for item, p in zip(items, priced):
        grouped[item["section_group"]].append((item, p))

    for label, _ in SECTION_ORDER:
        grp = grouped.get(label, [])
        if not grp:
            continue
        auto = sum(1 for _, p in grp if p["auto"])
        repro = sum(1 for _, p in grp if p["route"].startswith("③"))
        tots = sorted(p["tot"] for _, p in grp if p["tot"])
        mid = tots[len(tots) // 2] if tots else 0
        ws.append([label, len(grp), auto, repro, mid])

        sh = wb.create_sheet(label[:8])
        sh.append([f"{label} — 일위대가·품셈"])
        sh["A1"].font = Font(bold=True, size=12)
        sh.append([])
        hdr = ["파일", "행", "품명", "규격", "단위", "수량", "경로",
               "현재단가", "확정단가", "합계금액", "산출근거"]
        sh.append(hdr)
        for c in range(1, len(hdr) + 1):
            sh.cell(3, c).font = Font(bold=True)
            sh.cell(3, c).fill = HEADER_FILL
        for item, p in grp:
            amt = round(p["tot"] * item["qty"]) if p["tot"] else None
            sh.append([
                item["file"], item["row"], item["name"], item["spec"],
                item["unit"], item["qty"], p["route"],
                item["cur_tot"], p["tot"] if p["auto"] else "",
                amt if p["auto"] else "", p["basis"],
            ])
        for r in range(4, sh.max_row + 1):
            for c in (8, 9, 10):
                if isinstance(sh.cell(r, c).value, (int, float)):
                    sh.cell(r, c).number_format = MONEY
        sh.column_dimensions["A"].width = 11
        sh.column_dimensions["C"].width = 26
        sh.column_dimensions["K"].width = 58

    WORK.mkdir(parents=True, exist_ok=True)
    wb.save(POOMSEM_OUT)
    print(f"저장: {POOMSEM_OUT}")


def main() -> None:
    if not REVIEW_ALL.exists() and not (WORK / "검토_전체.xlsx").exists():
        print("검토_전체.xlsx 없음")
        sys.exit(1)

    prices = load_pool()
    print(f"단가 풀 {len(prices):,}건")

    items = load_review_items()
    print(f"공종별 검토 {len(items)}건")
    priced = [price_item(it, prices) for it in items]

    grp_cnt: dict[str, int] = defaultdict(int)
    for it in items:
        grp_cnt[it["section_group"]] += 1
    for label, _ in SECTION_ORDER:
        if grp_cnt.get(label):
            print(f"  {label}: {grp_cnt[label]}건")

    for route in sorted({p["route"] for p in priced}):
        n = sum(1 for p in priced if p["route"] == route)
        auto = sum(1 for p in priced if p["route"] == route and p["auto"])
        print(f"  {route}: {n}건 (자동 {auto})")

    write_ilwidae_workbook(items, priced)
    write_poomsem_workbook(items, priced)


if __name__ == "__main__":
    main()
